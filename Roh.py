import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import feedparser
import concurrent.futures
import hashlib
import os
import pyotp
import re
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

from SmartApi import SmartConnect
from datetime import datetime, timedelta, timezone, date
from email.utils import parsedate_to_datetime

try:
    import yfinance as yf
    _YF_AVAILABLE = True
except ImportError:
    _YF_AVAILABLE = False

# ── Server-side screen width detection ──────────────────────────
try:
    from streamlit_js_eval import streamlit_js_eval
    _screen_w = streamlit_js_eval(js_expressions="window.innerWidth", key="screen_w_detect", want_output=True)
    _IS_MOBILE = (isinstance(_screen_w, (int, float)) and _screen_w < 640)
except Exception:
    _IS_MOBILE = False

# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="Northeast Portfolio Tracker",
    layout="wide",
    page_icon="📈",
    initial_sidebar_state="expanded"
)

# =========================================================
# CUSTOM CSS
# =========================================================

st.markdown("""
<script>
(function(){
    if (!document.querySelector('meta[name="viewport"]')) {
        var m = document.createElement('meta');
        m.name = 'viewport';
        m.content = 'width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no';
        document.head.appendChild(m);
    }

    /* ── Resize Plotly charts to fit phone screen ── */
    function resizePlotlyMobile() {
        if (window.innerWidth > 600) return;           // desktop: do nothing
        var targetH = Math.min(Math.round(window.innerWidth * 0.58), 250);
        document.querySelectorAll('.js-plotly-plot').forEach(function(el) {
            try {
                // Use Plotly's own API so axes/labels re-render correctly
                if (window.Plotly) {
                    Plotly.relayout(el, { height: targetH });
                }
            } catch(e) {}
            // Also clamp the wrapper directly
            el.style.maxHeight = targetH + 'px';
            el.style.overflow  = 'hidden';
        });
    }

    // Fire whenever Streamlit adds new chart nodes
    var _obs = new MutationObserver(function(mutations) {
        mutations.forEach(function(m) {
            m.addedNodes.forEach(function(node) {
                if (node.nodeType === 1) {
                    // small delay so Plotly finishes rendering first
                    setTimeout(resizePlotlyMobile, 120);
                }
            });
        });
    });

    function _init() {
        resizePlotlyMobile();
        _obs.observe(document.body, { childList: true, subtree: true });
    }

    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', _init);
    } else {
        _init();
    }
    window.addEventListener('resize', resizePlotlyMobile);
})();
</script>
<style>
/* ── Google Font ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

/* =====================================================
   PROFESSIONAL DARK DASHBOARD — Northeast Portfolio
   ===================================================== */

/* ── CSS Variables ── */
:root {
    --bg-primary:   #0b0d17;
    --bg-card:      #131628;
    --bg-card-alt:  #181b2e;
    --bg-hover:     #1e2238;
    --border:       #252849;
    --border-light: #2e3355;
    --accent-blue:  #4f7ef8;
    --accent-green: #22d67b;
    --accent-red:   #f85454;
    --accent-gold:  #f5c842;
    --text-primary: #f0f2ff;
    --text-muted:   #7a7fa8;
    --text-dim:     #454870;
    --font: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    --radius-sm: 8px;
    --radius-md: 12px;
    --radius-lg: 16px;
}

/* ── Global App Styling ── */
html, body, [class*="css"], .stApp {
    font-family: var(--font) !important;
    background-color: var(--bg-primary) !important;
    color: var(--text-primary) !important;
    background-image:
        radial-gradient(ellipse 900px 500px at 0% 0%, rgba(79,126,248,0.06) 0%, transparent 60%),
        radial-gradient(ellipse 600px 400px at 100% 80%, rgba(167,139,250,0.05) 0%, transparent 60%),
        radial-gradient(ellipse 500px 300px at 70% 5%, rgba(34,199,214,0.03) 0%, transparent 60%) !important;
}

/* ── Main content padding ── */
.block-container {
    padding: 2.5rem 2rem 2rem 2rem !important;
    max-width: 100% !important;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: var(--bg-card) !important;
    border-right: 1px solid var(--border) !important;
    min-width: 240px !important;
}
[data-testid="stSidebar"] .block-container {
    padding: 1.5rem 1rem !important;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stMarkdown p {
    color: var(--text-muted) !important;
    font-size: 12px !important;
}

/* ── Sidebar nav buttons ── */
[data-testid="stSidebar"] [data-testid="stButton"] > button[kind="secondary"] {
    background: transparent !important;
    color: var(--text-muted) !important;
    border: 1px solid transparent !important;
    border-radius: var(--radius-sm) !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    padding: 7px 12px !important;
    text-align: left !important;
    justify-content: flex-start !important;
    transition: all 0.15s ease !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] > button[kind="secondary"]:hover {
    background: var(--bg-hover) !important;
    color: var(--text-primary) !important;
    border-color: var(--border) !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] > button[kind="primary"] {
    background: var(--accent-blue) !important;
    color: #fff !important;
    border: none !important;
    border-radius: var(--radius-sm) !important;
    font-size: 13px !important;
    font-weight: 700 !important;
    padding: 7px 12px !important;
    text-align: left !important;
    justify-content: flex-start !important;
}

/* ══════════════════════════════════════════════
   TABS — Desktop (> 600px): pill-bar style
   ══════════════════════════════════════════════ */
@media (min-width: 601px) {
    [data-testid="stTabs"] [role="tablist"] {
        gap: 4px;
        background: var(--bg-card) !important;
        border-radius: var(--radius-md) !important;
        padding: 6px !important;
        border: 1px solid var(--border) !important;
        overflow-x: auto;
        flex-wrap: nowrap;
        -webkit-overflow-scrolling: touch;
    }
    [data-testid="stTabs"] [role="tab"] {
        background: transparent !important;
        color: var(--text-muted) !important;
        border-radius: var(--radius-sm) !important;
        border: none !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        padding: 7px 14px !important;
        transition: all 0.2s ease !important;
        white-space: nowrap;
        letter-spacing: 0.2px;
    }
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        background: var(--accent-blue) !important;
        color: #fff !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 8px rgba(79,126,248,0.35) !important;
    }
    [data-testid="stTabs"] [role="tab"]:hover:not([aria-selected="true"]) {
        background: var(--bg-hover) !important;
        color: var(--text-primary) !important;
    }
}

/* ══════════════════════════════════════════════
   TABS — Mobile (≤ 600px): bottom nav dock style
   ══════════════════════════════════════════════ */
@media (max-width: 600px) {
    /* Sticky bottom navigation strip */
    [data-testid="stTabs"] [role="tablist"] {
        position: fixed !important;
        bottom: 0 !important;
        left: 0 !important;
        right: 0 !important;
        z-index: 9999 !important;
        display: flex !important;
        flex-direction: row !important;
        gap: 0 !important;
        background: rgba(13,15,30,0.97) !important;
        backdrop-filter: blur(20px) !important;
        -webkit-backdrop-filter: blur(20px) !important;
        border-top: 1px solid #252849 !important;
        border-radius: 0 !important;
        padding: 6px 4px 10px 4px !important;
        overflow-x: auto !important;
        overflow-y: hidden !important;
        -webkit-overflow-scrolling: touch !important;
        flex-wrap: nowrap !important;
        scrollbar-width: none !important;
        justify-content: flex-start !important;
        box-shadow: 0 -4px 24px rgba(0,0,0,0.5) !important;
    }
    [data-testid="stTabs"] [role="tablist"]::-webkit-scrollbar {
        display: none !important;
    }

    /* Each tab button */
    [data-testid="stTabs"] [role="tab"] {
        flex: 0 0 auto !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        min-width: 64px !important;
        padding: 5px 10px 3px 10px !important;
        background: transparent !important;
        color: #555577 !important;
        border: none !important;
        border-radius: 12px !important;
        font-size: 10px !important;
        font-weight: 600 !important;
        white-space: nowrap !important;
        transition: all 0.18s ease !important;
        position: relative !important;
        letter-spacing: 0.3px !important;
        line-height: 1.4 !important;
    }

    /* Active tab: glowing accent pill */
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        background: rgba(79,126,248,0.15) !important;
        color: #4f7ef8 !important;
        font-weight: 800 !important;
        border-radius: 12px !important;
        box-shadow: 0 0 12px rgba(79,126,248,0.2) !important;
    }

    /* Active indicator dot */
    [data-testid="stTabs"] [role="tab"][aria-selected="true"]::after {
        content: "" !important;
        display: block !important;
        width: 4px !important;
        height: 4px !important;
        background: #4f7ef8 !important;
        border-radius: 50% !important;
        margin: 2px auto 0 auto !important;
    }

    /* Hover state */
    [data-testid="stTabs"] [role="tab"]:hover:not([aria-selected="true"]) {
        background: rgba(255,255,255,0.05) !important;
        color: #a0a4c8 !important;
    }

    /* Push page content up so bottom nav doesn't cover it */
    [data-testid="stTabs"] [data-baseweb="tab-panel"],
    [data-testid="stTabs"] > div:last-child {
        padding-bottom: 80px !important;
    }
    .block-container {
        padding-bottom: 90px !important;
    }
}

/* ── Metrics ── */
[data-testid="stMetric"] {
    background: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius-md) !important;
    padding: 16px !important;
}
[data-testid="stMetricValue"] {
    font-size: 22px !important;
    font-weight: 700 !important;
    color: var(--text-primary) !important;
}
[data-testid="stMetricLabel"] {
    font-size: 11px !important;
    color: var(--text-muted) !important;
    text-transform: uppercase;
    letter-spacing: 0.8px;
}
[data-testid="stMetricDelta"] {
    font-size: 13px !important;
    font-weight: 600 !important;
}

/* ── Buttons ── */
[data-testid="stButton"] > button {
    background: var(--accent-blue) !important;
    color: #fff !important;
    border: none !important;
    border-radius: var(--radius-sm) !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    padding: 8px 18px !important;
    transition: opacity 0.2s ease !important;
}
[data-testid="stButton"] > button:hover {
    opacity: 0.85 !important;
}

/* ── Inputs ── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stSelectbox"] > div > div {
    background: var(--bg-card-alt) !important;
    border: 1px solid var(--border-light) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--text-primary) !important;
    font-size: 13px !important;
}
[data-testid="stSelectbox"],
[data-testid="stTextInput"],
[data-testid="stNumberInput"] {
    width: 100% !important;
}

/* ── DataFrames / Tables ── */
[data-testid="stDataFrame"] > div,
.stDataFrame > div,
div[class*="dataframe"] {
    overflow-x: auto !important;
    -webkit-overflow-scrolling: touch;
    border-radius: var(--radius-md) !important;
}
[data-testid="stDataFrame"] table {
    font-size: 12px !important;
    font-family: var(--font) !important;
}
[data-testid="stDataFrame"] th {
    background: var(--bg-card-alt) !important;
    color: var(--text-muted) !important;
    font-size: 11px !important;
    text-transform: uppercase;
    letter-spacing: 0.7px;
    padding: 8px 12px !important;
    border-bottom: 1px solid var(--border) !important;
}
[data-testid="stDataFrame"] td {
    padding: 7px 12px !important;
    border-bottom: 1px solid var(--border) !important;
    white-space: nowrap;
}

/* ── Expanders ── */
[data-testid="stExpander"] {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius-md) !important;
    background: var(--bg-card) !important;
    margin-bottom: 8px !important;
}
[data-testid="stExpander"] summary {
    font-size: 13px !important;
    font-weight: 600 !important;
    color: var(--text-primary) !important;
    padding: 10px 14px !important;
}

/* ── Portfolio tab: mobile-only collapsible sections ── */
/* On desktop (>600px): force .portfolio-mobile-expander open, hide toggle arrow */
@media (min-width: 601px) {
    .portfolio-mobile-expander > details {
        open: true !important;
    }
    .portfolio-mobile-expander details {
        pointer-events: none !important;
    }
    .portfolio-mobile-expander details > summary {
        display: none !important;
    }
    .portfolio-mobile-expander details > div {
        display: block !important;
    }
    /* Remove the expander border/bg on desktop so it looks like normal content */
    .portfolio-mobile-expander [data-testid="stExpander"] {
        border: none !important;
        background: transparent !important;
        padding: 0 !important;
        margin: 0 !important;
    }
    .portfolio-mobile-expander [data-testid="stExpander"] summary {
        display: none !important;
    }
}
/* On mobile (<=600px): show as normal collapsible expanders */
@media (max-width: 600px) {
    .portfolio-mobile-expander [data-testid="stExpander"] {
        border: 1px solid var(--border) !important;
        border-radius: var(--radius-md) !important;
        background: var(--bg-card) !important;
        margin-bottom: 10px !important;
    }
    .portfolio-mobile-expander [data-testid="stExpander"] summary {
        display: flex !important;
        font-size: 13px !important;
        font-weight: 700 !important;
        color: var(--text-primary) !important;
        padding: 12px 14px !important;
        pointer-events: auto !important;
        cursor: pointer !important;
    }
}

/* ── Collapsible desktop sections (Edit Holdings + Sell History) ── */
.collapsible-section {
    margin-bottom: 0px !important;
}
.collapsible-section [data-testid="stExpander"] {
    background: #0a0c1e !important;
    border: 1px solid #2a2e52 !important;
    border-radius: 14px !important;
    margin-bottom: 12px !important;
    box-shadow: 0 4px 32px rgba(79,126,248,0.08), 0 1px 3px rgba(0,0,0,0.4) !important;
    overflow: hidden !important;
}
/* Header row — identical gradient to holdings table header */
.collapsible-section [data-testid="stExpander"] summary {
    background: linear-gradient(90deg,#181b2e 0%,#1a1d33 100%) !important;
    font-family: 'Inter', -apple-system, sans-serif !important;
    font-size: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.9px !important;
    text-transform: uppercase !important;
    color: #6b7299 !important;
    padding: 12px 16px !important;
    border-bottom: 2px solid #2a2e52 !important;
    border-radius: 0 !important;
    min-height: 44px !important;
    display: flex !important;
    align-items: center !important;
}
.collapsible-section [data-testid="stExpander"] summary:hover {
    background: linear-gradient(90deg,#1e2238 0%,#1e2140 100%) !important;
    color: #c8cce8 !important;
    cursor: pointer !important;
}
.collapsible-section [data-testid="stExpander"] summary svg {
    color: #454870 !important;
    width: 16px !important;
    height: 16px !important;
}
/* Inner content area */
.collapsible-section [data-testid="stExpander"] > div[data-testid="stExpanderDetails"] {
    background: #0a0c1e !important;
    padding: 0 !important;
}
/* Style the data_editor inside to look less widget-like */
.collapsible-section [data-testid="stDataFrame"],
.collapsible-section [data-testid="stDataEditor"] {
    border: none !important;
    border-radius: 0 !important;
}

/* ── Info / Warning / Error boxes ── */
[data-testid="stInfo"],
.stInfo {
    background: rgba(79, 126, 248, 0.08) !important;
    border: 1px solid rgba(79, 126, 248, 0.25) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--text-primary) !important;
}

/* ── Dividers ── */
hr {
    border-color: var(--border) !important;
    margin: 20px 0 !important;
}

/* ── Plotly charts ── */
.js-plotly-plot, .plotly, .plot-container {
    width: 100% !important;
}
.js-plotly-plot .plotly .main-svg {
    max-width: 100% !important;
}

/* ── Download buttons ── */
[data-testid="stDownloadButton"] > button {
    background: var(--bg-card-alt) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border-light) !important;
    border-radius: var(--radius-sm) !important;
    font-size: 13px !important;
}
[data-testid="stDownloadButton"] > button:hover {
    border-color: var(--accent-blue) !important;
    color: var(--accent-blue) !important;
}

/* ── Overflow-safe HTML table wrapper ── */
.overflow-table-wrap {
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    border-radius: var(--radius-md);
}

/* ── Summary metric bar ── */
.summary-bar {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
    gap: 10px;
    padding: 0;
    margin-bottom: 22px;
}
.summary-item {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    padding: 14px 16px;
    text-align: center;
    transition: border-color 0.2s ease;
}
.summary-item:hover {
    border-color: var(--border-light);
}
.summary-label {
    font-size: 10px;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-bottom: 6px;
    font-weight: 500;
}
.summary-value {
    font-size: 13px;
    font-weight: 700;
    color: var(--text-primary);
    line-height: 1.2;
    white-space: nowrap;
}
.summary-value.positive { color: var(--accent-green); }
.summary-value.negative { color: var(--accent-red); }
.summary-divider { display: none; }
.pnl-positive { color: var(--accent-green) !important; font-weight: 600; }
.pnl-negative { color: var(--accent-red) !important; font-weight: 600; }

/* ── Responsive columns ── */
[data-testid="column"] { min-width: 0 !important; }

/* ── Headings ── */
h1, h2, h3 { font-family: var(--font) !important; }
h1 { font-size: 22px !important; font-weight: 700 !important; color: var(--text-primary) !important; }
h2 { font-size: 18px !important; font-weight: 600 !important; color: var(--text-primary) !important; }
h3 { font-size: 15px !important; font-weight: 600 !important; color: var(--text-primary) !important; }

/* ── Caption text ── */
[data-testid="stCaptionContainer"] {
    color: var(--text-muted) !important;
    font-size: 11px !important;
}

/* ============================================================
   TABLET  (≤ 900px)
   ============================================================ */
@media (max-width: 900px) {
    .summary-bar { grid-template-columns: repeat(3, 1fr); gap: 8px; }
    .summary-value { font-size: 12px; }
    .block-container { padding: 2.5rem 1rem 1.5rem 1rem !important; }
    [data-testid="stHorizontalBlock"] { flex-wrap: wrap !important; }
    [data-testid="stHorizontalBlock"] > [data-testid="column"] {
        flex: 1 1 45% !important; min-width: 140px !important;
    }
}

/* ============================================================
   MOBILE  (≤ 600px)
   ============================================================ */
@media (max-width: 600px) {
    .summary-bar { grid-template-columns: 1fr 1fr; gap: 6px; }
    .summary-value { font-size: 11px; }
    .summary-label { font-size: 9px; }
    .block-container { padding: 2.5rem 0.5rem 1rem 0.5rem !important; }

    /* Stack ALL side-by-side columns vertically */
    [data-testid="stHorizontalBlock"] { flex-wrap: wrap !important; gap: 6px !important; }
    [data-testid="stHorizontalBlock"] > [data-testid="column"] {
        flex: 1 1 100% !important; min-width: 100% !important;
    }

    [data-testid="stSidebar"][aria-expanded="true"] {
        min-width: 80vw !important; max-width: 90vw !important;
    }
    [data-testid="stDataFrame"] table { font-size: 11px !important; }
    [data-testid="stDataFrame"] th,
    [data-testid="stDataFrame"] td { padding: 4px 6px !important; }
    h1 { font-size: 18px !important; }
    h2 { font-size: 15px !important; }
    h3 { font-size: 13px !important; }

    /* ── Plotly charts: fit in one screen, no vertical overflow ── */
    /* The chart wrapper Streamlit injects */
    [data-testid="stPlotlyChart"] {
        height: 55vw !important;
        max-height: 260px !important;
        min-height: 180px !important;
        overflow: hidden !important;
    }
    /* The inner div Streamlit adds */
    [data-testid="stPlotlyChart"] > div {
        height: 100% !important;
    }
    /* The actual Plotly SVG container */
    [data-testid="stPlotlyChart"] .js-plotly-plot,
    [data-testid="stPlotlyChart"] .plot-container,
    [data-testid="stPlotlyChart"] .plotly {
        height: 100% !important;
        max-height: 260px !important;
    }
    [data-testid="stPlotlyChart"] svg.main-svg {
        height: 100% !important;
        max-height: 260px !important;
    }
    /* Shrink tick labels so bars don't get clipped */
    [data-testid="stPlotlyChart"] .xtick text,
    [data-testid="stPlotlyChart"] .ytick text { font-size: 9px !important; }
    [data-testid="stPlotlyChart"] .gtitle       { font-size: 11px !important; }
    [data-testid="stPlotlyChart"] .legendtext   { font-size: 9px !important; }

    /* Reduce dataframe table max-height so it doesn't push charts off screen */
    [data-testid="stDataFrame"] > div {
        max-height: 220px !important;
        overflow-y: auto !important;
    }
}

/* ── Mobile header classes ── */
.hdr-desktop { display: flex; }
.hdr-mobile  { display: none; }

@media (max-width: 600px) {
    .hdr-desktop { display: none !important; }
    .hdr-mobile  { display: block !important; }
}

/* ══════════════════════════════════════════════════════════
   CORPORATE ACTIONS — Mobile-only premium design (≤600px)
   ══════════════════════════════════════════════════════════ */
@media (max-width: 600px) {

  /* ── Summary bar: 2×2 KPI tiles ── */
  .ca-stat-grid {
    display: grid !important;
    grid-template-columns: 1fr 1fr !important;
    gap: 10px !important;
    margin-bottom: 18px !important;
  }
  .ca-stat-pill {
    position: relative !important;
    overflow: hidden !important;
    background: linear-gradient(145deg, #141729 0%, #0e1020 100%) !important;
    border: 1px solid #1e2340 !important;
    border-radius: 18px !important;
    padding: 16px 14px 14px 14px !important;
    text-align: center !important;
    box-shadow: 0 4px 16px rgba(0,0,0,0.35), inset 0 1px 0 rgba(255,255,255,0.04) !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease !important;
  }
  /* Glowing top accent bar (colour set inline) */
  .ca-stat-pill::before {
    content: "" !important;
    display: block !important;
    height: 3px !important;
    border-radius: 18px 18px 0 0 !important;
    position: absolute !important;
    top: 0; left: 0; right: 0 !important;
    opacity: 0.9 !important;
  }
  /* Subtle radial glow in background */
  .ca-stat-pill::after {
    content: "" !important;
    position: absolute !important;
    inset: 0 !important;
    background: radial-gradient(ellipse at 50% 0%, rgba(255,255,255,0.04) 0%, transparent 70%) !important;
    pointer-events: none !important;
  }
  .ca-stat-count {
    font-size: 30px !important;
    font-weight: 900 !important;
    line-height: 1.05 !important;
    letter-spacing: -1.5px !important;
    position: relative !important;
    z-index: 1 !important;
  }
  .ca-stat-label {
    font-size: 9px !important;
    color: #6b7099 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.9px !important;
    font-weight: 700 !important;
    margin-top: 5px !important;
    position: relative !important;
    z-index: 1 !important;
  }

  /* ── Urgent ex-date alert banner ── */
  .ca-alert-mob {
    background: linear-gradient(135deg, #1c0a0a 0%, #1a0d10 100%) !important;
    border: 1px solid rgba(248,84,84,0.3) !important;
    border-left: 4px solid #f85454 !important;
    border-radius: 16px !important;
    padding: 14px 16px !important;
    margin-bottom: 18px !important;
    box-shadow: 0 4px 20px rgba(248,84,84,0.1), 0 2px 8px rgba(0,0,0,0.3) !important;
  }
  .ca-alert-mob .alert-title {
    color: #ff6b6b !important;
    font-size: 12px !important;
    font-weight: 800 !important;
    margin-bottom: 10px !important;
    display: flex !important;
    align-items: center !important;
    gap: 7px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.6px !important;
  }
  .ca-alert-mob .alert-row {
    background: rgba(248,84,84,0.06) !important;
    border: 1px solid rgba(248,84,84,0.12) !important;
    border-radius: 12px !important;
    padding: 10px 12px !important;
    margin-bottom: 8px !important;
    display: flex !important;
    justify-content: space-between !important;
    align-items: center !important;
  }
  .ca-alert-mob .alert-row:last-child { margin-bottom: 0 !important; }
  .ca-alert-mob .alert-ticker {
    color: #f0f2ff !important;
    font-size: 14px !important;
    font-weight: 900 !important;
    letter-spacing: 0.2px !important;
  }
  .ca-alert-mob .alert-event {
    color: #8888aa !important;
    font-size: 11px !important;
    margin-top: 2px !important;
    font-weight: 500 !important;
  }
  .ca-alert-mob .alert-date {
    color: #f85454 !important;
    font-size: 13px !important;
    font-weight: 800 !important;
    text-align: right !important;
  }
  .ca-alert-mob .alert-daysaway {
    color: #ff9999 !important;
    font-size: 10px !important;
    text-align: right !important;
    margin-top: 2px !important;
    font-weight: 600 !important;
    opacity: 0.8 !important;
  }

  /* ── Section divider header ── */
  .ca-section-hdr {
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    font-size: 10px !important;
    font-weight: 800 !important;
    text-transform: uppercase !important;
    letter-spacing: 1.2px !important;
    color: #7a7fa8 !important;
    padding: 12px 0 8px 0 !important;
    border-bottom: 1px solid #1a1e38 !important;
    margin-bottom: 14px !important;
  }
  .ca-section-hdr::after {
    content: "" !important;
    flex: 1 !important;
    height: 1px !important;
    background: linear-gradient(90deg, #1e2238, transparent) !important;
    margin-left: 6px !important;
  }

  /* ══ Individual CA card — premium glass-morphism card ══ */
  .ca-card-mob {
    background: linear-gradient(145deg, #13162a 0%, #0d1020 100%) !important;
    border: 1px solid #1c2040 !important;
    border-radius: 20px !important;
    margin-bottom: 12px !important;
    overflow: hidden !important;
    box-shadow:
      0 4px 24px rgba(0,0,0,0.4),
      0 1px 0 rgba(255,255,255,0.035) inset !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease !important;
    position: relative !important;
  }

  /* Left accent bar (colour comes from border-top inline, repurposed as left bar below) */
  .ca-card-mob::before {
    content: "" !important;
    position: absolute !important;
    top: 0; left: 0 !important;
    width: 100% !important;
    height: 2px !important;
    opacity: 0.85 !important;
  }

  /* ── Card header: ticker + badge ── */
  .ca-card-mob .ca-card-header {
    display: flex !important;
    align-items: center !important;
    justify-content: space-between !important;
    padding: 14px 16px 12px 16px !important;
    border-bottom: 1px solid rgba(255,255,255,0.05) !important;
    background: rgba(255,255,255,0.025) !important;
  }
  .ca-card-mob .ca-ticker-block {
    display: flex !important;
    flex-direction: column !important;
    gap: 3px !important;
  }
  .ca-card-mob .ca-ticker {
    font-size: 17px !important;
    font-weight: 900 !important;
    color: #eef0ff !important;
    letter-spacing: 0.4px !important;
    line-height: 1 !important;
  }
  /* Event type label under ticker */
  .ca-card-mob .ca-company {
    font-size: 11px !important;
    font-weight: 600 !important;
    margin-top: 1px !important;
    opacity: 0.9 !important;
  }
  /* Priority badge — pill with glow */
  .ca-card-mob .ca-priority-badge {
    font-size: 10px !important;
    font-weight: 800 !important;
    padding: 5px 12px !important;
    border-radius: 50px !important;
    letter-spacing: 0.4px !important;
    white-space: nowrap !important;
    text-transform: uppercase !important;
    backdrop-filter: blur(4px) !important;
  }

  /* ── Card body: 2-col data grid ── */
  .ca-card-mob .ca-card-body {
    padding: 14px 16px 14px 16px !important;
    display: grid !important;
    grid-template-columns: 1fr 1fr !important;
    gap: 12px 10px !important;
  }
  .ca-card-mob .ca-field {
    display: flex !important;
    flex-direction: column !important;
    gap: 3px !important;
  }
  .ca-card-mob .ca-field-label {
    font-size: 8.5px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.9px !important;
    color: #454870 !important;
    font-weight: 700 !important;
  }
  .ca-card-mob .ca-field-value {
    font-size: 12.5px !important;
    font-weight: 700 !important;
    color: #c0c4e8 !important;
    line-height: 1.3 !important;
  }

  /* ── Ex-date countdown chip ── */
  .ca-ex-chip {
    display: inline-flex !important;
    align-items: center !important;
    gap: 4px !important;
    font-size: 9.5px !important;
    font-weight: 800 !important;
    padding: 2px 8px !important;
    border-radius: 50px !important;
    margin-top: 3px !important;
    letter-spacing: 0.3px !important;
  }

  /* ── Footer: details text ── */
  .ca-card-mob .ca-card-footer {
    padding: 10px 16px !important;
    background: rgba(0,0,0,0.25) !important;
    font-size: 10.5px !important;
    color: #6b7099 !important;
    border-top: 1px solid rgba(255,255,255,0.04) !important;
    display: flex !important;
    align-items: flex-start !important;
    gap: 7px !important;
    line-height: 1.5 !important;
  }

  /* ── Legend strip ── */
  .ca-legend-mob {
    background: linear-gradient(135deg, #0b0d1c 0%, #080a16 100%) !important;
    border: 1px solid #181c36 !important;
    border-radius: 14px !important;
    padding: 14px 16px !important;
    margin-top: 10px !important;
    margin-bottom: 80px !important;
    font-size: 10.5px !important;
    color: #555577 !important;
    line-height: 2.1 !important;
  }
  .ca-legend-mob b { color: #7a7fa8 !important; }
}
</style>
""", unsafe_allow_html=True)

# ── Auto-refresh every 60 s to keep clock current ───────────────
try:
    from streamlit_autorefresh import st_autorefresh
    st_autorefresh(interval=60000, key="clock_refresh")
except Exception:
    pass

# =========================================================
# GITHUB PERSISTENCE LAYER
# =========================================================

import base64, urllib.request, urllib.error

# ── Credentials from Streamlit secrets ───────────────────────────
_GH_TOKEN  = st.secrets.get("GITHUB_TOKEN",  "")
_GH_REPO   = st.secrets.get("GITHUB_REPO",   "devshah24m/stockdashboard")
_GH_BRANCH = st.secrets.get("GITHUB_BRANCH", "main")

def _gh_api(path):
    return f"https://api.github.com/repos/{_GH_REPO}/contents/{path}"

def _gh_headers():
    return {
        "Authorization": f"token {_GH_TOKEN}",
        "Content-Type": "application/json",
        "User-Agent": "northeast-portfolio-app"
    }

def gh_get_sha(path):
    """Get the current SHA of a file on GitHub (needed for updates). Returns None if not found."""
    try:
        req = urllib.request.Request(
            _gh_api(path) + f"?ref={_GH_BRANCH}",
            headers=_gh_headers()
        )
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read()).get("sha")
    except Exception:
        return None

def gh_read_file(path):
    """Read a file from GitHub. Returns (content_str, sha) or (None, None)."""
    try:
        req = urllib.request.Request(
            _gh_api(path) + f"?ref={_GH_BRANCH}",
            headers=_gh_headers()
        )
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            content = base64.b64decode(data["content"]).decode("utf-8")
            return content, data.get("sha")
    except Exception:
        return None, None

def gh_write_file(path, content_str, commit_msg=None):
    """Write/create/overwrite a file on GitHub. Returns True on success."""
    if not _GH_TOKEN:
        return False
    try:
        existing_sha = gh_get_sha(path)
        encoded = base64.b64encode(content_str.encode("utf-8")).decode()
        payload = {
            "message": commit_msg or f"Auto-save: {path}",
            "content": encoded,
            "branch":  _GH_BRANCH,
        }
        if existing_sha:
            payload["sha"] = existing_sha
        data = json.dumps(payload).encode()
        req = urllib.request.Request(
            _gh_api(path), data=data, method="PUT", headers=_gh_headers()
        )
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read())
            return "commit" in result
    except urllib.error.HTTPError as e:
        print(f"[GH] HTTP {e.code} writing {path}: {e.read().decode()[:200]}")
        return False
    except Exception as e:
        print(f"[GH] Error writing {path}: {e}")
        return False

def _gh_sync_to_local(gh_path, local_path):
    """Pull a file from GitHub → write to local disk. Returns True if pulled."""
    content, _ = gh_read_file(gh_path)
    if content:
        try:
            with open(local_path, "w", encoding="utf-8") as f:
                f.write(content)
            return True
        except Exception:
            pass
    return False

def _local_sync_to_gh(local_path, gh_path, commit_msg=None):
    """Read local file → push to GitHub. Returns True on success."""
    if not os.path.exists(local_path):
        return False
    try:
        with open(local_path, "r", encoding="utf-8") as f:
            content = f.read()
        return gh_write_file(gh_path, content, commit_msg)
    except Exception as e:
        print(f"[GH] Error syncing {local_path}: {e}")
        return False

def gh_save_csv(df, path, commit_msg=None):
    """Save a DataFrame as CSV directly to GitHub."""
    return gh_write_file(path, df.to_csv(index=False), commit_msg)

def gh_load_csv(path):
    """Load a CSV from GitHub into a DataFrame. Returns None if not found."""
    content, _ = gh_read_file(path)
    if content:
        try:
            return pd.read_csv(io.StringIO(content))
        except Exception:
            return None
    return None

def gh_sync_all_data_files():
    """Pull ALL data files (clients + all portfolio/trades CSVs) from GitHub to local disk."""
    import glob
    # Always sync clients
    _gh_sync_to_local("clients.json", "clients.json")
    # Sync any portfolio/trades CSVs already in the repo
    try:
        req = urllib.request.Request(
            f"https://api.github.com/repos/{_GH_REPO}/contents/?ref={_GH_BRANCH}",
            headers=_gh_headers()
        )
        with urllib.request.urlopen(req) as resp:
            files = json.loads(resp.read())
        for f in files:
            name = f.get("name", "")
            if (name.startswith("portfolio_") or name.startswith("trades_")) and name.endswith(".csv"):
                _gh_sync_to_local(name, name)
    except Exception as e:
        print(f"[GH] Could not list repo contents: {e}")

# ── On every startup: pull ALL data files from GitHub to local disk ──
gh_sync_all_data_files()

# =========================================================
# CLIENT MANAGEMENT — Client Code login, no visible client list
# =========================================================

CLIENTS_FILE = "clients.json"

# ── Developer credentials (only you know these) ───────────────────
DEV_CODE     = "DEVNORTHEAST"    # ← change to your private dev code
DEV_PASSWORD = "northeast@dev"   # ← change to your private dev password

def load_clients():
    import json
    if os.path.exists(CLIENTS_FILE):
        with open(CLIENTS_FILE) as f:
            data = json.load(f)
        # Migrate old format (name-keyed) → new format (code-keyed)
        migrated = {}
        changed = False
        for k, v in data.items():
            if not isinstance(v, dict):
                continue
            if "display_name" in v:
                migrated[k] = v
            else:
                # Old: key = display name, value = {password_hash: ...}
                code = re.sub(r"[^\w]", "", k).upper()[:8] or "C001"
                migrated[code] = {"display_name": k, "password_hash": v.get("password_hash", "")}
                changed = True
        if changed:
            with open(CLIENTS_FILE, "w") as f:
                json.dump(migrated, f, indent=2)
            _local_sync_to_gh(CLIENTS_FILE, CLIENTS_FILE, "Auto-migrate: clients.json")
        return migrated
    # First run — create default Rohith Sir account
    default = {
        "HO667": {
            "display_name":  "Rohith Sir",
            "password_hash": hashlib.sha256("123456".encode()).hexdigest()
        }
    }
    with open(CLIENTS_FILE, "w") as f:
        json.dump(default, f, indent=2)
    return default

def save_clients(d):
    import json
    with open(CLIENTS_FILE, "w") as f:
        json.dump(d, f, indent=2)
    # ── Also persist to GitHub so data survives Streamlit reboots ──
    _local_sync_to_gh(CLIENTS_FILE, CLIENTS_FILE, "Auto-save: clients.json")

def hash_password(pw: str) -> str:
    if not pw:
        return ""
    return hashlib.sha256(pw.encode()).hexdigest()

def client_portfolio_file(code):
    safe = re.sub(r"[^\w\-]", "_", code.strip().upper())
    path = f"portfolio_{safe}.csv"
    if not os.path.exists(path):
        _gh_sync_to_local(path, path)
    return path

def client_trades_file(code):
    safe = re.sub(r"[^\w\-]", "_", code.strip().upper())
    path = f"trades_{safe}.csv"
    if not os.path.exists(path):
        _gh_sync_to_local(path, path)
    return path

# ── Session state bootstrap ───────────────────────────────────────
for _k, _v in [
    ("auth_code",       None),
    ("auth_name",       None),
    ("is_dev",          False),
    ("login_err",       ""),
    ("nav_tab",         "Portfolio"),
    ("watchlist",       []),      # list of ticker strings
    ("wl_search_query", ""),
    ("wl_order",        None),    # pending order dict
    ("wl_expand",       None),    # ticker whose trade panel is open
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── Load clients ──────────────────────────────────────────────────
clients_dict = load_clients()

# =========================================================
# LOGIN PAGE — shown until authenticated
# Renders a centred card using pure Streamlit (no HTML overlay)
# =========================================================

if not st.session_state["auth_code"]:

    # Hide the sidebar entirely on the login page
    st.markdown("""
<style>
[data-testid="stSidebar"],
[data-testid="stSidebarNav"],
button[kind="header"] { display: none !important; }
.block-container { padding-top: 0 !important; }
</style>
""", unsafe_allow_html=True)

    # ── Vertical spacer + centred card ───────────────────────────
    st.markdown("<div style='height:60px'></div>", unsafe_allow_html=True)

    _l, _c, _r = st.columns([1, 1.05, 1])
    with _c:
        # Card shell
        st.markdown("""
<div style="
    background:#131628;
    border:1px solid #2e3355;
    border-radius:20px;
    padding:36px 32px 28px 32px;
    box-shadow:0 24px 64px rgba(0,0,0,0.6);
    text-align:center;
    margin-bottom:20px;
">
  <img src="https://www.northeastltd.com/wp-content/uploads/2020/06/logo_png.png"
       style="height:56px;margin-bottom:14px;object-fit:contain;"
       onerror="this.style.display='none'"/>
  <div style="font-size:22px;font-weight:800;color:#f0f2ff;
              letter-spacing:-0.3px;margin-bottom:4px;">
    Portfolio Tracker
  </div>
  <div style="font-size:12px;color:#7a7fa8;font-weight:500;margin-bottom:0;">
    Northeast Broking Services Limited
  </div>
</div>
""", unsafe_allow_html=True)

        # Inputs
        _code_val = st.text_input(
            "CLIENT CODE",
            placeholder="Enter your client code",
            key="_login_code",
            label_visibility="visible"
        )
        _pw_val = st.text_input(
            "PASSWORD",
            type="password",
            placeholder="Enter your password",
            key="_login_pw",
            label_visibility="visible"
        )

        # Error message (shown before button so it's close to inputs)
        if st.session_state["login_err"]:
            st.markdown(f"""
<div style="background:rgba(248,84,84,0.12);border:1px solid rgba(248,84,84,0.4);
            border-radius:8px;padding:10px 14px;color:#f85454;
            font-size:12px;font-weight:600;margin-bottom:8px;">
  ❌ {st.session_state["login_err"]}
</div>""", unsafe_allow_html=True)

        _btn = st.button("🔓  Sign In", key="_login_btn", use_container_width=True)

        st.markdown("""
<div style="font-size:10px;color:#454870;text-align:center;
            margin-top:14px;line-height:1.7;">
  Your credentials are private &amp; encrypted.<br>
  Contact your relationship manager for access.
</div>
""", unsafe_allow_html=True)

    # ── Handle login ─────────────────────────────────────────────
    if _btn:
        _code = _code_val.strip().upper()
        _pw   = _pw_val.strip()

        # Developer login
        if _code == DEV_CODE.upper() and _pw == DEV_PASSWORD:
            st.session_state["auth_code"]  = "__DEV__"
            st.session_state["auth_name"]  = "Developer"
            st.session_state["is_dev"]     = True
            st.session_state["login_err"]  = ""
            st.rerun()

        # Client login
        elif _code in clients_dict:
            _meta   = clients_dict[_code]
            _stored = _meta.get("password_hash", "")
            if _stored and hash_password(_pw) == _stored:
                st.session_state["auth_code"]  = _code
                st.session_state["auth_name"]  = _meta.get("display_name", _code)
                st.session_state["is_dev"]     = False
                st.session_state["login_err"]  = ""
                st.rerun()
            elif not _stored:          # no password set — allow blank
                st.session_state["auth_code"]  = _code
                st.session_state["auth_name"]  = _meta.get("display_name", _code)
                st.session_state["is_dev"]     = False
                st.session_state["login_err"]  = ""
                st.rerun()
            else:
                st.session_state["login_err"] = "Invalid Client Code or Password."
                st.rerun()
        else:
            st.session_state["login_err"] = "Invalid Client Code or Password."
            st.rerun()

    st.stop()   # Block the rest of the app

# ── Authenticated — resolve identities ───────────────────────────
selected_client   = st.session_state["auth_name"]
_auth_code        = st.session_state["auth_code"]
_is_dev           = st.session_state["is_dev"]

# ── Sidebar top bar (logged-in state) ────────────────────────────
with st.sidebar:
    if _is_dev:
        st.markdown("""
<div style="padding:4px 0 14px 0;border-bottom:1px solid #f5c84244;margin-bottom:14px;">
  <div style="font-size:11px;font-weight:700;color:#7a7fa8;text-transform:uppercase;
              letter-spacing:1.2px;margin-bottom:3px;">Northeast Broking</div>
  <div style="font-size:15px;font-weight:700;color:#f0f2ff;">Portfolio Tracker</div>
</div>
<div style="background:linear-gradient(135deg,#1a1400,#221900);
            border:2px solid #f5c842;border-radius:12px;
            padding:12px 14px;margin-bottom:14px;">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;">
    <span style="font-size:18px;">🛠</span>
    <span style="font-size:13px;font-weight:900;color:#f5c842;
                 text-transform:uppercase;letter-spacing:1.5px;">Developer</span>
  </div>
  <div style="font-size:11px;color:#a08020;font-weight:600;">
    Full access · All clients · No personal portfolio
  </div>
</div>
""", unsafe_allow_html=True)
    else:
        st.markdown(f"""
<div style="padding:4px 0 14px 0;border-bottom:1px solid #252849;margin-bottom:14px;">
  <div style="font-size:11px;font-weight:700;color:#7a7fa8;text-transform:uppercase;
              letter-spacing:1.2px;margin-bottom:3px;">Northeast Broking</div>
  <div style="font-size:15px;font-weight:700;color:#f0f2ff;">Portfolio Tracker</div>
</div>
<div style="background:#181b2e;border:1px solid #252849;border-radius:10px;
            padding:10px 13px;margin-bottom:12px;">
  <div style="font-size:10px;color:#7a7fa8;text-transform:uppercase;
              letter-spacing:0.9px;font-weight:600;margin-bottom:3px;">👤 Logged In As</div>
  <div style="font-size:14px;color:#f0f2ff;font-weight:700;">{selected_client}</div>
  <div style="font-size:10px;color:#454870;margin-top:2px;">Code: {_auth_code}</div>
</div>
""", unsafe_allow_html=True)

    if st.button("🚪 Log Out", key="logout_btn", use_container_width=True):
        st.session_state["auth_code"]  = None
        st.session_state["auth_name"]  = None
        st.session_state["is_dev"]     = False
        st.session_state["login_err"]  = ""
        st.session_state["nav_tab"]    = "Portfolio"
        st.rerun()

    # ── Sidebar Navigation (all users) ───────────────────────────
    st.markdown("""
<div style="font-size:11px;font-weight:700;color:#7a7fa8;text-transform:uppercase;
            letter-spacing:1px;margin:16px 0 8px 0;">📍 Navigation</div>
""", unsafe_allow_html=True)

    _NAV_ITEMS = [
        ("Watchlist",             "👁 Watchlist"),
        ("Portfolio",             "🏠 Portfolio"),
        ("Allocation",            "📊 Allocation"),
        ("Gainers / Losers",      "🔥 Gainers / Losers"),
        ("Market",                "📊 Market"),
        ("Portfolio News",        "📰 Portfolio News"),
        ("Corporate Actions",     "📅 Corporate Actions"),
        ("XIRR & Returns",        "📈 XIRR & Returns"),
        ("Capital Gains",         "⚖️ Capital Gains"),
        ("Reports & Export",      "📄 Reports & Export"),
        ("CRR Reconciliation",    "📁 CRR Reconciliation"),
        ("All Corporate Actions", "🏦 All Corporate Actions"),
    ]

    _cur_nav = st.session_state.get("nav_tab", "Portfolio")

    # ── Master Import (dev only) ──────────────────────────────
    if _is_dev:
        if st.button("📥 Master Import", key="nav_master_import", use_container_width=True,
                     type="primary" if _cur_nav == "Master Import" else "secondary"):
            st.session_state["nav_tab"] = "Master Import"
            st.rerun()
    for _nav_key, _nav_label in _NAV_ITEMS:
        _is_active = (_cur_nav == _nav_key)
        _btn_style = """
<style>
div[data-testid="stButton"] > button.nav-active {
    background: var(--accent-blue) !important;
}
</style>"""
        if st.button(
            _nav_label,
            key=f"nav_{_nav_key}",
            use_container_width=True,
            type="primary" if _is_active else "secondary",
        ):
            st.session_state["nav_tab"] = _nav_key
            st.rerun()

    st.divider()

    # ── Developer-only admin panel ────────────────────────────────
    if _is_dev:
        st.markdown("""
<div style="font-size:11px;font-weight:700;color:#f5c842;text-transform:uppercase;
            letter-spacing:1px;margin:14px 0 8px 0;">🛠 Admin Panel</div>
""", unsafe_allow_html=True)

        with st.expander("➕ Add New Client"):
            _nc_code = st.text_input("Client Code", key="adm_nc_code",
                                      placeholder="e.g. HO668").strip().upper()
            _nc_name = st.text_input("Display Name", key="adm_nc_name",
                                      placeholder="e.g. Suresh Kumar").strip()
            _nc_pw   = st.text_input("Password", type="password", key="adm_nc_pw")
            if st.button("Add Client", key="adm_add_btn"):
                if not _nc_code or not _nc_name or not _nc_pw:
                    st.error("All fields required.")
                elif _nc_code in clients_dict:
                    st.warning(f"Code '{_nc_code}' already exists.")
                elif _nc_code == DEV_CODE.upper():
                    st.error("Cannot use the developer code.")
                else:
                    clients_dict[_nc_code] = {
                        "display_name":  _nc_name,
                        "password_hash": hash_password(_nc_pw)
                    }
                    save_clients(clients_dict)
                    st.success(f"✅ Added {_nc_code} — {_nc_name}")

        with st.expander("🔑 Reset Client Password"):
            _rp_opts = [f"{c}  ({clients_dict[c]['display_name']})" for c in clients_dict]
            _rp_sel  = st.selectbox("Client", _rp_opts, key="adm_rp_sel")
            _rp_code = _rp_sel.split("  ")[0].strip() if _rp_sel else None
            _rp1 = st.text_input("New Password",     type="password", key="adm_rp1")
            _rp2 = st.text_input("Confirm Password", type="password", key="adm_rp2")
            if st.button("Update", key="adm_rp_btn"):
                if not _rp1:
                    st.error("Enter a password.")
                elif _rp1 != _rp2:
                    st.error("Passwords don't match.")
                elif _rp_code:
                    clients_dict[_rp_code]["password_hash"] = hash_password(_rp1)
                    save_clients(clients_dict)
                    st.success(f"✅ Password updated for {_rp_code}")

        with st.expander("🗑 Delete Client"):
            _dl_opts = [f"{c}  ({clients_dict[c]['display_name']})" for c in clients_dict]
            _dl_sel  = st.selectbox("Client to delete", _dl_opts, key="adm_dl_sel")
            _dl_code = _dl_sel.split("  ")[0].strip() if _dl_sel else None
            if st.button("Delete", key="adm_dl_btn"):
                if len(clients_dict) <= 1:
                    st.error("Cannot delete the last client.")
                elif _dl_code:
                    _dl_name = clients_dict[_dl_code]["display_name"]
                    del clients_dict[_dl_code]
                    save_clients(clients_dict)
                    for _fp in [client_portfolio_file(_dl_code), client_trades_file(_dl_code)]:
                        if os.path.exists(_fp):
                            os.remove(_fp)
                    st.success(f"✅ Deleted {_dl_code} — {_dl_name}")

        with st.expander("📋 All Client Codes"):
            for _c, _m in clients_dict.items():
                st.markdown(
                    f"<div style='font-size:12px;color:#b8bcd8;padding:5px 0;"
                    f"border-bottom:1px solid #252849;'>"
                    f"<span style='color:#f5c842;font-weight:700;'>{_c}</span>"
                    f" — {_m.get('display_name','?')}</div>",
                    unsafe_allow_html=True
                )

        # ── Developer: Client selector (manage any client's portfolio) ──
        st.markdown("""
<div style="font-size:11px;font-weight:700;color:#f5c842;text-transform:uppercase;
            letter-spacing:1px;margin:14px 0 8px 0;">📂 Manage Client Portfolio</div>
""", unsafe_allow_html=True)
        _client_opts = [f"{c}  —  {clients_dict[c]['display_name']}" for c in clients_dict]
        _dev_sel_default = st.session_state.get("dev_selected_client_opt", _client_opts[0] if _client_opts else None)
        _dev_sel_idx = _client_opts.index(_dev_sel_default) if _dev_sel_default in _client_opts else 0
        _dev_sel = st.selectbox(
            "Select Client to Manage",
            _client_opts,
            index=_dev_sel_idx,
            key="dev_client_selector"
        )
        if _dev_sel:
            _dev_client_code = _dev_sel.split("  —  ")[0].strip()
            _dev_client_name = clients_dict.get(_dev_client_code, {}).get("display_name", _dev_client_code)
            st.session_state["dev_selected_client_opt"] = _dev_sel
            st.session_state["dev_client_code"] = _dev_client_code
            st.session_state["dev_client_name"] = _dev_client_name
            st.markdown(f"""
<div style="background:#0f1a14;border:1px solid #22d67b44;border-radius:8px;
            padding:8px 12px;margin-top:6px;margin-bottom:4px;">
  <div style="font-size:10px;color:#22d67b;font-weight:700;margin-bottom:2px;">✅ Managing portfolio of:</div>
  <div style="font-size:13px;color:#f0f2ff;font-weight:700;">{_dev_client_name}</div>
  <div style="font-size:10px;color:#454870;">Code: {_dev_client_code}</div>
</div>
""", unsafe_allow_html=True)

    st.divider()

# ── Per-client file paths ─────────────────────────────────────────
if _is_dev:
    _dev_code = st.session_state.get("dev_client_code", list(clients_dict.keys())[0])
    PORTFOLIO_FILE = client_portfolio_file(_dev_code)
    TRADES_FILE    = client_trades_file(_dev_code)
    _managing_name = st.session_state.get("dev_client_name", clients_dict.get(_dev_code, {}).get("display_name", _dev_code))
else:
    PORTFOLIO_FILE = client_portfolio_file(_auth_code)
    TRADES_FILE    = client_trades_file(_auth_code)

# ── Header: Logo + Title + Live DateTime ─────────────────────────
_IST_TZ = timezone(timedelta(hours=5, minutes=30))
_now = datetime.now(_IST_TZ).strftime("%a, %d %b %Y  ·  %I:%M %p")
def _is_market_open_now():
    """Return True if Indian market (NSE/BSE) is currently open.
    Market hours: Mon–Fri 09:15–15:30 IST."""
    from datetime import time as dtime
    _IST = timezone(timedelta(hours=5, minutes=30))
    _now_ist = datetime.now(_IST)
    if _now_ist.weekday() >= 5:
        return False
    _t = _now_ist.time()
    return dtime(9, 15) <= _t <= dtime(15, 30)

_mkt_open = _is_market_open_now()
_mkt_badge = (
    '<span style="background:#22d67b22;color:#22d67b;font-size:10px;font-weight:700;'
    'padding:3px 10px;border-radius:20px;border:1px solid #22d67b44;margin-left:10px;">● MARKET OPEN</span>'
    if _mkt_open else
    '<span style="background:#f8545422;color:#f85454;font-size:10px;font-weight:700;'
    'padding:3px 10px;border-radius:20px;border:1px solid #f8545444;margin-left:10px;">● MARKET CLOSED</span>'
)
_header_name = (
    f"🛠 Developer &nbsp;<span style='font-size:13px;color:#f5c842;font-weight:600;'>"
    f"— Managing: {st.session_state.get('dev_client_name', selected_client)}</span>"
) if _is_dev else selected_client

st.markdown(f"""
<style>
@keyframes gradientShift {{
  0%   {{ background-position: 0% 50%; }}
  50%  {{ background-position: 100% 50%; }}
  100% {{ background-position: 0% 50%; }}
}}
.hdr-gradient-bar {{
  height: 3px;
  background: linear-gradient(90deg, #4f7ef8, #22d67b, #f5c842, #f85454, #a78bfa, #4f7ef8);
  background-size: 300% 100%;
  animation: gradientShift 5s ease infinite;
  box-shadow: 0 0 14px rgba(79,126,248,0.5), 0 0 30px rgba(79,126,248,0.2);
}}
.hdr-name-gradient {{
  background: linear-gradient(90deg, #4f7ef8, #22d67b, #f5c842, #f85454, #4f7ef8);
  background-size: 300% 100%;
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  background-clip: text;
  animation: gradientShift 4s ease infinite;
  font-weight: 800;
  letter-spacing: -0.3px;
  display: inline-block;
}}
</style>

<!-- ══════════ DESKTOP HEADER (hidden on mobile) ══════════ -->
<div class="hdr-desktop" style="align-items:center; justify-content:space-between;
            padding:10px 16px; margin-bottom:14px;
            background:#131628; border:1px solid {'#f5c84244' if _is_dev else '#252849'};
            border-radius:12px; overflow:hidden; position:relative;">
  <div class="hdr-gradient-bar" style="position:absolute;top:0;left:0;right:0;"></div>
  <div style="display:flex; align-items:center; gap:12px; padding-top:4px;">
    <img src="https://www.northeastltd.com/wp-content/uploads/2020/06/logo_png.png"
         style="height:34px; object-fit:contain; border-radius:5px;"
         onerror="this.style.display='none'"/>
    <div>
      <div style="display:flex; align-items:center; gap:8px;">
        <span class="hdr-name-gradient" style="font-size:15px;">
          {_header_name}
        </span>
        {_mkt_badge}
      </div>
      <div style="font-size:10px; color:#7a7fa8; margin-top:1px; font-weight:500;">
        Real-time Holdings &amp; P&amp;L Dashboard &nbsp;·&nbsp; Northeast Broking Services
      </div>
    </div>
  </div>
  <div style="text-align:right; padding-top:4px;">
    <div style="font-size:11px; color:#7a7fa8; font-weight:500;">🕐 {_now}</div>
    <div style="font-size:10px; color:#454870; margin-top:1px;">IST · Auto-refreshes every 60s</div>
  </div>
</div>

<!-- ══════════ MOBILE HEADER (hidden on desktop) ══════════ -->
<div class="hdr-mobile" style="margin-bottom:12px; border-radius:14px; overflow:hidden;
     background:linear-gradient(160deg,#131628 0%,#181d35 100%);
     border:1px solid {'#f5c84244' if _is_dev else '#2a2f52'};
     box-shadow:0 4px 20px rgba(79,126,248,0.10);">

  <!-- animated gradient top bar -->
  <div class="hdr-gradient-bar"></div>

  <!-- logo + name row -->
  <div style="display:flex; align-items:center; gap:10px; padding:10px 12px 8px 12px;">
    <img src="https://www.northeastltd.com/wp-content/uploads/2020/06/logo_png.png"
         style="height:34px; width:34px; object-fit:contain; border-radius:7px;
                background:#1e2340; padding:3px; flex-shrink:0;"
         onerror="this.style.display='none'"/>
    <div style="flex:1; min-width:0;">
      <div class="hdr-name-gradient" style="font-size:14px; line-height:1.2;
                  white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">
        {_header_name}
      </div>
      <div style="font-size:9px; color:#7a7fa8; margin-top:2px; font-weight:500;">
        Real-time Holdings &amp; P&amp;L · Northeast Broking
      </div>
    </div>
    <div style="flex-shrink:0;">
      {_mkt_badge}
    </div>
  </div>

  <!-- date/time strip -->
  <div style="padding:6px 12px 10px 12px; border-top:1px solid #252849;
              background:rgba(79,126,248,0.04); display:flex;
              align-items:center; justify-content:space-between;">
    <div style="font-size:10px; color:#7a7fa8; font-weight:500;">🕐 {_now}</div>
    <div style="font-size:9px; color:#454870;">Auto-refreshes every 60s</div>
  </div>
</div>
""", unsafe_allow_html=True)

# =========================================================
# ANGEL ONE CREDENTIALS
# =========================================================

CLIENT_CODE = "RRZLA8038"
PIN = "1432"
API_KEY = "nAUkqrs1"
TOTP_SECRET = "2T7LCR6V2MIJCOZ6HGSXTKA3SA"


# =========================================================
# ANGEL ONE LOGIN
# =========================================================

smartApi = None
auth_token = None

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

def _patch_smartapi_session(api_obj):
    try:
        retry = Retry(total=2, backoff_factor=1, status_forcelist=[500,502,503,504], allowed_methods=["GET","POST"])
        adapter = HTTPAdapter(max_retries=retry)
        if hasattr(api_obj, "session"):
            api_obj.session.mount("https://", adapter)
            api_obj.session.mount("http://", adapter)
    except Exception:
        pass

try:
    smartApi = SmartConnect(api_key=API_KEY)
    _patch_smartapi_session(smartApi)
    session_data = smartApi.generateSession(
        CLIENT_CODE,
        PIN,
        pyotp.TOTP(TOTP_SECRET).now()
    )
    if session_data and session_data.get("status"):
        auth_token = session_data.get("data", {}).get("jwtToken", "")
        st.sidebar.success("✅ Angel One Connected")
    else:
        msg = session_data.get("message", "Unknown error") if session_data else "No response"
        st.sidebar.warning(f"⚠️ Angel One: {msg} — running in offline mode")
        smartApi = None
except Exception as e:
    err = str(e)
    if "timed out" in err.lower() or "ConnectTimeout" in err or "Max retries" in err:
        st.sidebar.warning(
            "⚠️ Angel One unreachable (network timeout). "
            "Switch to **mobile hotspot** and restart. "
            "Running in offline mode — live prices unavailable."
        )
    elif "invalid" in err.lower() or "wrong" in err.lower():
        st.sidebar.error("❌ Invalid credentials. Check CLIENT_CODE / PIN / TOTP_SECRET.")
    else:
        st.sidebar.error(f"❌ Angel One Login Failed: {e}")
    smartApi = None

# =========================================================
# ANGEL ONE — DYNAMIC SCRIP MASTER TOKEN LOOKUP
# =========================================================

SCRIP_MASTER_URL = "https://margincalculator.angelbroking.com/OpenAPI_File/files/OpenAPIScripMaster.json"
SCRIP_MASTER_CACHE_FILE = "scrip_master.json"

@st.cache_data(ttl=3600, show_spinner=False)
def load_scrip_master():
    """
    Download Angel One scrip master JSON and build THREE lookup dicts:
      - equity_map:  NAME/SYMBOL -> {token, exch_seg, symbol}  (NSE EQ)
      - bse_map:     NAME/SYMBOL/SCRIPCODE -> {token, exch_seg, symbol}  (BSE EQ)
      - nfo_map:     SYMBOL_UPPER -> {token, exch_seg, symbol}  (NFO futures/options)
    Falls back to local cache file if download fails.
    """
    import json

    try:
        resp = requests.get(SCRIP_MASTER_URL, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        with open(SCRIP_MASTER_CACHE_FILE, "w") as f:
            json.dump(data, f)
    except Exception:
        if os.path.exists(SCRIP_MASTER_CACHE_FILE):
            with open(SCRIP_MASTER_CACHE_FILE) as f:
                data = json.load(f)
        else:
            return {}, {}, {}

    equity_map = {}
    bse_map    = {}
    nfo_map    = {}
    sym_to_name = {}   # NSE symbol (clean, no suffix) → proper-case company name

    for item in data:
        exch   = item.get("exch_seg", "")
        name   = item.get("name", "").strip()
        name_u = name.upper()
        sym    = item.get("symbol", "").strip().upper()
        token  = item.get("token", "")
        itype  = (item.get("instrumenttype", "") or "").strip().upper()

        entry = {"token": token, "exch_seg": exch, "symbol": sym, "name": name}

        # ── Equity (NSE cash segment) ──────────────────────────
        if exch == "NSE":
            if name_u and (name_u not in equity_map or itype == "EQ"):
                equity_map[name_u] = entry
            if sym:
                base = sym.split("-")[0].strip()
                if base and base not in equity_map:
                    equity_map[base] = entry
                # Build symbol → proper-case company name map (EQ entries take priority)
                if name and (base not in sym_to_name or itype == "EQ"):
                    sym_to_name[base] = name

        # ── Equity (BSE cash segment) ──────────────────────────
        if exch == "BSE":
            # Index by name
            if name and (name not in bse_map or itype == "EQ"):
                bse_map[name] = entry
            # Index by full symbol and base symbol (BSE symbols: "532540-A" etc.)
            if sym:
                bse_map[sym] = entry
                base = sym.split("-")[0].strip()
                if base and base not in bse_map:
                    bse_map[base] = entry
            # Index by token == BSE scrip code (numeric string like "532540")
            if token:
                bse_map[str(token)] = entry

        # ── NFO (futures & options) ────────────────────────────
        if exch == "NFO":
            if sym:
                nfo_map[sym] = entry
            if name and name not in nfo_map:
                nfo_map[name] = entry

    return equity_map, bse_map, nfo_map, sym_to_name


# ── F&O ticker parser ──────────────────────────────────────────────────────────
# User enters tickers like: "DRREDDY 26 MAY FUTURES", "ABB FUTURES", "NIFTY CE 24000"
# We need to map these to NFO scrip master symbols like "DRREDDY26MAY25FUT"

import calendar

MONTH_ABBR = {m.upper(): i for i, m in enumerate(calendar.month_abbr) if m}
# e.g. {"JAN":1, "FEB":2, ..., "DEC":12}

def parse_fno_ticker(raw):
    """
    Parse a user-entered F&O ticker string and return the most likely
    NFO symbol to search for in the scrip master.

    Examples:
      "DRREDDY 26 MAY FUTURES"  -> base="DRREDDY", expiry_hint="26MAY", kind="FUT"
      "ABB FUTURES"             -> base="ABB",      expiry_hint=None,    kind="FUT"
      "NIFTY 24000 CE"          -> base="NIFTY",    expiry_hint=None,    kind="CE", strike=24000
    """
    raw = raw.upper().replace(".NS", "").replace(".BO", "").strip()
    tokens = raw.split()

    kind       = "FUT"
    strike     = None
    expiry_day = None
    expiry_mon = None
    base_parts = []

    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("FUTURES", "FUT"):
            kind = "FUT"
        elif t in ("CE", "CALL"):
            kind = "CE"
        elif t in ("PE", "PUT"):
            kind = "PE"
        elif t in MONTH_ABBR:
            expiry_mon = t
        elif t.isdigit() and len(t) <= 2 and int(t) <= 31:
            # Could be expiry day
            expiry_day = t.zfill(2)
        elif t.isdigit() and len(t) >= 4:
            # Likely a strike price
            strike = t
        else:
            base_parts.append(t)
        i += 1

    base = "".join(base_parts)  # e.g. "DRREDDY", "ABB", "NIFTY"

    # Build expiry string: day + month + 2-digit year
    now = datetime.now()
    expiry_str = ""
    if expiry_day and expiry_mon:
        yr2 = str(now.year)[-2:]
        expiry_str = f"{expiry_day}{expiry_mon}{yr2}"  # e.g. "26MAY25"
    elif expiry_mon:
        yr2 = str(now.year)[-2:]
        expiry_str = f"{expiry_mon}{yr2}"              # e.g. "MAY25"

    return base, expiry_str, kind, strike


def resolve_nfo_token(raw_ticker, nfo_map):
    """
    Resolve F&O ticker to NFO scrip master entry.

    The user should enter tickers EXACTLY as Angel One NFO symbols:
      e.g.  ABB26MAY25FUT   DRREDDY26MAY25FUT   NIFTY26MAY2425000CE

    Fallback: parse human-readable names like "ABB FUTURE 26 MAY"
    and try to reconstruct the NFO symbol.

    Returns dict with {token, exch_seg, symbol} or None.
    Also returns the resolved_symbol string for debug display.
    """
    raw_up = raw_ticker.upper().replace(".NS", "").replace(".BO", "").strip()
    # Normalise spaces → no spaces (NFO symbols have no spaces)
    raw_nospace = re.sub(r"\s+", "", raw_up)

    # ── Step 1: Direct exact match on normalised symbol ───────────────
    for candidate in (raw_up, raw_nospace):
        if candidate in nfo_map:
            entry = dict(nfo_map[candidate])
            entry["_resolved_as"] = candidate
            return entry

    # ── Step 2: Parse human-readable name ─────────────────────────────
    base, expiry_str, kind, strike = parse_fno_ticker(raw_ticker)
    if not base:
        return None

    # Step 2a: base + expiry + kind   e.g. ABB26MAY25FUT
    if expiry_str:
        for c in (
            f"{base}{expiry_str}{kind}",
            f"{base}{expiry_str}{kind}",
        ):
            if c in nfo_map:
                entry = dict(nfo_map[c])
                entry["_resolved_as"] = c
                return entry
        # options with strike
        if strike:
            for c in (f"{base}{expiry_str}{strike}{kind}",
                      f"{base}{expiry_str}{kind}{strike}"):
                if c in nfo_map:
                    entry = dict(nfo_map[c])
                    entry["_resolved_as"] = c
                    return entry

    # ── Step 3: Scan for base + kind (FUT only — avoid CE/PE false hits) ──
    # Only match FUT contracts; do NOT do loose "any" matching for options
    # to avoid returning a CE when user wants FUT.
    fut_matches = sorted([
        (k, v) for k, v in nfo_map.items()
        if k.startswith(base) and k.endswith("FUT")
    ])
    if kind == "FUT" and fut_matches:
        entry = dict(fut_matches[0][1])
        entry["_resolved_as"] = fut_matches[0][0]
        return entry

    # CE/PE: must have expiry or strike to avoid wrong contract
    if kind in ("CE", "PE") and expiry_str:
        opt_matches = sorted([
            (k, v) for k, v in nfo_map.items()
            if k.startswith(base) and k.endswith(kind)
            and expiry_str in k
        ])
        if opt_matches:
            entry = dict(opt_matches[0][1])
            entry["_resolved_as"] = opt_matches[0][0]
            return entry

    return None


def debug_fno_resolution(portfolio_df, nfo_map):
    """Return a DataFrame showing what NFO symbol each F&O ticker resolved to."""
    rows = []
    fno = portfolio_df[portfolio_df["Asset_Type"] == "F&O"]
    for _, r in fno.iterrows():
        raw   = r["Ticker"].replace(".NS","").replace(".BO","").upper()
        info  = resolve_nfo_token(raw, nfo_map)
        rows.append({
            "Entered Ticker":    r["Ticker"],
            "Resolved NFO Symbol": info["_resolved_as"] if info else "❌ NOT FOUND",
            "Token":             info["token"]     if info else "—",
            "Exchange":          info["exch_seg"]  if info else "—",
        })
    return pd.DataFrame(rows)


# ── Portfolio ticker → Angel One scrip master symbol overrides ────────────
# When a stock shows nan for Daily P&L it means resolve_token() could not
# find it in equity_map. Add the correct Angel One symbol here as the value.
# Key = clean ticker (no .NS / .BO, uppercase).
ANGEL_ONE_ALIASES = {
    "GAYAPROJ":      "GAYAPROJFIN",
    "JAYBARMARU":    "JAYBARMARU",
    "KALAMANDIR":    "KALAMANDIR",
    "KAMATHOTEL":    "KAMATHOTEL",
    "LINCOLNPHARMA": "LINCOLNPHARMA",
    "LINCOLN":       "LINCOLNPHARMA",
    "OCCLLTD":       "OCCLLTD",
    "HEXT":          "HEXT",
}

def resolve_token(symbol_clean, equity_map, nfo_map, asset_type="Stock", bse_map=None):
    """Route to equity (NSE first, BSE fallback) or NFO lookup depending on asset type.
    For .BO tickers always go straight to BSE map.
    Applies ANGEL_ONE_ALIASES first so mismatched tickers resolve correctly."""
    if asset_type == "F&O":
        return resolve_nfo_token(symbol_clean, nfo_map)

    raw = symbol_clean.strip().upper()

    # ── Explicit .BO suffix → BSE only ────────────────────────────────
    if raw.endswith(".BO"):
        raw = raw[:-3]
        if bse_map:
            return _resolve_in_bse(raw, bse_map)
        return None

    # ── Explicit .NS suffix → NSE only ────────────────────────────────
    if raw.endswith(".NS"):
        raw = raw[:-3]

    # Equity / ETF — apply alias override before map lookup
    s  = ANGEL_ONE_ALIASES.get(raw, raw)
    # 1. Try NSE equity map
    result = _resolve_in_nse(s, equity_map)
    if result:
        return result
    # 2. Fallback: BSE equity map
    if bse_map:
        return _resolve_in_bse(s, bse_map)
    return None


def _resolve_in_nse(s, equity_map):
    """Look up a clean symbol in the NSE equity map."""
    if s in equity_map:
        return equity_map[s]
    s2 = re.sub(r"[^A-Z0-9]", "", s)
    if s2 in equity_map:
        return equity_map[s2]
    for key in equity_map:
        if key.startswith(s) or s.startswith(key):
            return equity_map[key]
    return None


def _resolve_in_bse(s, bse_map):
    """Look up a clean symbol / scrip-code in the BSE equity map."""
    if s in bse_map:
        return bse_map[s]
    s2 = re.sub(r"[^A-Z0-9]", "", s)
    if s2 in bse_map:
        return bse_map[s2]
    # Numeric scrip code entered directly (e.g. "532540")
    if s.isdigit() and s in bse_map:
        return bse_map[s]
    for key in bse_map:
        # Skip numeric-only keys to avoid spurious partial matches
        if key.isdigit():
            continue
        if key.startswith(s) or s.startswith(key):
            return bse_map[key]
    return None


def is_bse_ticker(ticker):
    """Return True if the ticker is explicitly a BSE ticker (.BO suffix)."""
    return str(ticker).upper().endswith(".BO")


def get_angel_one_price(symbol_clean, equity_map, nfo_map, asset_type="Stock", bse_map=None):
    """Fetch LTP from Angel One. Falls back to yfinance if Angel One is unavailable."""
    # ── 1. Angel One LTP ─────────────────────────────────────────────
    if smartApi is not None:
        info = resolve_token(symbol_clean, equity_map, nfo_map, asset_type, bse_map=bse_map)
        if info:
            try:
                ltp_data = smartApi.ltpData(
                    info["exch_seg"],
                    info["symbol"],
                    info["token"]
                )
                if ltp_data and ltp_data.get("status"):
                    v = _safe_float(ltp_data["data"]["ltp"])
                    if v and v > 0:
                        return v
            except Exception:
                pass

    # ── 2. yfinance fallback (when Angel One fails / offline) ────────
    if _YF_AVAILABLE:
        try:
            t = str(symbol_clean).strip()
            if not t.endswith(".NS") and not t.endswith(".BO"):
                t = t + ".NS"
            ticker_obj = yf.Ticker(t)
            fast = ticker_obj.fast_info
            lp = _safe_float(getattr(fast, "last_price", None))
            if lp and lp > 0:
                return lp
            hist = ticker_obj.history(period="1d", interval="1m")
            if not hist.empty:
                v = _safe_float(hist["Close"].iloc[-1])
                if v and v > 0:
                    return v
        except Exception:
            pass
    return None


# ── Unlisted Share CMP helper ─────────────────────────────────────────────────
def fetch_unlisted_cmp(ticker, equity_map, nfo_map, bse_map=None, session_cache=None):
    """
    Fetch a display-only CMP for an Unlisted Share ticker.
    Waterfall (stops at first non-zero result):
      1. Session-state manual override  — user typed a price in the sidebar
      2. Angel One equity lookup        — many 'unlisted' stocks are actually listed
      3. unlistedshares.com HTML scrape — genuine pre-IPO / unlisted names
      4. Returns 0.0 if all sources fail (shown as dash in the table)
    Buy_Price is NEVER read or written here.
    """
    # 1. Manual override from sidebar
    if session_cache:
        cached = session_cache.get(ticker, 0)
        if cached and float(cached) > 0:
            return float(cached)

    # 2. Try Angel One — some stocks tagged Unlisted are actually exchange-listed
    try:
        _clean_tkr = ticker.replace("#EQUITYSHARES", "").replace("EQEQ", "").strip()
        price = get_angel_one_price(_clean_tkr, equity_map, nfo_map, "Stock", bse_map=bse_map)
        if price and price > 0:
            return float(price)
    except Exception:
        pass

    # 3. Scrape unlistedshares.com
    try:
        import requests as _ureq, re as _ure
        _s = ticker.replace("#EQUITYSHARES", "").replace("EQEQ", "").replace("EQ", "").strip()
        _url  = "https://www.unlistedshares.com/?s=" + _s
        _hdrs = {
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/124.0 Safari/537.36"),
            "Accept": "text/html,application/xhtml+xml",
        }
        _r = _ureq.get(_url, headers=_hdrs, timeout=7)
        if _r.status_code == 200:
            _matches = _ure.findall(
                r'(?:₹|Rs\.?|INR)\s*([\d,]+(?:\.\d+)?)',
                _r.text
            )
            for _m in _matches:
                _v = float(_m.replace(",", ""))
                if _v > 0:
                    return _v
    except Exception:
        pass

    return 0.0


def get_angel_one_price_bulk(portfolio_df, equity_map, nfo_map, bse_map=None):
    """Fetch prices for all tickers using Angel One API concurrently.
    Uses Asset_Type column to route equity vs F&O lookups correctly.
    .BO tickers are routed to BSE; .NS / plain tickers try NSE then BSE."""
    prices = {}

    # Session-state CMP cache for unlisted shares (keyed by ticker, per client)
    _ul_cache = st.session_state.get("_ul_cmp_cache", {})

    def fetch_one(row):
        t     = row["Ticker"]
        atype = row.get("Asset_Type", "Stock")
        if atype == "Unlisted Share":
            # Use dedicated unlisted fetch — never reads Buy_Price
            return t, fetch_unlisted_cmp(t, equity_map, nfo_map,
                                         bse_map=bse_map, session_cache=_ul_cache)
        # Pass original ticker so resolve_token can detect .BO / .NS
        price = get_angel_one_price(t, equity_map, nfo_map, atype, bse_map=bse_map)
        return t, price

    rows = portfolio_df.to_dict("records")
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        for ticker, price in ex.map(fetch_one, rows):
            prices[ticker] = price

    return prices


def _safe_float(val):
    """Convert a value to float safely, stripping commas."""
    try:
        return float(str(val).replace(",", "").strip())
    except Exception:
        return None


def _prev_close_angel_one(t, info):
    """
    Fetch the official previous-day closing price from Angel One candle API.
    Strategy:
      • Go back 7 calendar days to ensure we always get at least 2 trading days.
      • During market hours  → candles[-2][4]  (yesterday's confirmed close).
      • After market close   → candles[-1][4]  (today's confirmed close).
      • Candle index 4 = official close price (OHLCV: 0=ts,1=O,2=H,3=L,4=C,5=V).
    """
    if smartApi is None or info is None:
        return None
    try:
        from datetime import date as _date
        today      = _date.today()
        from_date  = (today - timedelta(days=7)).strftime("%Y-%m-%d %H:%M")
        to_date    = today.strftime("%Y-%m-%d %H:%M")
        resp = smartApi.getCandleData({
            "exchange":    info["exch_seg"],
            "symboltoken": info["token"],
            "interval":    "ONE_DAY",
            "fromdate":    from_date,
            "todate":      to_date,
        })
        if resp and resp.get("status") and resp.get("data"):
            candles = resp["data"]
            if not candles:
                return None
            if _is_market_open_now():
                # Market is live → latest candle is today's intraday, not closed yet
                # Use previous completed day's close
                if len(candles) >= 2:
                    return _safe_float(candles[-2][4])
                return _safe_float(candles[-1][4])
            else:
                # Market closed → latest candle IS the official closing price
                return _safe_float(candles[-1][4])
    except Exception:
        pass
    return None


def _prev_close_nse_api(alias):
    """Fetch previous close from NSE quote API. Returns (prev_close, ltp)."""
    try:
        hdrs = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/json",
            "Referer": "https://www.nseindia.com",
        }
        s = requests.Session()
        s.get("https://www.nseindia.com", headers=hdrs, timeout=8)
        r = s.get(
            f"https://www.nseindia.com/api/quote-equity?symbol={alias}",
            headers=hdrs, timeout=8
        )
        if r.status_code == 200:
            d = r.json()
            pi = d.get("priceInfo", {}) or {}
            # NSE returns:
            #   priceInfo.previousClose  → official previous day close
            #   priceInfo.close          → today's official close (after 15:30)
            #   priceInfo.lastPrice      → current LTP
            if not _is_market_open_now():
                # After close: prefer today's official close if available
                close_today = _safe_float(pi.get("close"))
                if close_today and close_today > 0:
                    return close_today
            pc = _safe_float(pi.get("previousClose"))
            if pc and pc > 0:
                return pc
    except Exception:
        pass
    return None


@st.cache_data(ttl=86400, show_spinner=False)
def _fetch_sector_nse(symbol: str) -> str:
    """
    Robust sector fetch with 4-source waterfall.  Works for large-caps AND
    small/micro-cap stocks that may not have a full NSE quote-equity record.

    Source priority:
      1. NSE quote-equity API   – industryInfo.sector / metadata.industry
      2. BSE company detail API – Industry field
      3. yfinance Ticker.info   – sector / industry
      4. Scrip-master heuristic – instrument-type → broad category label
         (catches anything still missing: ETF, F&O, Index, etc.)
    """
    sym = symbol.upper().strip().replace(".NS", "").replace(".BO", "")
    hdrs_nse = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept":  "application/json",
        "Referer": "https://www.nseindia.com",
    }

    # ── Source 1: NSE quote-equity API ───────────────────────────────
    try:
        s1 = requests.Session()
        s1.get("https://www.nseindia.com", headers=hdrs_nse, timeout=8)
        r1 = s1.get(
            f"https://www.nseindia.com/api/quote-equity?symbol={sym}",
            headers=hdrs_nse, timeout=8
        )
        if r1.status_code == 200:
            d1 = r1.json()
            for src_key, field in [
                ("industryInfo", "sector"),
                ("industryInfo", "industry"),
                ("metadata",     "industry"),
                ("info",         "industry"),
            ]:
                sector = ((d1.get(src_key) or {}).get(field) or "").strip()
                if sector and sector.lower() not in ("", "na", "n/a", "-"):
                    return sector.title()
    except Exception:
        pass

    # ── Source 2: BSE company detail API ─────────────────────────────
    try:
        hdrs_bse = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept":  "application/json",
            "Referer": "https://www.bseindia.com",
        }
        # BSE search: find scrip code from symbol name
        s2 = requests.Session()
        s2.get("https://www.bseindia.com", headers=hdrs_bse, timeout=8)
        search_url = (
            f"https://api.bseindia.com/BseIndiaAPI/api/ListofScripData/w?"
            f"Group=&Scripcode=&scripname={sym}&segmentLink=9&strSearch=S"
        )
        r2 = s2.get(search_url, headers=hdrs_bse, timeout=8)
        if r2.status_code == 200:
            items = r2.json()
            if isinstance(items, list) and items:
                sc = str(items[0].get("SCRIP_CD", ""))
                if sc:
                    detail_url = (
                        f"https://api.bseindia.com/BseIndiaAPI/api/ComHeader/w?"
                        f"quotetype=EQ&scripcode={sc}&seriesid="
                    )
                    r2b = s2.get(detail_url, headers=hdrs_bse, timeout=8)
                    if r2b.status_code == 200:
                        d2 = r2b.json()
                        for fld in ("Industry", "Sector", "industry", "sector"):
                            sector = (d2.get(fld) or "").strip()
                            if sector and sector.lower() not in ("", "na", "n/a", "-"):
                                return sector.title()
    except Exception:
        pass

    # ── Source 3: yfinance Ticker.info ───────────────────────────────
    if _YF_AVAILABLE:
        for suffix in (".NS", ".BO"):
            try:
                info_yf = yf.Ticker(sym + suffix).info
                for fld in ("sector", "industry"):
                    sector = (info_yf.get(fld) or "").strip()
                    if sector and sector.lower() not in ("", "na", "n/a", "-",
                                                         "none", "unknown"):
                        return sector.title()
            except Exception:
                continue

    # ── Source 4: Scrip-master instrument-type heuristic ─────────────
    # classify_asset already inferred a category — promote it to a
    # human-readable "sector" label so the column is never blank.
    asset_cat = classify_asset(sym)
    _cat_to_sector = {
        "ETF":              "Index / Equity ETF",
        "Liquid ETF":       "Liquid / Money Market",
        "Commodity ETF":    "Commodities",
        "International ETF":"International Equity",
        "F&O":              "Derivatives",
        "Index F&O":        "Index Derivatives",
        "Currency F&O":     "Currency Derivatives",
        "REIT/InvIT":       "Real Estate / Infrastructure",
        "SGB":              "Sovereign Gold",
        "Bond/NCD":         "Fixed Income",
        "Mutual Fund":      "Mutual Fund",
        "Preference Share": "Preference Shares",
        "Unlisted Share":   "Unlisted Shares",
    }
    return _cat_to_sector.get(asset_cat, "")


def _prev_close_bse_api(bse_token, alias):
    """Fetch previous close from BSE quote API using scrip code or symbol."""
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Referer": "https://www.bseindia.com",
    }
    # Try 1: BSE scrip header data (most reliable, needs numeric scrip code)
    if bse_token and str(bse_token).isdigit():
        try:
            s = requests.Session()
            s.get("https://www.bseindia.com", headers=hdrs, timeout=8)
            url = (
                f"https://api.bseindia.com/BseIndiaAPI/api/getScripHeaderData/w?"
                f"Debtflag=&scripcode={bse_token}&seriesid="
            )
            r = s.get(url, headers=hdrs, timeout=8)
            if r.status_code == 200:
                d = r.json()
                # BSE field names vary — check all known variants
                for key in ("PrevClose", "Prev_Close", "prevClose",
                            "PREV_CLOSE", "prev_close"):
                    val = d.get(key) or (d.get("Header") or {}).get(key)
                    if val:
                        v = _safe_float(val)
                        if v and v > 0:
                            return v
                # Also check nested structure
                for section in ("Header", "Body", "Data"):
                    sub = d.get(section) or {}
                    for key in ("PrevClose", "Prev_Close", "prevClose", "CurrPrevClose"):
                        val = sub.get(key)
                        if val:
                            v = _safe_float(val)
                            if v and v > 0:
                                return v
        except Exception:
            pass

    # Try 2: BSE market data API by symbol
    try:
        s2 = requests.Session()
        s2.get("https://www.bseindia.com", headers=hdrs, timeout=8)
        url2 = (
            f"https://api.bseindia.com/BseIndiaAPI/api/StockTrading/w?"
            f"flag=0&quotetype=EQ&scripcode={bse_token}"
        )
        r2 = s2.get(url2, headers=hdrs, timeout=8)
        if r2.status_code == 200:
            d2 = r2.json()
            for key in ("PrevClose", "Prev_Close", "prevClose", "CurrPrevClose"):
                val = d2.get(key)
                if val:
                    v = _safe_float(val)
                    if v and v > 0:
                        return v
    except Exception:
        pass

    # Try 3: BSE marketdata endpoint (public, no auth)
    if alias:
        try:
            s3 = requests.Session()
            url3 = (
                f"https://marketdata.bseindia.com/BseIndiaAPI/api/StockReachGraph/w?"
                f"scripcode={bse_token}&flag=0&fromdate=&todate=&seriesid=EQ"
            )
            r3 = s3.get(url3, headers=hdrs, timeout=8)
            if r3.status_code == 200:
                d3 = r3.json()
                for key in ("PrevClose", "Prev_Close", "prevClose"):
                    val = d3.get(key)
                    if val:
                        v = _safe_float(val)
                        if v and v > 0:
                            return v
        except Exception:
            pass
    return None


def _prev_close_yfinance(ticker_original):
    """
    Fetch previous close via yfinance — most reliable, works for both
    NSE (.NS) and BSE (.BO) tickers. Returns official closing price.
    """
    if not _YF_AVAILABLE:
        return None
    try:
        t = ticker_original.strip()
        if not t.endswith(".NS") and not t.endswith(".BO"):
            t = t + ".NS"
        hist = yf.Ticker(t).history(period="5d", interval="1d", auto_adjust=True)
        if hist.empty:
            return None
        if _is_market_open_now():
            if len(hist) >= 2:
                return _safe_float(hist["Close"].iloc[-2])
            return _safe_float(hist["Close"].iloc[-1])
        else:
            return _safe_float(hist["Close"].iloc[-1])
    except Exception:
        pass
    return None


def get_prev_close_bulk(portfolio_df, equity_map, nfo_map, bse_map=None):
    """
    Fetch the official previous/latest closing price for every holding.

    Source waterfall for each ticker (stops at first valid non-zero result):
      1. Angel One candle API  — fastest, market-hours-aware
      2. NSE quote API         — official NSE close / previousClose
      3. BSE scrip header API  — for .BO tickers (tries 3 BSE endpoints)
      4. yfinance              — most reliable fallback, handles both .NS / .BO
    """
    prev_closes = {}

    _ul_cache_pc = st.session_state.get("_ul_cmp_cache", {})

    def fetch_prev(row):
        t     = row["Ticker"]
        atype = row.get("Asset_Type", "Stock")
        if atype == "Unlisted Share":
            # Use same unlisted CMP as current price (no intraday data for unlisted)
            return t, fetch_unlisted_cmp(t, equity_map, nfo_map,
                                         bse_map=bse_map, session_cache=_ul_cache_pc)

        clean = t.replace(".NS", "").replace(".BO", "").upper()
        alias = ANGEL_ONE_ALIASES.get(clean, clean)
        bse   = is_bse_ticker(t)

        # ── 1. Angel One candle API ───────────────────────────────
        info = resolve_token(t, equity_map, nfo_map, atype, bse_map=bse_map)
        pc = _prev_close_angel_one(t, info)
        if pc and pc > 0:
            return t, pc

        # ── 2. NSE quote API (NSE stocks / dual-listed) ───────────
        if not bse:
            pc = _prev_close_nse_api(alias)
            if pc and pc > 0:
                return t, pc

        # ── 3. BSE API (for .BO tickers) ─────────────────────────
        if bse and info:
            bse_token = info.get("token", "")
            pc = _prev_close_bse_api(bse_token, alias)
            if pc and pc > 0:
                return t, pc

        # ── 4. yfinance fallback (most reliable, slowest) ─────────
        pc = _prev_close_yfinance(t)
        if pc and pc > 0:
            return t, pc

        return t, None

    rows = portfolio_df.to_dict("records")
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        for ticker, pc in ex.map(fetch_prev, rows):
            prev_closes[ticker] = pc

    return prev_closes

# =====================================================
# ── XIRR CALCULATION ─────────────────────────────────
# =====================================================

def xirr(cashflows):
    """
    Compute XIRR given a list of (date, amount) tuples.
    Negative = outflow (buy), Positive = inflow (sell/current value).
    Returns annualised rate as a float (e.g. 0.18 = 18%) or None.

    Uses Newton-Raphson with bisection fallback.
    Hard-clamps result to [-99%, +500%] to prevent absurd values from
    very short holding periods or near-zero cost basis.
    """
    if not cashflows or len(cashflows) < 2:
        return None
    try:
        dates   = [cf[0] for cf in cashflows]
        amounts = [cf[1] for cf in cashflows]
        t0 = min(dates)
        days = [(d - t0).days for d in dates]

        # Need at least some time span to compute a meaningful annualised rate
        max_days = max(days)
        if max_days < 1:
            return None

        # Require mixed signs (outflows + inflows) for a valid XIRR
        has_neg = any(a < 0 for a in amounts)
        has_pos = any(a > 0 for a in amounts)
        if not (has_neg and has_pos):
            return None

        def npv(rate):
            return sum(a / ((1 + rate) ** (d / 365.0))
                       for a, d in zip(amounts, days))

        def npv_deriv(rate):
            return sum(-a * (d / 365.0) / ((1 + rate) ** (d / 365.0 + 1))
                       for a, d in zip(amounts, days))

        # Newton-Raphson — fast convergence for well-behaved cases
        rate = 0.10  # start at 10% guess
        for _ in range(100):
            f  = npv(rate)
            fp = npv_deriv(rate)
            if fp == 0:
                break
            new_rate = rate - f / fp
            # Keep in sane bounds during iteration
            new_rate = max(-0.9999, min(new_rate, 49.0))
            if abs(new_rate - rate) < 1e-7:
                rate = new_rate
                break
            rate = new_rate

        # Verify convergence; if NPV not near zero, fall back to bisection
        if abs(npv(rate)) > 1.0:
            lo, hi = -0.9999, 49.0
            # Ensure bracket contains a root
            if npv(lo) * npv(hi) > 0:
                return None
            for _ in range(200):
                mid = (lo + hi) / 2.0
                val = npv(mid)
                if abs(val) < 1e-6:
                    rate = mid
                    break
                if val > 0:
                    lo = mid
                else:
                    hi = mid
            else:
                rate = (lo + hi) / 2.0

        # Final NPV sanity check
        if abs(npv(rate)) > 10.0:
            return None

        # Hard clamp: never show more than 500% or less than -99%
        # Short-hold artefacts (< 7 days) get annualised figures that are
        # mathematically correct but practically meaningless — cap those too.
        rate = max(-0.9999, min(rate, 5.0))

        # If holding period < 30 days, annualised XIRR is misleading;
        # cap display at ±200% annualised to avoid shock figures.
        if max_days < 30:
            rate = max(-0.9999, min(rate, 2.0))

        return rate

    except Exception:
        return None


def compute_xirr_per_holding(calc_df, trades_df):
    """Return dict {ticker -> xirr_%} for each active holding."""
    results = {}
    for _, row in calc_df.iterrows():
        ticker = row["Ticker"]
        try:
            buy_date = pd.to_datetime(row.get("Buy_Date", None), errors="coerce")
            if pd.isna(buy_date):
                buy_date = datetime.today() - timedelta(days=365)
            buy_date = buy_date.date() if hasattr(buy_date, "date") else buy_date

            cf = [( buy_date, -float(row["Cost_Basis"]) )]

            # Add sell cashflows if any
            if not trades_df.empty and ticker in trades_df["Ticker"].values:
                for _, tr in trades_df[trades_df["Ticker"] == ticker].iterrows():
                    sd = pd.to_datetime(tr["Sell_Date"], errors="coerce")
                    if not pd.isna(sd):
                        sd = sd.date() if hasattr(sd, "date") else sd
                        cf.append((sd, float(tr["Sell_Qty"]) * float(tr["Sell_Price"])))

            # Current value as terminal cashflow today
            cf.append((datetime.today().date(), float(row["Value"])))

            r = xirr(cf)
            results[ticker] = round(r * 100, 2) if r is not None else None
        except Exception:
            results[ticker] = None
    return results


# =====================================================
# ── CAPITAL GAINS (STCG / LTCG) ─────────────────────
# =====================================================

def compute_capital_gains(trades_df, portfolio_df):
    """
    Classify each sell trade as STCG or LTCG.
    LTCG: holding period > 365 days. STCG: ≤ 365 days.
    Returns a DataFrame with gain classification.
    """
    if trades_df.empty:
        return pd.DataFrame()

    rows = []
    for _, tr in trades_df.iterrows():
        ticker = tr["Ticker"]
        sell_date = pd.to_datetime(tr.get("Sell_Date"), errors="coerce")
        # Find earliest buy date for this ticker
        buy_rows = portfolio_df[portfolio_df["Ticker"] == ticker]
        if not buy_rows.empty:
            buy_date = pd.to_datetime(buy_rows["Buy_Date"].iloc[0], errors="coerce")
        else:
            buy_date = sell_date

        holding_days = (sell_date - buy_date).days if not pd.isna(sell_date) and not pd.isna(buy_date) else 0
        gain_type = "LTCG" if holding_days > 365 else "STCG"
        booked = float(tr.get("Booked_PnL", 0))
        rows.append({
            "Ticker":        ticker,
            "Buy Date":      buy_date.strftime("%d-%b-%Y") if not pd.isna(buy_date) else "—",
            "Sell Date":     sell_date.strftime("%d-%b-%Y") if not pd.isna(sell_date) else "—",
            "Holding Days":  holding_days,
            "Gain Type":     gain_type,
            "Sell Qty":      float(tr.get("Sell_Qty", 0)),
            "Avg Buy ₹":     float(tr.get("Buy_Price_At_Sell", 0)),
            "Sell Price ₹":  float(tr.get("Sell_Price", 0)),
            "Booked P&L ₹":  booked,
        })
    return pd.DataFrame(rows)


# =====================================================
# ── PORTFOLIO BETA ────────────────────────────────────
# =====================================================

def compute_portfolio_beta(calc_df):
    """
    Estimate weighted portfolio beta using a hardcoded beta table.
    Returns overall beta and per-ticker beta.
    """
    BETA_TABLE = {
        # High beta
        "TATAMOTORS": 1.45, "ZOMATO": 1.60, "NYKAA": 1.55, "PAYTM": 1.70,
        "VEDL": 1.35, "ADANIENT": 1.50, "ADANIPORTS": 1.30,
        # Medium-high beta
        "DRREDDY": 0.85, "GVPIL": 1.10, "INTERARCH": 1.15, "VOLTAMP": 1.20,
        "ENRIN": 1.10, "RELIANCE": 0.95, "INFY": 0.90, "TCS": 0.85,
        "HDFCBANK": 0.80, "ICICIBANK": 0.90, "SBIN": 1.10, "AXISBANK": 1.05,
        "BAJFINANCE": 1.15, "KOTAKBANK": 0.80, "MARUTI": 0.90, "M&M": 0.95,
        "ABB": 1.10, "SIEMENS": 1.05, "BHEL": 1.25, "NTPC": 0.75,
        "POWERGRID": 0.65, "COALINDIA": 0.80, "ONGC": 0.90, "BPCL": 0.95,
        "SUNPHARMA": 0.75, "CIPLA": 0.80, "HINDUNILVR": 0.55, "ITC": 0.65,
        "TATASTEEL": 1.30, "JSWSTEEL": 1.25, "HINDALCO": 1.20,
        # ETFs (low beta)
        "NIFTYBEES": 1.00, "JUNIORBEES": 1.05, "GOLDBEES": 0.10, "LIQUIDBEES": 0.01,
    }
    rows = []
    total_value = calc_df["Value"].sum()
    for _, row in calc_df.iterrows():
        ticker_clean = row["Ticker"].replace(".NS","").replace(".BO","").upper()
        beta = BETA_TABLE.get(ticker_clean, 1.0)   # default 1.0 if unknown
        weight = float(row["Value"]) / total_value if total_value else 0
        rows.append({"Ticker": row["Ticker"], "Beta": beta, "Weight": weight,
                     "Weighted_Beta": round(beta * weight, 4)})
    beta_df = pd.DataFrame(rows)
    portfolio_beta = round(beta_df["Weighted_Beta"].sum(), 3)
    return portfolio_beta, beta_df


# =====================================================
# ── MARKET INTELLIGENCE: HIGH-IMPACT EVENTS FOR PORTFOLIO STOCKS ─────
# =====================================================

# Keywords that signal HIGH-IMPACT corporate events worth showing in the banner
_IMPACT_KEYWORDS = [
    # M&A
    "merger", "amalgamation", "acquisition", "takeover", "scheme of arrangement",
    "demerger", "spin-off", "spinoff", "divestment", "divestiture",
    # Capital actions
    "stock split", "sub-division", "bonus issue", "bonus share", "rights issue",
    "buyback", "buy-back", "share repurchase",
    # Dividends
    "dividend", "interim dividend", "final dividend", "special dividend",
    # Orders & contracts
    "order win", "order received", "large order", "major order", "wins order",
    "order book", "new order", "secures order", "contract win", "contract awarded",
    "letter of intent", "loi", "mou", "memorandum of understanding",
    "work order", "purchase order", "supply agreement",
    # Deals & shareholding
    "block deal", "bulk deal", "stake sale", "promoter buying", "promoter selling",
    "open market purchase", "insider buy", "insider sell",
    # Management
    "management change", "ceo change", "md change", "board change", "new ceo",
    "appoints", "resigns", "new managing director", "cfo", "director",
    "key managerial", "kmp", "board of directors",
    # Regulatory
    "delisting", "open offer", "preferential allotment", "sebi", "enforcement",
    "court order", "tribunal", "nclt", "nclat", "cci", "notice", "penalty",
    "pledge", "invocation", "default", "npa",
    # Results
    "quarterly result", "q1 result", "q2 result", "q3 result", "q4 result",
    "financial result", "declares result", "reported result", "result declared",
    "posts result", "announces result", "fy26", "fy25", "fy24",
    "net profit", "pat ", " pat,", "profit after tax",
    "net loss", "revenue from operations", "total income", "total revenue",
    "ebitda", "operating profit", "gross margin",
    "profit up", "profit down", "revenue up", "revenue down",
    "profit rises", "profit falls", "profit jumps", "profit surges", "profit drops",
    "revenue rises", "revenue falls", "revenue jumps", "revenue grows",
    "beats estimate", "misses estimate", "beats expectation",
    "q1fy", "q2fy", "q3fy", "q4fy",
    "annual result", "half year", "h1fy", "h2fy",
    "earnings", "turnover", "topline", "bottomline",
    # Board meetings / upcoming events
    "board meeting", "agm", "egm", "annual general meeting",
    "board to consider", "considers dividend", "considers buyback",
    "record date", "ex-date", "ex dividend", "book closure",
    # Analyst
    "upgrade", "downgrade", "target price", "rating change", "target raised", "target cut",
    "initiates coverage", "buy rating", "sell rating", "overweight", "underweight",
    "outperform", "underperform", "neutral", "accumulate", "reduce",
    # Fundraise
    "fund raise", "qip", "ncd", "ipo", "fpo", "fundraise", "fund infusion",
    "debt raise", "equity raise", "private placement",
    # Expansion & capex
    "capex", "expansion", "new plant", "joint venture", " jv ", "greenfield",
    "brownfield", "capacity addition", "new facility", "inaugurates",
    "commissioning", "commercial production",
    # General corporate announcements
    "announces", "declares", "approves", "recommends", "informed bse", "informed nse",
    "outcome of board", "outcome of agm", "corporate action", "disclosure",
    "allotment", "listing",
    # Macro / sector
    "interest rate", "rbi policy", "gst", "customs duty", "import duty", "tariff",
]

# Only block truly generic market-level noise (not company-specific at all)
_BANNER_BLOCK_KEYWORDS = [
    "market opens", "market closes", "pre-open session",
    "nifty opens", "sensex opens", "nifty 50 today", "sensex today",
    "top gainers", "top losers", "most active stocks",
]



def compute_alerts(calc_df, total_value):
    """
    Legacy function kept for compatibility — returns empty list.
    The banner is now handled by fetch_portfolio_intelligence().
    """
    return []


# =====================================================

# ── HTML DASHBOARD REPORT GENERATOR ──────────────────
# =====================================================

def generate_html_report(client_name, calc_df, summary_dict, trades_df, booked_pnl_map):
    """
    EDITABLE HTML Portfolio Dashboard.
    All cells for Qty, Avg Buy Price, LTP are inline-editable.
    All derived columns (Buy Value, Mkt Value, Unrealized P&L, P&L%, Daily MTM)
    and all KPI summary cards recompute live in the browser via JavaScript.
    Unlisted shares are excluded from P&L totals (same logic as before).
    Returns UTF-8 bytes.
    """
    import json, math

    IST      = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    s        = summary_dict
    xirr_map = s.get("xirr_map", {})
    cg_df    = s.get("cap_gains_df", pd.DataFrame())

    def _inr(v, dec=0):
        try:    return f"₹{float(v):,.{dec}f}"
        except: return "—"
    def _safe(v, fmt=str):
        try:    return fmt(v) if v is not None and (not isinstance(v,float) or not pd.isna(v)) else "—"
        except: return "—"
    def _col(v):  return "#22d67b" if float(v or 0) >= 0 else "#f85454"
    def _sign(v): return "+" if float(v or 0) >= 0 else ""

    _UL = "Unlisted Share"
    _FNO_TYPES = {"F&O", "Index F&O", "Currency F&O"}

    if calc_df.empty:
        listed_df = pd.DataFrame()
    elif "Asset_Type" in calc_df.columns:
        listed_df = calc_df[calc_df["Asset_Type"] != _UL].copy()
    else:
        listed_df = calc_df.copy()

    def _seg(df, types):
        if df.empty or "Asset_Type" not in df.columns: return pd.DataFrame()
        if isinstance(types, str): types = [types]
        return df[df["Asset_Type"].isin(types)].copy()

    seg_stocks   = _seg(listed_df, ["Stock"])
    seg_etf      = _seg(listed_df, ["ETF", "Liquid ETF", "Commodity ETF", "International ETF"])
    seg_fno      = _seg(listed_df, list(_FNO_TYPES))
    seg_reit     = _seg(listed_df, ["REIT/InvIT"])
    seg_sgb      = _seg(listed_df, ["SGB", "Bond/NCD"])

    if calc_df.empty or "Asset_Type" not in calc_df.columns:
        seg_unlisted = pd.DataFrame()
    else:
        seg_unlisted = calc_df[calc_df["Asset_Type"] == _UL].copy()

    if trades_df.empty:
        listed_trades   = pd.DataFrame()
        unlisted_trades = pd.DataFrame()
    elif "Asset_Type" in trades_df.columns:
        listed_trades   = trades_df[trades_df["Asset_Type"] != _UL].copy()
        unlisted_trades = trades_df[trades_df["Asset_Type"] == _UL].copy()
    else:
        listed_trades   = trades_df.copy()
        unlisted_trades = pd.DataFrame()

    try:
        total_inv_f  = float(listed_df["Cost_Basis"].sum()) if not listed_df.empty else 1.0
        total_val_f  = float(listed_df["Value"].sum())       if not listed_df.empty else 0.0
        total_pnl_f  = float(listed_df["Unrealized_PnL"].sum()) if not listed_df.empty and "Unrealized_PnL" in listed_df.columns else 0.0
        daily_pnl_f  = float(listed_df["Daily_PnL"].sum())   if not listed_df.empty and "Daily_PnL" in listed_df.columns else 0.0
        pnl_pct_f    = (total_pnl_f / total_inv_f * 100) if total_inv_f else 0.0
        daily_pct_f  = (daily_pnl_f / total_inv_f * 100) if total_inv_f else 0.0

        if not listed_trades.empty and "Booked_PnL" in listed_trades.columns:
            booked_f    = float(listed_trades["Booked_PnL"].sum())
            n_trades    = len(listed_trades)
            win_trades  = listed_trades[listed_trades["Booked_PnL"] > 0]
            win_rate_f  = len(win_trades) / n_trades * 100 if n_trades else 0
            win_rate    = f"{win_rate_f:.1f}%"
        else:
            booked_f = 0.0; n_trades = 0; win_rate = "—"; win_rate_f = 0

        n_stocks  = len(listed_df)
        n_sectors = listed_df["Asset_Type"].nunique() if "Asset_Type" in listed_df.columns else 1

        beta_f = float(s.get("portfolio_beta", 1.0) or 1.0)
        if   beta_f >= 1.4: risk_label, risk_col = "VERY HIGH RISK", "#f85454"
        elif beta_f >= 1.2: risk_label, risk_col = "HIGH RISK",      "#ff8c42"
        elif beta_f >= 0.9: risk_label, risk_col = "MODERATE RISK",  "#f5c842"
        elif beta_f >= 0.7: risk_label, risk_col = "MODERATE LOW",   "#7ed956"
        else:               risk_label, risk_col = "LOW RISK",       "#22d67b"

        fno_inv_f = float(seg_fno["Cost_Basis"].sum()) if not seg_fno.empty else 0.0
        fno_pct   = (fno_inv_f / total_inv_f * 100) if total_inv_f else 0.0

        _eq_df = listed_df[~listed_df.get("Asset_Type", pd.Series(dtype=str)).isin(_FNO_TYPES)].copy() \
                 if "Asset_Type" in listed_df.columns else listed_df.copy()
        if not _eq_df.empty and "PnL_%" in _eq_df.columns:
            _pser = pd.to_numeric(_eq_df["PnL_%"], errors="coerce")
            _bi = _pser.idxmax(); _wi = _pser.idxmin()
            top_ticker   = str(_eq_df.loc[_bi, "Ticker"])            if pd.notna(_bi) else "—"
            top_pct_v    = float(_pser[_bi])                          if pd.notna(_bi) else 0.0
            top_pnl_v    = float(_eq_df.loc[_bi, "Unrealized_PnL"])  if pd.notna(_bi) else 0.0
            worst_ticker = str(_eq_df.loc[_wi, "Ticker"])            if pd.notna(_wi) else "—"
            worst_pct_v  = float(_pser[_wi])                          if pd.notna(_wi) else 0.0
            worst_pnl_v  = float(_eq_df.loc[_wi, "Unrealized_PnL"])  if pd.notna(_wi) else 0.0
        else:
            top_ticker = worst_ticker = "—"
            top_pct_v = worst_pct_v = top_pnl_v = worst_pnl_v = 0.0

        def _seg_stats(df):
            if df.empty: return {"inv":0,"val":0,"pnl":0,"dpnl":0,"n":0}
            return {
                "inv":  float(df["Cost_Basis"].sum())    if "Cost_Basis"    in df.columns else 0,
                "val":  float(df["Value"].sum())          if "Value"         in df.columns else 0,
                "pnl":  float(df["Unrealized_PnL"].sum())if "Unrealized_PnL"in df.columns else 0,
                "dpnl": float(df["Daily_PnL"].sum())     if "Daily_PnL"     in df.columns else 0,
                "n":    len(df),
            }
        ss  = _seg_stats(seg_stocks)
        se  = _seg_stats(seg_etf)
        sf  = _seg_stats(seg_fno)
        sr  = _seg_stats(seg_reit)
        sb  = _seg_stats(seg_sgb)
        sul = _seg_stats(seg_unlisted)

    except Exception as ex:
        risk_label="DATA ERROR"; risk_col="#888"; beta_f=1.0
        total_inv_f=1; total_val_f=0; total_pnl_f=0; daily_pnl_f=0
        pnl_pct_f=0; daily_pct_f=0; booked_f=0; fno_pct=0; fno_inv_f=0
        n_stocks=0; n_sectors=1; n_trades=0; win_rate="—"; win_rate_f=0
        top_ticker=worst_ticker="—"; top_pct_v=worst_pct_v=top_pnl_v=worst_pnl_v=0.0
        ss=se=sf=sr=sb=sul={"inv":0,"val":0,"pnl":0,"dpnl":0,"n":0}

    # ── AI calls — all fired concurrently ────────────────────────────────
    import urllib.request as _ureq, json as _json2
    from concurrent.futures import ThreadPoolExecutor, as_completed

    ANTHROPIC_API_KEY2 = ""
    try:    ANTHROPIC_API_KEY2 = st.secrets.get("ANTHROPIC_API_KEY", "")
    except: pass
    if not ANTHROPIC_API_KEY2:
        ANTHROPIC_API_KEY2 = os.environ.get("ANTHROPIC_API_KEY", "")

    _API_HDRS = {
        "x-api-key": ANTHROPIC_API_KEY2,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    def _call_api(payload_dict, timeout=8):
        req = _ureq.Request(
            "https://api.anthropic.com/v1/messages",
            data=_json2.dumps(payload_dict).encode(),
            headers=_API_HDRS, method="POST",
        )
        with _ureq.urlopen(req, timeout=timeout) as r:
            return _json2.loads(r.read())

    psa_rows_list = []
    if ANTHROPIC_API_KEY2 and not listed_df.empty:
        _eq_for_psa = listed_df[~listed_df.get("Asset_Type", pd.Series(dtype=str)).isin(_FNO_TYPES)].copy() \
                      if "Asset_Type" in listed_df.columns else listed_df.copy()
        for _, row in _eq_for_psa.head(12).iterrows():
            psa_rows_list.append({
                "tk":  str(row.get("Ticker", "")),
                "cmp": float(row.get("Current_Price", 0) or 0),
                "bp":  float(row.get("Buy_Price", 0) or 0),
                "pp":  float(row.get("PnL_%", 0) or 0),
                "at":  str(row.get("Asset_Type", "Stock")),
            })

    def _fetch_stock(meta):
        tk, cmp, bp, pp, at = meta["tk"], meta["cmp"], meta["bp"], meta["pp"], meta["at"]
        d = _call_api({"model":"claude-haiku-4-5-20251001","max_tokens":200,
            "messages":[{"role":"user","content":
                f"For {tk} (NSE India, {at}), CMP ₹{cmp:.0f}, avg buy ₹{bp:.0f}, P&L {pp:+.1f}%. "
                f"Respond ONLY with a JSON object (no markdown): "
                f'{{"view":"BUY/HOLD/SELL","short_thesis":"...","key_risk":"...",'
                f'"price_target":<number>,"stop_loss":<number>,"holding_period":"..."}}'}]}, timeout=6)
        txt = d.get("content",[{}])[0].get("text","").strip()
        if txt.startswith("```"): txt = txt.split("```")[1].lstrip("json").strip()
        obj  = _json2.loads(txt)
        view = obj.get("view","HOLD").upper()
        vc   = "#22d67b" if view=="BUY" else ("#f85454" if view=="SELL" else "#f5c842")
        pt   = float(obj.get("price_target", cmp*1.15))
        sl   = float(obj.get("stop_loss",    cmp*0.92))
        return {"ticker":tk,"asset_type":at,"view":view,"view_color":vc,
                "short_thesis":obj.get("short_thesis","—")[:120],
                "key_risk":obj.get("key_risk","—")[:100],
                "price_target":pt,"stop_loss":sl,"upside_pct":((pt-cmp)/cmp*100) if cmp else 0,
                "holding_period":obj.get("holding_period","Medium term"),
                "pnl_pct":pp,"cmp":cmp,"avg_buy":bp}

    psa_data   = []
    ai_summary = None

    if ANTHROPIC_API_KEY2:
        gainers_list = ", ".join(listed_df[listed_df["Unrealized_PnL"]>0].nlargest(5,"Unrealized_PnL")["Ticker"].tolist()) \
                       if not listed_df.empty and "Unrealized_PnL" in listed_df.columns else "—"
        losers_list  = ", ".join(listed_df[listed_df["Unrealized_PnL"]<0].nsmallest(5,"Unrealized_PnL")["Ticker"].tolist()) \
                       if not listed_df.empty and "Unrealized_PnL" in listed_df.columns else "—"
        summary_payload = {"model":"claude-haiku-4-5-20251001","max_tokens":400,
            "messages":[{"role":"user","content":
                f"Write a 3-sentence professional portfolio summary in HTML (use <strong> for emphasis). "
                f"Client: {client_name}. Invested: {_inr(total_inv_f)}. Value: {_inr(total_val_f)}. "
                f"Unrealized P&L: {_sign(total_pnl_f)}{_inr(total_pnl_f)} ({pnl_pct_f:+.2f}%). "
                f"Top gainers: {gainers_list}. Top losers: {losers_list}. "
                f"Win rate: {win_rate}. Beta: {beta_f:.2f}. Be concise and professional."}]}

        with ThreadPoolExecutor(max_workers=min(14, 1+len(psa_rows_list))) as pool:
            futures = {pool.submit(_call_api, summary_payload, 8): "__summary__"}
            for meta in psa_rows_list:
                futures[pool.submit(_fetch_stock, meta)] = meta["tk"]
            _order = {m["tk"]: i for i, m in enumerate(psa_rows_list)}
            for fut in as_completed(futures):
                key = futures[fut]
                try:
                    res = fut.result()
                    if key == "__summary__":
                        ai_summary = res.get("content",[{}])[0].get("text", None)
                    else:
                        psa_data.append(res)
                except Exception:
                    pass
        psa_data.sort(key=lambda x: _order.get(x["ticker"], 999))

    if not ai_summary:
        gainers_list = ", ".join(listed_df[listed_df["Unrealized_PnL"]>0].nlargest(5,"Unrealized_PnL")["Ticker"].tolist()) \
                       if not listed_df.empty and "Unrealized_PnL" in listed_df.columns else "—"
        losers_list  = ", ".join(listed_df[listed_df["Unrealized_PnL"]<0].nsmallest(5,"Unrealized_PnL")["Ticker"].tolist()) \
                       if not listed_df.empty and "Unrealized_PnL" in listed_df.columns else "—"
        ai_summary = (
            f'Portfolio of <strong>{n_stocks} listed positions</strong> across <strong>{n_sectors} asset classes</strong>. '
            f'Deployed capital: <strong>{_inr(total_inv_f)}</strong> · Current market value: <strong>{_inr(total_val_f)}</strong>.<br><br>'
            f'Unrealized P&amp;L: <strong style="color:{_col(total_pnl_f)};">{_sign(total_pnl_f)}{_inr(total_pnl_f)} ({pnl_pct_f:+.2f}%)</strong> · '
            f'Today\'s MTM: <strong style="color:{_col(daily_pnl_f)};">{_sign(daily_pnl_f)}{_inr(daily_pnl_f)}</strong>.<br><br>'
            f'Top gainers: <strong style="color:#22d67b;">{gainers_list}</strong> · Top losers: <strong style="color:#f85454;">{losers_list}</strong>.'
        )

    # ── Chart data ────────────────────────────────────────────────────────
    seg_labels = []; seg_inv = []; seg_colors_chart = []
    for lbl, stat, col in [
        ("Stocks",    ss, "#4f7ef8"), ("ETFs", se, "#22d67b"),
        ("F&O",       sf, "#f5c842"), ("REIT/InvIT", sr, "#a78bfa"),
        ("SGB/Bond",  sb, "#22c7d6"),
    ]:
        if stat["n"] > 0:
            seg_labels.append(lbl); seg_inv.append(round(stat["inv"],0)); seg_colors_chart.append(col)

    bar_tickers = []; bar_pnl = []; bar_colors_list = []
    if not listed_df.empty:
        for _, row in listed_df.sort_values("Unrealized_PnL", ascending=False).head(20).iterrows():
            pv = float(row.get("Unrealized_PnL",0) or 0)
            bar_tickers.append(str(row.get("Ticker","")))
            bar_pnl.append(round(pv,0))
            bar_colors_list.append("#22d67b" if pv >= 0 else "#f85454")

    donut_js  = json.dumps(seg_labels)
    inv_js    = json.dumps(seg_inv)
    dcol_js   = json.dumps(seg_colors_chart)
    bt_js     = json.dumps(bar_tickers)
    bp_js     = json.dumps(bar_pnl)
    bc_js     = json.dumps(bar_colors_list)

    # ── Build JS rows data for editable tables ────────────────────────────
    # Each row: [ticker, asset_type, qty, ltp, avg_buy, prev_close, xirr]
    # All derived values computed in JS: buy_value=qty*avg_buy, mkt_value=qty*ltp
    # unrealized_pnl=mkt_value-buy_value, pnl_pct=unrealized_pnl/buy_value*100
    # daily_mtm=qty*(ltp-prev_close)

    def _df_to_js_rows(df):
        rows = []
        if df.empty: return rows
        for _, row in df.iterrows():
            qty       = float(row.get("Shares", 0) or 0)
            ltp       = float(row.get("Current_Price", 0) or row.get("CMP", 0) or 0)
            avg_buy   = float(row.get("Buy_Price", 0) or 0)
            prev_close= ltp - float(row.get("Daily_PnL", 0) or 0) / qty if qty else ltp
            ticker    = str(row.get("Ticker","—"))
            asset_type= str(row.get("Asset_Type","Stock"))
            xv        = xirr_map.get(ticker, "")
            try: xd = f"{float(xv):.1f}%" if xv != "" else "—"
            except: xd = "—"
            rows.append({
                "ticker": ticker,
                "asset_type": asset_type,
                "qty": round(qty, 4),
                "ltp": round(ltp, 2),
                "avg_buy": round(avg_buy, 2),
                "prev_close": round(prev_close, 4),
                "xirr": xd,
            })
        return rows

    all_rows_js      = json.dumps(_df_to_js_rows(listed_df))
    stocks_rows_js   = json.dumps(_df_to_js_rows(seg_stocks))
    etf_rows_js      = json.dumps(_df_to_js_rows(seg_etf))
    fno_rows_js      = json.dumps(_df_to_js_rows(seg_fno))
    reit_rows_js     = json.dumps(_df_to_js_rows(seg_reit))
    sgb_rows_js      = json.dumps(_df_to_js_rows(seg_sgb))

    # Unlisted rows
    def _df_to_js_rows_unlisted(df):
        rows = []
        if df.empty: return rows
        for _, row in df.iterrows():
            qty     = float(row.get("Shares", 0) or 0)
            ltp     = float(row.get("Current_Price", 0) or row.get("CMP", 0) or 0)
            avg_buy = float(row.get("Buy_Price", 0) or 0)
            rows.append({
                "ticker": str(row.get("Ticker","—")),
                "asset_type": "Unlisted Share",
                "qty": round(qty, 4),
                "ltp": round(ltp, 2),
                "avg_buy": round(avg_buy, 2),
                "prev_close": round(ltp, 4),
                "xirr": "—",
                "has_cmp": ltp > 0,
            })
        return rows
    unlisted_rows_js = json.dumps(_df_to_js_rows_unlisted(seg_unlisted))

    # Booked trades (not editable, static cards)
    def _booked_cards_html(asset_types, label, color, trades):
        if trades.empty: return ""
        if "Asset_Type" not in trades.columns: return ""
        if isinstance(asset_types, str): asset_types = [asset_types]
        df = trades[trades["Asset_Type"].isin(asset_types)]
        if df.empty: return ""
        total = float(df["Booked_PnL"].sum())
        tc = "#22d67b" if total >= 0 else "#f85454"
        cards = ""
        for _, tr in df.sort_values("Sell_Date", ascending=False).iterrows():
            bv  = float(tr.get("Booked_PnL", 0) or 0)
            bc  = "#22d67b" if bv >= 0 else "#f85454"
            sp  = float(tr.get("Sell_Price", 0) or 0)
            ab  = float(tr.get("Buy_Price_At_Sell", 0) or 0)
            qty = float(tr.get("Sell_Qty", 0) or 0)
            sd  = str(tr.get("Sell_Date", "—"))
            tk  = str(tr.get("Ticker","—"))
            pct = ((sp - ab) / ab * 100) if ab > 0 else 0
            cards += f"""<div class="bk-card">
  <div style="width:3px;height:100%;background:{bc};border-radius:3px 0 0 3px;flex-shrink:0;"></div>
  <div style="flex:1;padding:0 14px;">
    <div class="bk-top"><span class="bk-ticker">{tk}</span>
      <div class="bk-pnl" style="color:{bc};">{'+'if bv>=0 else ''}₹{abs(bv):,.2f} <span style="font-size:10px;color:#636899;">({pct:+.1f}%)</span></div></div>
    <div class="bk-meta">
      <span>Qty: <b>{qty:,.0f}</b></span><span>Sell: <b>₹{sp:,.2f}</b></span>
      <span>Avg Buy: <b>₹{ab:,.2f}</b></span><span>Date: <b style="color:#8890b8;">{sd}</b></span>
    </div>
  </div></div>"""
        return f"""<div class="bk-section">
  <div class="bk-sec-hdr" style="border-left:3px solid {color};">
    <span style="color:{color};font-weight:800;">{label}</span>
    <span style="color:{tc};font-weight:800;margin-left:auto;">Total: {'+'if total>=0 else ''}₹{total:,.0f}</span>
  </div>
  <div class="bk-cards">{cards}</div></div>"""

    stcg_total = ltcg_total = 0
    if not cg_df.empty and "Gain Type" in cg_df.columns and "Booked P&L ₹" in cg_df.columns:
        stcg_total = float(cg_df[cg_df["Gain Type"]=="STCG"]["Booked P&L ₹"].sum())
        ltcg_total = float(cg_df[cg_df["Gain Type"]=="LTCG"]["Booked P&L ₹"].sum())

    booked_tot_col = "#22d67b" if booked_f >= 0 else "#f85454"

    # PSA cards HTML
    def _psa_cards_for(asset_types):
        if isinstance(asset_types, str): asset_types = [asset_types]
        items = [a for a in psa_data if str(a.get("asset_type","")) in asset_types]
        if not items: return ""
        cards = ""
        for a in items:
            ticker = a.get("ticker",""); view = a.get("view","HOLD"); vc = a.get("view_color","#4f7ef8")
            thesis = a.get("short_thesis","—"); risk = a.get("key_risk","—")
            pt = a.get("price_target",0); sl = a.get("stop_loss",0); up = a.get("upside_pct",0)
            hz = a.get("holding_period","—"); pp = a.get("pnl_pct",0); cmp = a.get("cmp",0); bp = a.get("avg_buy",0)
            pc = "#22d67b" if pp >= 0 else "#f85454"; uc = "#22d67b" if up >= 0 else "#f85454"
            screener = f"https://www.screener.in/company/{ticker}/"
            gfin     = f"https://www.google.com/search?q={requests.utils.quote(ticker+' NSE equity research 2025')}"
            cards += f"""<div class="psa-card">
  <div class="psa-top"><div><span class="psa-ticker">{ticker}</span>
    <span class="psa-pnl" style="color:{pc};">{'▲'if pp>=0 else '▼'}{abs(pp):.1f}%</span></div>
    <span class="psa-badge" style="background:{vc}22;color:{vc};border:1px solid {vc}44;">{view}</span></div>
  <div class="psa-grid4">
    <div class="psa-cell"><div class="psa-cl">CMP</div><div class="psa-cv">₹{cmp:,.2f}</div></div>
    <div class="psa-cell"><div class="psa-cl">Target</div><div class="psa-cv" style="color:{vc};">₹{pt:,.2f}</div></div>
    <div class="psa-cell"><div class="psa-cl">Stop Loss</div><div class="psa-cv" style="color:#f85454;">₹{sl:,.2f}</div></div>
    <div class="psa-cell"><div class="psa-cl">Upside</div><div class="psa-cv" style="color:{uc};">{'▲'if up>=0 else '▼'}{abs(up):.1f}%</div></div>
  </div>
  <div class="psa-thesis">📋 {thesis}</div><div class="psa-risk">⚠️ <b>Risk:</b> {risk}</div>
  <div class="psa-hz">🕐 {hz} · Avg Buy ₹{bp:,.2f}</div>
  <div class="psa-links">
    <a href="{screener}" target="_blank" class="psa-link" style="color:#7ea8ff;">Screener.in</a>
    <a href="{gfin}" target="_blank" class="psa-link" style="color:#22d67b;">Research</a>
  </div></div>"""
        return f'<div class="psa-cards-grid">{cards}</div>'

    # Unlisted KPI values
    ul_inv  = float(seg_unlisted["Cost_Basis"].sum())    if not seg_unlisted.empty and "Cost_Basis"    in seg_unlisted.columns else 0
    ul_val  = float(seg_unlisted["Value"].sum())          if not seg_unlisted.empty and "Value"         in seg_unlisted.columns else 0
    ul_pnl  = float(seg_unlisted["Unrealized_PnL"].sum())if not seg_unlisted.empty and "Unrealized_PnL"in seg_unlisted.columns else 0
    ul_bkd  = float(unlisted_trades["Booked_PnL"].sum()) if not unlisted_trades.empty and "Booked_PnL" in unlisted_trades.columns else 0
    ul_n    = len(seg_unlisted); ul_n_tr = len(unlisted_trades)

    # ── FULL HTML ──────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Portfolio Report — {client_name} — {IST.strftime('%d %b %Y')}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0;}}
:root{{
  --bg:#060812;--bg2:#0b0e1e;--bg3:#0f1228;--bg4:#141832;
  --border:#181d36;--border2:#1e2440;--border3:#263060;
  --blue:#4f7ef8;--green:#22d67b;--red:#f85454;--gold:#f5c842;
  --purple:#a78bfa;--cyan:#22c7d6;--orange:#ff8c42;
  --text:#eef0ff;--muted:#636899;--dim:#303560;
  --r:14px;--r-sm:8px;
}}
html{{scroll-behavior:smooth;}}
body{{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);
  min-height:100vh;overflow-x:hidden;
  background-image:
    radial-gradient(ellipse 1100px 700px at 0% 0%,rgba(79,126,248,0.10) 0%,transparent 55%),
    radial-gradient(ellipse 800px 600px at 100% 100%,rgba(167,139,250,0.08) 0%,transparent 55%);
}}
.hdr{{background:linear-gradient(150deg,#06091e 0%,#0b1132 45%,#080d28 75%,#050818 100%);
  border-bottom:1px solid rgba(79,126,248,0.18);padding:40px 56px 32px;position:relative;overflow:hidden;}}
.hdr::before{{content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,transparent,#4f7ef8 30%,#a78bfa 55%,#22c7d6 75%,transparent);}}
.hdr-brand{{font-size:10px;font-weight:800;color:var(--blue);letter-spacing:3px;text-transform:uppercase;margin-bottom:14px;}}
.hdr-title{{font-size:38px;font-weight:900;letter-spacing:-1.5px;
  background:linear-gradient(100deg,#a0c8ff,#c8b4ff,#80eeff,#a0ffd8);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:8px;}}
.hdr-sub{{font-size:14px;color:var(--muted);margin-bottom:20px;}}
.chip{{display:inline-block;padding:4px 12px;border-radius:20px;font-size:11px;font-weight:600;
  background:rgba(79,126,248,0.10);border:1px solid rgba(79,126,248,0.22);color:#9ab4f8;margin-right:8px;margin-bottom:4px;}}
.chip-btn{{cursor:pointer;}}.chip-btn:hover{{background:rgba(79,126,248,0.20);}}
.nav{{display:flex;flex-wrap:wrap;gap:4px;padding:12px 56px;
  background:#090c1e;border-bottom:1px solid var(--border);position:sticky;top:0;z-index:100;}}
.nt{{padding:7px 16px;border-radius:20px;font-size:12px;font-weight:600;cursor:pointer;
  color:var(--muted);transition:all 0.15s;white-space:nowrap;}}
.nt:hover{{background:rgba(79,126,248,0.10);color:var(--text);}}
.nt.on{{background:var(--blue);color:#fff;box-shadow:0 0 16px rgba(79,126,248,0.35);}}
.main{{padding:24px 48px 60px;max-width:1600px;margin:0 auto;}}
.sec{{display:none;}}.sec.on{{display:block;}}
.sh{{display:flex;align-items:center;gap:14px;margin-bottom:24px;flex-wrap:wrap;}}
.sh-icon{{width:42px;height:42px;border-radius:var(--r-sm);background:rgba(79,126,248,0.12);
  display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0;}}
.sh-title{{font-size:20px;font-weight:800;color:var(--text);}}
.sh-sub{{font-size:12px;color:var(--muted);margin-top:2px;}}
.seg-pnl-pill{{padding:8px 18px;border-radius:10px;font-size:14px;font-weight:700;margin-left:auto;}}
.ga{{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:14px;margin-bottom:22px;}}
.kc{{position:relative;border-radius:var(--r);background:var(--bg3);border:1px solid var(--border2);
  padding:20px 20px 16px;overflow:hidden;transition:transform 0.15s;}}
.kc:hover{{transform:translateY(-2px);}}
.kc-top{{position:absolute;top:0;left:0;right:0;height:2px;background:var(--kline,var(--blue));opacity:0.9;}}
.kc-glow{{position:absolute;top:0;left:0;right:0;bottom:0;background:var(--kglow,transparent);pointer-events:none;}}
.kc-icon{{font-size:20px;margin-bottom:10px;}}
.kc-lbl{{font-size:11px;color:var(--muted);font-weight:600;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:6px;}}
.kc-val{{font-size:20px;font-weight:900;line-height:1;}}
.kc-sub{{font-size:12px;font-weight:700;margin-top:4px;}}
.divider{{height:1px;background:var(--border);margin:20px 0;}}
/* ── EDITABLE TABLE ── */
.ltp-wrap{{overflow-x:auto;border-radius:var(--r);border:1px solid var(--border2);margin-bottom:22px;}}
.ltp-table{{width:100%;border-collapse:collapse;font-size:13px;}}
.ltp-table thead tr{{background:linear-gradient(90deg,#0d1130,#111535);}}
.ltp-table th{{padding:10px 14px;font-size:10px;font-weight:800;color:var(--muted);
  text-transform:uppercase;letter-spacing:0.7px;white-space:nowrap;border-bottom:1px solid var(--border2);text-align:right;}}
.ltp-table th:first-child,.ltp-table th:nth-child(2){{text-align:left;}}
.ltp-row{{transition:background 0.1s;border-bottom:1px solid var(--border);}}
.ltp-row:hover{{background:rgba(79,126,248,0.06);}}
.ltp-ticker{{padding:10px 14px;font-weight:800;color:var(--text);font-size:13px;white-space:nowrap;}}
.ltp-atype{{padding:10px 8px;font-size:10px;color:var(--muted);white-space:nowrap;}}
.ltp-num{{padding:10px 14px;text-align:right;white-space:nowrap;color:#c0c4e0;font-size:13px;}}
/* Editable cells */
.editable-cell{{
  padding:6px 10px;border-radius:6px;min-width:80px;text-align:right;
  background:rgba(79,126,248,0.06);border:1px solid rgba(79,126,248,0.20);
  color:#a0c4ff;font-weight:700;font-size:13px;font-family:'Inter',sans-serif;
  cursor:pointer;transition:border-color 0.15s,background 0.15s;
  outline:none;width:90px;
}}
.editable-cell:hover{{background:rgba(79,126,248,0.12);border-color:rgba(79,126,248,0.45);}}
.editable-cell:focus{{background:rgba(79,126,248,0.18);border-color:#4f7ef8;box-shadow:0 0 0 2px rgba(79,126,248,0.20);color:#fff;}}
.edit-hint{{font-size:10px;color:var(--muted);margin-bottom:8px;padding:6px 12px;
  background:rgba(79,126,248,0.06);border-radius:6px;border-left:2px solid rgba(79,126,248,0.3);
  display:inline-block;}}
.reset-btn{{padding:5px 14px;border-radius:20px;border:1px solid rgba(79,126,248,0.3);
  background:rgba(79,126,248,0.08);color:#7ea8ff;font-size:11px;font-weight:700;cursor:pointer;
  transition:all 0.15s;margin-left:10px;}}
.reset-btn:hover{{background:rgba(79,126,248,0.20);border-color:#4f7ef8;}}
.seg-bar{{display:flex;align-items:center;flex-wrap:wrap;gap:20px;
  padding:12px 18px;background:var(--bg2);border-radius:var(--r-sm);margin-bottom:16px;}}
.seg-bar-label{{font-size:13px;font-weight:800;min-width:100px;}}
.seg-metric{{display:flex;flex-direction:column;gap:2px;}}
.seg-mlbl{{font-size:9px;text-transform:uppercase;letter-spacing:0.7px;color:var(--muted);font-weight:700;}}
.seg-mval{{font-size:13px;font-weight:700;color:var(--text);}}
.psa-sec-hdr{{font-size:14px;font-weight:800;color:var(--muted);margin:24px 0 12px;
  text-transform:uppercase;letter-spacing:0.7px;padding-left:12px;
  border-left:3px solid rgba(79,126,248,0.4);}}
.psa-cards-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:14px;margin-bottom:22px;}}
.psa-card{{background:var(--bg3);border:1px solid var(--border2);border-radius:var(--r);padding:18px;transition:transform 0.15s;}}
.psa-card:hover{{transform:translateY(-2px);}}
.psa-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;}}
.psa-ticker{{font-size:15px;font-weight:800;color:var(--text);}}
.psa-pnl{{font-size:12px;font-weight:700;margin-left:8px;}}
.psa-badge{{font-size:10px;font-weight:800;padding:3px 10px;border-radius:20px;letter-spacing:0.5px;}}
.psa-grid4{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;background:var(--bg2);border-radius:var(--r-sm);padding:10px 8px;margin-bottom:10px;}}
.psa-cell{{text-align:center;}}
.psa-cl{{font-size:9px;color:var(--muted);text-transform:uppercase;letter-spacing:0.6px;font-weight:700;margin-bottom:3px;}}
.psa-cv{{font-size:12px;font-weight:800;}}
.psa-thesis{{font-size:12px;color:#c0c4e0;padding:8px 10px;background:var(--bg2);border-radius:6px;border-left:2px solid #2a2e52;margin-bottom:7px;line-height:1.6;}}
.psa-risk{{font-size:11px;color:#9098bc;padding:7px 10px;background:rgba(248,84,84,0.05);border-radius:6px;border-left:2px solid rgba(248,84,84,0.3);margin-bottom:6px;}}
.psa-hz{{font-size:11px;color:#7a80a8;padding:5px 10px;background:rgba(245,200,66,0.04);border-radius:6px;border-left:2px solid rgba(245,200,66,0.2);margin-bottom:12px;}}
.psa-links{{display:flex;gap:8px;}}
.psa-link{{font-size:11px;font-weight:700;padding:5px 14px;border-radius:20px;text-decoration:none;background:rgba(79,126,248,0.10);border:1px solid rgba(79,126,248,0.25);transition:opacity 0.15s;}}
.psa-link:hover{{opacity:0.7;}}
.ai-card{{background:var(--bg3);border:1px solid var(--border2);border-radius:var(--r);padding:28px 32px;margin-bottom:22px;border-left:3px solid var(--blue);}}
.ai-label{{font-size:11px;color:var(--blue);font-weight:800;text-transform:uppercase;letter-spacing:1px;margin-bottom:14px;}}
.ai-text{{font-size:13.5px;line-height:1.85;color:#c0c4e0;}}
.bk-summary-strip{{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:12px;margin-bottom:24px;}}
.bk-sum-item{{background:var(--bg3);border:1px solid var(--border2);border-radius:var(--r-sm);padding:16px 18px;}}
.bk-sum-lbl{{font-size:10px;text-transform:uppercase;letter-spacing:0.7px;color:var(--muted);font-weight:700;margin-bottom:6px;}}
.bk-sum-val{{font-size:18px;font-weight:900;}}
.bk-section{{margin-bottom:24px;}}
.bk-sec-hdr{{display:flex;align-items:center;padding:10px 16px;background:var(--bg2);border-radius:var(--r-sm);margin-bottom:12px;font-size:13px;}}
.bk-cards{{display:flex;flex-direction:column;gap:8px;}}
.bk-card{{display:flex;background:var(--bg3);border:1px solid var(--border2);border-radius:var(--r-sm);overflow:hidden;transition:box-shadow 0.15s;}}
.bk-card:hover{{box-shadow:0 4px 16px rgba(0,0,0,0.3);}}
.bk-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;}}
.bk-ticker{{font-size:14px;font-weight:800;color:var(--text);}}
.bk-pnl{{font-size:15px;font-weight:800;}}
.bk-meta{{display:flex;gap:16px;flex-wrap:wrap;font-size:11px;color:var(--muted);}}
.bk-meta b{{color:var(--text);}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:22px;}}
.card{{background:var(--bg3);border:1px solid var(--border2);border-radius:var(--r);padding:22px 24px;}}
.card-title{{font-size:12px;font-weight:800;color:var(--muted);text-transform:uppercase;letter-spacing:0.8px;margin-bottom:14px;}}
.disclaimer{{font-size:11px;color:#454870;padding:14px 18px;background:var(--bg2);border-radius:var(--r-sm);border:1px solid var(--border);margin-top:32px;line-height:1.7;}}
@media(max-width:768px){{
  .hdr{{padding:24px 20px;}}.main{{padding:16px 12px 48px;}}.nav{{padding:10px 12px;gap:6px;}}
  .ga{{grid-template-columns:repeat(2,1fr);}}.g2{{grid-template-columns:1fr;}}
  .psa-cards-grid{{grid-template-columns:1fr;}}.psa-grid4{{grid-template-columns:repeat(2,1fr);}}
  .bk-summary-strip{{grid-template-columns:repeat(2,1fr);}}
}}
@media print{{.nav{{display:none;}}.sec{{display:block!important;}}body{{background:#fff;color:#000;}}}}
</style>
</head>
<body>

<div class="hdr">
  <div class="hdr-brand">Northeast Broking Services Limited · Portfolio Report</div>
  <div class="hdr-title">{client_name}</div>
  <div class="hdr-sub">Generated {IST.strftime('%d %b %Y, %I:%M %p IST')} · Listed Securities Only · Unlisted excluded from P&L · <span style="color:#22d67b;font-weight:700;">✏️ Editable — click any blue cell to modify</span></div>
  <div>
    <span class="chip">📅 {IST.strftime('%d %b %Y')}</span>
    <span class="chip">👤 {client_name}</span>
    <span class="chip" id="hdr-holdings-chip">📊 {n_stocks} Listed Holdings</span>
    <span class="chip chip-btn" onclick="window.print()">🖨️ Print / PDF</span>
    <span class="chip chip-btn" id="save-btn" onclick="saveEdits()" style="background:rgba(34,214,123,0.12);border-color:rgba(34,214,123,0.35);color:#22d67b;">💾 Save Edits</span>
    <span class="chip chip-btn" onclick="resetAll()">↩️ Reset All Edits</span>
  </div>
</div>

<nav class="nav">
  <div class="nt on"  onclick="sw('ov',this)">📊 Overview &amp; AI</div>
  <div class="nt"     onclick="sw('stocks',this)">📈 Stocks</div>
  <div class="nt"     onclick="sw('etf',this)">📦 ETFs</div>
  <div class="nt"     onclick="sw('fno',this)">🎯 F&amp;O</div>
  <div class="nt"     onclick="sw('reit',this)">🏢 REIT/InvIT</div>
  <div class="nt"     onclick="sw('sgb',this)">🥇 SGB/Bond</div>
  <div class="nt"     onclick="sw('unlisted',this)">🔒 Unlisted</div>
  <div class="nt"     onclick="sw('booked',this)">🏦 Booked P&amp;L</div>
</nav>

<main class="main">

<!-- ████ OVERVIEW ████ -->
<section class="sec on" id="sec-ov">
  <div class="sh">
    <div class="sh-icon">📊</div>
    <div>
      <div class="sh-title">Portfolio Overview — Listed Securities</div>
      <div class="sh-sub">Unlisted shares excluded from all figures below · <span style="color:#22d67b;">✏️ Edit Qty / Avg Buy / LTP in any tab — all values update live</span></div>
    </div>
  </div>

  <!-- KPI cards — all bound to JS, update on edit -->
  <div class="ga" id="ov-kpi-grid">
    <div class="kc" style="--kline:linear-gradient(90deg,#4f7ef8,#7ea8ff);--kglow:rgba(79,126,248,0.08);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">💰</div><div class="kc-lbl">Total Invested</div>
      <div class="kc-val" id="kpi-invested" style="color:#7ea8ff;">₹{total_inv_f:,.0f}</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,#22d67b,#4fffaa);--kglow:rgba(34,214,123,0.08);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">📈</div><div class="kc-lbl">Current Value</div>
      <div class="kc-val" id="kpi-value" style="color:#22d67b;">₹{total_val_f:,.0f}</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,{_col(total_pnl_f)},{_col(total_pnl_f)}88);--kglow:{_col(total_pnl_f)}12;">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">{'📉' if total_pnl_f<0 else '📈'}</div><div class="kc-lbl">Unrealized P&amp;L</div>
      <div class="kc-val" id="kpi-upnl" style="color:{_col(total_pnl_f)};">{_sign(total_pnl_f)}₹{abs(total_pnl_f):,.0f}</div>
      <div class="kc-sub" id="kpi-upnl-pct" style="color:{_col(total_pnl_f)};">{pnl_pct_f:+.2f}%</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,{_col(daily_pnl_f)},{_col(daily_pnl_f)}88);--kglow:{_col(daily_pnl_f)}12;">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">⚡</div><div class="kc-lbl">Today's MTM</div>
      <div class="kc-val" id="kpi-daily" style="color:{_col(daily_pnl_f)};">{_sign(daily_pnl_f)}₹{abs(daily_pnl_f):,.0f}</div>
      <div class="kc-sub" id="kpi-daily-pct" style="color:{_col(daily_pnl_f)};">{daily_pct_f:+.2f}%</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,{_col(booked_f)},{_col(booked_f)}88);--kglow:{_col(booked_f)}12;">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">🏦</div><div class="kc-lbl">Booked P&amp;L (Listed)</div>
      <div class="kc-val" style="color:{_col(booked_f)};">{_sign(booked_f)}₹{abs(booked_f):,.0f}</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,{risk_col},{risk_col}88);--kglow:{risk_col}12;">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">⚖️</div><div class="kc-lbl">Portfolio Beta</div>
      <div class="kc-val" style="color:{risk_col};">{beta_f:.2f}</div>
      <div class="kc-sub" style="color:{risk_col};">{risk_label}</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,{'#f5c842' if fno_pct>10 else '#22d67b'},{'#f5c842' if fno_pct>10 else '#22d67b'}88);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">🎯</div><div class="kc-lbl">F&amp;O Exposure</div>
      <div class="kc-val" style="color:{'#f5c842' if fno_pct>10 else '#22d67b'};">{fno_pct:.1f}%</div>
      <div class="kc-sub" style="color:#636899;">of capital</div>
    </div>
  </div>

  <!-- Segment summary bars — updated live -->
  <div style="margin-bottom:22px;" id="ov-seg-bars">
    <div style="font-size:11px;color:var(--muted);font-weight:700;text-transform:uppercase;letter-spacing:0.8px;margin-bottom:10px;">Asset Class Breakdown</div>
    <div class="seg-bar" id="seg-bar-stocks" style="border-left:3px solid #4f7ef8;">
      <span class="seg-bar-label" style="color:#4f7ef8;">📊 Stocks</span>
      <span class="seg-metric"><span class="seg-mlbl">Positions</span><span class="seg-mval" id="sb-stocks-n" style="color:#4f7ef8;">{ss['n']}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Invested</span><span class="seg-mval" id="sb-stocks-inv">₹{ss['inv']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Mkt Value</span><span class="seg-mval" id="sb-stocks-val">₹{ss['val']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Unreal P&amp;L</span><span class="seg-mval" id="sb-stocks-pnl" style="color:{'#22d67b' if ss['pnl']>=0 else '#f85454'};">{'+'if ss['pnl']>=0 else ''}₹{ss['pnl']:,.0f}</span></span>
    </div>
    <div class="seg-bar" id="seg-bar-etf" style="border-left:3px solid #22d67b;">
      <span class="seg-bar-label" style="color:#22d67b;">📦 ETFs</span>
      <span class="seg-metric"><span class="seg-mlbl">Positions</span><span class="seg-mval" id="sb-etf-n" style="color:#22d67b;">{se['n']}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Invested</span><span class="seg-mval" id="sb-etf-inv">₹{se['inv']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Mkt Value</span><span class="seg-mval" id="sb-etf-val">₹{se['val']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Unreal P&amp;L</span><span class="seg-mval" id="sb-etf-pnl" style="color:{'#22d67b' if se['pnl']>=0 else '#f85454'};">{'+'if se['pnl']>=0 else ''}₹{se['pnl']:,.0f}</span></span>
    </div>
    <div class="seg-bar" id="seg-bar-fno" style="border-left:3px solid #f5c842;">
      <span class="seg-bar-label" style="color:#f5c842;">🎯 F&amp;O</span>
      <span class="seg-metric"><span class="seg-mlbl">Positions</span><span class="seg-mval" id="sb-fno-n" style="color:#f5c842;">{sf['n']}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Invested</span><span class="seg-mval" id="sb-fno-inv">₹{sf['inv']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Mkt Value</span><span class="seg-mval" id="sb-fno-val">₹{sf['val']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Unreal P&amp;L</span><span class="seg-mval" id="sb-fno-pnl" style="color:{'#22d67b' if sf['pnl']>=0 else '#f85454'};">{'+'if sf['pnl']>=0 else ''}₹{sf['pnl']:,.0f}</span></span>
    </div>
    <div class="seg-bar" id="seg-bar-reit" style="border-left:3px solid #a78bfa;">
      <span class="seg-bar-label" style="color:#a78bfa;">🏢 REIT/InvIT</span>
      <span class="seg-metric"><span class="seg-mlbl">Positions</span><span class="seg-mval" id="sb-reit-n" style="color:#a78bfa;">{sr['n']}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Invested</span><span class="seg-mval" id="sb-reit-inv">₹{sr['inv']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Mkt Value</span><span class="seg-mval" id="sb-reit-val">₹{sr['val']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Unreal P&amp;L</span><span class="seg-mval" id="sb-reit-pnl" style="color:{'#22d67b' if sr['pnl']>=0 else '#f85454'};">{'+'if sr['pnl']>=0 else ''}₹{sr['pnl']:,.0f}</span></span>
    </div>
    <div class="seg-bar" id="seg-bar-sgb" style="border-left:3px solid #22c7d6;">
      <span class="seg-bar-label" style="color:#22c7d6;">🥇 SGB/Bond</span>
      <span class="seg-metric"><span class="seg-mlbl">Positions</span><span class="seg-mval" id="sb-sgb-n" style="color:#22c7d6;">{sb['n']}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Invested</span><span class="seg-mval" id="sb-sgb-inv">₹{sb['inv']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Mkt Value</span><span class="seg-mval" id="sb-sgb-val">₹{sb['val']:,.0f}</span></span>
      <span class="seg-metric"><span class="seg-mlbl">Unreal P&amp;L</span><span class="seg-mval" id="sb-sgb-pnl" style="color:{'#22d67b' if sb['pnl']>=0 else '#f85454'};">{'+'if sb['pnl']>=0 else ''}₹{sb['pnl']:,.0f}</span></span>
    </div>
  </div>

  <!-- Charts -->
  <div class="g2" style="margin-bottom:22px;">
    <div class="card">
      <div class="card-title">Asset Allocation (Invested Capital)</div>
      <div style="height:200px;position:relative;"><canvas id="donutChart" role="img" aria-label="Asset allocation donut chart">Asset allocation by invested capital.</canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Unrealized P&amp;L by Holding (Top 20 Listed)</div>
      <div style="height:240px;position:relative;"><canvas id="pnlBar" role="img" aria-label="Bar chart of unrealized P&L per holding">Unrealized P&L by holding.</canvas></div>
    </div>
  </div>

  <!-- Top / Worst performers -->
  <div class="g2" style="margin-bottom:22px;">
    <div class="card" style="border-left:3px solid #22d67b;">
      <div class="card-title" style="color:#22d67b;">🚀 Top Performer (Equity)</div>
      <div style="font-size:28px;font-weight:900;color:#22d67b;margin:8px 0;" id="top-ticker">{top_ticker}</div>
      <div style="font-size:16px;font-weight:700;color:#22d67b;" id="top-stats">{top_pct_v:+.2f}% · {_inr(top_pnl_v)} gain</div>
    </div>
    <div class="card" style="border-left:3px solid #f85454;">
      <div class="card-title" style="color:#f85454;">📉 Worst Performer (Equity)</div>
      <div style="font-size:28px;font-weight:900;color:#f85454;margin:8px 0;" id="worst-ticker">{worst_ticker}</div>
      <div style="font-size:16px;font-weight:700;color:#f85454;" id="worst-stats">{worst_pct_v:+.2f}% · {_inr(worst_pnl_v)} loss</div>
    </div>
  </div>

  <div class="ai-card">
    <div class="ai-label">🤖 AI Portfolio Summary — Northeast Analytics Engine</div>
    <div class="ai-text">{ai_summary}</div>
  </div>

  <div class="disclaimer">
    ⚠️ <b>Disclaimer:</b> All figures include listed securities only. Unlisted shares are excluded.
    Data is for informational purposes only and does not constitute investment advice.
    Northeast Broking Services Limited is not responsible for data discrepancies.
  </div>
</section>

<!-- ████ STOCKS ████ -->
<section class="sec" id="sec-stocks">
  <div class="sh">
    <div class="sh-icon">📈</div>
    <div><div class="sh-title">📊 Stocks</div>
      <div class="sh-sub" id="seg-sub-stocks">{ss['n']} positions · Invested ₹{ss['inv']:,.0f} · Mkt Value ₹{ss['val']:,.0f}</div></div>
    <div class="seg-pnl-pill" id="seg-pill-stocks"
      style="background:{'rgba(34,214,123,0.10)' if ss['pnl']>=0 else 'rgba(248,84,84,0.10)'};color:{'#22d67b' if ss['pnl']>=0 else '#f85454'};border:1px solid {'#22d67b44' if ss['pnl']>=0 else '#f8545444'};">
      {'+'if ss['pnl']>=0 else ''}₹{ss['pnl']:,.0f}</div>
  </div>
  <div class="edit-hint">✏️ Edit Qty, Avg Buy, or LTP — all formulas recalculate live</div>
  <button class="reset-btn" onclick="resetSegment('stocks')">↩️ Reset Stocks</button><br><br>
  <div id="tbl-stocks"></div>
  {"<div class='psa-sec-hdr'>🔬 AI Stock Analysis</div>" + _psa_cards_for(["Stock"]) if psa_data else ""}
</section>

<!-- ████ ETFs ████ -->
<section class="sec" id="sec-etf">
  <div class="sh">
    <div class="sh-icon">📦</div>
    <div><div class="sh-title">📦 ETFs</div>
      <div class="sh-sub" id="seg-sub-etf">{se['n']} positions · Invested ₹{se['inv']:,.0f} · Mkt Value ₹{se['val']:,.0f}</div></div>
    <div class="seg-pnl-pill" id="seg-pill-etf"
      style="background:{'rgba(34,214,123,0.10)' if se['pnl']>=0 else 'rgba(248,84,84,0.10)'};color:{'#22d67b' if se['pnl']>=0 else '#f85454'};border:1px solid {'#22d67b44' if se['pnl']>=0 else '#f8545444'};">
      {'+'if se['pnl']>=0 else ''}₹{se['pnl']:,.0f}</div>
  </div>
  <div class="edit-hint">✏️ Edit Qty, Avg Buy, or LTP — all formulas recalculate live</div>
  <button class="reset-btn" onclick="resetSegment('etf')">↩️ Reset ETFs</button><br><br>
  <div id="tbl-etf"></div>
  {"<div class='psa-sec-hdr'>🔬 AI ETF Analysis</div>" + _psa_cards_for(["ETF","Liquid ETF","Commodity ETF","International ETF"]) if psa_data else ""}
</section>

<!-- ████ F&O ████ -->
<section class="sec" id="sec-fno">
  <div class="sh">
    <div class="sh-icon">🎯</div>
    <div><div class="sh-title">🎯 F&amp;O / Derivatives</div>
      <div class="sh-sub" id="seg-sub-fno">{sf['n']} positions · Invested ₹{sf['inv']:,.0f} · Mkt Value ₹{sf['val']:,.0f}</div></div>
    <div class="seg-pnl-pill" id="seg-pill-fno"
      style="background:{'rgba(34,214,123,0.10)' if sf['pnl']>=0 else 'rgba(248,84,84,0.10)'};color:{'#22d67b' if sf['pnl']>=0 else '#f85454'};border:1px solid {'#22d67b44' if sf['pnl']>=0 else '#f8545444'};">
      {'+'if sf['pnl']>=0 else ''}₹{sf['pnl']:,.0f}</div>
  </div>
  <div class="edit-hint">✏️ Edit Qty, Avg Buy, or LTP — all formulas recalculate live</div>
  <button class="reset-btn" onclick="resetSegment('fno')">↩️ Reset F&amp;O</button><br><br>
  <div id="tbl-fno"></div>
</section>

<!-- ████ REIT ████ -->
<section class="sec" id="sec-reit">
  <div class="sh">
    <div class="sh-icon">🏢</div>
    <div><div class="sh-title">🏢 REIT / InvIT</div>
      <div class="sh-sub" id="seg-sub-reit">{sr['n']} positions · Invested ₹{sr['inv']:,.0f} · Mkt Value ₹{sr['val']:,.0f}</div></div>
    <div class="seg-pnl-pill" id="seg-pill-reit"
      style="background:{'rgba(34,214,123,0.10)' if sr['pnl']>=0 else 'rgba(248,84,84,0.10)'};color:{'#22d67b' if sr['pnl']>=0 else '#f85454'};border:1px solid {'#22d67b44' if sr['pnl']>=0 else '#f8545444'};">
      {'+'if sr['pnl']>=0 else ''}₹{sr['pnl']:,.0f}</div>
  </div>
  <div class="edit-hint">✏️ Edit Qty, Avg Buy, or LTP — all formulas recalculate live</div>
  <button class="reset-btn" onclick="resetSegment('reit')">↩️ Reset REIT</button><br><br>
  <div id="tbl-reit"></div>
</section>

<!-- ████ SGB ████ -->
<section class="sec" id="sec-sgb">
  <div class="sh">
    <div class="sh-icon">🥇</div>
    <div><div class="sh-title">🥇 SGB / Bond</div>
      <div class="sh-sub" id="seg-sub-sgb">{sb['n']} positions · Invested ₹{sb['inv']:,.0f} · Mkt Value ₹{sb['val']:,.0f}</div></div>
    <div class="seg-pnl-pill" id="seg-pill-sgb"
      style="background:{'rgba(34,214,123,0.10)' if sb['pnl']>=0 else 'rgba(248,84,84,0.10)'};color:{'#22d67b' if sb['pnl']>=0 else '#f85454'};border:1px solid {'#22d67b44' if sb['pnl']>=0 else '#f8545444'};">
      {'+'if sb['pnl']>=0 else ''}₹{sb['pnl']:,.0f}</div>
  </div>
  <div class="edit-hint">✏️ Edit Qty, Avg Buy, or LTP — all formulas recalculate live</div>
  <button class="reset-btn" onclick="resetSegment('sgb')">↩️ Reset SGB/Bond</button><br><br>
  <div id="tbl-sgb"></div>
</section>

<!-- ████ UNLISTED ████ -->
<section class="sec" id="sec-unlisted">
  <div class="sh">
    <div class="sh-icon">🔒</div>
    <div><div class="sh-title">Unlisted Shares</div>
      <div class="sh-sub">{ul_n} pre-IPO / off-market positions · CMP sourced from Planify.in · indicative only</div></div>
  </div>
  <div style="background:rgba(167,139,250,0.08);border:1px solid rgba(167,139,250,0.25);border-radius:10px;padding:14px 20px;margin-bottom:22px;display:flex;align-items:center;gap:14px;flex-wrap:wrap;">
    <div style="font-size:22px;">🚫</div>
    <div>
      <div style="font-size:13px;font-weight:800;color:#c4b5fd;margin-bottom:4px;">Excluded from all P&amp;L calculations</div>
      <div style="font-size:12px;color:#7a80a8;line-height:1.6;">Unlisted share values are <b>indicative only</b> and are <b>not included</b> in portfolio totals, unrealized P&amp;L, booked P&amp;L, or any metric on the Overview tab.</div>
    </div>
  </div>
  <div class="ga" style="margin-bottom:22px;">
    <div class="kc" style="--kline:linear-gradient(90deg,#a78bfa,#7c3aed);--kglow:rgba(167,139,250,0.10);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">📦</div><div class="kc-lbl">Holdings</div>
      <div class="kc-val" style="color:#c4b5fd;">{ul_n}</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,#a78bfa,#7c3aed);--kglow:rgba(167,139,250,0.08);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">💰</div><div class="kc-lbl">Invested</div>
      <div class="kc-val" id="ul-kpi-inv" style="color:#c4b5fd;">{_inr(ul_inv)}</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,#a78bfa,#7c3aed);--kglow:rgba(167,139,250,0.08);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">📊</div><div class="kc-lbl">Indicative Value</div>
      <div class="kc-val" id="ul-kpi-val" style="color:#c4b5fd;">{'₹{:,.0f}'.format(ul_val) if ul_val else '—'}</div>
      <div class="kc-sub" style="color:#7a80a8;font-size:10px;">Not in portfolio total</div>
    </div>
    <div class="kc" style="--kline:linear-gradient(90deg,#a78bfa,#7c3aed);--kglow:rgba(167,139,250,0.08);">
      <div class="kc-top"></div><div class="kc-glow"></div>
      <div class="kc-icon">{'📈' if ul_pnl>=0 else '📉'}</div><div class="kc-lbl">Indicative Unreal P&amp;L</div>
      <div class="kc-val" id="ul-kpi-pnl" style="color:{'#22d67b' if ul_pnl>=0 else '#f85454'};">{'₹{:,.0f}'.format(ul_pnl) if ul_val else '—'}</div>
      <div class="kc-sub" style="color:#7a80a8;font-size:10px;">Not in portfolio total</div>
    </div>
  </div>
  <div class="edit-hint">✏️ Edit Qty, Avg Buy, or LTP — indicative values update (not counted in main portfolio)</div>
  <button class="reset-btn" onclick="resetSegment('unlisted')">↩️ Reset Unlisted</button><br><br>
  <div id="tbl-unlisted"></div>
  <div class="disclaimer" style="margin-top:18px;">
    🔒 <b>Source:</b> CMP data from Planify.in (pre-IPO marketplace). Values are indicative and not real-time exchange prices.
    These positions are <b>strictly excluded</b> from all P&amp;L totals and performance metrics.
  </div>
</section>

<!-- ████ BOOKED P&L ████ -->
<section class="sec" id="sec-booked">
  <div class="sh">
    <div class="sh-icon">🏦</div>
    <div><div class="sh-title">Booked P&amp;L — Listed Securities</div>
      <div class="sh-sub">Unlisted shares excluded · {n_trades} closed trades · Win rate {win_rate}</div></div>
    <div class="seg-pnl-pill"
      style="background:{'rgba(34,214,123,0.10)' if booked_f>=0 else 'rgba(248,84,84,0.10)'};color:{booked_tot_col};border:1px solid {booked_tot_col}44;font-size:16px;font-weight:800;">
      {'+'if booked_f>=0 else ''}₹{booked_f:,.0f}</div>
  </div>
  <div class="bk-summary-strip">
    <div class="bk-sum-item"><div class="bk-sum-lbl">Total Booked (Listed)</div>
      <div class="bk-sum-val" style="color:{booked_tot_col};">{'+'if booked_f>=0 else ''}₹{booked_f:,.2f}</div></div>
    <div class="bk-sum-item"><div class="bk-sum-lbl">STCG</div>
      <div class="bk-sum-val" style="color:#f5c842;">{'+'if stcg_total>=0 else ''}₹{stcg_total:,.0f}</div></div>
    <div class="bk-sum-item"><div class="bk-sum-lbl">LTCG</div>
      <div class="bk-sum-val" style="color:#4f7ef8;">{'+'if ltcg_total>=0 else ''}₹{ltcg_total:,.0f}</div></div>
    <div class="bk-sum-item"><div class="bk-sum-lbl">Win Rate</div>
      <div class="bk-sum-val" style="color:{'#22d67b' if win_rate_f>=50 else '#f5c842'};">{win_rate}</div></div>
    <div class="bk-sum-item"><div class="bk-sum-lbl">Trades</div>
      <div class="bk-sum-val" style="color:#7ea8ff;">{n_trades}</div></div>
  </div>
  {_booked_cards_html(["Stock"],"📊 Stocks","#4f7ef8",listed_trades)}
  {_booked_cards_html(["ETF","Liquid ETF","Commodity ETF","International ETF"],"📦 ETFs","#22d67b",listed_trades)}
  {_booked_cards_html(list(_FNO_TYPES),"🎯 F&O / Derivatives","#f5c842",listed_trades)}
  {_booked_cards_html(["REIT/InvIT"],"🏢 REIT / InvIT","#a78bfa",listed_trades)}
  {_booked_cards_html(["SGB","Bond/NCD"],"🥇 SGB / Bond","#22c7d6",listed_trades)}
  <div class="disclaimer" style="margin-top:18px;">
    ✅ <b>Note:</b> All booked P&amp;L totals include <b>listed securities only</b>.
    Unlisted share trades are tracked in the 🔒 Unlisted tab and are <b>not</b> included in any total above.
  </div>
</section>

</main>

<script>
// ══════════════════════════════════════════════════════════════════
//  EDITABLE DASHBOARD ENGINE
//  Formulas:
//    Buy Value  = Qty × Avg Buy Price
//    Mkt Value  = Qty × LTP
//    Unreal P&L = Mkt Value − Buy Value
//    P&L %      = (Unreal P&L / Buy Value) × 100
//    Daily MTM  = Qty × (LTP − Prev Close)
// ══════════════════════════════════════════════════════════════════

// Raw data injected from Python
const ORIGINAL = {{
  ov:       {all_rows_js},
  stocks:   {stocks_rows_js},
  etf:      {etf_rows_js},
  fno:      {fno_rows_js},
  reit:     {reit_rows_js},
  sgb:      {sgb_rows_js},
  unlisted: {unlisted_rows_js},
}};

// Working copy — edits go here
let DATA = JSON.parse(JSON.stringify(ORIGINAL));

// ── Utility ────────────────────────────────────────────────────
function inr(v, dec) {{
  if (v === null || v === undefined || isNaN(v)) return '—';
  dec = dec !== undefined ? dec : 0;
  const abs = Math.abs(v), sign = v < 0 ? '-' : (dec === 0 ? '' : '');
  return sign + '₹' + abs.toLocaleString('en-IN', {{minimumFractionDigits: dec, maximumFractionDigits: dec}});
}}
function pnlInr(v) {{
  if (isNaN(v)) return '—';
  return (v >= 0 ? '+' : '') + inr(v, 0);
}}
function pnlCol(v) {{ return v >= 0 ? '#22d67b' : '#f85454'; }}
function pnlArrow(v) {{ return v >= 0 ? '▲' : '▼'; }}

// ── Per-row formula engine ─────────────────────────────────────
function calcRow(r) {{
  const qty      = parseFloat(r.qty)      || 0;
  const ltp      = parseFloat(r.ltp)      || 0;
  const avg_buy  = parseFloat(r.avg_buy)  || 0;
  const prev_cl  = parseFloat(r.prev_close) || ltp;
  const buy_val  = qty * avg_buy;          // invested / cost basis
  const mkt_val  = qty * ltp;             // current market value
  const upnl     = mkt_val - buy_val;     // unrealized P&L
  const pnl_pct  = buy_val ? (upnl / buy_val * 100) : 0;
  const daily    = qty * (ltp - prev_cl); // daily MTM
  return {{ qty, ltp, avg_buy, prev_cl, buy_val, mkt_val, upnl, pnl_pct, daily }};
}}

// ── Build one editable table ────────────────────────────────────
function buildTable(rows, containerId, segKey, isUnlisted) {{
  const c = document.getElementById(containerId);
  if (!c) return;
  if (!rows || rows.length === 0) {{
    c.innerHTML = '<div style="color:#454870;padding:18px;font-size:13px;">No positions in this segment.</div>';
    return;
  }}
  let thead = `<thead><tr>
    <th style="text-align:left;">Ticker</th>
    <th style="text-align:left;">Type</th>
    <th style="text-align:right;">Qty ✏️</th>
    <th style="text-align:right;">LTP ✏️</th>
    <th style="text-align:right;">Avg Buy ✏️</th>
    <th style="text-align:right;">Buy Value</th>
    <th style="text-align:right;">Mkt Value</th>
    <th style="text-align:right;">Return %</th>
    <th style="text-align:right;">Unreal P&amp;L</th>
    <th style="text-align:right;">Daily MTM</th>
    <th style="text-align:right;">XIRR</th>
  </tr></thead>`;

  let tbody = '<tbody>';
  rows.forEach(function(r, idx) {{
    const calc = calcRow(r);
    const pc = pnlCol(calc.upnl);
    const dc = pnlCol(calc.daily);
    const arrow = pnlArrow(calc.upnl);
    tbody += `<tr class="ltp-row" data-seg="${{segKey}}" data-idx="${{idx}}">
      <td class="ltp-ticker">${{r.ticker}}</td>
      <td class="ltp-atype">${{r.asset_type}}</td>
      <td class="ltp-num"><input type="number" class="editable-cell" data-field="qty"
        value="${{calc.qty}}" step="1" min="0"
        onchange="onEdit('${{segKey}}',${{idx}},this,'qty')"
        onclick="this.select()" title="Edit quantity"></td>
      <td class="ltp-num"><input type="number" class="editable-cell" data-field="ltp"
        value="${{calc.ltp.toFixed(2)}}" step="0.05" min="0"
        onchange="onEdit('${{segKey}}',${{idx}},this,'ltp')"
        onclick="this.select()" title="Edit LTP / Current Price"></td>
      <td class="ltp-num"><input type="number" class="editable-cell" data-field="avg_buy"
        value="${{calc.avg_buy.toFixed(2)}}" step="0.05" min="0"
        onchange="onEdit('${{segKey}}',${{idx}},this,'avg_buy')"
        onclick="this.select()" title="Edit Average Buy Price"></td>
      <td class="ltp-num" data-col="buy_val">${{inr(calc.buy_val)}}</td>
      <td class="ltp-num" style="font-weight:800;color:#e0e4ff;" data-col="mkt_val">${{inr(calc.mkt_val)}}</td>
      <td class="ltp-num" style="color:${{pc}};font-weight:700;" data-col="pnl_pct">
        ${{arrow}}${{Math.abs(calc.pnl_pct).toFixed(2)}}%</td>
      <td class="ltp-num" style="color:${{pc}};font-weight:700;" data-col="upnl">
        ${{pnlInr(calc.upnl)}}</td>
      <td class="ltp-num" style="color:${{dc}};" data-col="daily">
        ${{pnlInr(calc.daily)}}</td>
      <td class="ltp-num" style="color:#8890b8;">${{r.xirr || '—'}}</td>
    </tr>`;
  }});
  tbody += '</tbody>';
  c.innerHTML = `<div class="ltp-wrap"><table class="ltp-table">${{thead}}${{tbody}}</table></div>`;
}}

// ── Handle cell edit ────────────────────────────────────────────
function onEdit(segKey, idx, input, field) {{
  const val = parseFloat(input.value);
  if (isNaN(val) || val < 0) {{ input.value = DATA[segKey][idx][field]; return; }}
  DATA[segKey][idx][field] = val;
  refreshRow(segKey, idx);
  refreshOverviewFromData();
  if (segKey !== 'ov') refreshSegmentSummary(segKey);
  // Keep 'ov' table in sync if editing a segment table
  if (segKey !== 'ov') syncOvRow(segKey, idx, field, val);
}}

// ── Refresh a single row's computed cells ──────────────────────
function refreshRow(segKey, idx) {{
  const r = DATA[segKey][idx];
  const calc = calcRow(r);
  const pc = pnlCol(calc.upnl);
  const dc = pnlCol(calc.daily);
  const arrow = pnlArrow(calc.upnl);
  const rows = document.querySelectorAll(`[data-seg="${{segKey}}"][data-idx="${{idx}}"]`);
  rows.forEach(function(row) {{
    const setCell = function(col, html, style) {{
      const td = row.querySelector(`[data-col="${{col}}"]`);
      if (td) {{ td.innerHTML = html; if (style) td.style.color = style; }}
    }};
    setCell('buy_val', inr(calc.buy_val));
    setCell('mkt_val', inr(calc.mkt_val), '#e0e4ff');
    setCell('pnl_pct', arrow + Math.abs(calc.pnl_pct).toFixed(2) + '%', pc);
    setCell('upnl', pnlInr(calc.upnl), pc);
    setCell('daily', pnlInr(calc.daily), dc);
  }});
}}

// ── Sync ov table row when editing a segment table ─────────────
function syncOvRow(segKey, idx, field, val) {{
  const ticker = DATA[segKey][idx].ticker;
  DATA['ov'].forEach(function(r, i) {{
    if (r.ticker === ticker) {{
      DATA['ov'][i][field] = val;
      refreshRow('ov', i);
    }}
  }});
}}

// ── Recalculate overview KPI cards from live DATA ───────────────
function refreshOverviewFromData() {{
  let totInv=0, totVal=0, totDaily=0;
  DATA['ov'].forEach(function(r) {{
    const c = calcRow(r);
    totInv   += c.buy_val;
    totVal   += c.mkt_val;
    totDaily += c.daily;
  }});
  const totPnl    = totVal - totInv;
  const pnlPct    = totInv ? (totPnl / totInv * 100) : 0;
  const dailyPct  = totInv ? (totDaily / totInv * 100) : 0;

  function setKPI(id, text, color) {{
    const el = document.getElementById(id);
    if (el) {{ el.textContent = text; if (color) el.style.color = color; }}
  }}
  setKPI('kpi-invested', inr(totInv), '#7ea8ff');
  setKPI('kpi-value',    inr(totVal), pnlCol(totPnl));
  setKPI('kpi-upnl',     (totPnl>=0?'+':'')+inr(totPnl), pnlCol(totPnl));
  setKPI('kpi-upnl-pct', pnlPct.toFixed(2)+'%', pnlCol(totPnl));
  setKPI('kpi-daily',    (totDaily>=0?'+':'')+inr(totDaily), pnlCol(totDaily));
  setKPI('kpi-daily-pct',dailyPct.toFixed(2)+'%', pnlCol(totDaily));
}}

// ── Segment header summary bar update ──────────────────────────
function refreshSegmentSummary(segKey) {{
  const segMap = {{
    stocks: {{inv:'sb-stocks-inv',val:'sb-stocks-val',pnl:'sb-stocks-pnl',n:'sb-stocks-n'}},
    etf:    {{inv:'sb-etf-inv',   val:'sb-etf-val',   pnl:'sb-etf-pnl',   n:'sb-etf-n'}},
    fno:    {{inv:'sb-fno-inv',   val:'sb-fno-val',   pnl:'sb-fno-pnl',   n:'sb-fno-n'}},
    reit:   {{inv:'sb-reit-inv',  val:'sb-reit-val',  pnl:'sb-reit-pnl',  n:'sb-reit-n'}},
    sgb:    {{inv:'sb-sgb-inv',   val:'sb-sgb-val',   pnl:'sb-sgb-pnl',   n:'sb-sgb-n'}},
  }};
  const ids = segMap[segKey];
  if (!ids) return;
  let inv=0, val=0, pnl=0;
  DATA[segKey].forEach(function(r) {{
    const c = calcRow(r);
    inv += c.buy_val; val += c.mkt_val; pnl += c.upnl;
  }});
  function setEl(id, text, color) {{
    const el = document.getElementById(id);
    if (el) {{ el.textContent = text; if (color) el.style.color = color; }}
  }}
  setEl(ids.inv, inr(inv));
  setEl(ids.val, inr(val));
  setEl(ids.pnl, pnlInr(pnl), pnlCol(pnl));
  // Also update seg header subtitle & pill
  const sub = document.getElementById('seg-sub-'+segKey);
  if (sub) sub.textContent = DATA[segKey].length + ' positions · Invested ' + inr(inv) + ' · Mkt Value ' + inr(val);
  const pill = document.getElementById('seg-pill-'+segKey);
  if (pill) {{ pill.textContent = pnlInr(pnl); pill.style.color = pnlCol(pnl); }}
}}

// ── Reset functions ─────────────────────────────────────────────
function resetAll() {{
  DATA = JSON.parse(JSON.stringify(ORIGINAL));
  ['ov','stocks','etf','fno','reit','sgb','unlisted'].forEach(function(seg) {{
    buildTable(DATA[seg], 'tbl-'+seg, seg, seg==='unlisted');
  }});
  refreshOverviewFromData();
  ['stocks','etf','fno','reit','sgb'].forEach(refreshSegmentSummary);
}}

function resetSegment(segKey) {{
  DATA[segKey] = JSON.parse(JSON.stringify(ORIGINAL[segKey]));
  buildTable(DATA[segKey], 'tbl-'+segKey, segKey, segKey==='unlisted');
  // Sync back to ov
  DATA[segKey].forEach(function(r, idx) {{
    DATA['ov'].forEach(function(ov, i) {{
      if (ov.ticker === r.ticker) DATA['ov'][i] = JSON.parse(JSON.stringify(r));
    }});
  }});
  buildTable(DATA['ov'], 'tbl-ov', 'ov', false);
  refreshOverviewFromData();
  if (segKey !== 'ov') refreshSegmentSummary(segKey);
}}

// ── Initial render ──────────────────────────────────────────────
window.addEventListener('DOMContentLoaded', function() {{
  buildTable(DATA['ov'],       'tbl-ov',       'ov',       false);
  buildTable(DATA['stocks'],   'tbl-stocks',   'stocks',   false);
  buildTable(DATA['etf'],      'tbl-etf',      'etf',      false);
  buildTable(DATA['fno'],      'tbl-fno',      'fno',      false);
  buildTable(DATA['reit'],     'tbl-reit',     'reit',     false);
  buildTable(DATA['sgb'],      'tbl-sgb',      'sgb',      false);
  buildTable(DATA['unlisted'], 'tbl-unlisted', 'unlisted', true);

  // Donut chart
  var dCtx = document.getElementById('donutChart');
  if (dCtx) {{
    new Chart(dCtx, {{
      type:'doughnut',
      data:{{ labels:{donut_js}, datasets:[{{ data:{inv_js}, backgroundColor:{dcol_js},
        borderColor:'#060812', borderWidth:2, hoverOffset:6 }}] }},
      options:{{ plugins:{{ legend:{{ position:'right',labels:{{ color:'#7a80a8',font:{{size:11,family:'Inter'}},padding:10,boxWidth:10 }} }},
        tooltip:{{ callbacks:{{ label: function(c){{ return c.label+': ₹'+c.parsed.toLocaleString('en-IN'); }} }} }} }},
        cutout:'65%', responsive:true, maintainAspectRatio:false }}
    }});
  }}

  // Bar chart
  var bCtx = document.getElementById('pnlBar');
  if (bCtx) {{
    new Chart(bCtx, {{
      type:'bar',
      data:{{ labels:{bt_js}, datasets:[{{ data:{bp_js}, backgroundColor:{bc_js},
        borderRadius:4, borderSkipped:false }}] }},
      options:{{ plugins:{{ legend:{{ display:false }},
        tooltip:{{ callbacks:{{ label: function(c){{ return '₹'+c.parsed.y.toLocaleString('en-IN'); }} }} }} }},
        scales:{{ x:{{ ticks:{{ color:'#636899',font:{{size:10}},maxRotation:45,autoSkip:false }}, grid:{{ color:'rgba(255,255,255,0.03)' }} }},
          y:{{ ticks:{{ color:'#636899',font:{{size:10}},callback: function(v){{ return '₹'+v.toLocaleString('en-IN'); }} }},
            grid:{{ color:'rgba(255,255,255,0.05)' }} }} }},
        responsive:true, maintainAspectRatio:false }}
    }});
  }}
}});

// ── Tab switcher ─────────────────────────────────────────────────
function sw(id, el) {{
  document.querySelectorAll('.sec').forEach(s => s.classList.remove('on'));
  document.querySelectorAll('.nt').forEach(n => n.classList.remove('on'));
  var sec = document.getElementById('sec-' + id);
  if (sec) sec.classList.add('on');
  el.classList.add('on');
  window.scrollTo({{top:0,behavior:'smooth'}});
  setTimeout(function(){{ window.dispatchEvent(new Event('resize')); }}, 50);
}}

// ── Save Edits — bakes current DATA into ORIGINAL so edits persist on reopen ─
function saveEdits() {{
  var btn = document.getElementById('save-btn');
  var origText = btn.textContent;
  btn.textContent = '⏳ Saving…';
  btn.style.opacity = '0.7';
  btn.style.pointerEvents = 'none';
  setTimeout(function() {{
    try {{
      var html = document.documentElement.outerHTML;
      var dataJson = JSON.stringify(DATA);
      html = html.replace(
        /const ORIGINAL = \{{[\s\S]*?\}};(\s*\/\/ Working copy)/,
        'const ORIGINAL = ' + dataJson + ';$1'
      );
      var blob = new Blob([html], {{type: 'text/html;charset=utf-8'}});
      var url  = URL.createObjectURL(blob);
      var a    = document.createElement('a');
      var now  = new Date();
      var dd   = String(now.getDate()).padStart(2,'0');
      var mm   = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][now.getMonth()];
      var yy   = now.getFullYear();
      a.href     = url;
      a.download = 'Portfolio_{client_name.replace(" ","_")}_' + dd + mm + yy + '_saved.html';
      document.body.appendChild(a);
      a.click();
      document.body.removeChild(a);
      URL.revokeObjectURL(url);
      btn.textContent = '✅ Saved!';
      btn.style.background = 'rgba(34,214,123,0.25)';
      setTimeout(function() {{
        btn.textContent = origText;
        btn.style.background = 'rgba(34,214,123,0.12)';
        btn.style.opacity = '1';
        btn.style.pointerEvents = 'auto';
      }}, 2500);
    }} catch(e) {{
      btn.textContent = '❌ Error';
      btn.style.opacity = '1';
      btn.style.pointerEvents = 'auto';
      setTimeout(function() {{
        btn.textContent = origText;
        btn.style.background = 'rgba(34,214,123,0.12)';
      }}, 2500);
    }}
  }}, 60);
}}
</script>
</body>
</html>"""
    return html.encode("utf-8")


def generate_html_report_mobile(client_name, calc_df, summary_dict, trades_df, booked_pnl_map,
                                _prebuilt_desktop_bytes=None):
    """
    Mobile-optimised portfolio HTML dashboard.
    Post-processes generate_html_report() output:
      - Injects mobile-first CSS (card layout, 2-col KPIs, scrollable nav)
      - Replaces tables with swipe-friendly editable cards via JS injection
      - Charts resized for single-column mobile layout
      - Save Edits / all interactivity identical to desktop version

    Pass _prebuilt_desktop_bytes to skip regenerating the desktop report
    (avoids duplicate AI + LTP calls when both reports are requested).
    """
    if _prebuilt_desktop_bytes:
        desktop_bytes = _prebuilt_desktop_bytes
    else:
        desktop_bytes = generate_html_report(client_name, calc_df, summary_dict, trades_df, booked_pnl_map)
    if not desktop_bytes:
        return None

    html = desktop_bytes.decode("utf-8")

    # ── 1. Mobile CSS injected before </style> ───────────────────────────
    mobile_css = """
/* ═══════ MOBILE OVERRIDE — generate_html_report_mobile() ═══════ */
body { font-size:14px; }
.hdr { padding:20px 16px 16px !important; }
.hdr-brand { font-size:9px !important; letter-spacing:2px !important; margin-bottom:8px !important; }
.hdr-title { font-size:24px !important; letter-spacing:-0.5px !important; margin-bottom:6px !important; }
.hdr-sub { font-size:11px !important; line-height:1.5 !important; }
.chip { font-size:10px !important; padding:3px 9px !important; margin-right:4px !important; }
.nav {
  padding:8px 10px !important; gap:4px !important;
  flex-wrap:nowrap !important; overflow-x:auto !important;
  -webkit-overflow-scrolling:touch !important; scrollbar-width:none !important;
}
.nav::-webkit-scrollbar { display:none !important; }
.nt { padding:6px 12px !important; font-size:11px !important; white-space:nowrap !important; flex-shrink:0 !important; }
.main { padding:14px 12px 60px !important; max-width:100% !important; }
.ga { grid-template-columns:repeat(2,1fr) !important; gap:10px !important; }
.kc { padding:14px 14px 12px !important; }
.kc-icon { font-size:16px !important; margin-bottom:6px !important; }
.kc-lbl  { font-size:9px !important; }
.kc-val  { font-size:16px !important; }
.kc-sub  { font-size:10px !important; }
.g2 { grid-template-columns:1fr !important; gap:12px !important; }
.seg-bar { flex-direction:column !important; align-items:flex-start !important; gap:8px !important; padding:12px 14px !important; }
.seg-bar-label { font-size:14px !important; }
.seg-metric { flex-direction:row !important; gap:6px !important; align-items:center !important; }
.seg-mlbl { min-width:70px !important; }
.ai-card { padding:18px 16px !important; }
.ai-text { font-size:13px !important; line-height:1.7 !important; }
.sh { margin-bottom:16px !important; }
.sh-title { font-size:17px !important; }
.bk-summary-strip { grid-template-columns:repeat(2,1fr) !important; gap:8px !important; }
.bk-sum-val { font-size:15px !important; }
.psa-cards-grid { grid-template-columns:1fr !important; }
.psa-grid4 { grid-template-columns:repeat(2,1fr) !important; }
.disclaimer { font-size:10px !important; }
/* Hide desktop table, show mobile cards */
.ltp-wrap { display:none !important; }
.m-cards  { display:flex !important; flex-direction:column; gap:10px; }
/* ── Mobile stock card ─────────────────────────────── */
.m-card {
  background:var(--bg3); border:1px solid var(--border2);
  border-radius:14px; padding:14px 14px 12px;
  position:relative; overflow:hidden;
}
.m-card::before {
  content:''; position:absolute; top:0; left:0; right:0;
  height:3px; background:var(--mc-line, var(--blue));
}
.m-card-top { display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:10px; }
.m-card-ticker { font-size:16px; font-weight:900; color:var(--text); letter-spacing:-0.2px; }
.m-card-type   { font-size:10px; color:var(--muted); font-weight:600; text-transform:uppercase; letter-spacing:0.5px; margin-top:2px; }
.m-card-pnl    { text-align:right; }
.m-card-pnl-val { font-size:15px; font-weight:800; }
.m-card-pnl-pct { font-size:11px; font-weight:700; margin-top:2px; }
.m-card-stats {
  display:grid; grid-template-columns:1fr 1fr 1fr;
  gap:6px; background:var(--bg2); border-radius:10px;
  padding:9px 10px; margin-bottom:10px;
}
.m-card-stat-lbl { font-size:9px; text-transform:uppercase; letter-spacing:0.5px; color:var(--muted); font-weight:700; margin-bottom:3px; }
.m-card-stat-val { font-size:12px; font-weight:700; color:var(--text); }
.m-card-inputs   { display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; margin-bottom:8px; }
.m-card-field label {
  font-size:9px; text-transform:uppercase; letter-spacing:0.5px;
  color:#a0c4ff; font-weight:700; display:block; margin-bottom:4px;
}
.m-card-input {
  width:100%; padding:7px 6px; border-radius:8px;
  background:rgba(79,126,248,0.08); border:1px solid rgba(79,126,248,0.25);
  color:#a0c4ff; font-weight:700; font-size:13px; font-family:'Inter',sans-serif;
  outline:none; text-align:center; box-sizing:border-box;
}
.m-card-input:focus {
  background:rgba(79,126,248,0.18); border-color:#4f7ef8;
  box-shadow:0 0 0 2px rgba(79,126,248,0.18); color:#fff;
}
.m-card-footer {
  display:flex; justify-content:space-between; align-items:center;
  border-top:1px solid var(--border); padding-top:8px;
}
.m-card-footer-lbl { font-size:9px; color:var(--muted); text-transform:uppercase; letter-spacing:0.5px; font-weight:700; }
.m-card-footer-val { font-size:12px; font-weight:700; }
"""
    html = html.replace("</style>", mobile_css + "\n</style>", 1)

    # ── 2. Inject mobile JS before the buildTable function ───────────────
    mobile_js = """
// ── Mobile card builder (injected by generate_html_report_mobile) ──
function buildMobileCards(rows, segKey) {
  var cardId  = 'tbl-' + segKey + '-mcards';
  var tableEl = document.getElementById('tbl-' + segKey);
  if (!tableEl) return;
  var cardDiv = document.getElementById(cardId);
  if (!cardDiv) {
    cardDiv = document.createElement('div');
    cardDiv.id = cardId;
    cardDiv.className = 'm-cards';
    tableEl.parentNode.insertBefore(cardDiv, tableEl.nextSibling);
  }
  if (!rows || rows.length === 0) {
    cardDiv.innerHTML = '<div style="color:#454870;padding:18px;font-size:13px;">No positions.</div>';
    return;
  }
  var h = '';
  rows.forEach(function(r, idx) {
    var calc = calcRow(r);
    var pc   = pnlCol(calc.upnl);
    var dc   = pnlCol(calc.daily);
    var arr  = pnlArrow(calc.upnl);
    var lc   = calc.upnl >= 0 ? '#22d67b' : '#f85454';
    h += '<div class="m-card" style="--mc-line:' + lc + ';" data-seg="' + segKey + '" data-idx="' + idx + '">'
      + '<div class="m-card-top">'
      +   '<div><div class="m-card-ticker">' + r.ticker + '</div>'
      +   '<div class="m-card-type">' + r.asset_type + '</div></div>'
      +   '<div class="m-card-pnl">'
      +     '<div class="m-card-pnl-val" style="color:' + pc + ';" data-col="upnl">' + pnlInr(calc.upnl) + '</div>'
      +     '<div class="m-card-pnl-pct" style="color:' + pc + ';" data-col="pnl_pct">' + arr + Math.abs(calc.pnl_pct).toFixed(2) + '%</div>'
      +   '</div>'
      + '</div>'
      + '<div class="m-card-stats">'
      +   '<div><div class="m-card-stat-lbl">Buy Value</div><div class="m-card-stat-val" data-col="buy_val">' + inr(calc.buy_val) + '</div></div>'
      +   '<div><div class="m-card-stat-lbl">Mkt Value</div><div class="m-card-stat-val" style="color:#e0e4ff;" data-col="mkt_val">' + inr(calc.mkt_val) + '</div></div>'
      +   '<div><div class="m-card-stat-lbl">XIRR</div><div class="m-card-stat-val" style="color:#8890b8;">' + (r.xirr || '—') + '</div></div>'
      + '</div>'
      + '<div class="m-card-inputs">'
      +   '<div class="m-card-field"><label>Qty ✏</label>'
      +     '<input type="number" class="m-card-input" data-field="qty" value="' + calc.qty + '" step="1" min="0" onchange="onEdit(\\''+segKey+'\\','+idx+',this,\\'qty\\')" onclick="this.select()"></div>'
      +   '<div class="m-card-field"><label>LTP ✏</label>'
      +     '<input type="number" class="m-card-input" data-field="ltp" value="' + calc.ltp.toFixed(2) + '" step="0.05" min="0" onchange="onEdit(\\''+segKey+'\\','+idx+',this,\\'ltp\\')" onclick="this.select()"></div>'
      +   '<div class="m-card-field"><label>Avg Buy ✏</label>'
      +     '<input type="number" class="m-card-input" data-field="avg_buy" value="' + calc.avg_buy.toFixed(2) + '" step="0.05" min="0" onchange="onEdit(\\''+segKey+'\\','+idx+',this,\\'avg_buy\\')" onclick="this.select()"></div>'
      + '</div>'
      + '<div class="m-card-footer">'
      +   '<div><span class="m-card-footer-lbl">Daily MTM </span><span class="m-card-footer-val" style="color:' + dc + ';" data-col="daily">' + pnlInr(calc.daily) + '</span></div>'
      +   '<div><span class="m-card-footer-lbl">Return </span><span class="m-card-footer-val" style="color:' + pc + ';">' + arr + Math.abs(calc.pnl_pct).toFixed(2) + '%</span></div>'
      + '</div>'
      + '</div>';
  });
  cardDiv.innerHTML = h;
}

// Monkey-patch refreshRow to sync mobile cards too
(function() {
  var _orig = refreshRow;
  refreshRow = function(segKey, idx) {
    _orig(segKey, idx);
    var cardDiv = document.getElementById('tbl-' + segKey + '-mcards');
    if (!cardDiv) return;
    var r    = DATA[segKey][idx];
    var calc = calcRow(r);
    var pc   = pnlCol(calc.upnl);
    var dc   = pnlCol(calc.daily);
    var arr  = pnlArrow(calc.upnl);
    var cards = cardDiv.querySelectorAll('[data-seg="'+segKey+'"][data-idx="'+idx+'"]');
    cards.forEach(function(card) {
      function sc(col, html, color) {
        var el = card.querySelector('[data-col="'+col+'"]');
        if (el) { el.innerHTML = html; if (color) el.style.color = color; }
      }
      sc('buy_val', inr(calc.buy_val));
      sc('mkt_val', inr(calc.mkt_val), '#e0e4ff');
      sc('pnl_pct', arr + Math.abs(calc.pnl_pct).toFixed(2) + '%', pc);
      sc('upnl',    pnlInr(calc.upnl), pc);
      sc('daily',   pnlInr(calc.daily), dc);
      card.querySelectorAll('.m-card-input').forEach(function(inp) {
        var f = inp.getAttribute('data-field');
        if (f==='qty')     inp.value = calc.qty;
        if (f==='ltp')     inp.value = calc.ltp.toFixed(2);
        if (f==='avg_buy') inp.value = calc.avg_buy.toFixed(2);
      });
    });
  };
})();
"""
    html = html.replace(
        "function buildTable(rows, containerId, segKey, isUnlisted) {",
        mobile_js + "\nfunction buildTable(rows, containerId, segKey, isUnlisted) {"
    )

    # ── 3. Patch DOMContentLoaded init to also call buildMobileCards ──────
    init_mobile = """
  // Build mobile cards for all segments
  ['stocks','etf','fno','reit','sgb','unlisted'].forEach(function(seg) {
    buildMobileCards(DATA[seg], seg);
  });
"""
    html = html.replace(
        "buildTable(DATA['unlisted'], 'tbl-unlisted', 'unlisted', true);",
        "buildTable(DATA['unlisted'], 'tbl-unlisted', 'unlisted', true);\n" + init_mobile
    )

    # ── 4. Patch resetAll / resetSegment to rebuild mobile cards ──────────
    reset_mobile = """
  // Rebuild mobile cards after reset
  ['stocks','etf','fno','reit','sgb','unlisted'].forEach(function(s) {
    buildMobileCards(DATA[s], s);
  });
"""
    # Insert after the resetAll DATA copy
    html = html.replace(
        "// Working copy\n  buildTable(DATA['ov'],",
        "// Working copy\n" + reset_mobile + "  buildTable(DATA['ov'],"
    )

    # ── 5. Compact chart heights for single-column mobile layout ──────────
    html = html.replace('style="height:200px;position:relative;"', 'style="height:170px;position:relative;"')
    html = html.replace('style="height:240px;position:relative;"', 'style="height:190px;position:relative;"')
    html = html.replace('style="height:260px;position:relative;"', 'style="height:190px;position:relative;"')

    # ── 6. Tag the filename so mobile save is clearly labelled ────────────
    html = html.replace("_saved.html'", "_mobile_saved.html'")
    html = html.replace(
        "Northeast Broking Services Limited &middot; Portfolio Report",
        "Northeast Broking Services Limited &middot; Portfolio Report &middot; Mobile"
    )

    return html.encode("utf-8")




def _arc_xy(angle_deg, cx, cy, r):
    """SVG arc endpoint for gauge at angle_deg (0=left, 180=right)."""
    import math
    rad = math.radians(angle_deg)
    x = cx + r * math.cos(math.pi - rad)
    y = cy - r * math.sin(math.pi - rad)
    large = 1 if angle_deg > 90 else 0
    return f"{x:.1f},{y:.1f}"


def _zone_ticks():
    """Return SVG tick mark lines at zone boundaries of gauge."""
    import math
    ticks = []
    cx, cy, r_out, r_in = 150, 150, 131, 118
    for deg in [0, 36, 72, 108, 144, 180]:
        rad = math.radians(deg)
        x1 = cx + r_out * math.cos(math.pi - rad)
        y1 = cy - r_out * math.sin(math.pi - rad)
        x2 = cx + r_in  * math.cos(math.pi - rad)
        y2 = cy - r_in  * math.sin(math.pi - rad)
        ticks.append(f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" stroke="#060812" stroke-width="2.5"/>')
    return ticks

# =====================================================
# ── EXCEL EXPORT ─────────────────────────────────────
# =====================================================

def generate_excel_report(client_name, calc_df, trades_df, summary_dict):
    """
    Generate a professional multi-sheet Excel dashboard.
    Sheets: 1-Cover  2-Dashboard  3-Holdings  4-AI Analysis  5-Capital Gains  6-Trade History
    Returns bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                     GradientFill)
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint

        wb  = openpyxl.Workbook()
        IST = datetime.now(timezone(timedelta(hours=5, minutes=30)))
        s   = summary_dict
        xirr_map = s.get("xirr_map", {})
        cg_df    = s.get("cap_gains_df", pd.DataFrame())

        # ─── palette ──────────────────────────────────────────────────
        C_NAVY    = "0D1B2A"
        C_BLUE    = "1565C0"
        C_ACCENT  = "1976D2"
        C_GREEN   = "1B5E20"
        C_GREEN_L = "E8F5E9"
        C_RED     = "B71C1C"
        C_RED_L   = "FFEBEE"
        C_GOLD    = "F57F17"
        C_GOLD_L  = "FFFDE7"
        C_GRAY    = "37474F"
        C_LGRAY   = "ECEFF1"
        C_WHITE   = "FFFFFF"
        C_HROW    = "E3F2FD"   # alternating row tint

        # ─── style factories ─────────────────────────────────────────
        def _fill(hex6):
            return PatternFill("solid", fgColor=hex6)

        def _font(bold=False, size=10, color=C_NAVY, italic=False, name="Arial"):
            return Font(name=name, bold=bold, size=size, color=color, italic=italic)

        def _border(style="thin"):
            s = Side(border_style=style, color="B0BEC5")
            return Border(left=s, right=s, top=s, bottom=s)

        def _al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def _inr(v, decimals=2):
            try:
                f = float(v)
                return f"₹{f:,.{decimals}f}"
            except:
                return "—"

        def _pct(v):
            try:
                return f"{float(v):+.2f}%"
            except:
                return "—"

        def _num(v, decimals=2):
            try:
                return round(float(v), decimals) if pd.notna(v) else ""
            except:
                return str(v) if v is not None else ""

        def _set_col_widths(ws, widths_dict):
            for col_letter, w in widths_dict.items():
                ws.column_dimensions[col_letter].width = w

        def _auto_width(ws, extra=4, cap=40):
            for col in ws.columns:
                ml = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + extra, cap)

        def _style_header_row(ws, row_num, fill_hex, font_color=C_WHITE,
                               bold=True, size=10, height=22):
            ws.row_dimensions[row_num].height = height
            for cell in ws[row_num]:
                cell.fill      = _fill(fill_hex)
                cell.font      = _font(bold=bold, size=size, color=font_color)
                cell.alignment = _al("center")
                cell.border    = _border()

        def _style_data_row(ws, row_num, even=True, height=18):
            ws.row_dimensions[row_num].height = height
            bg = C_HROW if even else C_WHITE
            for cell in ws[row_num]:
                if cell.fill.patternType not in ("solid",) or cell.fill.fgColor.rgb in ("00000000", C_WHITE, C_HROW):
                    cell.fill = _fill(bg)
                cell.alignment = _al("center")
                cell.border    = _border()

        def _merge_title(ws, cell_range, text, fill_hex, font_color=C_WHITE,
                          size=12, bold=True, height=28):
            ws.merge_cells(cell_range)
            c = ws[cell_range.split(":")[0]]
            c.value     = text
            c.fill      = _fill(fill_hex)
            c.font      = _font(bold=bold, size=size, color=font_color)
            c.alignment = _al("center")
            row_num = int(''.join(filter(str.isdigit, cell_range.split(":")[0])))
            ws.row_dimensions[row_num].height = height

        # ──────────────────────────────────────────────────────────────
        # SHEET 1 — COVER PAGE
        # ──────────────────────────────────────────────────────────────
        ws_cov = wb.active
        ws_cov.title = "Cover"
        ws_cov.sheet_view.showGridLines = False
        _set_col_widths(ws_cov, {c: 18 for c in "ABCDE"})
        ws_cov.column_dimensions["A"].width = 4
        ws_cov.column_dimensions["F"].width = 4

        for r in range(1, 35):
            ws_cov.row_dimensions[r].height = 20

        # top color band
        for row in range(1, 6):
            for col in range(1, 10):
                ws_cov.cell(row=row, column=col).fill = _fill(C_NAVY)

        ws_cov.merge_cells("B2:E2")
        c = ws_cov["B2"]
        c.value = "NORTHEAST BROKING SERVICES LIMITED"
        c.font  = _font(bold=True, size=13, color="90CAF9")
        c.alignment = _al("center")

        ws_cov.merge_cells("B3:E3")
        c = ws_cov["B3"]
        c.value = "PORTFOLIO STATEMENT"
        c.font  = _font(bold=True, size=20, color=C_WHITE)
        c.alignment = _al("center")

        ws_cov.merge_cells("B4:E4")
        c = ws_cov["B4"]
        c.value = f"Client: {client_name}"
        c.font  = _font(bold=False, size=11, color="B3E5FC")
        c.alignment = _al("center")

        ws_cov.row_dimensions[7].height = 6

        # KPI boxes  (row 8-12)
        total_inv = s.get("total_invested", 0)
        total_val = s.get("total_value", 0)
        total_pnl = s.get("total_pnl", 0)
        pnl_pct   = s.get("total_pnl_pct", 0)
        daily_pnl = s.get("daily_pnl", 0)
        booked    = s.get("total_booked_pnl", 0)
        beta      = s.get("portfolio_beta", 1.0)

        kpis = [
            ("Total Invested",   _inr(total_inv),   C_ACCENT),
            ("Current Value",    _inr(total_val),   C_BLUE),
            ("Unrealized P&L",   f"{_inr(total_pnl)}  ({_pct(pnl_pct)})",
             C_GREEN if float(total_pnl or 0) >= 0 else C_RED),
            ("Today's P&L",      _inr(daily_pnl),
             C_GREEN if float(daily_pnl or 0) >= 0 else C_RED),
            ("Booked P&L",       _inr(booked),
             C_GREEN if float(booked or 0) >= 0 else C_RED),
            ("Portfolio Beta",   str(beta),          C_GOLD),
        ]
        kpi_cols = ["B", "C", "D", "E", "B", "C"]
        kpi_rows = [8,   8,   8,   8,  11,  11]
        kpi_mcol = ["B", "C", "D", "E", "B", "C"]

        # 2x3 grid
        positions = [("B8","C9"), ("C8","D9"), ("D8","E9"),
                     ("B11","C12"), ("C11","D12"), ("D11","E12")]
        for (start_cell, end_cell), (label, value, color) in zip(positions, kpis):
            ws_cov.merge_cells(f"{start_cell}:{end_cell}")
            c = ws_cov[start_cell]
            c.value     = f"{label}\n{value}"
            c.fill      = _fill(color)
            c.font      = _font(bold=True, size=10, color=C_WHITE)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _border()
            row_num = int(''.join(filter(str.isdigit, start_cell)))
            ws_cov.row_dimensions[row_num].height = 36

        ws_cov.row_dimensions[10].height = 6

        # allocation breakdown (row 14-22)
        ws_cov.row_dimensions[13].height = 8
        ws_cov.merge_cells("B14:E14")
        c = ws_cov["B14"]
        c.value = "ASSET ALLOCATION BREAKDOWN"
        c.font  = _font(bold=True, size=11, color=C_NAVY)
        c.fill  = _fill(C_LGRAY)
        c.alignment = _al("center")
        ws_cov.row_dimensions[14].height = 22

        alloc_data = [
            ("Stocks",    s.get("stocks_invested", 0),  C_BLUE),
            ("ETFs",      s.get("etf_invested", 0),     "0288D1"),
            ("F&O",       s.get("fno_invested", 0),     C_GOLD),
            ("REIT/InvIT",s.get("reit_invested", 0),    "6A1B9A"),
            ("SGB/Bond",  s.get("other_invested", 0),   "2E7D32"),
            ("Unlisted",  s.get("unlisted_invested", 0),"455A64"),
        ]
        ws_cov["B15"].value = "Asset Class"
        ws_cov["C15"].value = "Invested (₹)"
        ws_cov["D15"].value = "% of Portfolio"
        ws_cov["E15"].value = "Indicator"
        _style_header_row(ws_cov, 15, C_NAVY, C_WHITE, height=20)

        for i, (aname, ainv, acolor) in enumerate(alloc_data):
            r = 16 + i
            pct_alloc = (float(ainv) / float(total_inv) * 100) if total_inv else 0
            ws_cov.cell(r, 2).value = aname
            ws_cov.cell(r, 3).value = _inr(ainv)
            ws_cov.cell(r, 4).value = f"{pct_alloc:.1f}%"
            # bar indicator via cell background width trick
            ws_cov.cell(r, 5).value = "█" * int(pct_alloc / 5) if pct_alloc > 0 else "—"
            ws_cov.cell(r, 5).font  = _font(color=acolor, bold=True)
            ws_cov.row_dimensions[r].height = 18
            for col in range(2, 6):
                cell = ws_cov.cell(r, col)
                cell.fill      = _fill(C_HROW if i % 2 == 0 else C_WHITE)
                cell.alignment = _al("center")
                cell.border    = _border()
                cell.font      = _font(size=9)

        # footer
        footer_row = 24
        ws_cov.merge_cells(f"B{footer_row}:E{footer_row}")
        c = ws_cov[f"B{footer_row}"]
        c.value = f"Generated on {IST.strftime('%d %b %Y, %I:%M %p IST')}  |  Northeast Broking Services Limited  |  Confidential"
        c.font  = _font(italic=True, size=8, color="90A4AE")
        c.alignment = _al("center")

        # ──────────────────────────────────────────────────────────────
        # SHEET 2 — HOLDINGS DASHBOARD
        # ──────────────────────────────────────────────────────────────
        ws_h = wb.create_sheet("Holdings")
        ws_h.sheet_view.showGridLines = False
        ws_h.freeze_panes = "A3"

        _merge_title(ws_h, "A1:L1",
                     f"HOLDINGS  ·  {client_name}  ·  {IST.strftime('%d %b %Y')}",
                     C_NAVY, C_WHITE, size=12, height=30)

        h_headers = ["#", "Ticker", "Asset Type", "Qty", "Avg Buy ₹",
                     "CMP ₹", "Cost Basis ₹", "Mkt Value ₹",
                     "Unreal P&L ₹", "P&L %", "Daily P&L ₹", "XIRR %"]
        ws_h.append(h_headers)
        _style_header_row(ws_h, 2, C_BLUE, C_WHITE, height=22)

        sorted_df = calc_df.sort_values("Value", ascending=False).reset_index(drop=True)
        for idx_r, row in sorted_df.iterrows():
            row_num = idx_r + 3
            pnl_v   = float(row.get("Unrealized_PnL", 0) or 0)
            dpnl_v  = float(row.get("Daily_PnL", 0) or 0)
            xirr_v  = xirr_map.get(row.get("Ticker", ""), "")
            try:
                xirr_disp = f"{float(xirr_v):.1f}%" if xirr_v != "" else "—"
            except:
                xirr_disp = "—"

            row_data = [
                idx_r + 1,
                row.get("Ticker", ""),
                row.get("Asset_Type", ""),
                _num(row.get("Shares"), 2),
                _num(row.get("Buy_Price"), 2),
                _num(row.get("Current_Price"), 2),
                _num(row.get("Cost_Basis"), 0),
                _num(row.get("Value"), 0),
                round(pnl_v, 0),
                f"{float(row.get('PnL_%', 0) or 0):+.2f}%",
                round(dpnl_v, 0),
                xirr_disp,
            ]
            ws_h.append(row_data)
            _style_data_row(ws_h, row_num, even=(idx_r % 2 == 0))

            # P&L colour
            pnl_fill  = _fill(C_GREEN_L) if pnl_v  >= 0 else _fill(C_RED_L)
            dpnl_fill = _fill(C_GREEN_L) if dpnl_v >= 0 else _fill(C_RED_L)
            pnl_font  = _font(bold=True, color=C_GREEN if pnl_v  >= 0 else C_RED)
            dpnl_font = _font(bold=True, color=C_GREEN if dpnl_v >= 0 else C_RED)
            for col_i in [9, 10]:
                ws_h.cell(row_num, col_i).fill = pnl_fill
                ws_h.cell(row_num, col_i).font = pnl_font
            ws_h.cell(row_num, 11).fill = dpnl_fill
            ws_h.cell(row_num, 11).font = dpnl_font

        # totals row
        t_row = len(sorted_df) + 3
        totals = [
            "", "TOTAL", "", "",  "", "",
            f"₹{calc_df['Cost_Basis'].sum():,.0f}",
            f"₹{calc_df['Value'].sum():,.0f}",
            f"₹{calc_df['Unrealized_PnL'].sum():,.0f}",
            _pct(s.get("total_pnl_pct", 0)),
            f"₹{calc_df['Daily_PnL'].sum():,.0f}",
            "",
        ]
        ws_h.append(totals)
        ws_h.row_dimensions[t_row].height = 22
        for col_i in range(1, 13):
            c = ws_h.cell(t_row, col_i)
            c.fill   = _fill(C_NAVY)
            c.font   = _font(bold=True, color=C_WHITE, size=10)
            c.alignment = _al("center")
            c.border = _border()

        _set_col_widths(ws_h, {"A":5,"B":14,"C":16,"D":8,"E":12,"F":12,
                                "G":16,"H":16,"I":16,"J":10,"K":14,"L":10})

        # ──────────────────────────────────────────────────────────────
        # SHEET 3 — AI PORTFOLIO ANALYSIS
        # ──────────────────────────────────────────────────────────────
        ws_ai = wb.create_sheet("AI Analysis")
        ws_ai.sheet_view.showGridLines = False
        _set_col_widths(ws_ai, {"A":3, "B":28, "C":60, "D":3})
        for r in range(1, 120):
            ws_ai.row_dimensions[r].height = 18

        _merge_title(ws_ai, "B1:C1",
                     "🤖  AI PORTFOLIO ANALYSIS  —  NORTHEAST BROKING",
                     C_NAVY, C_WHITE, size=13, height=34)

        ws_ai["B2"].value = f"Client: {client_name}     |     Report Date: {IST.strftime('%d %b %Y')}     |     Powered by Northeast Analytics Engine"
        ws_ai["B2"].font  = _font(italic=True, size=9, color=C_GRAY)
        ws_ai.merge_cells("B2:C2")
        ws_ai["B2"].alignment = _al("center")
        ws_ai.row_dimensions[2].height = 20

        current_row = [4]  # mutable to use inside nested func

        def _section(title, color_hex):
            r = current_row[0]
            ws_ai.merge_cells(f"B{r}:C{r}")
            c = ws_ai[f"B{r}"]
            c.value     = title
            c.fill      = _fill(color_hex)
            c.font      = _font(bold=True, size=11, color=C_WHITE)
            c.alignment = _al("left")
            c.border    = _border()
            ws_ai.row_dimensions[r].height = 24
            current_row[0] += 1

        def _row(label, text, label_color=C_NAVY, even=True):
            r = current_row[0]
            ws_ai[f"B{r}"].value     = label
            ws_ai[f"B{r}"].font      = _font(bold=True, size=10, color=label_color)
            ws_ai[f"B{r}"].fill      = _fill(C_HROW if even else C_WHITE)
            ws_ai[f"B{r}"].alignment = _al("left")
            ws_ai[f"B{r}"].border    = _border()
            ws_ai[f"C{r}"].value     = text
            ws_ai[f"C{r}"].font      = _font(size=10, color=C_GRAY)
            ws_ai[f"C{r}"].fill      = _fill(C_HROW if even else C_WHITE)
            ws_ai[f"C{r}"].alignment = _al("left", wrap=True)
            ws_ai[f"C{r}"].border    = _border()
            ws_ai.row_dimensions[r].height = 20
            current_row[0] += 1

        def _blank():
            current_row[0] += 1

        # ── compute derived stats ──────────────────────────────────────
        try:
            total_inv_f  = float(total_inv) if total_inv else 1
            total_val_f  = float(total_val) if total_val else 0
            total_pnl_f  = float(total_pnl) if total_pnl else 0
            daily_pnl_f  = float(daily_pnl) if daily_pnl else 0
            booked_f     = float(booked)    if booked    else 0
            beta_f       = float(beta)      if beta      else 1.0

            # concentration: top-5 by value
            top5     = sorted_df.nlargest(5, "Value") if not sorted_df.empty else pd.DataFrame()
            top5_pct = (top5["Value"].sum() / total_val_f * 100) if total_val_f else 0
            top5_list = ", ".join(top5["Ticker"].tolist()) if not top5.empty else "—"

            # Exclude F&O from gainers/losers (equity/ETF/Bond only)
            _FNO_TYPES_XL = {"F&O", "Index F&O", "Currency F&O"}
            _eq_df_xl = calc_df[~calc_df.get("Asset_Type", pd.Series(dtype=str)).isin(_FNO_TYPES_XL)].copy() \
                        if "Asset_Type" in calc_df.columns else calc_df.copy()
            if _eq_df_xl.empty:
                _eq_df_xl = calc_df.copy()

            # worst losers (equity only)
            losers   = _eq_df_xl[_eq_df_xl["Unrealized_PnL"] < 0].sort_values("Unrealized_PnL")
            losers_list = ", ".join(losers["Ticker"].head(5).tolist()) if not losers.empty else "None"
            losers_pnl  = losers["Unrealized_PnL"].sum() if not losers.empty else 0

            # best gainers (equity only)
            gainers  = _eq_df_xl[_eq_df_xl["Unrealized_PnL"] > 0].sort_values("Unrealized_PnL", ascending=False)
            gainers_list = ", ".join(gainers["Ticker"].head(5).tolist()) if not gainers.empty else "None"
            gainers_pnl  = gainers["Unrealized_PnL"].sum() if not gainers.empty else 0

            # diversification score (0-100)
            n_stocks   = len(calc_df)
            n_sectors  = calc_df["Asset_Type"].nunique() if "Asset_Type" in calc_df.columns else 1
            div_score  = min(100, int((n_stocks * 4) + (n_sectors * 5) + (100 - top5_pct)))

            # riskometer
            if   beta_f >= 1.4:  risk_label, risk_color = "⚠️  VERY HIGH RISK",  C_RED
            elif beta_f >= 1.2:  risk_label, risk_color = "🔶  HIGH RISK",        C_RED
            elif beta_f >= 0.9:  risk_label, risk_color = "🟡  MODERATE RISK",    C_GOLD
            elif beta_f >= 0.7:  risk_label, risk_color = "🟢  MODERATE-LOW",     C_GREEN
            else:                risk_label, risk_color = "🛡️  LOW RISK",          C_GREEN

            # dilution check: F&O + high-beta % of portfolio
            fno_inv_f  = float(s.get("fno_invested", 0) or 0)
            fno_pct    = (fno_inv_f / total_inv_f * 100) if total_inv_f else 0

            # XIRR stats
            xirr_values = [v for v in xirr_map.values() if isinstance(v, (int, float)) and not pd.isna(v)]
            avg_xirr    = sum(xirr_values) / len(xirr_values) if xirr_values else 0
            best_xirr   = max(xirr_map, key=lambda k: xirr_map.get(k, -999) if isinstance(xirr_map.get(k), (int,float)) else -999, default="—")
            worst_xirr  = min(xirr_map, key=lambda k: xirr_map.get(k, 999)  if isinstance(xirr_map.get(k), (int,float)) else 999,  default="—")

            # trade stats
            if not trades_df.empty:
                n_trades    = len(trades_df)
                win_trades  = trades_df[trades_df["Booked_PnL"] > 0] if "Booked_PnL" in trades_df.columns else pd.DataFrame()
                win_rate    = f"{len(win_trades)/n_trades*100:.1f}%" if n_trades else "—"
            else:
                n_trades, win_rate = 0, "—"

            # mistakes list
            mistakes = []
            if top5_pct > 60:
                mistakes.append(f"High concentration — top 5 stocks ({top5_list}) = {top5_pct:.1f}% of portfolio. Over-exposure to single-stock risk.")
            if fno_pct > 20:
                mistakes.append(f"Excessive F&O allocation at {fno_pct:.1f}% of portfolio. Derivatives magnify losses; reduce to < 20%.")
            if beta_f > 1.3:
                mistakes.append(f"Portfolio beta {beta_f:.2f} is high — the portfolio moves {beta_f:.2f}x the broader market. Consider adding defensive stocks (FMCG, Pharma, Utilities).")
            if not losers.empty and abs(losers_pnl) > 0.1 * total_inv_f:
                mistakes.append(f"Large unrealised losses of {_inr(losers_pnl)} from {losers_list}. Review stop-losses and exit weak positions.")
            if n_stocks < 8:
                mistakes.append(f"Under-diversified — only {n_stocks} holdings. Aim for 15-20 diversified positions to reduce unsystematic risk.")
            if avg_xirr < 8 and avg_xirr != 0:
                mistakes.append(f"Average XIRR of {avg_xirr:.1f}% is below the risk-free rate. Revisit position sizing and stock selection.")
            if not mistakes:
                mistakes.append("No significant structural mistakes detected — portfolio is well-composed.")

            # recommendations
            recs = []
            if top5_pct > 50:
                recs.append("Consider trimming top concentrated positions and deploying into 2-3 under-represented sectors.")
            if beta_f > 1.2:
                recs.append("Add low-beta defensive counters (HINDUNILVR, ITC, POWERGRID, NTPC) to buffer downside volatility.")
            if fno_pct > 15:
                recs.append("Reduce F&O exposure. Use hedging (protective puts) instead of speculative F&O to limit downside.")
            if not losers.empty:
                recs.append(f"Set strict stop-losses on {losers_list}. Consider averaging down only after fundamental confirmation.")
            if n_stocks < 10:
                recs.append("Diversify into new sectors — consider IT, Healthcare, or Infrastructure to spread risk.")
            if avg_xirr > 15:
                recs.append(f"Strong XIRR of {avg_xirr:.1f}% — book partial profits on best performers and redeploy into laggards or safer assets.")
            recs.append("Review portfolio quarterly and rebalance if any single position exceeds 15% of total value.")
            if not recs:
                recs.append("Maintain current allocation. Continue monitoring beta and concentration metrics quarterly.")

        except Exception as e:
            risk_label, risk_color = "— DATA ERROR —", C_GRAY
            div_score = 0
            mistakes  = [f"Error computing analysis: {e}"]
            recs      = ["Please re-check data."]
            top5_list = "—"; top5_pct = 0; beta_f = 1.0
            n_stocks = 0; n_trades = 0; win_rate = "—"
            losers_list = "—"; avg_xirr = 0
            best_xirr = "—"; worst_xirr = "—"
            fno_pct = 0; booked_f = 0; gainers_list = "—"
            xirr_values = []

        # ── write AI analysis ─────────────────────────────────────────
        _section("  📊  PORTFOLIO SNAPSHOT", C_NAVY)
        _row("Client",          client_name, even=True)
        _row("Report Date",     IST.strftime("%d %b %Y, %I:%M %p IST"), even=False)
        _row("Total Invested",  _inr(total_inv), even=True)
        _row("Current Value",   _inr(total_val), even=False)
        _row("Unrealized P&L",  f"{_inr(total_pnl)}  ({_pct(s.get('total_pnl_pct',0))})", even=True)
        _row("Today's P&L",     _inr(daily_pnl), even=False)
        _row("Booked P&L",      _inr(booked),    even=True)
        _row("No. of Holdings", str(n_stocks),   even=False)
        _blank()

        _section("  🎯  RISKOMETER", C_BLUE)
        _row("Portfolio Beta",    f"{beta_f:.3f}", even=True)
        _row("Risk Rating",       risk_label,      even=False)
        _row("Top-5 Concentration", f"{top5_list}  →  {top5_pct:.1f}% of portfolio", even=True)
        _row("F&O Allocation",    f"{fno_pct:.1f}% of portfolio", even=False)
        _row("Diversification Score", f"{div_score}/100  {'(Good)' if div_score>=60 else '(Needs Work)'}", even=True)
        _row("No. of Asset Types",str(n_sectors if 'n_sectors' in dir() else '—'), even=False)
        _blank()

        _section("  📈  PERFORMANCE ANALYSIS", "1565C0")
        _row("Best Gainers",     gainers_list, even=True)
        _row("Total Gain (Gainers)", _inr(gainers_pnl), even=False)
        _row("Worst Losers",     losers_list,  even=True)
        _row("Total Loss (Losers)",  _inr(losers_pnl), even=False)
        _row("Avg Portfolio XIRR",  f"{avg_xirr:.1f}%" if xirr_values else "—", even=True)
        _row("Best XIRR Holding",   str(best_xirr), even=False)
        _row("Worst XIRR Holding",  str(worst_xirr), even=True)
        _row("Total Sell Trades",   str(n_trades), even=False)
        _row("Win Rate (Trades)",   win_rate, even=True)
        _blank()

        _section("  ⚠️  MISTAKES IDENTIFIED", C_RED)
        for i, m in enumerate(mistakes):
            _row(f"Issue {i+1}", m, label_color=C_RED, even=(i%2==0))
        _blank()

        _section("  💡  RECOMMENDATIONS", C_GREEN)
        for i, r in enumerate(recs):
            _row(f"Action {i+1}", r, label_color=C_GREEN, even=(i%2==0))
        _blank()

        _section("  ⚖️  DISCLAIMER", C_GRAY)
        _row("Note", ("This report is generated for informational purposes only by Northeast Broking Services Limited. "
                       "All data is sourced from Angel One. This does not constitute investment advice. "
                       "Please consult a SEBI-registered advisor before making investment decisions."), even=True)

        # ──────────────────────────────────────────────────────────────
        # SHEET 4 — CAPITAL GAINS
        # ──────────────────────────────────────────────────────────────
        ws_cg = wb.create_sheet("Capital Gains")
        ws_cg.sheet_view.showGridLines = False
        ws_cg.freeze_panes = "A3"

        _merge_title(ws_cg, "A1:H1",
                     f"CAPITAL GAINS REPORT  ·  {client_name}",
                     C_NAVY, C_WHITE, size=12, height=30)

        # summary block first
        if not cg_df.empty:
            stcg = cg_df[cg_df["Gain Type"] == "STCG"]["Booked P&L ₹"].sum() if "Gain Type" in cg_df.columns else 0
            ltcg = cg_df[cg_df["Gain Type"] == "LTCG"]["Booked P&L ₹"].sum() if "Gain Type" in cg_df.columns else 0
        else:
            stcg = ltcg = 0

        ws_cg.append(["", "STCG Total", _inr(stcg), "", "LTCG Total", _inr(ltcg), "", ""])
        r = ws_cg.max_row
        ws_cg.row_dimensions[r].height = 24
        for col, color in [(2, C_GOLD), (3, C_GOLD), (5, C_BLUE), (6, C_BLUE)]:
            ws_cg.cell(r, col).fill      = _fill(color)
            ws_cg.cell(r, col).font      = _font(bold=True, color=C_WHITE, size=11)
            ws_cg.cell(r, col).alignment = _al("center")
            ws_cg.cell(r, col).border    = _border()

        ws_cg.append([])

        cg_headers = ["#", "Ticker", "Sell Qty", "Sell Price ₹", "Avg Buy ₹",
                       "Sell Date", "Holding Days", "Gain Type", "Booked P&L ₹"]
        ws_cg.append(cg_headers)
        _style_header_row(ws_cg, ws_cg.max_row, C_BLUE, C_WHITE)

        if not cg_df.empty:
            for i, (_, tr) in enumerate(cg_df.iterrows()):
                booked_v = float(tr.get("Booked P&L ₹", 0) or 0)
                ws_cg.append([
                    i + 1,
                    tr.get("Ticker", ""),
                    _num(tr.get("Sell Qty"), 2),
                    _inr(tr.get("Sell Price ₹")),
                    _inr(tr.get("Avg Buy ₹")),
                    str(tr.get("Sell Date", "")),
                    tr.get("Holding Days", ""),
                    tr.get("Gain Type", ""),
                    round(booked_v, 2),
                ])
                r = ws_cg.max_row
                _style_data_row(ws_cg, r, even=(i % 2 == 0))
                pnl_col = ws_cg.cell(r, 9)
                pnl_col.fill = _fill(C_GREEN_L) if booked_v >= 0 else _fill(C_RED_L)
                pnl_col.font = _font(bold=True, color=C_GREEN if booked_v >= 0 else C_RED)

        _set_col_widths(ws_cg, {"A":5,"B":14,"C":10,"D":14,"E":14,"F":14,"G":14,"H":10,"I":16})

        # ──────────────────────────────────────────────────────────────
        # SHEET 5 — TRADE HISTORY
        # ──────────────────────────────────────────────────────────────
        ws_tr = wb.create_sheet("Trade History")
        ws_tr.sheet_view.showGridLines = False
        ws_tr.freeze_panes = "A3"

        _merge_title(ws_tr, "A1:H1",
                     f"COMPLETE TRADE HISTORY  ·  {client_name}",
                     C_NAVY, C_WHITE, size=12, height=30)

        tr_headers = ["#", "Ticker", "Asset Type", "Sell Qty",
                       "Sell Price ₹", "Avg Buy ₹", "Sell Date", "Booked P&L ₹"]
        ws_tr.append(tr_headers)
        _style_header_row(ws_tr, ws_tr.max_row, C_BLUE, C_WHITE)

        if not trades_df.empty:
            for i, (_, tr) in enumerate(trades_df.iterrows()):
                booked_v = float(tr.get("Booked_PnL", 0) or 0)
                ws_tr.append([
                    i + 1,
                    tr.get("Ticker", ""),
                    tr.get("Asset_Type", ""),
                    _num(tr.get("Sell_Qty"), 2),
                    _inr(tr.get("Sell_Price")),
                    _inr(tr.get("Buy_Price_At_Sell")),
                    str(tr.get("Sell_Date", "")),
                    round(booked_v, 2),
                ])
                r = ws_tr.max_row
                _style_data_row(ws_tr, r, even=(i % 2 == 0))
                pnl_cell = ws_tr.cell(r, 8)
                pnl_cell.fill = _fill(C_GREEN_L) if booked_v >= 0 else _fill(C_RED_L)
                pnl_cell.font = _font(bold=True, color=C_GREEN if booked_v >= 0 else C_RED)

            # totals
            t_r = ws_tr.max_row + 1
            total_booked_trades = trades_df["Booked_PnL"].sum() if "Booked_PnL" in trades_df.columns else 0
            ws_tr.append(["", "TOTAL", "", "", "", "", "", round(float(total_booked_trades), 2)])
            for col in range(1, 9):
                c = ws_tr.cell(t_r, col)
                c.fill   = _fill(C_NAVY)
                c.font   = _font(bold=True, color=C_WHITE)
                c.alignment = _al("center")
                c.border = _border()

        _set_col_widths(ws_tr, {"A":5,"B":14,"C":16,"D":10,"E":16,"F":16,"G":14,"H":16})

        # ──────────────────────────────────────────────────────────────
        # TAB ORDER + COLOURS
        # ──────────────────────────────────────────────────────────────
        tab_colors = {"Cover":"0D1B2A", "Holdings":"1565C0",
                      "AI Analysis":"1B5E20", "Capital Gains":"F57F17",
                      "Trade History":"37474F"}
        for sname, color in tab_colors.items():
            if sname in wb.sheetnames:
                wb[sname].sheet_properties.tabColor = color

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    except Exception as exc:
        # fallback: return None so caller can show error
        import traceback
        traceback.print_exc()
        return None


# =====================================================
# ── WHATSAPP MESSAGE GENERATOR ───────────────────────
# =====================================================

def generate_whatsapp_message(client_name, summary_dict, calc_df):
    """Return a formatted WhatsApp portfolio update message."""
    s = summary_dict
    def _inr(v):
        try:
            v = float(v)
            if abs(v) >= 1e7:
                return f"₹{v/1e7:.2f}Cr"
            elif abs(v) >= 1e5:
                return f"₹{v/1e5:.2f}L"
            else:
                return f"₹{v:,.0f}"
        except:
            return "—"
    def _pct(v):
        try: return f"{float(v):+.2f}%"
        except: return "—"

    daily_pnl = float(s.get("daily_pnl", 0))
    total_pnl = float(s.get("total_pnl", 0))
    daily_emoji = "📈" if daily_pnl >= 0 else "📉"
    overall_emoji = "🟢" if total_pnl >= 0 else "🔴"

    # Top gainers today
    top_gainers = ""
    if not calc_df.empty:
        gdf = calc_df.dropna(subset=["Daily_PnL_%"]).nlargest(3, "Daily_PnL_%")
        for _, g in gdf.iterrows():
            top_gainers += f"   • {g['Ticker']}: {float(g['Daily_PnL_%']):+.2f}%\n"

    msg = f"""🏢 *Northeast Broking Services*
📋 *Portfolio Update — {client_name}*
📅 {datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime('%d %b %Y, %I:%M %p IST')}

━━━━━━━━━━━━━━━━━━━
💰 *Invested:* {_inr(s.get('total_invested',0))}
   ├ Holdings: {_inr(s.get('stocks_invested',0))}
   ├ ETF: {_inr(s.get('etf_invested',0))}
   ├ F&O: {_inr(s.get('fno_invested',0))}
   ├ REIT/InvIT: {_inr(s.get('reit_invested',0))}
   └ Other: {_inr(s.get('other_invested',0))}

📦 *Portfolio Value:* {_inr(s.get('total_value',0))}
{daily_emoji} *Today's P&L:* {_inr(s.get('daily_pnl',0))} ({_pct(s.get('daily_pnl_pct',0))})
{overall_emoji} *Total P&L:* {_inr(s.get('total_pnl',0))} ({_pct(s.get('total_pnl_pct',0))})
✅ *Booked P&L:* {_inr(s.get('total_booked_pnl',0))}
━━━━━━━━━━━━━━━━━━━
📊 *Top Movers Today:*
{top_gainers if top_gainers else '   Data unavailable'}
━━━━━━━━━━━━━━━━━━━
_For queries, contact your advisor._
"""
    return msg.strip()


# =====================================================
# ── CRR FILE UPLOAD & RECONCILIATION ─────────────────
# =====================================================

def reconcile_crr(crr_df, portfolio_df):
    """
    Compare CRR holdings vs portfolio holdings.
    Returns a DataFrame highlighting matches, mismatches and missing entries.
    """
    crr_df = crr_df.copy()
    port   = portfolio_df.copy()

    # Normalise column names — look for common variants
    crr_df.columns = [c.strip() for c in crr_df.columns]
    col_map = {}
    for c in crr_df.columns:
        cl = c.lower().replace(" ","").replace("_","")
        if cl in ("symbol","scrip","scripname","stock","ticker","isin"):
            col_map[c] = "Symbol"
        elif cl in ("qty","quantity","netqty","holdings","shares"):
            col_map[c] = "CRR_Qty"
        elif cl in ("avgcost","avgbuyprice","avgprice","costprice","buyprice"):
            col_map[c] = "CRR_AvgBuy"
    crr_df = crr_df.rename(columns=col_map)

    if "Symbol" not in crr_df.columns:
        return None, "Could not find Symbol/Ticker column in CRR file."
    if "CRR_Qty" not in crr_df.columns:
        return None, "Could not find Qty column in CRR file."

    crr_df["Symbol"] = crr_df["Symbol"].astype(str).str.upper().str.strip()
    crr_df["CRR_Qty"] = pd.to_numeric(crr_df["CRR_Qty"], errors="coerce")

    port["Symbol"] = port["Ticker"].str.upper().str.replace(".NS","").str.replace(".BO","")
    port["Port_Qty"] = pd.to_numeric(port["Shares"], errors="coerce")

    merged = pd.merge(crr_df[["Symbol","CRR_Qty"] + (["CRR_AvgBuy"] if "CRR_AvgBuy" in crr_df.columns else [])],
                      port[["Symbol","Port_Qty","Buy_Price"]],
                      on="Symbol", how="outer")

    def _status(row):
        crr = row.get("CRR_Qty")
        prt = row.get("Port_Qty")
        if pd.isna(crr):
            return "⚠️ In Portfolio Only"
        if pd.isna(prt):
            return "❌ In CRR Only"
        diff = abs(float(crr) - float(prt))
        return "✅ Match" if diff < 0.01 else f"🔴 Mismatch (diff={diff:.2f})"

    merged["Status"] = merged.apply(_status, axis=1)
    merged["Qty Diff"] = (merged["CRR_Qty"].fillna(0) - merged["Port_Qty"].fillna(0)).round(2)
    return merged, None


# =====================================================
# =========================================================
# ASSET TYPE CLASSIFIER  (comprehensive — v2)
# =========================================================

# ── F&O indicators (checked first — highest priority) ──────────────
_FNO_SUFFIXES   = ("FUT", "CE", "PE")
_FNO_SUBSTRINGS = ("-FUT", "-CE", "-PE", "FUTURES", "FUTURE", "CALLS", "PUTS")

# ── Well-known index / commodity / currency F&O bases ───────────────
_INDEX_FNO_BASES = {
    "NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX",
    "BANKEX", "NIFTYNXT50", "INDIA VIX", "INDIAVIX",
    "USDINR", "EURINR", "GBPINR", "JPYINR",         # currency futures
}

# ── ETF patterns (keywords anywhere in ticker) ───────────────────────
_ETF_KEYWORDS = (
    "BEES", "ETF", "IETF",
    "GOLD", "SILVER", "COPPER", "ZINC", "LEAD", "NICKEL",   # commodity ETFs
    "LIQUIDBEES", "LIQUALGO",
    "N50", "100ETF", "MIDCAP", "SMALLCAP", "CPSE", "PSU",
    "MOMENTUM", "ALPHA", "QUALIT", "VALUE",                  # factor ETFs
    "IT", "INFRA", "AUTO", "PHARMA", "FMCG", "BANK", "REALTY",  # sector ETFs
    "HANG", "NASDAQ", "SP500", "MAFANG",                     # international ETFs
)
# Explicit ETF tickers that wouldn't match keywords
_ETF_EXPLICIT = {
    "MON100", "NIFTY50", "JUNIORBEES", "NIFTYBEES", "SETFNIF50",
    "ICICINIFTY", "HDFCNIFTY", "KOTAK50", "SBINIFTY",
    "CPSEETF", "PSUBNKBEES",
}

# ── REIT / InvIT ──────────────────────────────────────────────────────
_REIT_KEYWORDS  = ("REIT", "INVIT", "NXTREIT", "BROOKFIELD", "MINDSPACE", "NEXUS")
_REIT_EXPLICIT  = {
    "EMBASSY", "MINDSPACE", "BROOKFIELD", "NEXUS", "POWERGRID", "INDIGRID",
    "IRBINVIT", "NHPCINVIT",
}

# ── Sovereign Gold Bond ───────────────────────────────────────────────
_SGB_KEYWORDS   = ("SGBFEB", "SGBMAR", "SGBAPR", "SGBMAY", "SGBJUN",
                   "SGBJUL", "SGBAUG", "SGBSEP", "SGBOCT", "SGBNOV",
                   "SGBDEC", "SGBJAN", "SGOLD")

# ── Bonds / NCDs / Debentures ────────────────────────────────────────
_BOND_KEYWORDS  = ("NCD", "BOND", "DEBENTURE", "TBILL", "GSEC",
                   "N2", "N3", "N5", "N10")   # G-Sec ETFs sometimes use Nx codes

# ── Mutual Fund units (BSE-listed) ───────────────────────────────────
_MF_KEYWORDS    = ("MF", "MUTUALFUND", "NAVI", "GROWW", "MIRAE",
                   "PARAG", "MOTILAL")

# ── Preference shares ────────────────────────────────────────────────
_PREF_KEYWORDS  = ("PREF", "-PR", "PREFSHARE")

_INSTRUMENT_ORDER = [
    "F&O", "Index F&O", "Currency F&O",
    "ETF", "Liquid ETF", "Commodity ETF", "International ETF",
    "REIT/InvIT", "SGB", "Bond/NCD", "Mutual Fund",
    "Preference Share", "Unlisted Share", "Stock",
]

def classify_asset(ticker: str) -> str:
    """
    Classify any Indian market instrument into a detailed category.

    Returns one of:
      F&O              – single-stock futures / options
      Index F&O        – index-based futures / options (NIFTY, BANKNIFTY, …)
      Currency F&O     – currency futures / options (USDINR, EURINR, …)
      ETF              – equity / index ETF
      Liquid ETF       – overnight / liquid fund ETFs (LIQUIDBEES, etc.)
      Commodity ETF    – gold / silver / commodity ETFs
      International ETF– NASDAQ, Hang Seng, etc. ETFs
      REIT/InvIT       – listed real estate / infrastructure trusts
      SGB              – Sovereign Gold Bonds
      Bond/NCD         – listed bonds, NCDs, G-Secs, T-bills
      Mutual Fund      – BSE-listed MF units
      Preference Share – preference / convertible preference shares
      Stock            – equity (default)
    """
    raw = str(ticker).strip().upper()
    t   = raw.replace(".NS", "").replace(".BO", "").replace(".MCX", "").strip()

    # ── 1. F&O detection ────────────────────────────────────────────
    if any(t.endswith(s) for s in _FNO_SUFFIXES) or \
       any(s in t for s in _FNO_SUBSTRINGS):
        # Identify if it's index or currency F&O
        base = t
        for s in _FNO_SUFFIXES:
            if base.endswith(s):
                base = base[:-len(s)]
                break
        # Strip date-like suffix (e.g. "26MAY25")
        base_clean = re.sub(r'\d{2}[A-Z]{3}\d{2}', '', base)
        base_clean = re.sub(r'[A-Z]{3}\d{2}', '', base_clean)
        if any(base_clean.startswith(idx) for idx in _INDEX_FNO_BASES) or \
           base_clean in _INDEX_FNO_BASES:
            if any(c in base_clean for c in ("USDINR","EURINR","GBPINR","JPYINR")):
                return "Currency F&O"
            return "Index F&O"
        if any(c in base_clean for c in ("USDINR","EURINR","GBPINR","JPYINR")):
            return "Currency F&O"
        return "F&O"

    # ── 2. SGB ──────────────────────────────────────────────────────
    if any(t.startswith(k) for k in _SGB_KEYWORDS) or "SGOLD" in t:
        return "SGB"

    # ── 3. REIT / InvIT ─────────────────────────────────────────────
    if any(k in t for k in _REIT_KEYWORDS) or t in _REIT_EXPLICIT:
        return "REIT/InvIT"

    # ── 4. Bond / NCD ───────────────────────────────────────────────
    if any(k in t for k in _BOND_KEYWORDS):
        return "Bond/NCD"

    # ── 5. ETF sub-classification ───────────────────────────────────
    is_etf = t in _ETF_EXPLICIT or any(k in t for k in _ETF_KEYWORDS)
    if is_etf:
        if "LIQUID" in t or "OVERNIGHT" in t:
            return "Liquid ETF"
        if any(k in t for k in ("GOLD","SILVER","COPPER","ZINC","NICKEL","LEAD")):
            return "Commodity ETF"
        if any(k in t for k in ("HANG","NASDAQ","SP500","MAFANG","MON100")):
            return "International ETF"
        return "ETF"

    # ── 6. Mutual Fund ──────────────────────────────────────────────
    if any(k in t for k in _MF_KEYWORDS):
        return "Mutual Fund"

    # ── 7. Preference Share ─────────────────────────────────────────
    if any(k in t for k in _PREF_KEYWORDS):
        return "Preference Share"

    # ── 8. Default: Equity Stock ────────────────────────────────────
    return "Stock"


def asset_type_emoji(asset_type: str) -> str:
    """Return a display emoji for each asset type."""
    return {
        "Stock":            "🏦",
        "ETF":              "📊",
        "Liquid ETF":       "💧",
        "Commodity ETF":    "🪙",
        "International ETF":"🌏",
        "F&O":              "⚡",
        "Index F&O":        "📈",
        "Currency F&O":     "💱",
        "REIT/InvIT":       "🏢",
        "SGB":              "🥇",
        "Bond/NCD":         "📄",
        "Mutual Fund":      "🌱",
        "Preference Share": "🔵",
        "Unlisted Share":   "🔒",
    }.get(asset_type, "🔹")


def is_equity_like(asset_type: str) -> bool:
    """Return True for asset types that behave like equities for P&L / beta calculations."""
    return asset_type in ("Stock", "ETF", "Liquid ETF", "Commodity ETF",
                          "International ETF", "REIT/InvIT", "SGB",
                          "Preference Share")


def get_exchange(ticker):
    """Return 'BSE', 'NSE', or '' based on ticker suffix."""
    t = str(ticker).upper().strip()
    if t.endswith(".BO"):
        return "BSE"
    if t.endswith(".NS"):
        return "NSE"
    return ""

# =========================================================
# PORTFOLIO FILE
# =========================================================

# PORTFOLIO_FILE and TRADES_FILE are set dynamically above based on selected client

if os.path.exists(PORTFOLIO_FILE):
    df = pd.read_csv(PORTFOLIO_FILE)
else:
    df = pd.DataFrame(
        columns=["Ticker", "Shares", "Buy_Price", "Buy_Date"]
    )
    df.to_csv(PORTFOLIO_FILE, index=False)
    _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')

# ── Normalise Asset_Type from bulk-import variants → canonical names ──
# Maps values like "Bond", "bond", "NCD", "Equity", "MF" etc.
# to the exact strings Roh.py tab system expects.
_ASSET_TYPE_NORMALISE = {
    # Stocks
    "equity": "Stock", "equities": "Stock", "shares": "Stock",
    "eq": "Stock", "cash equity": "Stock", "nse": "Stock", "bse": "Stock",
    # ETF
    "exchange traded fund": "ETF", "exchange-traded fund": "ETF", "index etf": "ETF",
    # Mutual Fund
    "mf": "Mutual Fund", "fund": "Mutual Fund", "sip": "Mutual Fund",
    "direct fund": "Mutual Fund", "mutual fund": "Mutual Fund",
    # Bond/NCD — ALL variants collapse to "Bond/NCD"
    "bond": "Bond/NCD", "bonds": "Bond/NCD",
    "ncd": "Bond/NCD", "debenture": "Bond/NCD",
    "g-sec": "Bond/NCD", "gsec": "Bond/NCD",
    "government bond": "Bond/NCD", "govt bond": "Bond/NCD",
    "t-bill": "Bond/NCD", "tbill": "Bond/NCD",
    "corporate bond": "Bond/NCD", "fixed income": "Bond/NCD", "debt": "Bond/NCD",
    # SGB
    "sovereign gold bond": "SGB", "sg bond": "SGB",
    # REIT/InvIT
    "reit": "REIT/InvIT", "reits": "REIT/InvIT",
    "real estate investment trust": "REIT/InvIT",
    "invit": "REIT/InvIT", "infrastructure investment trust": "REIT/InvIT",
    # F&O
    "fo": "F&O", "futures": "F&O", "future": "F&O",
    "options": "F&O", "option": "F&O", "derivatives": "F&O",
    # ETF sub-types
    "commodity etf": "Commodity ETF", "gold etf": "Commodity ETF",
    "liquid etf": "Liquid ETF",
    # Preference Share
    "preference share": "Preference Share", "pref share": "Preference Share",
    "unlisted share": "Unlisted Share", "unlisted shares": "Unlisted Share",
    "unlisted": "Unlisted Share", "pre-ipo": "Unlisted Share", "pre ipo": "Unlisted Share",
}

if "Asset_Type" in df.columns:
    df["Asset_Type"] = df["Asset_Type"].apply(
        lambda v: _ASSET_TYPE_NORMALISE.get(str(v).strip().lower(), str(v).strip())
        if pd.notna(v) and str(v).strip() else v
    )

# ── Sell Trades ────────────────────────────────────────────────────
if os.path.exists(TRADES_FILE):
    trades_df = pd.read_csv(TRADES_FILE)
else:
    trades_df = pd.DataFrame(
        columns=["Ticker", "Sell_Qty", "Sell_Price", "Sell_Date",
                 "Buy_Price_At_Sell", "Booked_PnL", "Asset_Type"]
    )
    trades_df.to_csv(TRADES_FILE, index=False)
    _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')

# ── Auto-classify Asset_Type for trades that lack it ─────────────
import re as _re_mod

def _classify_ticker_to_asset_type(ticker: str) -> str:
    """Infer asset type from ticker symbol patterns (F&O, ETF, Stock, etc.)."""
    t = str(ticker).strip().upper()
    # F&O option patterns: SYMBOL + DDMMMYYYY or DDMONYR + strike + CE/PE
    if _re_mod.search(r'\d{2}[A-Z]{3}\d{2,4}\d+[CP]E?$', t):
        return "F&O"
    # F&O futures pattern: SYMBOL + DDMMMYYYY + FUT
    if _re_mod.search(r'\d{2}[A-Z]{3}\d{2,4}FUT$', t):
        return "F&O"
    # NSE F&O monthly expiry shorthand: SYMBOLYYMMMSTRIKECE/PE
    if _re_mod.search(r'\d{2}[A-Z]{3}\d+[CP]E?$', t):
        return "F&O"
    # ETF keywords
    if any(k in t for k in ["ETF","BEES","NIFTYETF","BANKBEES","GOLDBEES","LIQUIDBEES","ICICINIFTY","MAFANG","N50","JUNIORBEES"]):
        return "ETF"
    # SGB
    if "SGB" in t or "SGBFEB" in t or t.startswith("SGB"):
        return "SGB"
    # REIT/InvIT
    if any(k in t for k in ["REIT","INVIT","EMBASSY","MINDSPACE","NEXUS","INDIGRID","POWERGRID","BROOKFIELD"]):
        return "REIT/InvIT"
    # Bond indicators
    if any(k in t for k in ["BOND","NCD","GSEC","TBILL","DEBENTURE"]):
        return "Bond/NCD"
    return "Stock"

if "Asset_Type" not in trades_df.columns:
    trades_df["Asset_Type"] = ""

# Build ticker→asset_type lookup from portfolio for enrichment
_pf_asset_map = {}
if "Asset_Type" in df.columns and "Ticker" in df.columns:
    for _, _pf_row in df.iterrows():
        _t = str(_pf_row.get("Ticker","")).strip()
        _a = str(_pf_row.get("Asset_Type","")).strip()
        if _t and _a:
            _pf_asset_map[_t] = _a

# Fill missing Asset_Type in trades: portfolio lookup first, then regex
_trades_changed = False
for _idx, _tr in trades_df.iterrows():
    if not trades_df.at[_idx, "Asset_Type"] or str(trades_df.at[_idx, "Asset_Type"]).strip() in ("", "nan", "None"):
        _tk = str(_tr.get("Ticker","")).strip()
        _at = _pf_asset_map.get(_tk) or _classify_ticker_to_asset_type(_tk)
        trades_df.at[_idx, "Asset_Type"] = _at
        _trades_changed = True
if _trades_changed and os.path.exists(TRADES_FILE):
    trades_df.to_csv(TRADES_FILE, index=False)
    _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')


def compute_net_holdings(portfolio_df, trades_df):
    """
    Compute net remaining quantity per ticker by subtracting total sold qty
    from total buy qty.  Buy_Price is NEVER touched — sell trades have zero
    effect on the weighted average buy cost.

    Strategy: collapse all buy lots for a ticker into one aggregate row so
    the Shares column simply reflects (total_bought - total_sold).  The
    weighted average Buy_Price across all lots is preserved on that row.

    Returns:
        net_df        — one row per ticker (Buy_Price = wtd avg, Shares = net)
        booked_pnl_map — {ticker: total_booked_pnl}
    """
    if portfolio_df.empty:
        out = portfolio_df.copy()
        out["Sold_Qty"] = 0.0
        return out, {}

    # ── Step 1: collapse buy lots into one row per ticker (weighted avg price) ──
    rows = []
    for ticker, grp in portfolio_df.groupby("Ticker", sort=False):
        total_qty = float(grp["Shares"].sum())
        if total_qty == 0:
            continue
        wtd_price = round(
            (grp["Shares"] * grp["Buy_Price"]).sum() / total_qty, 4
        )
        # Keep first row's metadata (Buy_Date, Asset_Type, etc.)
        base = grp.iloc[0].copy()
        base["Shares"]    = total_qty
        base["Buy_Price"] = wtd_price
        rows.append(base)

    if not rows:
        out = portfolio_df.iloc[:0].copy()
        out["Sold_Qty"] = 0.0
        return out, {}

    net_df = pd.DataFrame(rows).reset_index(drop=True)
    net_df["Sold_Qty"] = 0.0

    if trades_df.empty:
        return net_df, {}

    booked_pnl_map = {}

    # ── Step 2: subtract sold qty — Buy_Price stays untouched ────────────────
    for ticker, t_group in trades_df.groupby("Ticker"):
        total_sold = float(t_group["Sell_Qty"].sum())
        booked_pnl_map[ticker] = float(t_group["Booked_PnL"].sum())

        mask = net_df["Ticker"] == ticker
        if not mask.any():
            continue

        idx = net_df[mask].index[0]
        held = float(net_df.at[idx, "Shares"])
        net_df.at[idx, "Shares"]   = round(max(held - total_sold, 0.0), 6)
        net_df.at[idx, "Sold_Qty"] = round(total_sold, 6)

    return net_df, booked_pnl_map

# =========================================================
# SIDEBAR — MANAGE PORTFOLIO (DEVELOPER ONLY)
# =========================================================

with st.sidebar:

    if _is_dev:
        # ── Developer: show which client portfolio is being managed ──
        _managing_name = st.session_state.get("dev_client_name", "—")
        st.markdown(f"""
<div style="font-size:11px;font-weight:700;color:#f5c842;text-transform:uppercase;
            letter-spacing:1px;margin-bottom:10px;">
  ✏️ Add / Delete / Sell — {_managing_name}
</div>
""", unsafe_allow_html=True)

        # ── 📂 IMPORT FROM EXCEL ─────────────────────────────────────
        with st.expander("📂 Import Holdings from Excel / CSV", expanded=False):

            # ── Column guide (collapsible) ────────────────────────────
            with st.expander("📋 Accepted Column Names — click to expand", expanded=False):
                st.markdown("""
<div style="background:#0f1535;border:1px solid #2e3355;border-radius:10px;
            padding:12px 14px;font-size:11px;line-height:2.2;">
  <b style="color:#f0f2ff;">Any of these column names are auto-detected:</b><br>
  <table style="width:100%;border-collapse:collapse;color:#b8bcd8;margin-top:6px;">
    <tr style="border-bottom:1px solid #1e2340;">
      <th style="text-align:left;padding:4px 10px 4px 0;color:#f5c842;font-size:10px;letter-spacing:.8px;text-transform:uppercase;">Field</th>
      <th style="text-align:left;padding:4px 0;color:#f5c842;font-size:10px;letter-spacing:.8px;text-transform:uppercase;">Accepted Column Names</th>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#f85454;font-weight:700;white-space:nowrap;">Ticker ★</td>
      <td style="color:#7a7fa8;">Scrip Name, Scrip, Symbol, Ticker, Stock, Company, Name, Stock Name, ISIN</td>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#f85454;font-weight:700;white-space:nowrap;">Quantity ★</td>
      <td style="color:#7a7fa8;">Quantity, Qty, Shares, Lots, Units, No of Shares, Volume, Holding, Net Qty</td>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#f85454;font-weight:700;white-space:nowrap;">Buy Price ★</td>
      <td style="color:#7a7fa8;">Buy Price, Avg Buy, Avg Cost, Purchase Price, Cost Price, Avg Price, Rate, Average Price</td>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#22d67b;font-weight:700;white-space:nowrap;">Asset Class</td>
      <td style="color:#7a7fa8;">Asset Class, Asset Type, Type, Category, Instrument, Segment — values: Stock, ETF, F&amp;O, REIT/InvIT, SGB, Bond/NCD, Mutual Fund…</td>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#4f7ef8;font-weight:700;white-space:nowrap;">Sell Price</td>
      <td style="color:#7a7fa8;">Sell Price, Selling Price, Exit Price, Sale Price — fill only for sold rows</td>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#4f7ef8;font-weight:700;white-space:nowrap;">Buy Date</td>
      <td style="color:#7a7fa8;">Buy Date, Purchase Date, Date of Purchase, Date Bought, Date</td>
    </tr>
    <tr>
      <td style="padding:3px 10px 3px 0;color:#4f7ef8;font-weight:700;white-space:nowrap;">Sell Date</td>
      <td style="color:#7a7fa8;">Sell Date, Date of Sale, Exit Date, Date Sold</td>
    </tr>
  </table>
  <div style="margin-top:8px;color:#454870;font-size:10px;">
    ★ Required &nbsp;·&nbsp;
    <b style="color:#22d67b;">Asset Class strongly recommended</b> — without it the app auto-detects from ticker name.
    Multi-sheet Excel files are supported; you can pick which sheet to import.
  </div>
</div>
""", unsafe_allow_html=True)

            # ── Upload widget ─────────────────────────────────────────
            _xl_file = st.file_uploader(
                "📎 Drop your Excel or CSV file here",
                type=["xlsx", "xls", "csv"],
                key=f"xl_import_{_managing_name}",
                help="Supports .xlsx, .xls, and .csv files. Any column order is fine."
            )

            # ── Once a file is uploaded, show preview + options ───────
            if _xl_file:
                try:
                    import io as _io_mod

                    # ── 1. Read raw file ─────────────────────────────
                    _fname = _xl_file.name.lower()
                    _xl_file.seek(0)
                    _raw_bytes = _xl_file.read()

                    if _fname.endswith(".csv"):
                        # Try multiple encodings for CSV
                        _xl_raw = None
                        for _enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                            try:
                                _xl_raw = pd.read_csv(_io_mod.BytesIO(_raw_bytes), encoding=_enc,
                                                      dtype=str, keep_default_na=False)
                                break
                            except Exception:
                                continue
                        if _xl_raw is None:
                            st.error("❌ Could not read CSV. Try saving as Excel (.xlsx) instead.")
                            st.stop()
                        _sheet_names = ["Sheet1"]
                        _selected_sheet = "Sheet1"
                        _sheet_dfs = {"Sheet1": _xl_raw}

                    else:
                        # Excel — enumerate sheets
                        _xf = pd.ExcelFile(_io_mod.BytesIO(_raw_bytes))
                        _sheet_names = _xf.sheet_names
                        _sheet_dfs = {}
                        for _sn in _sheet_names:
                            try:
                                _sheet_dfs[_sn] = pd.read_excel(_io_mod.BytesIO(_raw_bytes),
                                                                  sheet_name=_sn, dtype=str,
                                                                  keep_default_na=False)
                            except Exception:
                                pass

                        # Sheet selector (only shown if multiple sheets)
                        if len(_sheet_names) > 1:
                            st.markdown(
                                f"<div style='font-size:12px;color:#7a7fa8;margin-bottom:4px;'>"
                                f"📑 <b>{len(_sheet_names)} sheets found</b> — pick which one to import:</div>",
                                unsafe_allow_html=True
                            )
                            _selected_sheet = st.selectbox(
                                "Sheet", _sheet_names, key="xl_sheet_sel",
                                label_visibility="collapsed"
                            )
                        else:
                            _selected_sheet = _sheet_names[0]

                    _xl_raw = _sheet_dfs.get(_selected_sheet, pd.DataFrame())

                    # ── 2. Show raw preview ──────────────────────────
                    st.markdown(
                        f"<div style='font-size:12px;color:#22d67b;font-weight:700;margin:10px 0 4px;'>"
                        f"✅ File loaded — <b>{len(_xl_raw)}</b> rows · <b>{len(_xl_raw.columns)}</b> columns"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                    with st.expander("👁 Raw file preview (first 8 rows)", expanded=True):
                        st.dataframe(_xl_raw.head(8), use_container_width=True, hide_index=True)

                    # ── 3. Normalise column names ────────────────────
                    _xl_df = _xl_raw.copy()
                    _xl_df.columns = [
                        str(c).strip().lower()
                        .replace(" ", "_").replace("-", "_").replace(".", "_").replace("/", "_")
                        for c in _xl_df.columns
                    ]

                    # ── 4. Comprehensive column alias map ────────────
                    _ASSET_ALIAS_XL = {
                        "equity": "Stock", "shares": "Stock", "cash": "Stock",
                        "fno": "F&O", "f_and_o": "F&O", "f&o": "F&O",
                        "futures": "F&O", "options": "F&O", "derivatives": "F&O",
                        "reit": "REIT/InvIT", "invit": "REIT/InvIT",
                        "bond": "Bond/NCD", "ncd": "Bond/NCD", "gsec": "Bond/NCD",
                        "mf": "Mutual Fund", "mutual": "Mutual Fund",
                        "liquid": "Liquid ETF",
                        "commodity": "Commodity ETF",
                        "international": "International ETF", "global": "International ETF",
                        "index_f&o": "Index F&O", "index_fno": "Index F&O",
                        "currency_f&o": "Currency F&O", "currency": "Currency F&O",
                        "pref": "Preference Share", "preference": "Preference Share",
                        "unlisted": "Unlisted Share", "unlisted share": "Unlisted Share",
                        "pre-ipo": "Unlisted Share",
                        "sgb": "SGB", "sovereign": "SGB",
                    }

                    _TICKER_COLS   = {"scrip_name","scrip","symbol","ticker","stock","company",
                                      "name","stock_name","scripname","script_name","isin",
                                      "instrument","security","security_name","trading_symbol",
                                      "trading_sym","scrip_code","nse_symbol","bse_symbol",
                                      "share_name","stock_symbol","sym"}
                    _QTY_COLS      = {"quantity","qty","shares","lots","units","no_of_shares",
                                      "no__of_shares","volume","holding","net_qty","balance_qty",
                                      "available_qty","net_holding","holdings","lot","no_of_units",
                                      "number_of_shares","netqty"}
                    _BUYPRICE_COLS = {"buy_price","avg_buy","avg_cost","purchase_price","cost_price",
                                      "avg_price","buyprice","average_price","average_buy_price",
                                      "rate","ltp","price","weighted_avg_price","average_cost",
                                      "avg_buy_price","investment_price","cost_per_unit","avg_rate"}
                    _SELLPRICE_COLS= {"sell_price","selling_price","exit_price","sale_price",
                                      "sellprice","sold_price","sp","realized_price","close_price"}
                    _BUYDATE_COLS  = {"buy_date","purchase_date","date_of_purchase","buydate",
                                      "date_bought","date","trade_date","order_date","entry_date",
                                      "transaction_date","bought_date","allotment_date"}
                    _SELLDATE_COLS = {"sell_date","date_of_sale","exit_date","selldate",
                                      "date_sold","closing_date","sold_date"}
                    _ASSET_COLS    = {"asset_class","asset_type","type","category","instrument_type",
                                      "class","asset","segment","asset_category","product_type",
                                      "security_type","instrument_category","series"}

                    _auto_col_map = {}
                    for _c in _xl_df.columns:
                        if _c in _TICKER_COLS    and "Ticker"         not in _auto_col_map.values(): _auto_col_map[_c] = "Ticker"
                        elif _c in _QTY_COLS     and "Shares"         not in _auto_col_map.values(): _auto_col_map[_c] = "Shares"
                        elif _c in _BUYPRICE_COLS and "Buy_Price"      not in _auto_col_map.values(): _auto_col_map[_c] = "Buy_Price"
                        elif _c in _SELLPRICE_COLS and "Sell_Price"    not in _auto_col_map.values(): _auto_col_map[_c] = "Sell_Price"
                        elif _c in _BUYDATE_COLS  and "Buy_Date"       not in _auto_col_map.values(): _auto_col_map[_c] = "Buy_Date"
                        elif _c in _SELLDATE_COLS and "Sell_Date_xl"   not in _auto_col_map.values(): _auto_col_map[_c] = "Sell_Date_xl"
                        elif _c in _ASSET_COLS    and "Asset_Class_xl" not in _auto_col_map.values(): _auto_col_map[_c] = "Asset_Class_xl"

                    # ── 5. Manual column mapping UI ──────────────────
                    st.markdown(
                        "<div style='font-size:12px;font-weight:700;color:#f0f2ff;"
                        "margin:14px 0 6px;'>🗺️ Column Mapping</div>",
                        unsafe_allow_html=True
                    )
                    st.caption("Auto-detected below. Change any mapping if incorrect.")

                    _raw_cols_opts = ["— skip —"] + list(_xl_raw.columns)
                    _FIELD_LABELS = [
                        ("Ticker ★",    "Ticker",         True),
                        ("Quantity ★",  "Shares",         True),
                        ("Buy Price ★", "Buy_Price",      True),
                        ("Asset Class", "Asset_Class_xl", False),
                        ("Sell Price",  "Sell_Price",      False),
                        ("Buy Date",    "Buy_Date",        False),
                        ("Sell Date",   "Sell_Date_xl",   False),
                    ]

                    # Build reverse map: standard → raw col
                    _rev = {v: k for k, v in _auto_col_map.items()}

                    _user_col_map = {}
                    _map_cols = st.columns(2)
                    for _fi, (_flabel, _fkey, _freq) in enumerate(_FIELD_LABELS):
                        _default_raw = _rev.get(_fkey, "— skip —")
                        _default_idx = _raw_cols_opts.index(_default_raw) if _default_raw in _raw_cols_opts else 0
                        _chosen = _map_cols[_fi % 2].selectbox(
                            _flabel,
                            _raw_cols_opts,
                            index=_default_idx,
                            key=f"xl_map_{_fkey}"
                        )
                        if _chosen != "— skip —":
                            _user_col_map[_chosen] = _fkey

                    # ── 6. Validate required fields mapped ───────────
                    _mapped_targets = set(_user_col_map.values())
                    _required = {"Ticker", "Shares", "Buy_Price"}
                    _not_mapped = _required - _mapped_targets

                    if _not_mapped:
                        st.warning(
                            f"⚠️ Map these required fields first: **{', '.join(_not_mapped)}**"
                        )
                    else:
                        # ── 7. Defaults & options ────────────────────
                        _dc1, _dc2, _dc3 = st.columns(3)
                        _xl_date      = _dc1.date_input("Default Buy Date",  value=datetime.now().date(), key="xl_buy_date")
                        _xl_sell_date = _dc2.date_input("Default Sell Date", value=datetime.now().date(), key="xl_sell_date")
                        _xl_mode      = _dc3.radio(
                            "Import Mode",
                            ["🔄 Replace (clear existing)", "➕ Append to existing"],
                            key="xl_import_mode",
                            help="Replace wipes the client's current portfolio. Append adds new rows."
                        )

                        # ── 8. Build mapped dataframe ────────────────
                        # Select only the mapped raw columns and rename
                        _mapped_raw_cols = list(_user_col_map.keys())
                        _xl_mapped = _xl_raw[[c for c in _mapped_raw_cols if c in _xl_raw.columns]].copy()
                        _xl_mapped = _xl_mapped.rename(columns=_user_col_map)

                        # Numeric cleaning
                        for _nc in ("Shares", "Buy_Price", "Sell_Price"):
                            if _nc in _xl_mapped.columns:
                                _xl_mapped[_nc] = (
                                    _xl_mapped[_nc].astype(str)
                                    .str.replace(",", "").str.replace("₹", "").str.strip()
                                )
                                _xl_mapped[_nc] = pd.to_numeric(_xl_mapped[_nc], errors="coerce")

                        # Ticker cleaning
                        _xl_mapped["Ticker"] = (
                            _xl_mapped["Ticker"].astype(str)
                            .str.strip().str.upper()
                            .str.replace(r"\s+", "", regex=True)
                        )

                        # Sell price column
                        if "Sell_Price" not in _xl_mapped.columns:
                            _xl_mapped["Sell_Price"] = float("nan")

                        # Drop completely empty / invalid rows
                        _xl_mapped = _xl_mapped.dropna(subset=["Ticker", "Buy_Price", "Shares"])
                        _xl_mapped = _xl_mapped[_xl_mapped["Ticker"].str.strip().str.upper().isin(["", "NAN", "NONE", "NA", "NULL"]) == False]
                        _xl_mapped = _xl_mapped[_xl_mapped["Shares"] > 0]
                        _xl_mapped = _xl_mapped[_xl_mapped["Buy_Price"] > 0]

                        # Buy date
                        if "Buy_Date" in _xl_mapped.columns:
                            _xl_mapped["Buy_Date"] = pd.to_datetime(_xl_mapped["Buy_Date"], errors="coerce", dayfirst=True).dt.strftime("%Y-%m-%d")
                            _xl_mapped["Buy_Date"] = _xl_mapped["Buy_Date"].fillna(str(_xl_date))
                        else:
                            _xl_mapped["Buy_Date"] = str(_xl_date)

                        # Sell date
                        if "Sell_Date_xl" in _xl_mapped.columns:
                            _xl_mapped["Sell_Date_xl"] = pd.to_datetime(_xl_mapped["Sell_Date_xl"], errors="coerce", dayfirst=True).dt.strftime("%Y-%m-%d")
                            _xl_mapped["Sell_Date_xl"] = _xl_mapped["Sell_Date_xl"].fillna(str(_xl_sell_date))
                        else:
                            _xl_mapped["Sell_Date_xl"] = str(_xl_sell_date)

                        # Asset type resolution
                        def _resolve_asset_type_v2(row):
                            if "Asset_Class_xl" in row.index:
                                raw = str(row["Asset_Class_xl"]).strip().lower()
                                if raw and raw not in ("nan", "", "none", "-", "null"):
                                    mapped = _ASSET_ALIAS_XL.get(raw)
                                    if mapped:
                                        return mapped
                                    for vt in ["Stock","ETF","Liquid ETF","Commodity ETF",
                                               "International ETF","F&O","Index F&O",
                                               "Currency F&O","REIT/InvIT","SGB",
                                               "Bond/NCD","Mutual Fund","Preference Share",
                                               "Unlisted Share"]:
                                        if raw == vt.lower():
                                            return vt
                                    if "unlisted" in raw or "pre-ipo" in raw or "pre ipo" in raw: return "Unlisted Share"
                                    if "stock" in raw or "equity" in raw:     return "Stock"
                                    if "liquid" in raw:                        return "Liquid ETF"
                                    if "commodity" in raw:                     return "Commodity ETF"
                                    if "international" in raw or "global" in raw: return "International ETF"
                                    if "etf" in raw:                           return "ETF"
                                    if "index" in raw:                         return "Index F&O"
                                    if "currency" in raw:                      return "Currency F&O"
                                    if "f&o" in raw or "fno" in raw or "future" in raw or "option" in raw:
                                        return "F&O"
                                    if "reit" in raw or "invit" in raw:        return "REIT/InvIT"
                                    if "sgb" in raw or "sovereign" in raw:     return "SGB"
                                    if "bond" in raw or "ncd" in raw:          return "Bond/NCD"
                                    if "mutual" in raw or raw == "mf":         return "Mutual Fund"
                                    if "pref" in raw:                          return "Preference Share"
                                    return row["Asset_Class_xl"].strip()
                            return classify_asset(row["Ticker"])

                        _xl_mapped["Asset_Type"] = _xl_mapped.apply(_resolve_asset_type_v2, axis=1)

                        # ── 9. Split active vs sold ───────────────────
                        _sold_mask  = _xl_mapped["Sell_Price"].notna() & (_xl_mapped["Sell_Price"] > 0)
                        _xl_sold    = _xl_mapped[_sold_mask].copy()
                        _xl_active  = _xl_mapped[~_sold_mask].copy()

                        # ── 10. Show mapped preview ───────────────────
                        st.markdown(
                            f"<div style='font-size:12px;font-weight:700;color:#f0f2ff;"
                            f"margin:14px 0 6px;'>📋 Mapped Preview — "
                            f"<span style='color:#22d67b;'>{len(_xl_active)} holdings</span>"
                            f" &nbsp;+&nbsp; "
                            f"<span style='color:#f5c842;'>{len(_xl_sold)} sell trades</span>"
                            f"</div>",
                            unsafe_allow_html=True
                        )

                        _prev_cols = ["Ticker","Shares","Buy_Price","Buy_Date","Asset_Type"]
                        if not _xl_active.empty:
                            _preview_df = _xl_active[[c for c in _prev_cols if c in _xl_active.columns]].copy()
                            _preview_df.insert(0, "Status", "🟢 Active")
                            if not _xl_sold.empty:
                                _sold_prev = _xl_sold[[c for c in _prev_cols if c in _xl_sold.columns]].copy()
                                _sold_prev.insert(0, "Status", "🔴 Sold")
                                _combined_prev = pd.concat([_preview_df, _sold_prev], ignore_index=True)
                            else:
                                _combined_prev = _preview_df
                            st.dataframe(_combined_prev.head(20), use_container_width=True, hide_index=True)
                        elif not _xl_sold.empty:
                            _sold_prev = _xl_sold[[c for c in _prev_cols if c in _xl_sold.columns]].copy()
                            _sold_prev.insert(0, "Status", "🔴 Sold")
                            st.dataframe(_sold_prev.head(20), use_container_width=True, hide_index=True)

                        # Asset type breakdown
                        _type_summary = _xl_mapped.groupby("Asset_Type")["Ticker"].apply(list).to_dict()
                        if _type_summary:
                            _has_xl_class = "Asset_Class_xl" in _xl_mapped.columns
                            _src_note = "from your column" if _has_xl_class else "**auto-detected**"
                            _chips_html = ""
                            for _atype, _tickers in _type_summary.items():
                                _clr = {
                                    "Stock":"#4f7ef8","ETF":"#a78bfa","F&O":"#f5c842",
                                    "Liquid ETF":"#38bdf8","Commodity ETF":"#fbbf24",
                                    "REIT/InvIT":"#22d67b","SGB":"#f5c842",
                                    "Bond/NCD":"#9ca3af","Mutual Fund":"#34d399",
                                    "Index F&O":"#fb923c","Currency F&O":"#c084fc",
                                    "Unlisted Share":"#a78bfa","Preference Share":"#60a5fa",
                                }.get(_atype, "#7a7fa8")
                                _chips_html += (
                                    f'<span style="background:{_clr}22;color:{_clr};border:1px solid {_clr}44;'
                                    f'border-radius:20px;padding:2px 10px;font-size:11px;font-weight:700;'
                                    f'margin:2px 4px 2px 0;display:inline-block;">'
                                    f'{_atype} ({len(_tickers)})</span>'
                                )
                            st.markdown(
                                f"<div style='margin:8px 0 2px;'>Asset types {_src_note}: {_chips_html}</div>",
                                unsafe_allow_html=True
                            )

                        # ── 11. Final import button ───────────────────
                        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                        _imp_btn = st.button(
                            f"🚀 Confirm & Import {len(_xl_active)} Holdings"
                            + (f" + {len(_xl_sold)} Sell Trades" if not _xl_sold.empty else ""),
                            key="xl_import_btn",
                            type="primary",
                            use_container_width=True
                        )

                        if _imp_btn:
                            # ── Build new portfolio rows (active holdings) ──────────────
                            # Take EXACTLY what the Excel says — no price recalculation,
                            # no weighted-average merging for buys, no FIFO, nothing.
                            _new_portfolio = _xl_active[["Ticker","Shares","Buy_Price","Buy_Date","Asset_Type"]].copy()

                            if "➕ Append" in _xl_mode:
                                import os as _os_mod
                                if _os_mod.path.exists(PORTFOLIO_FILE):
                                    _existing_pf = pd.read_csv(PORTFOLIO_FILE)
                                    # Only skip rows that are EXACT duplicates on all key fields
                                    _merged_pf = pd.concat([_existing_pf, _new_portfolio], ignore_index=True)
                                    _merged_pf = _merged_pf.drop_duplicates(subset=["Ticker","Buy_Price","Shares"], keep="last")
                                    _merged_pf.to_csv(PORTFOLIO_FILE, index=False)
                                    _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                                    _saved_count = len(_new_portfolio)
                                else:
                                    _new_portfolio.to_csv(PORTFOLIO_FILE, index=False)
                                    _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                                    _saved_count = len(_new_portfolio)
                            else:
                                # Replace mode — save exactly what Excel has
                                _new_portfolio.to_csv(PORTFOLIO_FILE, index=False)
                                _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                                _saved_count = len(_new_portfolio)

                            # ── Build trades rows (sold entries) ────────────────────────
                            # Rule: merge rows that share the SAME ticker into one trade row
                            # using weighted average sell price and summed qty/P&L.
                            # Buy_Price_At_Sell comes directly from Excel's Buy_Price column —
                            # never recalculated, never overwritten.
                            if not _xl_sold.empty:
                                _new_trades_rows = []
                                for _sticker, _sg in _xl_sold.groupby("Ticker", sort=False):
                                    _total_sq = float(_sg["Sell_Qty"].sum()) if "Sell_Qty" in _sg.columns else float(_sg["Shares"].sum())
                                    # Weighted average sell price across all rows for this ticker
                                    _sq_col = "Sell_Qty" if "Sell_Qty" in _sg.columns else "Shares"
                                    _wtd_sp = round(
                                        (_sg[_sq_col].astype(float) * _sg["Sell_Price"].astype(float)).sum()
                                        / _sg[_sq_col].astype(float).sum(), 4
                                    ) if _sg[_sq_col].astype(float).sum() > 0 else float(_sg["Sell_Price"].iloc[0])
                                    # Weighted average buy price from Excel (user-provided — never recalculate)
                                    _wtd_bp = round(
                                        (_sg[_sq_col].astype(float) * _sg["Buy_Price"].astype(float)).sum()
                                        / _sg[_sq_col].astype(float).sum(), 4
                                    ) if _sg[_sq_col].astype(float).sum() > 0 else float(_sg["Buy_Price"].iloc[0])
                                    _booked = round((_wtd_sp - _wtd_bp) * _total_sq, 2)
                                    _tr_atype = str(_sg["Asset_Type"].iloc[0]) if "Asset_Type" in _sg.columns else ""
                                    if not _tr_atype or _tr_atype in ("nan", "None"):
                                        _tr_atype = _classify_ticker_to_asset_type(str(_sticker))
                                    # Sell date: use most recent sell date among merged rows
                                    _sd_col = "Sell_Date_xl"
                                    _sell_date_val = str(_sg[_sd_col].iloc[-1]) if _sd_col in _sg.columns else str(_xl_sell_date)
                                    _new_trades_rows.append({
                                        "Ticker":            str(_sticker),
                                        "Sell_Qty":          _total_sq,
                                        "Sell_Price":        _wtd_sp,
                                        "Sell_Date":         _sell_date_val,
                                        "Buy_Price_At_Sell": _wtd_bp,
                                        "Booked_PnL":        _booked,
                                        "Asset_Type":        _tr_atype,
                                    })
                                _new_trades_df = pd.DataFrame(_new_trades_rows)
                                if "➕ Append" in _xl_mode and _os_mod.path.exists(TRADES_FILE):
                                    _existing_tr = pd.read_csv(TRADES_FILE)
                                    _merged_tr = pd.concat([_existing_tr, _new_trades_df], ignore_index=True)
                                    _merged_tr.to_csv(TRADES_FILE, index=False)
                                    _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')
                                else:
                                    _new_trades_df.to_csv(TRADES_FILE, index=False)
                                    _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')
                            elif "🔄 Replace" in _xl_mode:
                                pd.DataFrame(
                                    columns=["Ticker","Sell_Qty","Sell_Price","Sell_Date",
                                             "Buy_Price_At_Sell","Booked_PnL","Asset_Type"]
                                ).to_csv(TRADES_FILE, index=False)
                                _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')

                            st.success(
                                f"✅ Import complete! "
                                f"**{_saved_count}** holdings saved · "
                                f"**{len(_xl_sold)}** sell trades recorded."
                            )
                            st.rerun()

                except Exception as _xl_err:
                    st.error(f"❌ Import failed: {_xl_err}")
                    st.caption(f"Detail: `{type(_xl_err).__name__}: {_xl_err}`")
                    st.caption("Tip: Make sure your file has at least 3 columns — Ticker/Scrip Name, Quantity, and Buy Price.")

        st.divider()

        with st.form("add_form"):
            ticker = st.text_input(
                "Ticker Symbol",
                placeholder="RELIANCE.NS  or  532540.BO"
            ).upper()

            asset_type_input = st.selectbox(
                "Asset Type",
                ["Stock", "ETF", "Liquid ETF", "Commodity ETF", "International ETF",
                 "F&O", "Index F&O", "Currency F&O",
                 "REIT/InvIT", "SGB", "Bond/NCD", "Mutual Fund", "Preference Share", "Unlisted Share"],
                index=0,
                help="Auto-detected from ticker. Override here if needed."
            )

            c1, c2 = st.columns(2)
            shares = c1.number_input("Qty / Lots", min_value=0.01, step=1.0)
            buy_price = c2.number_input("Buy Price ₹", min_value=0.01, step=0.01)
            buy_date = st.date_input("Buy Date", datetime.now().date())

            if st.form_submit_button("➕ Add Holding"):
                if not ticker:
                    st.error("Please enter a ticker symbol.")
                else:
                    if "Asset_Type" not in df.columns:
                        df["Asset_Type"] = df["Ticker"].apply(classify_asset)

                    existing = df[df["Ticker"] == ticker]

                    if not existing.empty:
                        idx = existing.index[0]
                        old_qty   = float(df.at[idx, "Shares"])
                        old_price = float(df.at[idx, "Buy_Price"])
                        new_qty   = old_qty + shares
                        avg_price = round((old_qty * old_price + shares * buy_price) / new_qty, 4)

                        df.at[idx, "Shares"]    = new_qty
                        df.at[idx, "Buy_Price"] = avg_price
                        df.at[idx, "Buy_Date"]  = str(buy_date)
                        df.to_csv(PORTFOLIO_FILE, index=False)
                        _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                        st.success(
                            f"✅ {ticker} updated — "
                            f"New Qty: {new_qty:,.2f} | "
                            f"Avg Price: ₹{avg_price:,.2f}"
                        )
                    else:
                        new_row = pd.DataFrame(
                            [[ticker, shares, buy_price, buy_date, asset_type_input]],
                            columns=["Ticker", "Shares", "Buy_Price", "Buy_Date", "Asset_Type"]
                        )
                        df = pd.concat([df, new_row], ignore_index=True)
                        df.to_csv(PORTFOLIO_FILE, index=False)
                        _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                        st.success(f"✅ {ticker} added at ₹{buy_price:,.2f}")
                    st.rerun()

        # Add Asset_Type column if missing
        if "Asset_Type" not in df.columns and not df.empty:
            df["Asset_Type"] = df["Ticker"].apply(classify_asset)
            df.to_csv(PORTFOLIO_FILE, index=False)
            _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')

        # ── UPDATE UNLISTED SHARE CMP ──────────────────────────────
        _unlisted_holdings = df[df.get("Asset_Type", pd.Series()) == "Unlisted Share"] if "Asset_Type" in df.columns else pd.DataFrame()
        if not _unlisted_holdings.empty:
            # Initialise session cache if not present
            if "_ul_cmp_cache" not in st.session_state:
                st.session_state["_ul_cmp_cache"] = {}
            st.markdown("""
<div style="margin:14px 0 6px 0;">
  <span style="font-size:13px;font-weight:700;color:#a78bfa;">🔒 Manual CMP Override — Unlisted Shares</span>
  <span style="font-size:10px;color:#6b7299;margin-left:6px;">Buy Price is kept unchanged · CMP stored separately</span>
</div>""", unsafe_allow_html=True)
            with st.form("unlisted_cmp_form"):
                _ul_ticker = st.selectbox(
                    "Select Unlisted Holding",
                    _unlisted_holdings["Ticker"].tolist(),
                    key="ul_ticker_sel"
                )
                _ul_cur = st.session_state["_ul_cmp_cache"].get(_ul_ticker, 0.0)
                _ul_cmp = st.number_input(
                    "Current Market Price ₹",
                    min_value=0.0, step=0.01, value=float(_ul_cur),
                    help="Fetched automatically from unlistedshares.com · override here if needed · Buy Price is never changed"
                )
                if st.form_submit_button("💾 Save Manual CMP"):
                    st.session_state["_ul_cmp_cache"][_ul_ticker] = round(float(_ul_cmp), 4)
                    st.success(f"✅ Manual CMP for {_ul_ticker} set to ₹{_ul_cmp:,.2f}  (Buy Price unchanged)")
                    st.rerun()
            st.divider()

        # DELETE HOLDING
        if not df.empty:
            st.header("🗑 Delete Holding")
            idx = st.selectbox(
                "Select Holding",
                df.index,
                format_func=lambda x: f"{df.loc[x,'Ticker']} ({df.loc[x,'Shares']} shares)"
            )
            if st.button("Delete Selected"):
                df = df.drop(idx).reset_index(drop=True)
                df.to_csv(PORTFOLIO_FILE, index=False)
                _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                st.success("Holding Deleted")
                st.rerun()

        # ── RECORD A SELL ─────────────────────────────────────────
        st.header("💰 Record a Sell")

        # Selectbox OUTSIDE form — only show tickers with available (unsold) qty > 0
        sell_ticker = None
        if not df.empty:
            _net_tmp_all, _ = compute_net_holdings(df, trades_df)
            _ticker_opts = [
                t for t in df["Ticker"].unique().tolist()
                if float(_net_tmp_all[_net_tmp_all["Ticker"] == t]["Shares"].sum()) > 0
            ]
            if not _ticker_opts:
                st.info("No holdings with available quantity to sell.")
            else:
                sell_ticker = st.selectbox(
                    "Select Holding to Sell",
                    _ticker_opts,
                    key="sell_ticker_sel"
                )

        # Compute available qty (re-evaluates on every ticker change)
        if not df.empty and sell_ticker:
            net_tmp, _ = compute_net_holdings(df, trades_df)
            avail = float(net_tmp[net_tmp["Ticker"] == sell_ticker]["Shares"].sum())
        else:
            avail = 0.0

        if sell_ticker:
            st.caption(f"Available Qty: **{avail:,.2f}**")

        if sell_ticker:
            # Show default buy rate (weighted avg) as hint, but user can override
            _hint_rows = df[df["Ticker"] == sell_ticker]
            _hint_buy = round(
                (_hint_rows["Shares"] * _hint_rows["Buy_Price"]).sum()
                / _hint_rows["Shares"].sum(), 2
            ) if _hint_rows["Shares"].sum() > 0 else 0.0

            with st.form("sell_form"):
                sc1, sc2 = st.columns(2)
                sell_qty        = sc1.number_input("Sell Qty",      min_value=0.01, step=1.0,  max_value=float(max(avail, 0.01)))
                sell_price      = sc2.number_input("Sell Price ₹",  min_value=0.01, step=0.01)
                sc3, sc4        = st.columns(2)
                buy_rate_input  = sc3.number_input(
                    "Your Buy Rate ₹",
                    min_value=0.00, step=0.01, value=float(_hint_buy),
                    help="Enter the buy rate for this sell. Defaults to your weighted avg buy price."
                )
                sell_date       = sc4.date_input("Sell Date", datetime.now().date(), key="sell_date")

                if st.form_submit_button("✅ Record Sell"):
                    if sell_qty > avail:
                        st.error(f"❌ Sell qty ({sell_qty}) exceeds available ({avail:.2f})")
                    else:
                        # Each sell is always recorded as a NEW separate row — no merging
                        booked = round((sell_price - buy_rate_input) * sell_qty, 2)
                        _sell_asset_type = _pf_asset_map.get(sell_ticker) or _classify_ticker_to_asset_type(sell_ticker)
                        new_trade = pd.DataFrame([[
                            sell_ticker, sell_qty, sell_price,
                            str(sell_date), round(buy_rate_input, 2), booked, _sell_asset_type
                        ]], columns=["Ticker","Sell_Qty","Sell_Price","Sell_Date",
                                     "Buy_Price_At_Sell","Booked_PnL","Asset_Type"])
                        trades_df = pd.concat([trades_df, new_trade], ignore_index=True)
                        trades_df.to_csv(TRADES_FILE, index=False)
                        _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')
                        st.success(
                            f"✅ Sold {sell_qty:,.0f} × {sell_ticker} @ ₹{sell_price:,.2f} | "
                            f"Buy Rate: ₹{buy_rate_input:,.2f} | "
                            f"Booked P&L: ₹{booked:,.2f}"
                        )
                        st.rerun()

        # ── DELETE A SELL TRADE (sidebar) ─────────────────────────
        if not trades_df.empty:
            st.header("🗑 Delete a Sell Trade")
            with st.expander("Undo / Remove a sell record", expanded=False):
                _del_idx = st.selectbox(
                    "Select trade to delete",
                    trades_df.index,
                    format_func=lambda x: (
                        f"{trades_df.loc[x,'Ticker']} | "
                        f"Qty {trades_df.loc[x,'Sell_Qty']} | "
                        f"@ \u20b9{trades_df.loc[x,'Sell_Price']} | "
                        f"{trades_df.loc[x,'Sell_Date']}"
                    ),
                    key="del_trade_sidebar"
                )
                if st.button("🗑️ Delete Trade", key="del_trade_sidebar_btn"):
                    trades_df = trades_df.drop(_del_idx).reset_index(drop=True)
                    trades_df.to_csv(TRADES_FILE, index=False)
                    _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')
                    st.success("✅ Trade deleted.")
                    st.rerun()

    else:
        # ── Client: show change password only ─────────────────────
        with st.expander("\U0001f511 Change My Password"):
            _cp_old  = st.text_input("Current Password", type="password", key="cp_old")
            _cp_new1 = st.text_input("New Password",     type="password", key="cp_new1")
            _cp_new2 = st.text_input("Confirm Password", type="password", key="cp_new2")
            if st.button("Update Password", key="cp_btn"):
                _stored = clients_dict.get(_auth_code, {}).get("password_hash", "")
                if _stored and hash_password(_cp_old) != _stored:
                    st.error("\u274c Current password incorrect.")
                elif not _cp_new1:
                    st.error("\u274c New password cannot be blank.")
                elif _cp_new1 != _cp_new2:
                    st.error("\u274c Passwords do not match.")
                else:
                    clients_dict[_auth_code]["password_hash"] = hash_password(_cp_new1)
                    save_clients(clients_dict)
                    st.success("\u2705 Password updated!")

# =========================================================

# =========================================================
# MASTER IMPORT — Asset Classes, P&L Logic, Template
# =========================================================

_MI_ASSET_CLASSES = [
    "Equity", "F&O - Futures", "F&O - Options",
    "Mutual Fund", "ETF", "SGBs (Gold Bonds)", "Currency", "Commodity",
]
_MI_SHORT_SUPPORTED = {"Equity", "F&O - Futures", "F&O - Options", "Currency", "Commodity"}
_MI_REQUIRED_COLS   = ["Stock Name", "Asset Class", "Buy Price", "Buy Qty", "Buy Date"]
_MI_OPTIONAL_COLS   = ["Sell Qty", "Sell Price", "Sell Date"]
_MI_ALL_COLS        = _MI_REQUIRED_COLS + _MI_OPTIONAL_COLS


def _mi_generate_template() -> bytes:
    """Generate a combined Excel template: Client info + portfolio trades in one sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Portfolio Import"

    # ── Header colours ──
    client_fill  = PatternFill("solid", start_color="1A3A5C", end_color="1A3A5C")   # dark blue  → client cols
    trade_fill   = PatternFill("solid", start_color="1A4A2E", end_color="1A4A2E")   # dark green → trade cols
    hfont        = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    short_fill   = PatternFill("solid", start_color="3D0000", end_color="3D0000")
    alt_fill     = PatternFill("solid", start_color="EEF2FF", end_color="EEF2FF")
    white_fill   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="2E4A6F"), right=Side(style="thin", color="2E4A6F"),
        top=Side(style="thin",  color="2E4A6F"), bottom=Side(style="thin", color="2E4A6F"),
    )

    # ── Combined columns: 3 client cols + 8 trade cols ──
    client_headers = ["Client ID", "Client Name", "Password"]
    trade_headers  = ["Stock Name", "Asset Class", "Buy Price", "Buy Qty", "Buy Date",
                      "Sell Qty", "Sell Price", "Sell Date"]
    all_headers = client_headers + trade_headers

    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = client_fill if col <= 3 else trade_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # ── Sample rows: 2 clients × multiple trades ──
    sample_rows = [
        # Client 1 — Suresh Kumar (3 trades)
        ["HO668", "Suresh Kumar",  "suresh@123",  "RELIANCE",     "Equity",            2450.50,  10, "01-01-2024", 5,   2600.00, "15-03-2024"],
        ["HO668", "Suresh Kumar",  "suresh@123",  "TCS",          "Equity",            3500.00,   5, "10-02-2024", "",  "",      ""],
        ["HO668", "Suresh Kumar",  "suresh@123",  "GOLDBEES",     "ETF",                 55.00, 100, "12-04-2024", "",  "",      ""],
        # Client 2 — Priya Sharma (3 trades)
        ["HO701", "Priya Sharma",  "priya@456",   "HDFCBANK",     "Equity",            1600.00,  20, "05-03-2024", "",  "",      ""],
        ["HO701", "Priya Sharma",  "priya@456",   "NIFTYFUT",     "F&O - Futures",    22000.00,  -1, "10-03-2024", -1, 21500.00, "25-03-2024"],
        ["HO701", "Priya Sharma",  "priya@456",   "SGB2028",      "SGBs (Gold Bonds)", 5800.00,   5, "01-05-2024", "",  "",      ""],
        # Client 3 — Ramesh Gupta (2 trades)
        ["HO802", "Ramesh Gupta",  "ramesh@789",  "INFY",         "Equity",            1500.00,  15, "15-01-2024", 15, 1650.00, "20-04-2024"],
        ["HO802", "Ramesh Gupta",  "ramesh@789",  "HDFCNIFTY50",  "Mutual Fund",        180.00,  50, "15-05-2024", "",  "",      ""],
    ]

    for r_idx, row in enumerate(sample_rows, 2):
        is_short = isinstance(row[6], (int, float)) and row[6] < 0
        fill = short_fill if is_short else (alt_fill if r_idx % 2 == 0 else white_fill)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10, color="FFB3B3" if is_short else "000000")
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # ── Column widths ──
    widths = [14, 22, 15,   20, 20, 12, 10, 14, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    # ── Instructions sheet ──
    notes = wb.create_sheet("Instructions")
    notes["A1"] = "COMBINED CLIENT + PORTFOLIO IMPORT — INSTRUCTIONS"
    notes["A1"].font = Font(bold=True, size=13, color="1E3A5F", name="Arial")
    notes["A3"] = "Column"; notes["B3"] = "Description"
    notes["A3"].font = Font(bold=True, name="Arial"); notes["B3"].font = Font(bold=True, name="Arial")
    instr = [
        ("Client ID",   "Unique client code. e.g. HO668. Must match for all rows of same client."),
        ("Client Name", "Full display name. e.g. Suresh Kumar. Same for all rows of same client."),
        ("Password",    "Login password. Same for all rows of same client."),
        ("Stock Name",  "NSE/BSE ticker or name. e.g. RELIANCE, GOLDBEES"),
        ("Asset Class", "One of: " + ", ".join(_MI_ASSET_CLASSES)),
        ("Buy Price",   "Entry price per unit. Always POSITIVE."),
        ("Buy Qty",     "POSITIVE = Long/Buy. NEGATIVE = Short Sell / Option Writing."),
        ("Buy Date",    "Entry date. Format: DD-MM-YYYY"),
        ("Sell Qty",    "For long: positive qty. For short: negative qty. Leave blank if still open."),
        ("Sell Price",  "Exit price. Leave blank if still open."),
        ("Sell Date",   "Exit date. Leave blank if still open."),
    ]
    notes["A5"] = "HOW IT WORKS"
    notes["A5"].font = Font(bold=True, name="Arial", color="1A3A5C")
    notes["B5"] = "Each row = one trade for one client. Repeat Client ID/Name/Password for every trade row. All trades with the same Client ID go into that client\'s portfolio."
    notes["B5"].font = Font(name="Arial", color="1A3A5C")
    for i, (c, d) in enumerate(instr, 7):
        notes.cell(row=i, column=1, value=c).font = Font(name="Arial", bold=True, color="1E3A5F")
        notes.cell(row=i, column=2, value=d).font  = Font(name="Arial", size=10)
    notes["A19"] = "⚠️ Notes"
    notes["A19"].font = Font(bold=True, name="Arial", color="CC0000")
    notes["B19"] = "Duplicate Client IDs will be skipped (existing clients not overwritten by default). Dev code cannot be used as Client ID."
    notes["B19"].font = Font(name="Arial", color="CC0000")
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 85

    # ── Asset Class Guide sheet ──
    ref = wb.create_sheet("Asset Class Guide")
    ref["A1"] = "Asset Class Reference"
    ref["A1"].font = Font(bold=True, size=13, color="1E3A5F", name="Arial")
    ref["A3"] = "Asset Class"; ref["B3"] = "Examples"; ref["C3"] = "Short Sell?"
    for cell in [ref["A3"], ref["B3"], ref["C3"]]:
        cell.font = Font(bold=True, name="Arial")
    guide = [
        ("Equity",            "RELIANCE, TCS, HDFCBANK",        "Yes"),
        ("F&O - Futures",     "NIFTYFUT, BANKNIFTYFUT",          "Yes"),
        ("F&O - Options",     "NIFTY25000CE, NIFTY24000PE",      "Yes (negative qty = option writing)"),
        ("Mutual Fund",       "HDFCNIFTY50, PARAG FLEXI",        "No"),
        ("ETF",               "GOLDBEES, NIFTYBEES",             "No"),
        ("SGBs (Gold Bonds)", "SGB2028, SGB2030",                "No"),
        ("Currency",          "USDINR, EURINR",                  "Yes"),
        ("Commodity",         "CRUDEOIL, GOLD, SILVER",          "Yes"),
    ]
    for i, (ac, ex, sh) in enumerate(guide, 4):
        ref.cell(row=i, column=1, value=ac).font  = Font(name="Arial", bold=True)
        ref.cell(row=i, column=2, value=ex).font  = Font(name="Arial")
        ref.cell(row=i, column=3, value=sh).font  = Font(name="Arial", color="006600" if "Yes" in sh else "CC0000")
    ref.column_dimensions["A"].width = 22
    ref.column_dimensions["B"].width = 35
    ref.column_dimensions["C"].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mi_generate_template_old_unused() -> bytes:
    """OLD single-client template — kept for reference only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Trades"
    header_fill  = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    short_fill   = PatternFill("solid", start_color="3D0000", end_color="3D0000")
    alt_fill     = PatternFill("solid", start_color="EEF2FF", end_color="EEF2FF")
    white_fill   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    border = Border(
        left=Side(style="thin",color="2E4A6F"), right=Side(style="thin",color="2E4A6F"),
        top=Side(style="thin",color="2E4A6F"),  bottom=Side(style="thin",color="2E4A6F"),
    )
    for col, h in enumerate(_MI_ALL_COLS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sample_rows = [
        ["RELIANCE",     "Equity",            2450.50,  10,  "01-01-2024", 5,   2600.00,  "15-03-2024"],
        ["TCS",          "Equity",            3500.00,  5,   "10-02-2024", "",  "",       ""],
        ["NIFTY25000CE", "F&O - Options",      120.00,  50,  "05-03-2024", 50,  180.00,   "20-03-2024"],
        ["NIFTY25000CE", "F&O - Options",      150.00, -75,  "06-03-2024", -75,  90.00,   "21-03-2024"],
        ["NIFTYFUT",     "F&O - Futures",    22000.00,  -1,  "10-03-2024", -1, 21500.00,  "25-03-2024"],
        ["GOLDBEES",     "ETF",                 55.00,  100, "12-04-2024", "",  "",       ""],
        ["SGB2028",      "SGBs (Gold Bonds)", 5800.00,  5,   "01-05-2024", "",  "",       ""],
        ["USDINR",       "Currency",             83.00, -10, "02-05-2024", -10,  82.50,   "10-05-2024"],
        ["CRUDEOIL",     "Commodity",          6500.00,  2,  "03-05-2024", "",  "",       ""],
        ["HDFCNIFTY50",  "Mutual Fund",          180.00, 50, "15-05-2024", "",  "",       ""],
    ]
    for r_idx, row in enumerate(sample_rows, 2):
        is_short = isinstance(row[3], (int, float)) and row[3] < 0
        fill = short_fill if is_short else (alt_fill if r_idx % 2 == 0 else white_fill)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10, color="FFB3B3" if is_short else "000000")
            cell.fill = fill; cell.border = border
            cell.alignment = Alignment(horizontal="center")
    widths = [20, 20, 12, 12, 14, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    notes = wb.create_sheet("Instructions")
    notes["A1"] = "MASTER IMPORT TEMPLATE — INSTRUCTIONS"
    notes["A1"].font = Font(bold=True, size=13, color="1E3A5F", name="Arial")
    notes["A3"] = "Column"; notes["B3"] = "Description"
    notes["A3"].font = Font(bold=True, name="Arial"); notes["B3"].font = Font(bold=True, name="Arial")
    notes["A5"] = "SHORT SELL RULE"
    notes["A5"].font = Font(bold=True, name="Arial", color="CC0000")
    notes["B5"] = "Use NEGATIVE Buy Qty for short positions. P&L = (Entry Price − Cover Price) × |Qty|"
    notes["B5"].font = Font(name="Arial", color="CC0000")
    notes["A6"] = "LONG RULE"
    notes["A6"].font = Font(bold=True, name="Arial", color="006600")
    notes["B6"] = "P&L = (Sell Price − Buy Price) × Qty. Profit when price rises."
    notes["B6"].font = Font(name="Arial", color="006600")
    instr = [
        ("Stock Name",  "NSE/BSE ticker or name. e.g. RELIANCE, NIFTY25000CE, GOLDBEES"),
        ("Asset Class", "One of: " + ", ".join(_MI_ASSET_CLASSES)),
        ("Buy Price",   "Entry price per unit. Always POSITIVE."),
        ("Buy Qty",     "POSITIVE = Long/Buy. NEGATIVE = Short Sell / Option Writing."),
        ("Buy Date",    "Entry date. Format: DD-MM-YYYY"),
        ("Sell Qty",    "For long: positive exit qty. For short: negative exit qty."),
        ("Sell Price",  "Exit/Cover price. Always POSITIVE."),
        ("Sell Date",   "Exit date. Format: DD-MM-YYYY"),
    ]
    for i, (c, d) in enumerate(instr, 8):
        notes.cell(row=i, column=1, value=c).font = Font(name="Arial", bold=True, color="1E3A5F")
        notes.cell(row=i, column=2, value=d).font  = Font(name="Arial", size=10)
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 80

    ref = wb.create_sheet("Asset Class Guide")
    ref["A1"] = "Asset Class Reference"
    ref["A1"].font = Font(bold=True, size=13, color="1E3A5F", name="Arial")
    ref["A3"] = "Asset Class"; ref["B3"] = "Examples"; ref["C3"] = "Short Sell?"
    for cell in [ref["A3"], ref["B3"], ref["C3"]]:
        cell.font = Font(bold=True, name="Arial")
    guide = [
        ("Equity",            "RELIANCE, TCS, HDFCBANK",        "Yes"),
        ("F&O - Futures",     "NIFTYFUT, BANKNIFTYFUT",          "Yes"),
        ("F&O - Options",     "NIFTY25000CE, NIFTY24000PE",      "Yes (negative qty = option writing)"),
        ("Mutual Fund",       "HDFCNIFTY50, PARAG FLEXI",        "No"),
        ("ETF",               "GOLDBEES, NIFTYBEES",             "No"),
        ("SGBs (Gold Bonds)", "SGB2028, SGB2030",                "No"),
        ("Currency",          "USDINR, EURINR",                  "Yes"),
        ("Commodity",         "CRUDEOIL, GOLD, SILVER",          "Yes"),
    ]
    for i, (ac, ex, sh) in enumerate(guide, 4):
        ref.cell(row=i, column=1, value=ac).font  = Font(name="Arial", bold=True)
        ref.cell(row=i, column=2, value=ex).font  = Font(name="Arial")
        ref.cell(row=i, column=3, value=sh).font  = Font(name="Arial", color="006600" if "Yes" in sh else "CC0000")
    ref.column_dimensions["A"].width = 22
    ref.column_dimensions["B"].width = 35
    ref.column_dimensions["C"].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mi_calc_pnl(row):
    bp, sp, bq, sq = row["Buy Price"], row["Sell Price"], row["Buy Qty"], row["Sell Qty"]
    if pd.isna(sp) or pd.isna(sq) or pd.isna(bp):
        return None, None
    qty_abs = abs(sq)
    pnl = (bp - sp) * qty_abs if bq < 0 else (sp - bp) * qty_abs
    cost = bp * qty_abs
    return round(pnl, 2), round((pnl / cost) * 100, 2) if cost else None


def _mi_validate(df: pd.DataFrame):
    errors = []
    missing = [c for c in _MI_REQUIRED_COLS if c not in df.columns]
    if missing:
        errors.append(f"❌ Missing required columns: {', '.join(missing)}")
        return df, errors
    for c in _MI_OPTIONAL_COLS:
        if c not in df.columns:
            df[c] = None
    df = df[_MI_ALL_COLS].copy()
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    for i, row in df.iterrows():
        rn   = i + 2
        name = str(row["Stock Name"]).strip() if pd.notna(row["Stock Name"]) else f"Row {rn}"
        if pd.isna(row["Stock Name"]) or str(row["Stock Name"]).strip() == "":
            errors.append(f"Row {rn}: Stock Name is empty.")
        ac = str(row["Asset Class"]).strip() if pd.notna(row["Asset Class"]) else ""
        if ac not in _MI_ASSET_CLASSES:
            errors.append(f"Row {rn} ({name}): Asset Class '{ac}' invalid.")
        try:
            bp = float(row["Buy Price"])
            if bp <= 0: errors.append(f"Row {rn} ({name}): Buy Price must be > 0.")
        except: errors.append(f"Row {rn} ({name}): Buy Price is not a valid number.")
        try:
            bq = int(float(row["Buy Qty"]))
            if bq == 0: errors.append(f"Row {rn} ({name}): Buy Qty cannot be 0.")
            if bq < 0 and ac not in _MI_SHORT_SUPPORTED:
                errors.append(f"Row {rn} ({name}): Short sell not supported for '{ac}'.")
        except: errors.append(f"Row {rn} ({name}): Buy Qty is not a valid integer.")
        if pd.isna(row["Buy Date"]) or str(row["Buy Date"]).strip() == "":
            errors.append(f"Row {rn} ({name}): Buy Date is empty.")
        sq, sp, sd = row["Sell Qty"], row["Sell Price"], row["Sell Date"]
        has = [pd.notna(x) and str(x).strip() != "" for x in [sq, sp, sd]]
        if any(has) and not all(has):
            errors.append(f"Row {rn} ({name}): Sell Qty, Sell Price & Sell Date must all be filled or all empty.")
        if all(has):
            try:
                bq = int(float(row["Buy Qty"])); sqv = int(float(sq))
                if (bq < 0 and sqv > 0) or (bq > 0 and sqv < 0):
                    errors.append(f"Row {rn} ({name}): Sell Qty sign must match Buy Qty sign.")
                if abs(sqv) > abs(bq):
                    errors.append(f"Row {rn} ({name}): |Sell Qty| cannot exceed |Buy Qty|.")
            except: errors.append(f"Row {rn} ({name}): Sell Qty invalid.")
    return df, errors


def _mi_enrich(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Buy Price"]  = pd.to_numeric(df["Buy Price"],  errors="coerce")
    df["Buy Qty"]    = pd.to_numeric(df["Buy Qty"],    errors="coerce").astype("Int64")
    df["Sell Qty"]   = pd.to_numeric(df["Sell Qty"],   errors="coerce")
    df["Sell Price"] = pd.to_numeric(df["Sell Price"], errors="coerce")
    df["Position"]   = df["Buy Qty"].apply(lambda q: "🔴 Short" if pd.notna(q) and q < 0 else "🟢 Long")
    df["Capital (₹)"] = (df["Buy Price"] * df["Buy Qty"].abs()).round(2)
    pnl = df.apply(_mi_calc_pnl, axis=1, result_type="expand")
    df["P&L (₹)"] = pnl[0]; df["P&L %"] = pnl[1]
    df["Status"]   = df["Sell Qty"].apply(lambda x: "✅ Closed" if pd.notna(x) else "🔵 Open")
    return df


def _mi_client_template() -> bytes:
    """Generate client bulk-import Excel template."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    headers = ["Client ID", "Client Name", "Password"]
    hf = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    border = Border(
        left=Side(style="thin",color="2E4A6F"), right=Side(style="thin",color="2E4A6F"),
        top=Side(style="thin",color="2E4A6F"),  bottom=Side(style="thin",color="2E4A6F"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sample = [
        ["HO668", "Suresh Kumar",  "suresh@123"],
        ["HO701", "Priya Sharma",  "priya@456"],
        ["HO802", "Ramesh Gupta",  "ramesh@789"],
        ["HO903", "Anita Reddy",   "anita@321"],
    ]
    alt_fill  = PatternFill("solid", start_color="EEF2FF", end_color="EEF2FF")
    wht_fill  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    for r_idx, row in enumerate(sample, 2):
        fill = alt_fill if r_idx % 2 == 0 else wht_fill
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill; cell.border = border
            cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate([18, 25, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    notes = wb.create_sheet("Instructions")
    notes["A1"] = "CLIENT IMPORT TEMPLATE — INSTRUCTIONS"
    notes["A1"].font = Font(bold=True, size=13, color="1E3A5F", name="Arial")
    instr = [
        ("Client ID",   "Unique client code. e.g. HO668. Will be converted to UPPERCASE."),
        ("Client Name", "Full display name. e.g. Suresh Kumar"),
        ("Password",    "Initial login password. Client can change after login."),
    ]
    notes["A3"] = "Column"; notes["B3"] = "Description"
    for cell in [notes["A3"], notes["B3"]]:
        cell.font = Font(bold=True, name="Arial")
    for i, (c, d) in enumerate(instr, 4):
        notes.cell(row=i, column=1, value=c).font = Font(name="Arial", bold=True, color="1E3A5F")
        notes.cell(row=i, column=2, value=d).font  = Font(name="Arial", size=10)
    notes["A8"] = "⚠️ Note"
    notes["A8"].font = Font(bold=True, name="Arial", color="CC0000")
    notes["B8"] = "Duplicate Client IDs will be SKIPPED. Dev code cannot be used as Client ID."
    notes["B8"].font = Font(name="Arial", color="CC0000")
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 70
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def show_master_import_tab(clients_dict_ref, save_clients_fn, hash_pw_fn,
                            client_portfolio_file_fn, client_trades_file_fn,
                            dev_code_ref):
    """Combined Master Import: one Excel creates clients + their portfolios in one shot."""

    st.markdown("""
<div style="background:linear-gradient(135deg,#0f0f23,#1a1a3e);border:1px solid #2a2a5a;
            border-radius:12px;padding:16px 24px;margin-bottom:18px;">
  <div style="font-size:20px;font-weight:800;color:#f0f2ff;">📥 Master Import</div>
  <div style="font-size:12px;color:#8888aa;margin-top:4px;">
    One Excel file → creates all clients + their full portfolios in a single import
  </div>
</div>
""", unsafe_allow_html=True)

    # ── How it works callout ──────────────────────────────────────────
    st.info(
        "**How it works:** Each row = one trade for one client. "
        "Repeat the Client ID / Name / Password for every trade of that client. "
        "Importing will **create the client account** and **save their portfolio** automatically."
    )

    col_dl, _ = st.columns([1, 3])
    with col_dl:
        st.download_button(
            "⬇️ Download Combined Template (.xlsx)",
            data=_mi_generate_template(),
            file_name="Northeast_Master_Import_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="mi_dl_combined_template",
        )

    st.markdown("---")
    uploaded = st.file_uploader(
        "Upload filled Excel file (Client ID + Client Name + Password + Trades)",
        type=["xlsx", "xls"], key="mi_combined_upload"
    )
    if uploaded is None:
        st.info("👆 Upload your filled Excel to preview and import all clients + portfolios.")
        return

    # ── Read file ────────────────────────────────────────────────────
    try:
        raw_df = pd.read_excel(uploaded, dtype=str)
        raw_df.columns = [str(c).strip() for c in raw_df.columns]
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    # ── Check required columns ───────────────────────────────────────
    required_cols = ["Client ID", "Client Name", "Password"] + _MI_REQUIRED_COLS
    missing = [c for c in required_cols if c not in raw_df.columns]
    if missing:
        st.error(f"❌ Missing columns: {', '.join(missing)}")
        st.markdown("Make sure your Excel has all these columns: " + ", ".join(required_cols))
        st.dataframe(raw_df.head(), use_container_width=True, hide_index=True)
        return

    # Fill optional cols if absent
    for c in _MI_OPTIONAL_COLS:
        if c not in raw_df.columns:
            raw_df[c] = None

    raw_df = raw_df.dropna(how="all").reset_index(drop=True)

    # ── Validate each row ────────────────────────────────────────────
    errors = []
    valid_rows = []
    for i, row in raw_df.iterrows():
        rn = i + 2
        cid  = str(row.get("Client ID", "")).strip().upper()
        cname= str(row.get("Client Name", "")).strip()
        cpw  = str(row.get("Password", "")).strip()
        sname= str(row.get("Stock Name", "")).strip()
        ac   = str(row.get("Asset Class", "")).strip()

        if not cid:   errors.append(f"Row {rn}: Client ID is empty.")
        if not cname: errors.append(f"Row {rn}: Client Name is empty.")
        if not cpw:   errors.append(f"Row {rn}: Password is empty.")
        if cid == dev_code_ref.upper():
            errors.append(f"Row {rn}: '{cid}' is the developer code — cannot use as Client ID.")
        if not sname: errors.append(f"Row {rn}: Stock Name is empty.")
        if ac not in _MI_ASSET_CLASSES:
            errors.append(f"Row {rn} ({sname}): Asset Class '{ac}' is invalid.")
        try:
            bp = float(row["Buy Price"])
            if bp <= 0: errors.append(f"Row {rn} ({sname}): Buy Price must be > 0.")
        except:
            errors.append(f"Row {rn} ({sname}): Buy Price is not a valid number.")
        try:
            bq = int(float(row["Buy Qty"]))
            if bq == 0: errors.append(f"Row {rn} ({sname}): Buy Qty cannot be 0.")
            if bq < 0 and ac not in _MI_SHORT_SUPPORTED:
                errors.append(f"Row {rn} ({sname}): Short sell not supported for '{ac}'.")
        except:
            errors.append(f"Row {rn} ({sname}): Buy Qty is not a valid number.")
        if not str(row.get("Buy Date", "")).strip():
            errors.append(f"Row {rn} ({sname}): Buy Date is empty.")

        sq = row.get("Sell Qty", None);  sp = row.get("Sell Price", None);  sd = row.get("Sell Date", None)
        has_sell = [pd.notna(x) and str(x).strip() not in ("", "nan") for x in [sq, sp, sd]]
        if any(has_sell) and not all(has_sell):
            errors.append(f"Row {rn} ({sname}): Sell Qty, Sell Price & Sell Date must all be filled or all empty.")

        if cid and cname and cpw and sname and ac in _MI_ASSET_CLASSES:
            valid_rows.append({
                "Client ID": cid, "Client Name": cname, "Password": cpw,
                "Stock Name": sname, "Asset Class": ac,
                "Buy Price": row["Buy Price"], "Buy Qty": row["Buy Qty"], "Buy Date": row["Buy Date"],
                "Sell Qty": row.get("Sell Qty"), "Sell Price": row.get("Sell Price"), "Sell Date": row.get("Sell Date"),
            })

    if errors:
        st.error(f"**{len(errors)} issue(s) found — fix before importing:**")
        for err in errors[:20]:
            st.markdown(f"- {err}")
        if len(errors) > 20:
            st.markdown(f"*...and {len(errors)-20} more. Fix all issues and re-upload.*")
        return

    if not valid_rows:
        st.error("No valid rows found.")
        return

    valid_df = pd.DataFrame(valid_rows)

    # ── Summary by client ────────────────────────────────────────────
    client_groups = valid_df.groupby("Client ID")
    client_summary = []
    for cid, grp in client_groups:
        cname = grp["Client Name"].iloc[0]
        cpw   = grp["Password"].iloc[0]
        is_new = cid not in clients_dict_ref
        client_summary.append({
            "Client ID": cid,
            "Client Name": cname,
            "Trades": len(grp),
            "Status": "🆕 New" if is_new else "⚠️ Already exists"
        })

    summary_df = pd.DataFrame(client_summary)
    new_count  = (summary_df["Status"] == "🆕 New").sum()
    exist_count= (summary_df["Status"] == "⚠️ Already exists").sum()

    st.success(f"✅ {len(valid_df)} valid trade(s) across {len(summary_df)} client(s) found. Preview:")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    m1, m2, m3 = st.columns(3)
    m1.metric("Total Clients",   len(summary_df))
    m2.metric("New Clients",     new_count)
    m3.metric("Already Exist",   exist_count)

    with st.expander("📋 Full Trade Preview", expanded=False):
        st.dataframe(
            valid_df[["Client ID","Client Name","Stock Name","Asset Class","Buy Price","Buy Qty","Buy Date","Sell Qty","Sell Price","Sell Date"]],
            use_container_width=True, hide_index=True
        )

    st.markdown("---")

    def _do_import(overwrite_existing: bool):
        imported_clients  = 0
        skipped_clients   = 0
        imported_trades   = 0

        for cid, grp in client_groups:
            cname = grp["Client Name"].iloc[0]
            cpw   = grp["Password"].iloc[0]

            # ── Create / update client account ──
            if cid not in clients_dict_ref:
                clients_dict_ref[cid] = {
                    "display_name":  cname,
                    "password_hash": hash_pw_fn(cpw),
                }
                imported_clients += 1
            elif overwrite_existing:
                clients_dict_ref[cid] = {
                    "display_name":  cname,
                    "password_hash": hash_pw_fn(cpw),
                }
                imported_clients += 1
            else:
                skipped_clients += 1

            # ── Save portfolio CSV ──
            trade_cols = ["Stock Name","Asset Class","Buy Price","Buy Qty","Buy Date","Sell Qty","Sell Price","Sell Date"]
            trade_df = grp[trade_cols].copy()

            # Rename to match app's internal column names
            # NOTE: Buy_Qty must be saved as "Shares" — that is the app's internal column name
            trade_df = trade_df.rename(columns={
                "Stock Name":  "Ticker",
                "Buy Price":   "Buy_Price",
                "Buy Qty":     "Shares",
                "Buy Date":    "Buy_Date",
                "Sell Qty":    "Sell_Qty",
                "Sell Price":  "Sell_Price",
                "Sell Date":   "Sell_Date",
                "Asset Class": "Asset_Type",
            })

            pf_path = client_portfolio_file_fn(cid)

            if os.path.exists(pf_path) and not overwrite_existing:
                # Merge / deduplicate
                existing_pf = pd.read_csv(pf_path)
                merged = pd.concat([existing_pf, trade_df], ignore_index=True).drop_duplicates(
                    subset=["Ticker","Asset_Type","Buy_Date","Buy_Price","Shares"], keep="last"
                )
                merged.to_csv(pf_path, index=False)
                _local_sync_to_gh(pf_path, pf_path, f"Import: portfolio for {cid}")
            else:
                trade_df.to_csv(pf_path, index=False)
                _local_sync_to_gh(pf_path, pf_path, f"Import: portfolio for {cid}")

            imported_trades += len(trade_df)

        # ── Save clients.json ──
        save_clients_fn(clients_dict_ref)

        return imported_clients, skipped_clients, imported_trades

    ca1, ca2 = st.columns(2)
    with ca1:
        if st.button(
            f"➕ Import {new_count} New Client(s) + Their Portfolios",
            use_container_width=True, type="primary", key="mi_combined_import_btn",
            disabled=(new_count == 0)
        ):
            ic, sc, it = _do_import(overwrite_existing=False)
            st.success(f"✅ Done! {ic} new client(s) created, {sc} skipped (already exist), {it} trade(s) saved.")
            st.balloons()
            st.info("💾 All data saved to GitHub — will persist after reboots.")
            st.rerun()

    with ca2:
        if st.button(
            f"🔄 Import & Overwrite All ({len(summary_df)} clients)",
            use_container_width=True, type="secondary", key="mi_combined_overwrite_btn"
        ):
            ic, sc, it = _do_import(overwrite_existing=True)
            st.warning(f"⚠️ {ic} client(s) imported/updated, {it} trade(s) saved (existing data overwritten).")
            st.info("💾 All data saved to GitHub — will persist after reboots.")
            st.rerun()



# MAIN DASHBOARD
# =========================================================

# ── Active navigation tab ─────────────────────────────────────────
_nav_tab = st.session_state.get("nav_tab", "Portfolio")


# =========================================================
# WATCHLIST TAB — renders before portfolio data is needed
# =========================================================
if _nav_tab == "Watchlist":

    # ── helpers ──────────────────────────────────────────────────
    _WL_INDICES = {
        # ── Indian indices ──────────────────────────────────────
        "^NSEI":   ("NIFTY 50",      "🇮🇳"),
        "^BSESN":  ("SENSEX",        "🇮🇳"),
        "^NSEBANK":("BANKNIFTY",     "🇮🇳"),
        "NIFTYMIDCAP150.NS": ("MIDCAP 150", "🇮🇳"),
        "^CNXIT":  ("NIFTY IT",      "🇮🇳"),
        "^CNXAUTO":("NIFTY AUTO",    "🇮🇳"),
        "^CNXPHARMA":("NIFTY PHARMA","🇮🇳"),
        "^CNXFMCG": ("NIFTY FMCG",  "🇮🇳"),
        "^CNXMETAL":("NIFTY METAL",  "🇮🇳"),
        "^CNXREALTY":("NIFTY REALTY","🇮🇳"),
        # ── Global indices ───────────────────────────────────────
        "^GSPC":   ("S&P 500",       "🇺🇸"),
        "^DJI":    ("DOW JONES",     "🇺🇸"),
        "^IXIC":   ("NASDAQ",        "🇺🇸"),
        "^FTSE":   ("FTSE 100",      "🇬🇧"),
        "^GDAXI":  ("DAX",           "🇩🇪"),
        "^N225":   ("NIKKEI 225",    "🇯🇵"),
        "^HSI":    ("HANG SENG",     "🇭🇰"),
        "000001.SS":("SHANGHAI",     "🇨🇳"),
    }

    # ── Build full BSE + NSE ticker list from scrip master for instant search ──
    # Each entry: {"sym": "RELIANCE", "exch": "NSE", "name": "Reliance Industries Ltd"}
    @st.cache_data(ttl=7200, show_spinner=False)
    def _build_all_exchange_tickers():
        """
        Load scrip master and extract all equity symbols from NSE + BSE.
        Returns a list of dicts: {sym, exch, name, display} for the search UI.
        Also returns a dict: sym_upper -> {exch, name} for quick lookup.
        """
        import json
        _url = "https://margincalculator.angelbroking.com/OpenAPI_File/files/OpenAPIScripMaster.json"
        _cache = "scrip_master.json"
        try:
            resp = requests.get(_url, timeout=20)
            resp.raise_for_status()
            data = resp.json()
        except Exception:
            if os.path.exists(_cache):
                with open(_cache) as f:
                    data = json.load(f)
            else:
                data = []

        seen = {}   # key = (sym, exch) to avoid dupes
        results = []
        for item in data:
            exch  = item.get("exch_seg", "")
            itype = (item.get("instrumenttype", "") or "").strip().upper()
            sym_raw = (item.get("symbol", "") or "").strip().upper()
            name    = (item.get("name",   "") or "").strip()
            if exch not in ("NSE", "BSE"):
                continue
            # Only EQ / equity instruments (skip F&O, Index, CD etc.)
            if itype and itype not in ("", "EQ", "BE", "ST"):
                continue
            # Strip BSE series suffix like "-A", "-EQ"
            sym = sym_raw.split("-")[0].strip()
            if not sym or len(sym) > 20:
                continue
            # Skip pure numeric tokens (BSE script code rows without a symbol)
            if sym.isdigit():
                continue
            key = (sym, exch)
            if key not in seen:
                seen[key] = True
                results.append({"sym": sym, "exch": exch, "name": name})

        # Sort: NSE first, then BSE, alphabetically within each
        results.sort(key=lambda x: (0 if x["exch"] == "NSE" else 1, x["sym"]))
        return results

    _ALL_EXCHANGE_TICKERS = _build_all_exchange_tickers()
    # Quick name lookup: sym_upper -> name (prefer NSE entry)
    _SYM_NAME_MAP = {}
    for _et in _ALL_EXCHANGE_TICKERS:
        _k = _et["sym"]
        if _k not in _SYM_NAME_MAP or _et["exch"] == "NSE":
            _SYM_NAME_MAP[_k] = {"name": _et["name"], "exch": _et["exch"]}

    # Keep backward-compat reference used in search below
    _NSE_TICKERS = sorted({e["sym"] for e in _ALL_EXCHANGE_TICKERS})

    def _fetch_index_prices():
        """Fetch all index prices using yfinance. Returns dict ticker→(price, change_pct)."""
        results = {}
        if not _YF_AVAILABLE:
            return results
        try:
            tickers_str = " ".join(_WL_INDICES.keys())
            data = yf.download(tickers_str, period="2d", interval="1d",
                               progress=False, group_by="ticker", auto_adjust=True)
            for sym in _WL_INDICES:
                try:
                    if len(_WL_INDICES) > 1 and sym in data.columns.get_level_values(0):
                        closes = data[sym]["Close"].dropna()
                    else:
                        closes = data["Close"].dropna()
                    if len(closes) >= 2:
                        cur, prev = float(closes.iloc[-1]), float(closes.iloc[-2])
                        results[sym] = (cur, (cur - prev) / prev * 100)
                    elif len(closes) == 1:
                        results[sym] = (float(closes.iloc[-1]), 0.0)
                except Exception:
                    pass
        except Exception:
            pass
        return results

    def _fetch_wl_prices(tickers_list):
        """Fetch LTP + prev close for watchlist tickers. Returns dict ticker→(price, chg_pct)."""
        results = {}
        if not _YF_AVAILABLE or not tickers_list:
            return results
        for sym in tickers_list:
            try:
                t = sym if (sym.endswith(".NS") or sym.endswith(".BO")) else sym + ".NS"
                obj = yf.Ticker(t)
                fast = obj.fast_info
                lp = _safe_float(getattr(fast, "last_price", None))
                pc = _safe_float(getattr(fast, "previous_close", None))
                if lp and lp > 0:
                    chg = ((lp - pc) / pc * 100) if pc and pc > 0 else 0.0
                    results[sym] = (lp, chg)
            except Exception:
                pass
        return results

    def _save_watchlist():
        """Persist watchlist for this client."""
        _code = st.session_state.get("auth_code", "")
        if _code:
            try:
                _fname = f"watchlist_{re.sub(r'[^\\w\\-]', '_', _code.upper())}.csv"
                pd.DataFrame({"ticker": st.session_state["watchlist"]}).to_csv(_fname, index=False)
            except Exception:
                pass

    def _load_watchlist():
        """Load persisted watchlist for this client."""
        _code = st.session_state.get("auth_code", "")
        if _code:
            try:
                _fname = f"watchlist_{re.sub(r'[^\\w\\-]', '_', _code.upper())}.csv"
                if os.path.exists(_fname):
                    _df = pd.read_csv(_fname)
                    return _df["ticker"].dropna().tolist()
            except Exception:
                pass
        return []

    # load from disk once per session
    if "wl_loaded" not in st.session_state:
        st.session_state["watchlist"] = _load_watchlist()
        st.session_state["wl_loaded"] = True

    # ── CSS injection — Upstox / Groww professional style ────────
    st.markdown("""
<style>
/* ═══════════════════════════════════════════════════
   WATCHLIST — UPSTOX / GROWW MOBILE DESIGN SYSTEM
   ═══════════════════════════════════════════════════ */

/* ── Mobile-only scope ── */
@media (min-width: 641px) {
  .wl-mobile-only { display: none !important; }
}

/* ── Scrolling market ticker bar ── */
.wl-ticker-wrap {
    width: 100%;
    overflow: hidden;
    background: #0b0d17;
    border-bottom: 1px solid #1e2238;
    border-radius: 10px;
    padding: 0;
    margin-bottom: 14px;
    position: relative;
}
.wl-ticker-wrap::before, .wl-ticker-wrap::after {
    content: '';
    position: absolute;
    top: 0; bottom: 0;
    width: 40px;
    z-index: 2;
    pointer-events: none;
}
.wl-ticker-wrap::before { left: 0;  background: linear-gradient(to right,#0b0d17,transparent); }
.wl-ticker-wrap::after  { right: 0; background: linear-gradient(to left,#0b0d17,transparent); }
.wl-ticker-track {
    display: flex;
    width: max-content;
    animation: wl-scroll 55s linear infinite;
}
.wl-ticker-track:hover { animation-play-state: paused; }
@keyframes wl-scroll {
    0%   { transform: translateX(0); }
    100% { transform: translateX(-50%); }
}
.wl-ticker-item {
    display: flex;
    flex-direction: column;
    justify-content: center;
    padding: 8px 20px;
    border-right: 1px solid #1a1d30;
    white-space: nowrap;
    min-width: 132px;
    cursor: default;
    transition: background 0.2s;
}
.wl-ticker-item:hover { background: #131628; }
.wl-ti-header { display:flex; align-items:center; gap:4px; margin-bottom:2px; }
.wl-ti-flag  { font-size:11px; }
.wl-ti-name  { font-size:9px; font-weight:700; color:#555880;
               text-transform:uppercase; letter-spacing:.8px; }
.wl-ti-price { font-size:14px; font-weight:700; color:#f0f2ff; line-height:1.1; }
.wl-ti-chg   { font-size:10px; font-weight:600; margin-top:1px; }
.wl-ti-chg.up   { color:#00c853; }
.wl-ti-chg.down { color:#ff4444; }
.wl-ti-chg.flat { color:#555880; }

/* ── Search box section header ── */
.wl-section-label {
    font-size: 11px;
    font-weight: 700;
    color: #555880;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin: 0 0 8px 2px;
}

/* ── Search result rows (Groww-style) ── */
.wl-sr-card {
    display: flex;
    align-items: center;
    background: #13162a;
    border: 1px solid #1e2238;
    border-radius: 12px;
    padding: 11px 14px;
    margin-bottom: 5px;
    gap: 10px;
    transition: background 0.12s, border-color 0.12s;
}
.wl-sr-card:hover { background: #1a1d35; border-color: #2e3355; }
.wl-sr-icon {
    width: 36px; height: 36px;
    border-radius: 10px;
    background: linear-gradient(135deg, #1e2545, #252d55);
    display: flex; align-items: center; justify-content: center;
    font-size: 14px; font-weight: 800; color: #4f7ef8;
    flex-shrink: 0; letter-spacing: -1px;
    border: 1px solid #2e3560;
}
.wl-sr-body { display:flex; flex-direction:column; flex:1; min-width:0; gap:1px; }
.wl-sr-sym  { font-size:14px; font-weight:800; color:#f0f2ff;
              white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.wl-sr-name { font-size:11px; color:#555880; white-space:nowrap;
              overflow:hidden; text-overflow:ellipsis; }
.wl-sr-exch-badge {
    font-size: 9px; font-weight: 700;
    padding: 2px 7px; border-radius: 20px;
    letter-spacing: .5px;
    flex-shrink: 0;
}
.wl-sr-exch-nse { background:#4f7ef822; color:#4f7ef8; border:1px solid #4f7ef844; }
.wl-sr-exch-bse { background:#f5c84222; color:#f5c842; border:1px solid #f5c84244; }

/* ══ WATCHLIST CARDS — Groww / Upstox row cards ══ */
.wl-card {
    background: #13162a;
    border: 1px solid #1e2238;
    border-radius: 14px;
    padding: 13px 14px 11px 14px;
    margin-bottom: 7px;
    transition: background 0.12s, border-color 0.12s;
    position: relative;
    overflow: hidden;
}
.wl-card:hover { background: #1a1d38; border-color: #2e3358; }
.wl-card-top {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    gap: 8px;
}
.wl-card-left { display:flex; align-items:center; gap:10px; min-width:0; }
.wl-card-avatar {
    width: 38px; height: 38px; border-radius: 10px; flex-shrink: 0;
    background: linear-gradient(135deg,#1e2545,#252d55);
    border: 1px solid #2a3060;
    display: flex; align-items: center; justify-content: center;
    font-size: 13px; font-weight: 900; color: #4f7ef8; letter-spacing: -1px;
}
.wl-card-info  { display:flex; flex-direction:column; gap:1px; min-width:0; }
.wl-card-sym   { font-size:15px; font-weight:800; color:#f0f2ff;
                 white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.wl-card-co    { font-size:11px; color:#555880; white-space:nowrap;
                 overflow:hidden; text-overflow:ellipsis; max-width:160px; }
.wl-card-right { display:flex; flex-direction:column; align-items:flex-end; gap:3px; flex-shrink:0; }
.wl-card-price { font-size:16px; font-weight:800; color:#f0f2ff; line-height:1; }
.wl-card-pill  {
    display: inline-flex;
    align-items: center;
    gap: 3px;
    font-size: 11px;
    font-weight: 700;
    padding: 3px 9px;
    border-radius: 20px;
    line-height: 1;
}
.wl-pill-up   { background:#00c85318; color:#00c853; border:1px solid #00c85340; }
.wl-pill-down { background:#ff444418; color:#ff4444; border:1px solid #ff444440; }
.wl-pill-flat { background:#55588018; color:#7a7fa8; border:1px solid #55588040; }

/* Exchange micro-badge */
.wl-exch-dot {
    font-size: 9px; font-weight: 700;
    padding: 1px 6px; border-radius: 4px;
    margin-left: 4px; vertical-align: middle;
    letter-spacing: .4px;
}
.wl-exch-nse { background:#4f7ef822; color:#4f7ef8; border:1px solid #4f7ef833; }
.wl-exch-bse { background:#f5c84222; color:#f5c842; border:1px solid #f5c84233; }

/* Action buttons row */
.wl-card-actions {
    display: flex;
    gap: 6px;
    margin-top: 10px;
    padding-top: 10px;
    border-top: 1px solid #1a1d30;
}
.wl-act-btn {
    flex: 1;
    padding: 8px 0;
    border-radius: 9px;
    font-size: 12px;
    font-weight: 700;
    text-align: center;
    cursor: pointer;
    border: none;
    line-height: 1;
    letter-spacing: .3px;
}
.wl-act-buy  { background:#00c85320; color:#00c853; border:1px solid #00c85355 !important; }
.wl-act-sell { background:#ff444420; color:#ff4444; border:1px solid #ff444455 !important; }
.wl-act-del  { background:transparent; color:#555880; border:1px solid #252849 !important;
               flex: 0 0 36px; }
.wl-act-del:hover { color:#ff4444; border-color:#ff444466 !important; }

/* Empty state */
.wl-empty {
    background: #13162a;
    border: 1.5px dashed #1e2238;
    border-radius: 16px;
    padding: 36px 20px;
    text-align: center;
}

/* Compact section counter badge */
.wl-count-badge {
    background: #1e2238;
    color: #7a7fa8;
    font-size: 10px;
    font-weight: 700;
    padding: 2px 10px;
    border-radius: 20px;
    border: 1px solid #252849;
    letter-spacing: .4px;
    display: inline-block;
    vertical-align: middle;
    margin-left: 6px;
}
</style>
""", unsafe_allow_html=True)

    # ── Header — Upstox style ─────────────────────────────────────
    st.markdown(f"""
<div style="display:flex;align-items:center;justify-content:space-between;
            margin-bottom:14px;">
  <div>
    <div style="font-size:20px;font-weight:800;color:#f0f2ff;letter-spacing:-0.3px;">
      Watchlist
    </div>
    <div style="font-size:11px;color:#555880;margin-top:2px;">
      Market indices &amp; your stocks
    </div>
  </div>
  <div style="background:#1e2238;border:1px solid #252849;border-radius:20px;
              padding:4px 12px;font-size:10px;font-weight:700;color:#7a7fa8;">
    {'🟢 MARKET OPEN' if _mkt_open else '🔴 MARKET CLOSED'}
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Section 1: Scrolling Index Ticker ────────────────────────
    with st.spinner("Loading indices…"):
        _idx_prices = _fetch_index_prices()

    # Build ticker items (duplicated for seamless loop)
    def _build_ticker_items():
        items = ""
        for _sym, (_lbl, _flag) in _WL_INDICES.items():
            _price, _chg = _idx_prices.get(_sym, (None, 0.0))
            _price_str = f"{_price:,.1f}" if _price else "—"
            _chg_cls = "up" if _chg > 0 else ("down" if _chg < 0 else "flat")
            _chg_arrow = "▲" if _chg > 0 else ("▼" if _chg < 0 else "—")
            _chg_str = f"{_chg_arrow} {abs(_chg):.2f}%" if _price else "—"
            items += f"""<div class="wl-ticker-item">
  <div class="wl-ti-header">
    <span class="wl-ti-flag">{_flag}</span>
    <span class="wl-ti-name">{_lbl}</span>
  </div>
  <span class="wl-ti-price">{_price_str}</span>
  <span class="wl-ti-chg {_chg_cls}">{_chg_str}</span>
</div>"""
        return items

    _ticker_items = _build_ticker_items()
    st.markdown(f"""
<div class="wl-ticker-wrap">
  <div class="wl-ticker-track">
    {_ticker_items}
    {_ticker_items}
  </div>
</div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ── Extra CSS for search results + compact rows + new mobile design ──
    st.markdown("""
<style>
/* ── Custom search wrapper with clear button ── */
.wl-search-wrap {
    position: relative;
    margin-bottom: 8px;
}
.wl-search-wrap input {
    width: 100% !important;
    background: #13162a !important;
    border: 1.5px solid #252d55 !important;
    border-radius: 12px !important;
    color: #f0f2ff !important;
    font-size: 14px !important;
    padding: 11px 42px 11px 14px !important;
    outline: none !important;
    box-sizing: border-box;
    transition: border-color 0.15s;
}
.wl-search-wrap input:focus { border-color: #4f7ef8 !important; }
.wl-search-clear {
    position: absolute;
    right: 12px; top: 50%;
    transform: translateY(-50%);
    background: #252849;
    border: none;
    color: #7a7fa8;
    width: 22px; height: 22px;
    border-radius: 50%;
    font-size: 13px;
    font-weight: 700;
    cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    line-height: 1;
    transition: background 0.12s, color 0.12s;
    padding: 0;
    z-index: 10;
}
.wl-search-clear:hover { background:#f8545430; color:#f85454; }

/* Search result row */
.wl-sr-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background: #131628;
    border: 1px solid #252849;
    border-radius: 8px;
    padding: 9px 14px;
    margin-bottom: 4px;
}
.wl-sr-left { display:flex; flex-direction:column; gap:1px; }
.wl-sr-sym-t  { font-size: 13px; font-weight: 700; color: #f0f2ff; }
.wl-sr-exch { font-size: 10px; color: #555880; }

/* ── Watchlist row — new Upstox-style single line ── */
.wl-row {
    display: flex;
    align-items: center;
    background: #13162a;
    border: 1px solid #1e2238;
    border-radius: 13px;
    padding: 13px 13px;
    margin-bottom: 6px;
    gap: 10px;
    overflow: hidden;
    cursor: pointer;
    transition: background 0.12s, border-color 0.12s;
    position: relative;
}
.wl-row:active, .wl-row.wl-row-open {
    background: #181c38;
    border-color: #4f7ef855;
}
/* Avatar circle */
.wl-row-avatar {
    width: 40px; height: 40px;
    border-radius: 11px;
    background: linear-gradient(135deg, #1e2545, #252d55);
    border: 1px solid #2a3060;
    display: flex; align-items: center; justify-content: center;
    font-size: 13px; font-weight: 900; color: #4f7ef8;
    letter-spacing: -1px;
    flex-shrink: 0;
}
/* Left block: ticker + exchange */
.wl-row-left {
    display: flex;
    flex-direction: column;
    flex: 1 1 auto;
    min-width: 0;
    gap: 1px;
}
.wl-row-sym  {
    font-size: 14px; font-weight: 800; color: #f0f2ff;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.wl-row-co   {
    font-size: 10px; color: #555880;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    max-width: 140px;
}
/* Right price block: price + change, right-aligned */
.wl-row-price-block {
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    flex-shrink: 0;
    gap: 2px;
}
.wl-row-price { font-size: 15px; font-weight: 800; color: #f0f2ff; white-space: nowrap; }
.wl-row-pill  {
    display: inline-flex; align-items: center;
    font-size: 10px; font-weight: 700;
    padding: 2px 8px; border-radius: 20px; white-space: nowrap;
}
.wl-pill-up   { background:#00c85318; color:#00c853; border:1px solid #00c85340; }
.wl-pill-down { background:#ff444418; color:#ff4444; border:1px solid #ff444440; }
.wl-pill-flat { background:#55588018; color:#7a7fa8; border:1px solid #55588040; }

/* Exchange micro-badge on row */
.wl-row-exch {
    font-size: 9px; font-weight: 700;
    padding: 1px 6px; border-radius: 4px;
    letter-spacing: .4px;
    flex-shrink: 0;
    display: inline-block;
    margin-left: 5px;
    vertical-align: middle;
}
.wl-row-exch-nse { background:#4f7ef822; color:#4f7ef8; border:1px solid #4f7ef833; }
.wl-row-exch-bse { background:#f5c84222; color:#f5c842; border:1px solid #f5c84233; }

/* ── Trade action row — appears below row when open ── */
.wl-trade-row {
    display: flex;
    gap: 7px;
    padding: 0 2px 4px 2px;
    margin-top: -3px;
    margin-bottom: 6px;
    animation: wlSlide 0.15s ease;
}
@keyframes wlSlide {
    from { opacity:0; transform:translateY(-6px); }
    to   { opacity:1; transform:translateY(0); }
}
.wl-tb {
    flex: 1;
    padding: 9px 0;
    border-radius: 9px;
    font-size: 13px;
    font-weight: 700;
    cursor: pointer;
    text-align: center;
    border: none;
    line-height: 1;
    letter-spacing: .3px;
    transition: opacity 0.15s;
}
.wl-tb:active { opacity: 0.75; }
.wl-tb-buy  { background:#00c85320; color:#00c853; border:1px solid #00c85366 !important; }
.wl-tb-sell { background:#ff444420; color:#ff4444; border:1px solid #ff444466 !important; }
.wl-tb-del  {
    background: transparent; color: #555880;
    border: 1px solid #252849 !important;
    flex: 0 0 44px;
}
.wl-tb-del:hover { color:#f85454; }

/* Plus button in search results */
.wl-plus-btn {
    flex: 0 0 auto;
    width: 32px; height: 32px;
    border-radius: 50%;
    border: 1.5px solid #4f7ef8;
    background: transparent;
    color: #4f7ef8;
    font-size: 20px;
    font-weight: 700;
    cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    transition: all 0.15s;
    padding: 0;
}
.wl-plus-btn:hover { background:#4f7ef822; }

/* Order panel */
.wl-order-panel {
    background: #0d1020;
    border-radius: 14px;
    padding: 16px;
    margin-bottom: 12px;
}
.wl-order-title {
    font-size: 14px; font-weight: 800;
    margin-bottom: 12px;
}
/* ── Search result item (new single-row card style) ── */
.wl-sr-item {
    display: flex;
    align-items: center;
    padding: 11px 12px;
    gap: 10px;
    background: #13162a;
    border: 1px solid #1e2238;
    border-radius: 12px;
    margin-bottom: 6px;
    transition: background 0.1s, border-color 0.1s;
    cursor: default;
}
.wl-sr-item:hover { background: #181c38; border-color: #2e3355; }
.wl-sr-icon2 {
    width: 38px; height: 38px;
    border-radius: 10px;
    background: linear-gradient(135deg, #1e2545, #252d55);
    border: 1px solid #2e3560;
    display: flex; align-items: center; justify-content: center;
    font-size: 13px; font-weight: 800; color: #4f7ef8;
    flex-shrink: 0; letter-spacing: -1px;
}
.wl-sr-info { flex: 1; min-width: 0; }
.wl-sr-sym2  { font-size: 14px; font-weight: 800; color: #f0f2ff; line-height: 1.3; }
.wl-sr-name2 { font-size: 11px; color: #555880; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; margin-top: 1px; }
/* ＋ button inside the HTML row — no longer used, kept for compat */
.wl-sr-plus { display: none; }

/* ── Fuse the info column + button column into one visual row ── */
.wl-sr-row-wrap > div[data-testid="stHorizontalBlock"] {
    gap: 0 !important;
    margin-bottom: 6px !important;
    align-items: stretch !important;
}
/* Info column: card fills left, no margin */
.wl-sr-row-wrap > div[data-testid="stHorizontalBlock"] > div:first-child {
    flex: 1 !important;
}
.wl-sr-row-wrap .wl-sr-item {
    border-radius: 12px 0 0 12px !important;
    border-right: none !important;
    margin-bottom: 0 !important;
    height: 100%;
}
/* Button column: snaps flush to the right of the card */
.wl-sr-row-wrap > div[data-testid="stHorizontalBlock"] > div:last-child {
    flex: 0 0 52px !important;
    min-width: 52px !important;
    max-width: 52px !important;
}
.wl-sr-row-wrap > div[data-testid="stHorizontalBlock"] > div:last-child [data-testid="stButton"] > button {
    height: 100% !important;
    min-height: 62px !important;
    border-radius: 0 12px 12px 0 !important;
    border: 1px solid #1e2238 !important;
    border-left: 1px solid #2a3060 !important;
    background: #1a1f3a !important;
    color: #4f7ef8 !important;
    font-size: 22px !important;
    font-weight: 300 !important;
    padding: 0 !important;
    margin: 0 !important;
    line-height: 1 !important;
    width: 100% !important;
}
.wl-sr-row-wrap > div[data-testid="stHorizontalBlock"] > div:last-child [data-testid="stButton"] > button:hover {
    background: #4f7ef825 !important;
    color: #7fa8ff !important;
}

/* ── Mobile: ✕ clear button compact ── */
@media (max-width: 640px) {
    border-radius: 50% !important;
    padding: 4px 6px !important;
    min-height: 38px !important;
    font-size: 14px !important;
    font-weight: 700 !important;
    background: #252849 !important;
    color: #7a7fa8 !important;
    border: 1px solid #2e3355 !important;
  }
  /* Tap-zone button — fully invisible overlay */
  button[title*="Tap to trade"],
  button[title*="trade"] {
    opacity: 0 !important;
    position: relative !important;
    margin-top: -64px !important;
    height: 64px !important;
    z-index: 5 !important;
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    cursor: pointer !important;
  }
}
</style>
""", unsafe_allow_html=True)

    # ── Section 2: Stock Search — Upstox style with clear button ──
    # We use a JS-powered search that stores in session_state via URL param trick
    st.markdown('<div class="wl-section-label">🔍 Add to Watchlist</div>',
                unsafe_allow_html=True)

    # Inject JS to wire up the clear button
    st.markdown("""
<script>
(function() {
  function _wireClear() {
    var clearBtn = document.getElementById('wl-clear-btn');
    var inp = document.querySelector('[data-testid="stTextInput"] input#wl_search_input');
    if (!clearBtn || !inp) {
      // Also try by placeholder
      inp = document.querySelector('input[placeholder*="RELIANCE"]') ||
            document.querySelector('input[placeholder*="Search by symbol"]');
    }
    if (clearBtn && inp) {
      clearBtn.addEventListener('click', function(e) {
        e.preventDefault();
        e.stopPropagation();
        // Set native input value to '' and dispatch React-compatible events
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(inp, '');
        inp.dispatchEvent(new Event('input', { bubbles: true }));
        inp.dispatchEvent(new Event('change', { bubbles: true }));
        inp.focus();
      });
    }
  }
  // Retry a few times until DOM is ready
  var _tries = 0;
  var _iv = setInterval(function(){
    _wireClear();
    if (++_tries > 20) clearInterval(_iv);
  }, 300);
})();
</script>
""", unsafe_allow_html=True)

    # ── Search input — use a reset-flag to clear without touching
    #    the widget key after it has been instantiated (avoids
    #    StreamlitAPIException).
    # If a clear was requested on the previous run, use value="" and
    # drop the flag BEFORE the widget is created.
    _wl_search_default = ""
    if st.session_state.get("_wl_search_reset"):
        st.session_state.pop("_wl_search_reset", None)
        # Overwrite the stored value while the widget hasn't rendered yet
        if "wl_search_box" in st.session_state:
            del st.session_state["wl_search_box"]
        _wl_search_default = ""

    _search_col, _clear_col = st.columns([9, 1])
    with _search_col:
        _search_q = st.text_input(
            label="",
            placeholder="Search symbol or company…  e.g. RELIANCE, Infosys",
            key="wl_search_box",
            label_visibility="collapsed",
        ).strip().upper()
    with _clear_col:
        if _search_q:
            if st.button("✕", key="wl_search_clear", use_container_width=True,
                         help="Clear search"):
                # Set flag — actual clear happens at top of next run
                st.session_state["_wl_search_reset"] = True
                st.rerun()

    if _search_q:
        # Match against sym OR company name (case-insensitive)
        _already = set(t.replace(".NS","").replace(".BO","").upper()
                       for t in st.session_state["watchlist"])
        _matches = []
        for _et in _ALL_EXCHANGE_TICKERS:
            _s = _et["sym"]
            _n = (_et["name"] or "").upper()
            if _search_q in _s or _search_q in _n:
                _matches.append(_et)
            if len(_matches) >= 20:
                break
        _shown = [m for m in _matches if m["sym"] not in _already]

        if _shown:
            for _m in _shown[:12]:
                _sym  = _m["sym"]
                _name = _m.get("name", "") or ""
                _exch = _m.get("exch", "NSE")
                _initials = (_sym[:2] if len(_sym) >= 2 else _sym).upper()
                _badge_cls = "wl-sr-exch-nse" if _exch == "NSE" else "wl-sr-exch-bse"
                _name_disp = (_name[:32] + "…") if len(_name) > 32 else (_name or _sym)

                # Wrap each row in a unique container so we can CSS-target it
                _row_key = f"{_sym}_{_exch}"
                st.markdown(f'<div class="wl-sr-row-wrap" id="wlwrap-{_row_key}">', unsafe_allow_html=True)
                _ci, _cb = st.columns([6, 1])
                with _ci:
                    st.markdown(f"""
<div class="wl-sr-item">
  <div class="wl-sr-icon2">{_initials}</div>
  <div class="wl-sr-info">
    <div class="wl-sr-sym2">{_sym}&nbsp;<span class="wl-row-exch {_badge_cls}">{_exch}</span></div>
    <div class="wl-sr-name2">{_name_disp}</div>
  </div>
</div>""", unsafe_allow_html=True)
                with _cb:
                    if st.button("＋", key=f"wladd_{_row_key}", use_container_width=True):
                        _ticker_to_add = _sym + (".NS" if _exch == "NSE" else ".BO")
                        if _ticker_to_add not in st.session_state["watchlist"]:
                            st.session_state["watchlist"].append(_ticker_to_add)
                        _save_watchlist()
                        st.session_state["_wl_search_reset"] = True
                        st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
        elif _matches:
            st.caption("All matching stocks already in your watchlist.")
        else:
            st.caption("No stocks found. Try a different symbol or company name.")

    st.markdown("---")

    # ── Section 3: My Watchlist — Upstox-style tap-to-trade ────────
    _wl = st.session_state["watchlist"]
    st.markdown(f"""
<div style="display:flex;align-items:center;margin:14px 0 10px 2px;">
  <span style="font-size:12px;font-weight:700;color:#555880;
               text-transform:uppercase;letter-spacing:1px;">⭐ My Watchlist</span>
  <span class="wl-count-badge">{len(_wl)}</span>
  <span style="font-size:10px;color:#454870;margin-left:auto;margin-right:2px;">
    Tap stock to trade
  </span>
</div>""", unsafe_allow_html=True)

    if not _wl:
        st.markdown("""
<div class="wl-empty">
  <div style="font-size:32px;margin-bottom:10px;">📋</div>
  <div style="font-size:14px;font-weight:700;color:#f0f2ff;margin-bottom:4px;">
    Your watchlist is empty
  </div>
  <div style="font-size:12px;color:#555880;">
    Search above and tap <b style="color:#4f7ef8;">＋</b> to add a stock
  </div>
</div>""", unsafe_allow_html=True)
    else:
        with st.spinner("Fetching live prices…"):
            _wl_prices = _fetch_wl_prices(_wl)

        # ── Inline order panel ────────────────────────────────────
        if st.session_state.get("wl_order"):
            _ord      = st.session_state["wl_order"]
            _ord_sym  = _ord["sym"]
            _ord_side = _ord["side"]
            _ord_color = "#00c853" if _ord_side == "BUY" else "#ff4444"
            _ltp_val   = _wl_prices.get(_ord_sym, (None, 0))[0]

            st.markdown(f"""
<div style="background:#0d1020;border:2px solid {_ord_color}44;border-radius:14px;
            padding:16px;margin-bottom:14px;">
  <div style="font-size:14px;font-weight:800;color:{_ord_color};margin-bottom:12px;">
    {'🟢' if _ord_side=='BUY' else '🔴'} {_ord_side} — {_ord_sym}
  </div>
</div>""", unsafe_allow_html=True)

            _oc1, _oc2 = st.columns(2)
            with _oc1:
                _ord_qty = st.number_input("Qty", min_value=1, value=1,
                                            key="wl_ord_qty", step=1)
            with _oc2:
                _ord_price = st.number_input("Price (₹)", min_value=0.01,
                                              value=float(_ltp_val) if _ltp_val else 100.0,
                                              key="wl_ord_price", step=0.05, format="%.2f")

            # Show buy rate input only for SELL orders
            if _ord_side == "SELL":
                _clean_for_hint = _ord_sym.replace(".NS","").replace(".BO","").upper()
                try:
                    _pf_hint = pd.read_csv(PORTFOLIO_FILE) if os.path.exists(PORTFOLIO_FILE) else pd.DataFrame()
                    _pf_row  = _pf_hint[_pf_hint["Ticker"] == _clean_for_hint] if not _pf_hint.empty else pd.DataFrame()
                    _buy_hint = round(
                        (_pf_row["Shares"] * _pf_row["Buy_Price"]).sum() / _pf_row["Shares"].sum(), 2
                    ) if not _pf_row.empty and _pf_row["Shares"].sum() > 0 else 0.01
                except Exception:
                    _buy_hint = 0.01
                _ord_buy_rate = st.number_input(
                    "Your Buy Rate ₹",
                    min_value=0.01, value=float(_buy_hint), step=0.01,
                    key="wl_ord_buy_rate",
                    help="Buy rate used to calculate booked P&L for this sell."
                )
            else:
                _ord_buy_rate = 0.0

            st.markdown(
                f"<div style='font-size:11px;color:#7a7fa8;margin:6px 0 10px;'>"
                f"Order Value: <b style='color:#f0f2ff;'>₹{_ord_qty * _ord_price:,.2f}</b></div>",
                unsafe_allow_html=True)

            _ob1, _ob2 = st.columns([3, 1])
            with _ob1:
                if st.button(f"✅ Confirm {_ord_side}", key="wl_ord_confirm",
                             use_container_width=True, type="primary"):

                    _clean_ticker = _ord_sym.replace(".NS","").replace(".BO","").upper()
                    _today_str    = datetime.now().strftime("%Y-%m-%d")

                    # ── Reload latest portfolio from disk ─────────────
                    if os.path.exists(PORTFOLIO_FILE):
                        _pf = pd.read_csv(PORTFOLIO_FILE)
                    else:
                        _pf = pd.DataFrame(columns=["Ticker","Shares","Buy_Price","Buy_Date","Asset_Type"])

                    if _ord_side == "BUY":
                        _existing = _pf[_pf["Ticker"] == _clean_ticker]
                        if not _existing.empty:
                            _idx      = _existing.index[0]
                            _old_qty  = float(_pf.at[_idx, "Shares"])
                            _old_px   = float(_pf.at[_idx, "Buy_Price"])
                            _new_qty  = _old_qty + _ord_qty
                            _avg_px   = round((_old_qty * _old_px + _ord_qty * _ord_price) / _new_qty, 4)
                            _pf.at[_idx, "Shares"]    = _new_qty
                            _pf.at[_idx, "Buy_Price"] = _avg_px
                            _pf.at[_idx, "Buy_Date"]  = _today_str
                            _pf.to_csv(PORTFOLIO_FILE, index=False)
                            _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                            st.success(
                                f"✅ BUY recorded — {_clean_ticker} · "
                                f"New Qty: {_new_qty:,.2f} · Avg Price: ₹{_avg_px:,.2f}"
                            )
                        else:
                            # Classify asset type
                            try:
                                _atype = classify_asset(_clean_ticker)
                            except Exception:
                                _atype = "Equity"
                            _new_row = pd.DataFrame(
                                [[_clean_ticker, _ord_qty, _ord_price, _today_str, _atype]],
                                columns=["Ticker","Shares","Buy_Price","Buy_Date","Asset_Type"]
                            )
                            _pf = pd.concat([_pf, _new_row], ignore_index=True)
                            _pf.to_csv(PORTFOLIO_FILE, index=False)
                            _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                            st.success(
                                f"✅ BUY recorded — {_clean_ticker} added · "
                                f"Qty: {_ord_qty} · Price: ₹{_ord_price:.2f}"
                            )

                    elif _ord_side == "SELL":
                        # Reload trades file too
                        if os.path.exists(TRADES_FILE):
                            _tf = pd.read_csv(TRADES_FILE)
                        else:
                            _tf = pd.DataFrame(columns=["Ticker","Sell_Qty","Sell_Price",
                                                          "Sell_Date","Buy_Price_At_Sell","Booked_PnL","Asset_Type"])

                        _existing = _pf[_pf["Ticker"] == _clean_ticker]
                        if _existing.empty:
                            st.error(f"❌ {_clean_ticker} not found in your portfolio.")
                        else:
                            _wl_atype = str(_pf.at[_existing.index[0], "Asset_Type"]) if "Asset_Type" in _pf.columns else _classify_ticker_to_asset_type(_clean_ticker)
                            # Use the buy rate the user entered
                            _pnl = round((_ord_price - _ord_buy_rate) * _ord_qty, 2)
                            # Each sell = new separate row (no merging, no deduction from portfolio Shares)
                            _sell_row = pd.DataFrame(
                                [[_clean_ticker, _ord_qty, _ord_price,
                                  _today_str, round(_ord_buy_rate, 2), _pnl, _wl_atype]],
                                columns=["Ticker","Sell_Qty","Sell_Price",
                                         "Sell_Date","Buy_Price_At_Sell","Booked_PnL","Asset_Type"]
                            )
                            _tf = pd.concat([_tf, _sell_row], ignore_index=True)
                            _tf.to_csv(TRADES_FILE, index=False)
                            _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')
                            _pnl_str = f"{'▲' if _pnl >= 0 else '▼'} ₹{abs(_pnl):,.2f}"
                            st.success(
                                f"✅ SELL recorded — {_clean_ticker} · "
                                f"Qty: {_ord_qty} · Price: ₹{_ord_price:.2f} · "
                                f"Buy Rate: ₹{_ord_buy_rate:.2f} · P&L: {_pnl_str}"
                            )

                    st.session_state["wl_order"] = None
                    st.rerun()
            with _ob2:
                if st.button("✕", key="wl_ord_cancel", use_container_width=True):
                    st.session_state["wl_order"] = None
                    st.rerun()

            st.markdown("---")

        # ── Watchlist rows — tap to expand BUY/SELL ────────────────
        _to_remove    = None
        _expand_trade = st.session_state.get("wl_expand", None)

        for _wi, _wsym in enumerate(_wl):
            # Resolve display info
            _clean_sym = _wsym.replace(".NS","").replace(".BO","").upper()
            _is_bse    = _wsym.endswith(".BO")
            _exch_label = "BSE" if _is_bse else "NSE"
            _exch_cls   = "wl-row-exch-bse" if _is_bse else "wl-row-exch-nse"
            _co_name    = (_SYM_NAME_MAP.get(_clean_sym) or {}).get("name", "") or ""
            _initials   = (_clean_sym[:2] if len(_clean_sym) >= 2 else _clean_sym)

            _wlp, _wlc = _wl_prices.get(_wsym, (None, 0.0))
            _price_disp = f"₹{_wlp:,.2f}" if _wlp else "—"
            _chg_cls   = "wl-pill-up" if (_wlc or 0) > 0 else ("wl-pill-down" if (_wlc or 0) < 0 else "wl-pill-flat")
            _chg_arrow = "▲" if (_wlc or 0) > 0 else ("▼" if (_wlc or 0) < 0 else "—")
            _chg_disp  = f"{_chg_arrow} {abs(_wlc):.2f}%" if _wlp else "—"

            _is_open = (_expand_trade == _wsym)
            _open_cls = "wl-row-open" if _is_open else ""

            # ── Stock row — tap = toggle trade panel ─────────────
            st.markdown(f"""
<div class="wl-row {_open_cls}">
  <div class="wl-row-avatar">{_initials}</div>
  <div class="wl-row-left">
    <span class="wl-row-sym">{_clean_sym}
      <span class="wl-row-exch {_exch_cls}">{_exch_label}</span>
    </span>
    <span class="wl-row-co">{_co_name if _co_name else "—"}</span>
  </div>
  <div class="wl-row-price-block">
    <span class="wl-row-price">{_price_disp}</span>
    <span class="wl-row-pill {_chg_cls}">{_chg_disp}</span>
  </div>
</div>""", unsafe_allow_html=True)

            # ── Trade buttons — only shown when row is open ───────
            if _is_open:
                # BUY | SELL | ✕ — all in one line
                _tb1, _tb2, _tb3 = st.columns([5, 5, 1.8])
                with _tb1:
                    if st.button("🟢 BUY", key=f"wlbuy_{_wi}_{_wsym}",
                                 use_container_width=True):
                        st.session_state["wl_order"]  = {"sym": _wsym, "side": "BUY"}
                        st.session_state["wl_expand"] = None
                        st.rerun()
                with _tb2:
                    if st.button("🔴 SELL", key=f"wlsell_{_wi}_{_wsym}",
                                 use_container_width=True):
                        st.session_state["wl_order"]  = {"sym": _wsym, "side": "SELL"}
                        st.session_state["wl_expand"] = None
                        st.rerun()
                with _tb3:
                    if st.button("✕", key=f"wlclose_{_wi}_{_wsym}",
                                 use_container_width=True):
                        st.session_state["wl_expand"] = None
                        st.rerun()
            else:
                # Invisible tap zone — uses a tiny button styled as the card
                _ta1, _ta2 = st.columns([9, 1])
                with _ta1:
                    if st.button("​", key=f"wltap_{_wi}_{_wsym}",  # zero-width space label
                                 use_container_width=True,
                                 help=f"Tap to trade {_clean_sym}"):
                        # Toggle expand
                        if st.session_state.get("wl_expand") == _wsym:
                            st.session_state["wl_expand"] = None
                        else:
                            st.session_state["wl_expand"] = _wsym
                        st.rerun()
                with _ta2:
                    if st.button("✕", key=f"wldel_{_wi}_{_wsym}",
                                 use_container_width=True,
                                 help=f"Remove {_clean_sym}"):
                        _to_remove = _wsym

            # Hide the ugly Streamlit tap button — it overlaps the card HTML above
            st.markdown("""
<style>
/* Make the invisible tap button completely transparent so card HTML is what you see */
[data-testid="stButton"] > button[title*="Tap to trade"],
[data-testid="stButton"] > button[title*="trade"] {
    position: absolute !important;
    opacity: 0 !important;
    height: 62px !important;
    margin-top: -68px !important;
    z-index: 5 !important;
    cursor: pointer !important;
    border-radius: 13px !important;
    border: none !important;
    background: transparent !important;
    box-shadow: none !important;
}
</style>""", unsafe_allow_html=True)

        if _to_remove:
            st.session_state["watchlist"] = [x for x in _wl if x != _to_remove]
            if st.session_state.get("wl_expand") == _to_remove:
                st.session_state["wl_expand"] = None
            _save_watchlist()
            st.rerun()

        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
        if st.button("🔄 Refresh Prices", key="wl_refresh"):
            st.rerun()

    st.stop()

# ── end of Watchlist tab ──────────────────────────────────────────


# Ensure Asset_Type column exists
if "Asset_Type" not in df.columns:
    df["Asset_Type"] = df["Ticker"].apply(classify_asset)

# ── Apply sold quantities → net holdings ──────────────────────────
df_net, booked_pnl_map = compute_net_holdings(df, trades_df)
# Keep long positions (Shares > 0) AND open short positions (Shares < 0)
df_active = df_net[df_net["Shares"] != 0].copy()
# Tag short positions for downstream rendering
if "Is_Short" not in df_active.columns:
    df_active["Is_Short"] = df_active["Shares"] < 0
else:
    df_active["Is_Short"] = (
        df_active["Is_Short"].astype(str).str.lower().isin(["true","1","yes"])
        | (df_active["Shares"] < 0)
    )

tickers = df["Ticker"].unique().tolist()  # original list for news/calendar

# =====================================================
# FETCH LIVE PRICES — ANGEL ONE ONLY
# =====================================================

with st.spinner("📡 Loading Angel One Scrip Master..."):
    equity_map, bse_map, nfo_map, sym_to_name = load_scrip_master()
    if not equity_map and not bse_map:
        st.sidebar.warning("⚠️ Scrip master unavailable — prices may be missing")

with st.spinner("📡 Fetching Live Prices from Angel One..."):
    prices = get_angel_one_price_bulk(df_active, equity_map, nfo_map, bse_map=bse_map)

with st.spinner("📡 Fetching Previous Close Prices..."):
    prev_closes = get_prev_close_bulk(df_active, equity_map, nfo_map, bse_map=bse_map)

# =====================================================
# CALCULATIONS
# =====================================================

calc = df_active.copy()
calc["Current_Price"] = calc["Ticker"].map(prices)

# Short positions: profit = (Sell_Init_Price - CMP) x |Qty|
# (Buy_Price column stores the price at which the short was initiated)
import numpy as _np_short
_is_short_mask = calc.get("Is_Short", __import__('pandas').Series(False, index=calc.index)).astype(bool)
_abs_shares = calc["Shares"].abs()

# Value and Cost_Basis use absolute qty for both long and short
calc["Value"] = (_abs_shares * calc["Current_Price"]).round(2)
calc["Cost_Basis"] = (_abs_shares * calc["Buy_Price"]).round(2)

# Unrealized P&L:
#   Long:  (CMP - Buy_Price) * Shares
#   Short: (Buy_Price - CMP) * abs(Shares)  -- profit when CMP falls below short price
calc["Unrealized_PnL"] = _np_short.where(
    _is_short_mask,
    ((calc["Buy_Price"] - calc["Current_Price"]) * _abs_shares).round(2),
    (calc["Value"] - calc["Cost_Basis"]).round(2)
)
calc["PnL_%"] = (calc["Unrealized_PnL"] / calc["Cost_Basis"].replace(0, float("nan")) * 100).round(2)

calc["Prev_Close"]  = calc["Ticker"].map(prev_closes)
# Daily P&L: longs gain on price rise; shorts gain on price fall
calc["Daily_PnL"] = _np_short.where(
    _is_short_mask,
    ((calc["Prev_Close"] - calc["Current_Price"]) * _abs_shares).round(2),
    ((calc["Current_Price"] - calc["Prev_Close"]) * calc["Shares"]).round(2)
)
calc["Daily_PnL_%"] = ((calc["Current_Price"] - calc["Prev_Close"]) / calc["Prev_Close"] * 100).round(2)

# Carry Sold_Qty into calc for display
calc["Sold_Qty"] = df_net.set_index(df_net.index)["Sold_Qty"].reindex(calc.index).fillna(0)

# ── Separate unlisted shares — excluded from P&L totals ──────────────────
unlisted_df = calc[calc["Asset_Type"] == "Unlisted Share"].copy()
# Zero out all P&L columns for unlisted so they can never pollute totals,
# regardless of what CMP was fetched or what Buy_Price is stored.
# CMP (Current_Price) and Value are kept for display-only purposes.
for _ul_col in ["Cost_Basis", "Unrealized_PnL", "PnL_%", "Daily_PnL", "Daily_PnL_%"]:
    if _ul_col in unlisted_df.columns:
        unlisted_df[_ul_col] = 0.0
# For total P&L calcs, exclude unlisted entirely
calc_listed = calc[calc["Asset_Type"] != "Unlisted Share"].copy()

# Split by type — all instrument categories
stocks_df = calc_listed[calc_listed["Asset_Type"] == "Stock"].copy()
etf_df    = calc_listed[calc_listed["Asset_Type"].isin([
    "ETF", "Liquid ETF", "Commodity ETF", "International ETF"
])].copy()
fno_df    = calc_listed[calc_listed["Asset_Type"].isin([
    "F&O", "Index F&O", "Currency F&O"
])].copy()
reit_df   = calc_listed[calc_listed["Asset_Type"] == "REIT/InvIT"].copy()
other_df  = calc_listed[calc_listed["Asset_Type"].isin([
    "SGB", "Bond/NCD", "Mutual Fund", "Preference Share"
])].copy()

# =====================================================
# SUMMARY METRICS
# =====================================================

total_invested  = calc_listed["Cost_Basis"].sum()
total_value     = calc_listed["Value"].sum()
total_pnl       = calc_listed["Unrealized_PnL"].sum()
total_pnl_pct   = (total_pnl / total_invested * 100) if total_invested else 0

# Unlisted: separate invested & value (not added to main totals)
unlisted_invested = unlisted_df["Cost_Basis"].sum() if not unlisted_df.empty else 0.0
unlisted_value    = unlisted_df["Value"].sum()       if not unlisted_df.empty else 0.0

# Invested value broken down by asset type
stocks_invested = stocks_df["Cost_Basis"].sum() if not stocks_df.empty else 0.0
etf_invested    = etf_df["Cost_Basis"].sum()    if not etf_df.empty    else 0.0
fno_invested    = fno_df["Cost_Basis"].sum()    if not fno_df.empty    else 0.0
reit_invested   = reit_df["Cost_Basis"].sum()   if not reit_df.empty   else 0.0
other_invested  = other_df["Cost_Basis"].sum()  if not other_df.empty  else 0.0

# Real Daily P&L from prev close (listed only)
daily_pnl       = calc_listed["Daily_PnL"].sum() if calc_listed["Daily_PnL"].notna().any() else 0.0
daily_invested  = (calc_listed["Prev_Close"] * calc_listed["Shares"]).sum()
daily_pnl_pct   = (daily_pnl / daily_invested * 100) if daily_invested and daily_invested > 0 else 0.0

# Booked P&L from sells
total_booked_pnl = sum(booked_pnl_map.values()) if booked_pnl_map else 0.0

# ── New metrics ───────────────────────────────────────────────────
portfolio_beta, beta_df = compute_portfolio_beta(calc_listed)
xirr_map = compute_xirr_per_holding(calc_listed, trades_df)
alerts   = compute_alerts(calc_listed, total_value)
cap_gains_df = compute_capital_gains(trades_df, df)

# Build summary dict (used by PDF / Excel / WhatsApp)
summary_dict = {
    "total_invested":   total_invested,
    "stocks_invested":  stocks_invested,
    "etf_invested":     etf_invested,
    "fno_invested":     fno_invested,
    "reit_invested":    reit_invested,
    "other_invested":   other_invested,
    "total_value":      total_value,
    "total_pnl":        total_pnl,
    "total_pnl_pct":    total_pnl_pct,
    "daily_pnl":        daily_pnl,
    "daily_pnl_pct":    daily_pnl_pct,
    "total_booked_pnl": total_booked_pnl,
    "portfolio_beta":   portfolio_beta,
    "xirr_map":         xirr_map,
    "cap_gains_df":     cap_gains_df,
    "unlisted_invested": unlisted_invested,
    "unlisted_value":    unlisted_value,
}

def fmt_inr(val):
    return f"₹{val:,.2f}" if pd.notna(val) else "N/A"

def pnl_cls(val):
    return "positive" if val >= 0 else "negative"

# =========================================================
# SHARED STYLE HELPERS — available to ALL nav tabs
# =========================================================

def style_pnl(val):
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ""
        color = "#22d67b" if float(val) >= 0 else "#f85454"
        return f"color: {color}; font-weight: 600"
    except Exception:
        return ""

def safe_inr(x):
    try:
        return f"\u20b9{x:,.2f}" if x is not None and not pd.isna(x) else "\u2014"
    except Exception:
        return "\u2014"

def safe_pct(x):
    try:
        return f"{x:+.2f}%" if x is not None and not pd.isna(x) else "\u2014"
    except Exception:
        return "\u2014"

# =========================================================
# NAV-GATED RENDERING — only show the active screen
# =========================================================

if _nav_tab == "Portfolio":

    _total_pnl_col   = "#22d67b" if total_pnl  >= 0 else "#f85454"
    _daily_pnl_col   = "#22d67b" if daily_pnl  >= 0 else "#f85454"
    _booked_pnl_col  = "#22d67b" if total_booked_pnl >= 0 else "#f85454"
    _total_pnl_arrow = "▲" if total_pnl  > 0 else ("▼" if total_pnl  < 0 else "")
    _daily_pnl_arrow = "▲" if daily_pnl  > 0 else ("▼" if daily_pnl  < 0 else "")

    st.markdown(f"""
<style>
/* ── MOBILE summary header (Upstox-style) — hidden everywhere ── */
.mob-portfolio-header {{
    display: none !important;
}}
/* ── Mini bar inside tabs — hidden on mobile ── */
.mini-bar-wrap {{
    display: flex; gap:10px; flex-wrap:wrap; margin-bottom:14px;
}}
@media (max-width: 640px) {{
    .mini-bar-wrap {{ display: none !important; }}
    .desktop-summary-bar {{ display: block !important; }}
    .summary-bar {{ grid-template-columns: 1fr 1fr !important; gap: 8px !important; }}
    .summary-value {{ font-size: 11px !important; white-space: normal !important; }}
    .summary-label {{ font-size: 9px !important; }}
    .summary-item {{ padding: 10px 6px !important; }}
}}

/* ── DESKTOP grid ── */
.summary-bar {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
    gap: 12px;
    padding: 0;
    margin-bottom: 24px;
}}
.summary-item {{
    background: linear-gradient(145deg, #131628, #0f1120);
    border: 1px solid #252849;
    border-radius: 16px;
    padding: 12px 8px;
    text-align: center;
    transition: all 0.2s ease;
    position: relative;
    overflow: visible;
    min-width: 0;
}}
.summary-item::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #4f7ef8, #7c4dff);
    opacity: 0;
    transition: opacity 0.2s;
}}
.summary-item:hover {{ border-color: #2e3355; transform: translateY(-1px); }}
.summary-item:hover::before {{ opacity: 1; }}
.summary-label {{
    font-size: 9px;
    color: #6b7299;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-bottom: 8px;
    font-weight: 700;
}}
.summary-value {{
    font-size: 13px;
    font-weight: 800;
    color: #f0f2ff;
    line-height: 1.2;
    white-space: nowrap;
}}
.summary-value.positive {{ color: #22d67b; }}
.summary-value.negative {{ color: #f85454; }}
</style>

<!-- ════ MOBILE HEADER (Upstox-style) ════ -->
<div class="mob-portfolio-header">
  <!-- Invested / Current row -->
  <div style="display:flex; justify-content:space-between; margin-bottom:14px;">
    <div>
      <div style="font-size:11px; color:#5a5f88; font-weight:600; margin-bottom:3px;">Invested</div>
      <div style="font-size:22px; font-weight:800; color:#f0f2ff; letter-spacing:-0.5px;">{fmt_inr(total_invested)}</div>
    </div>
    <div style="text-align:right;">
      <div style="font-size:11px; color:#5a5f88; font-weight:600; margin-bottom:3px;">Current</div>
      <div style="font-size:22px; font-weight:800; color:#f0f2ff; letter-spacing:-0.5px;">{fmt_inr(total_value)}</div>
    </div>
  </div>
  <!-- Total return / Today's return row -->
  <div style="display:flex; justify-content:space-between; border-top:1px solid #1e2240; padding-top:12px;">
    <div>
      <div style="font-size:10px; color:#5a5f88; font-weight:600; margin-bottom:3px;">Total return</div>
      <div style="font-size:15px; font-weight:800; color:{_total_pnl_col};">
        {_total_pnl_arrow} {fmt_inr(total_pnl)}
        <span style="font-size:12px; font-weight:700;">({total_pnl_pct:+.2f}%)</span>
      </div>
    </div>
    <div style="text-align:right;">
      <div style="font-size:10px; color:#5a5f88; font-weight:600; margin-bottom:3px;">Today's return</div>
      <div style="font-size:15px; font-weight:800; color:{_daily_pnl_col};">
        {_daily_pnl_arrow} {fmt_inr(daily_pnl)}
        <span style="font-size:12px; font-weight:700;">({daily_pnl_pct:+.2f}%)</span>
      </div>
    </div>
  </div>
  <!-- Extra row: Booked P&L + Beta -->
  <div style="display:flex; gap:10px; margin-top:12px; border-top:1px solid #1e2240; padding-top:12px;">
    <div style="flex:1; background:#0b0d1e; border-radius:10px; padding:10px 12px; text-align:center;">
      <div style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700; margin-bottom:4px;">✅ Booked P&L</div>
      <div style="font-size:13px; font-weight:800; color:{_booked_pnl_col};">{fmt_inr(total_booked_pnl)}</div>
    </div>
    <div style="flex:1; background:#0b0d1e; border-radius:10px; padding:10px 12px; text-align:center;">
      <div style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700; margin-bottom:4px;">📐 Beta</div>
      <div style="font-size:13px; font-weight:800; color:#f5c842;">{portfolio_beta:.3f}</div>
    </div>
  </div>
</div>

<!-- ════ DESKTOP GRID ════ -->
<div class="desktop-summary-bar">
<div class="summary-bar">
  <div class="summary-item" style="border-color:#4f7ef844; background:linear-gradient(145deg,#111a2e,#0d1220);">
    <div class="summary-label">💰 Total Invested</div>
    <div class="summary-value" style="color:#4f7ef8;">{fmt_inr(total_invested)}</div>
  </div>
  <div class="summary-item" style="border-color:#4f7ef844; background:linear-gradient(145deg,#111a2e,#0d1220);">
    <div class="summary-label">📦 Portfolio Value</div>
    <div class="summary-value" style="color:#4f7ef8;">{fmt_inr(total_value)}</div>
  </div>
  <div class="summary-item" style="border-color:{'#22d67b' if daily_pnl >= 0 else '#f85454'}44; background:{'linear-gradient(145deg,#0f1a14,#0c1410)' if daily_pnl >= 0 else 'linear-gradient(145deg,#1a0f0f,#1a0c0c)'};">
    <div class="summary-label">📅 Today's P&amp;L</div>
    <div class="summary-value {pnl_cls(daily_pnl)}">{fmt_inr(daily_pnl)}<br><span style="font-size:12px;font-weight:700;">({daily_pnl_pct:+.2f}%)</span></div>
  </div>
  <div class="summary-item" style="border-color:{'#22d67b' if total_pnl >= 0 else '#f85454'}44; background:{'linear-gradient(145deg,#0f1a14,#0c1410)' if total_pnl >= 0 else 'linear-gradient(145deg,#1a0f0f,#1a0c0c)'};">
    <div class="summary-label">📈 Unrealized P&amp;L</div>
    <div class="summary-value {pnl_cls(total_pnl)}">{fmt_inr(total_pnl)}<br><span style="font-size:12px;font-weight:700;">({total_pnl_pct:+.2f}%)</span></div>
  </div>
  <div class="summary-item" style="border-color:{'#22d67b' if total_booked_pnl >= 0 else '#f85454'}44;">
    <div class="summary-label">✅ Booked P&amp;L</div>
    <div class="summary-value {pnl_cls(total_booked_pnl)}">{fmt_inr(total_booked_pnl)}</div>
  </div>
  {f'''<div class="summary-item" style="border-color:#a78bfa44; background:linear-gradient(145deg,#170f22,#12091c);">
    <div class="summary-label">🔒 Unlisted (excl.)</div>
    <div class="summary-value" style="color:#a78bfa;">{fmt_inr(unlisted_invested)}</div>
    <div style="font-size:9px;color:#6b4fa8;margin-top:3px;">P&amp;L not included</div>
  </div>''' if unlisted_invested > 0 else ''}
  <div class="summary-item" style="border-color:#f5c84244; background:linear-gradient(145deg,#1a1505,#12100a);">
    <div class="summary-label">📐 Portfolio Beta</div>
    <div class="summary-value" style="color:#f5c842;">{portfolio_beta:.3f}</div>
  </div>
</div>
</div>
""", unsafe_allow_html=True)

    # ── MARKET INTELLIGENCE TICKER BANNER ──────────────────────────────
    # Pulls ONLY from already-cached dashboard data:
    #   1. get_filtered_news()      → portfolio news items
    #   2. get_results_calendar()   → portfolio corporate actions / ex-dates
    # Zero extra HTTP calls — instant, always in sync with the dashboard.
    # ─────────────────────────────────────────────────────────────────────

    _portfolio_tickers_for_intel = tuple(
        t for t in (list(df["Ticker"].unique()) if not df.empty else [])
        if not any(x in t.upper() for x in ["FUT","CE","PE"])
    )

    # ── Live Intel: always fetch independently from the internet ──────────────
    # This ensures the ticker and expanded view have data even if the Portfolio News
    # or Corporate Actions tabs have never been visited in the current session.
    # fetch_all_news / get_results_calendar are @st.cache_data so costs are free
    # after the first call; any tab that also calls them just hits the cache.
    if _portfolio_tickers_for_intel:
        if "_ticker_news" not in st.session_state:
            try:
                st.session_state["_ticker_news"] = get_filtered_news(_portfolio_tickers_for_intel)
            except Exception:
                st.session_state["_ticker_news"] = []
        if "_ticker_cal_df" not in st.session_state:
            try:
                st.session_state["_ticker_cal_df"] = get_results_calendar(_portfolio_tickers_for_intel)
            except Exception:
                st.session_state["_ticker_cal_df"] = pd.DataFrame()
        # Force re-fetch if session keys exist but are empty (e.g. first load returned nothing)
        if not st.session_state.get("_ticker_news") and not st.session_state.get("_intel_fetched"):
            try:
                _fresh = get_filtered_news(_portfolio_tickers_for_intel)
                if _fresh:
                    st.session_state["_ticker_news"] = _fresh
                    st.session_state["_intel_fetched"] = True
            except Exception:
                pass
        if (st.session_state.get("_ticker_cal_df") is None or
                (hasattr(st.session_state.get("_ticker_cal_df"), "empty") and
                 st.session_state["_ticker_cal_df"].empty)) and \
                not st.session_state.get("_intel_cal_fetched"):
            try:
                _fresh_cal = get_results_calendar(_portfolio_tickers_for_intel)
                if not _fresh_cal.empty:
                    st.session_state["_ticker_cal_df"] = _fresh_cal
                    st.session_state["_intel_cal_fetched"] = True
            except Exception:
                pass

    def _build_ticker_items_from_session(raw_news, cal_df):
        """
        Build ticker items from data already fetched by tabs (stored in session_state).
        raw_news  = list from get_filtered_news(), stored by tab4
        cal_df    = DataFrame from get_results_calendar(), stored by tab5
        Returns list of dicts: {ticker, headline, category, cat_color, age_str, url}
        """
        items = []
        seen_headlines = set()
        _now_utc = datetime.now(timezone.utc)
        IST_OFFSET = timedelta(hours=5, minutes=30)

        def _age_str(pub_dt):
            if pub_dt is None:
                return "recently"
            try:
                if pub_dt.tzinfo is None:
                    pub_dt = pub_dt.replace(tzinfo=timezone.utc)
                delta = _now_utc - pub_dt
                if delta.total_seconds() < 0:
                    d = abs(delta.days)
                    return "today" if d == 0 else ("tomorrow" if d == 1 else f"in {d}d")
                if delta.days == 0:
                    h = delta.seconds // 3600
                    return f"{h}h ago" if h > 0 else "just now"
                return f"{delta.days}d ago"
            except Exception:
                return "recently"

        # ── 1. HIGH-priority portfolio news (from session_state) ──────────
        try:
            for n in (raw_news or []):
                prio = n.get("priority", "")
                headline = n.get("title") or n.get("headline") or ""
                if not headline:
                    continue
                hkey = headline[:80].lower()
                if hkey in seen_headlines:
                    continue
                # Only HIGH priority in ticker; normal news stays in tab
                if prio != "🔴 HIGH":
                    continue
                seen_headlines.add(hkey)
                ticker_sym = (n.get("stock_tag") or n.get("ticker") or "").upper().replace(".NS","").replace(".BO","")
                cat  = n.get("category") or "📌 Corporate Event"
                col  = n.get("cat_color") or "#60a5fa"
                pub_dt = n.get("published_dt")
                items.append({
                    "ticker":    ticker_sym,
                    "headline":  headline[:100],
                    "category":  cat,
                    "cat_color": col,
                    "age_str":   _age_str(pub_dt),
                    "url":       n.get("url") or n.get("link") or "",
                    "pub_dt":    pub_dt,
                    "_sort_key": (pub_dt or datetime.min.replace(tzinfo=timezone.utc)),
                })
        except Exception:
            pass

        # ── 2. Corporate actions / ex-dates (from session_state) ──────────
        try:
            cal = cal_df if cal_df is not None else pd.DataFrame()
            if not cal.empty:
                today_ts = pd.Timestamp.now().normalize()
                for _, row in cal.iterrows():
                    sym     = str(row.get("Ticker","") or row.get("Symbol","")).replace(".NS","").replace(".BO","").upper()
                    event   = str(row.get("Event","") or row.get("Action","") or row.get("Subject","")).strip()
                    exd_raw = row.get("Ex-Date","")
                    days_aw = str(row.get("Days Away","")).strip()
                    prio    = str(row.get("Priority","")).strip()

                    if not sym or not event:
                        continue

                    # Only put in ticker if within 7 days or HIGH priority
                    is_urgent = (
                        days_aw in ("Today", "Tomorrow") or
                        prio == "🔴 HIGH" or
                        any(d in days_aw for d in ["1d","2d","3d","4d","5d","6d","7d"])
                    )
                    # Also check ex-date directly
                    try:
                        if exd_raw and exd_raw != "—":
                            ex_ts = pd.to_datetime(str(exd_raw), dayfirst=True)
                            if 0 <= (ex_ts - today_ts).days <= 7:
                                is_urgent = True
                    except Exception:
                        pass

                    if not is_urgent:
                        continue

                    # Build a short headline
                    if exd_raw and exd_raw != "—":
                        hl = f"{event} · Ex-Date: {exd_raw}"
                        if days_aw == "Today":
                            hl = f"⚡ Ex-Date TODAY: {event}"
                        elif days_aw == "Tomorrow":
                            hl = f"⚡ Ex-Date Tomorrow: {event}"
                    else:
                        hl = f"{event} · {days_aw}"

                    hkey = f"{sym}|{hl[:60].lower()}"
                    if hkey in seen_headlines:
                        continue
                    seen_headlines.add(hkey)

                    # Choose color/category by event type
                    el = event.lower()
                    if "dividend" in el:
                        cat, col = "💰 Dividend", "#90caf9"
                    elif "bonus" in el:
                        cat, col = "🎁 Bonus", "#22d67b"
                    elif "split" in el:
                        cat, col = "✂️ Split", "#69f0ae"
                    elif "rights" in el:
                        cat, col = "📜 Rights", "#ffcc02"
                    elif "buyback" in el or "buy back" in el:
                        cat, col = "🔁 Buyback", "#ff9800"
                    elif "result" in el or "earnings" in el:
                        cat, col = "📊 Results", "#fb923c"
                    elif "agm" in el or "egm" in el:
                        cat, col = "🏛 AGM/EGM", "#a78bfa"
                    else:
                        cat, col = "📌 Corp Action", "#60a5fa"

                    # Build a pub_dt for sorting (use ex-date if available, else today)
                    try:
                        _pdt = pd.to_datetime(str(exd_raw), dayfirst=True).to_pydatetime().replace(tzinfo=timezone.utc) if exd_raw and exd_raw != "—" else _now_utc
                    except Exception:
                        _pdt = _now_utc

                    items.append({
                        "ticker":    sym,
                        "headline":  hl[:100],
                        "category":  cat,
                        "cat_color": col,
                        "age_str":   days_aw if days_aw else "upcoming",
                        "url":       "",
                        "pub_dt":    _pdt,
                        "_sort_key": _pdt,
                    })
        except Exception:
            pass

        # Sort: today's ex-dates first, then by recency
        items.sort(key=lambda x: x["_sort_key"], reverse=True)
        return items


    # ── Ticker reads from session_state populated by tabs ─────────────
    # Tab4 stores: st.session_state["_ticker_news"]  = list of news items
    # Tab5 stores: st.session_state["_ticker_cal_df"] = cal_df DataFrame
    # On first load session_state is empty → show placeholder;
    # after first tab render the ticker fills in automatically on rerun.
    if _portfolio_tickers_for_intel:
        _ss_news   = st.session_state.get("_ticker_news", [])
        _ss_cal_df = st.session_state.get("_ticker_cal_df", pd.DataFrame())
        _intel_items = _build_ticker_items_from_session(_ss_news, _ss_cal_df)

        if _intel_items:
            _now_utc = datetime.now(timezone.utc)
            _ticker_parts = []
            for item in _intel_items:
                _is_upcoming = item.get("pub_dt") and item["pub_dt"] > _now_utc
                _age_color = "#fbbf24" if _is_upcoming else "#555577"
                _age_prefix = "🔔 " if _is_upcoming else ""
                _ticker_parts.append(
                    f'<span style="color:{item["cat_color"]};font-weight:700;margin-right:4px;">'
                    f'{item["category"]}</span>'
                    f'<span style="color:#c8c8e8;font-weight:600;">{item["ticker"]}</span>'
                    f'<span style="color:#8888aa;margin:0 4px;">·</span>'
                    f'<span style="color:#b8bcd8;">{item["headline"]}</span>'
                    f'<span style="color:{_age_color};margin:0 6px;font-size:10px;">'
                    f'({_age_prefix}{item["age_str"]})</span>'
                    f'<span style="color:#2a2a5a;margin:0 16px;">｜</span>'
                )
            _ticker_text = "".join(_ticker_parts)
            if len(_intel_items) < 4:
                _ticker_text = _ticker_text * 4
            else:
                _ticker_text = _ticker_text * 2

            _anim_dur = max(26, len(_intel_items) * 6.5)

            st.markdown(f"""
    <style>
    @keyframes scroll-ticker {{
        0%   {{ transform: translateX(0); }}
        100% {{ transform: translateX(-50%); }}
    }}
    .intel-ticker-wrap {{
        background: linear-gradient(90deg, #0a0a1e 0%, #0d0d28 100%);
        border: 1px solid #2a2a5a;
        border-radius: 10px;
        padding: 0;
        overflow: hidden;
        display: flex;
        align-items: stretch;
        margin-bottom: 12px;
        height: 38px;
    }}
    .intel-ticker-label {{
        flex-shrink: 0;
        background: linear-gradient(135deg, #7c3aed, #4f46e5);
        color: #fff;
        font-size: 11px;
        font-weight: 800;
        letter-spacing: 0.8px;
        padding: 0 14px;
        display: flex;
        align-items: center;
        white-space: nowrap;
        border-radius: 9px 0 0 9px;
        text-transform: uppercase;
    }}
    .intel-ticker-scroll {{
        flex: 1;
        overflow: hidden;
        display: flex;
        align-items: center;
        position: relative;
    }}
    .intel-ticker-inner {{
        display: flex;
        align-items: center;
        white-space: nowrap;
        font-size: 12.5px;
        line-height: 1;
        animation: scroll-ticker {_anim_dur}s linear infinite;
        padding-left: 24px;
    }}
    .intel-ticker-inner:hover {{
        animation-play-state: paused;
    }}
    </style>
    <div class="intel-ticker-wrap">
      <div class="intel-ticker-label">📡 Live Intel</div>
      <div class="intel-ticker-scroll">
        <div class="intel-ticker-inner">{_ticker_text}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

            # ── Expandable full list below the ticker ────────────────────
            # Split into two buckets: HIGH-priority news vs upcoming corp actions
            _high_news_items  = [x for x in _intel_items if x.get("category","") not in
                                 ("💰 Dividend","🎁 Bonus","✂️ Split","📜 Rights","🔁 Buyback",
                                  "📊 Results","🏛 AGM/EGM","📌 Corp Action")]
            _corp_action_items = [x for x in _intel_items if x.get("category","") in
                                  ("💰 Dividend","🎁 Bonus","✂️ Split","📜 Rights","🔁 Buyback",
                                   "📊 Results","🏛 AGM/EGM","📌 Corp Action")]

            _expander_label = (
                f"📋 Live Intel — {len(_high_news_items)} high-impact news"
                + (f" · {len(_corp_action_items)} upcoming events" if _corp_action_items else "")
                + " — click to expand"
            )
            with st.expander(_expander_label):
                # ── Section 1: HIGH-priority market news ─────────────────
                if _high_news_items:
                    st.markdown(
                        '<div style="font-size:13px;font-weight:900;color:#ff5252;'
                        'border-left:4px solid #ff5252;padding:6px 12px;'
                        'background:#ff525211;border-radius:0 8px 8px 0;margin:8px 0 10px;">'
                        '🔴 HIGH-IMPACT NEWS &nbsp;<span style="color:#8888aa;font-weight:400;'
                        f'font-size:10px;">{len(_high_news_items)} item{"s" if len(_high_news_items)>1 else ""}'
                        '</span></div>',
                        unsafe_allow_html=True
                    )
                    for item in _high_news_items:
                        cat_color = item.get("cat_color","#f85454")
                        _link = (f'<a href="{item["url"]}" target="_blank" style="color:{cat_color};'
                                 f'text-decoration:none;font-size:11px;margin-left:6px;">↗ Read</a>') \
                                 if item.get("url") else ""
                        _is_up = item.get("pub_dt") and item["pub_dt"] > _now_utc
                        _age_col = "#fbbf24" if _is_up else "#6b7280"
                        _age_bg  = "#fbbf2422" if _is_up else "#1a1a2a"
                        _prio_badge = (
                            '<span style="background:#ff5252;color:#fff;font-size:9px;font-weight:800;'
                            'border-radius:3px;padding:1px 6px;margin-right:6px;">HIGH</span>'
                        )
                        _cat_pill = (
                            f'<span style="background:{cat_color}22;color:{cat_color};font-size:9px;'
                            f'font-weight:700;border-radius:3px;padding:1px 6px;margin-right:6px;">'
                            f'{item["category"]}</span>'
                        )
                        st.markdown(
                            f'<div style="display:flex;align-items:flex-start;gap:10px;'
                            f'padding:9px 12px;border-bottom:1px solid #1a1a3a;margin-bottom:2px;'
                            f'background:#131628;border-radius:6px;margin-bottom:4px;">'
                            f'<span style="color:#00e676;font-weight:800;font-size:12px;'
                            f'min-width:90px;flex-shrink:0;padding-top:2px;">{item["ticker"]}</span>'
                            f'<div style="flex:1;line-height:1.5;">'
                            f'{_prio_badge}{_cat_pill}'
                            f'<span style="color:#e8e8ff;font-size:12px;">{item["headline"]}</span>'
                            f'{_link}</div>'
                            f'<span style="color:{_age_col};font-size:10px;white-space:nowrap;'
                            f'flex-shrink:0;background:{_age_bg};padding:2px 6px;border-radius:4px;">'
                            f'{item["age_str"]}</span>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                else:
                    st.markdown(
                        '<div style="color:#454870;font-size:11px;padding:8px 12px;">'
                        '🔴 No high-impact news found for your portfolio stocks right now.</div>',
                        unsafe_allow_html=True
                    )

                # ── Section 2: Upcoming corporate actions / results ───────
                if _corp_action_items:
                    st.markdown(
                        '<div style="font-size:13px;font-weight:900;color:#fbbf24;'
                        'border-left:4px solid #fbbf24;padding:6px 12px;'
                        'background:#fbbf2411;border-radius:0 8px 8px 0;margin:18px 0 10px;">'
                        '📅 UPCOMING RESULTS & CORPORATE ACTIONS &nbsp;'
                        f'<span style="color:#8888aa;font-weight:400;font-size:10px;">'
                        f'{len(_corp_action_items)} event{"s" if len(_corp_action_items)>1 else ""}'
                        '</span></div>',
                        unsafe_allow_html=True
                    )
                    # Group by category
                    _ca_groups = {}
                    for item in _corp_action_items:
                        _ca_groups.setdefault(item["category"], []).append(item)

                    for cat, items_in_cat in _ca_groups.items():
                        cat_color = items_in_cat[0]["cat_color"]
                        st.markdown(
                            f'<div style="font-size:11px;font-weight:800;color:{cat_color};'
                            f'padding:4px 10px;background:{cat_color}18;border-radius:4px;'
                            f'margin:8px 0 4px;display:inline-block;">'
                            f'{cat} &nbsp;<span style="color:#8888aa;font-weight:400;font-size:10px;">'
                            f'{len(items_in_cat)}</span></div>',
                            unsafe_allow_html=True
                        )
                        for item in items_in_cat:
                            _is_today   = item.get("age_str","") in ("Today","0h ago","just now")
                            _is_tmrw    = item.get("age_str","") == "Tomorrow"
                            _is_up      = item.get("pub_dt") and item["pub_dt"] > _now_utc
                            _age_col    = "#f85454" if _is_today else ("#ffaa00" if _is_tmrw else
                                          ("#fbbf24" if _is_up else "#6b7280"))
                            _urgency_badge = ""
                            if _is_today:
                                _urgency_badge = (
                                    '<span style="background:#ff5252;color:#fff;font-size:9px;'
                                    'font-weight:800;border-radius:3px;padding:1px 6px;margin-right:6px;">'
                                    '⚡ TODAY</span>'
                                )
                            elif _is_tmrw:
                                _urgency_badge = (
                                    '<span style="background:#ffaa00;color:#0a0a0a;font-size:9px;'
                                    'font-weight:800;border-radius:3px;padding:1px 6px;margin-right:6px;">'
                                    '🔔 TOMORROW</span>'
                                )
                            elif _is_up:
                                _urgency_badge = (
                                    '<span style="background:#fbbf24;color:#0a0a0a;font-size:9px;'
                                    'font-weight:800;border-radius:3px;padding:1px 6px;margin-right:6px;">'
                                    'UPCOMING</span>'
                                )
                            st.markdown(
                                f'<div style="display:flex;align-items:flex-start;gap:10px;'
                                f'padding:8px 12px;border-bottom:1px solid #1a1a3a;'
                                f'background:#131628;border-radius:6px;margin-bottom:4px;">'
                                f'<span style="color:{cat_color};font-weight:800;font-size:12px;'
                                f'min-width:90px;flex-shrink:0;padding-top:2px;">{item["ticker"]}</span>'
                                f'<div style="flex:1;line-height:1.5;">'
                                f'{_urgency_badge}'
                                f'<span style="color:#e8e8ff;font-size:12px;">{item["headline"]}</span>'
                                f'</div>'
                                f'<span style="color:{_age_col};font-size:10px;white-space:nowrap;'
                                f'flex-shrink:0;font-weight:700;padding:2px 6px;border-radius:4px;'
                                f'background:{_age_col}22;">{item["age_str"]}</span>'
                                f'</div>',
                                unsafe_allow_html=True
                            )
                else:
                    st.markdown(
                        '<div style="color:#454870;font-size:11px;padding:8px 12px;margin-top:10px;">'
                        '📅 No upcoming corporate actions or results within 7 days.</div>',
                        unsafe_allow_html=True
                    )
        else:
            st.markdown("""
    <div style="background:#0b0d17;border:1px solid #252849;border-radius:10px;
                padding:8px 18px;margin-bottom:12px;display:flex;align-items:center;gap:10px;">
      <span style="color:#454870;font-size:12px;font-weight:700;">📡 LIVE INTEL</span>
      <span style="color:#454870;font-size:11px;">No high-impact events or upcoming ex-dates for your stocks.
      Results, dividends, splits, bonus and order wins will appear here automatically.</span>
    </div>
    """, unsafe_allow_html=True)

    # =====================================================
    # COLORED TABLE HELPER
    # =====================================================
    # style_pnl / safe_inr / safe_pct are defined at module level above

    def _c(val, fmt_fn=None):
        """Format a numeric value; return (formatted_str, colour)."""
        try:
            v = float(val)
            s = fmt_fn(v) if fmt_fn else f"{v:,.2f}"
            colour = "#22d67b" if v >= 0 else "#f85454"
            return s, colour
        except Exception:
            return "—", "#8888aa"

    def render_section_table(section_df, show_sold=False, section_key=""):
        if section_df.empty:
            st.info("No holdings in this category.")
            return

        # ── EDITABLE top table: Ticker | Asset Type | Qty | Avg Buy ─────
        edit_df = section_df[["Ticker", "Asset_Type", "Shares", "Buy_Price"]].copy()
        edit_df["Shares"]    = pd.to_numeric(edit_df["Shares"],    errors="coerce")
        edit_df["Buy_Price"] = pd.to_numeric(edit_df["Buy_Price"], errors="coerce")

        col_config_edit = {
            "Ticker":     st.column_config.TextColumn("Ticker",           disabled=True),
            "Asset_Type": st.column_config.TextColumn("Asset Type",       disabled=True),
            "Shares":     st.column_config.NumberColumn("Qty / Lots",     min_value=0.0001, step=1.0,  format="%.2f"),
            "Buy_Price":  st.column_config.NumberColumn("Wtd. Avg Price ₹", min_value=0.0001, step=0.01, format="%.2f"),
        }

        if _is_dev:
            st.markdown('<div class="collapsible-section">', unsafe_allow_html=True)
            with st.expander("✏️  Edit Holdings — Qty & Weighted Avg Price", expanded=False):
                st.caption("Click any **Qty / Lots** or **Wtd. Avg Price ₹** cell to edit · Use the ✕ row button to delete a holding · Changes save automatically")
                _orig_tickers = list(edit_df["Ticker"])
                edited = st.data_editor(
                    edit_df,
                    column_config=col_config_edit,
                    use_container_width=True,
                    hide_index=True,
                    key=f"editor_{section_key}",
                    num_rows="dynamic",
                )

                # ── Detect deleted rows ───────────────────────────────────
                _edited_tickers = list(edited["Ticker"]) if "Ticker" in edited.columns else []
                _deleted_tickers = [t for t in _orig_tickers if t not in _edited_tickers]
                if _deleted_tickers:
                    _mask = df["Ticker"].isin(_deleted_tickers)
                    df.drop(df[_mask].index, inplace=True)
                    df.reset_index(drop=True, inplace=True)
                    df.to_csv(PORTFOLIO_FILE, index=False)
                    _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                    st.success(f"🗑️ Deleted: {', '.join(_deleted_tickers)}. Refreshing...")
                    st.rerun()

                # ── Detect & save edits to Qty / Avg Buy ─────────────────
                changed_rows = []
                for i, (orig_idx, orig_row) in enumerate(section_df.iterrows()):
                    if i >= len(edited):
                        break
                    new_shares    = edited.iloc[i]["Shares"]
                    new_buy_price = edited.iloc[i]["Buy_Price"]
                    if (abs(float(new_shares)    - float(orig_row["Shares"]))    > 0.0001 or
                        abs(float(new_buy_price) - float(orig_row["Buy_Price"])) > 0.0001):
                        changed_rows.append((orig_idx, new_shares, new_buy_price))

                if changed_rows:
                    for orig_idx, new_shares, new_buy_price in changed_rows:
                        df.at[orig_idx, "Shares"]    = round(float(new_shares),    4)
                        df.at[orig_idx, "Buy_Price"] = round(float(new_buy_price), 4)
                    df.to_csv(PORTFOLIO_FILE, index=False)
                    _local_sync_to_gh(PORTFOLIO_FILE, PORTFOLIO_FILE, 'Auto-save: portfolio')
                    st.success(f"✅ {len(changed_rows)} holding(s) updated. Refreshing...")
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        # ── COLOURED HTML table — all P&L columns on same line ───────────
        extra_cols = ["Sold_Qty", "Booked_PnL"] if show_sold else []

        # Prepare booked P&L lookup
        booked_lookup = booked_pnl_map if show_sold else {}

        # Build header
        extra_headers = ""
        if show_sold:
            extra_headers = "<th style='padding:10px 14px; text-align:right; white-space:nowrap;'>Sold Qty</th><th style='padding:10px 14px; text-align:right; white-space:nowrap;'>Booked P&amp;L</th>"

        header = f"""
        <tr style="background:linear-gradient(90deg,#181b2e 0%,#1a1d33 100%); color:#6b7299; font-size:10px; text-transform:uppercase; letter-spacing:0.9px; font-weight:700;">
          <th style="padding:12px 16px; text-align:left; border-bottom:2px solid #2a2e52; white-space:nowrap;">Holding <span style="font-size:9px; color:#5a5f88; font-weight:400;">(click row to expand lots)</span></th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Qty / Lots</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Wtd. Avg Buy ₹</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">CMP ₹</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Cost Basis</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Curr. Value</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Unreal P&amp;L</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">P&amp;L %</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Today P&amp;L</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">Day %</th>
          <th style="padding:12px 14px; text-align:right; border-bottom:2px solid #2a2e52; white-space:nowrap;">XIRR %</th>
          {extra_headers}
        </tr>"""

        # ── Helpers (defined once, used for both desktop rows and mobile cards) ──
        def fmt_inr_v(v):
            try: return f"₹{float(v):,.2f}"
            except: return "—"
        def fmt_pct_v(v):
            try: return f"{float(v):+.2f}%"
            except: return "—"
        def fmt_num_v(v):
            try: return f"{float(v):,.2f}"
            except: return "—"

        def coloured_cell(val, fmt_fn, padding="10px 14px"):
            try:
                fv = float(val)
                if fv > 0:
                    col = "#22d67b"
                    bg_cell = "rgba(34,214,123,0.06)"
                elif fv < 0:
                    col = "#f85454"
                    bg_cell = "rgba(248,84,84,0.06)"
                else:
                    col = "#7a7fa8"
                    bg_cell = "transparent"
                arrow = "▲" if fv > 0 else ("▼" if fv < 0 else "")
                return (f'<td style="padding:{padding}; text-align:right; color:{col}; font-weight:700;'
                        f' border-bottom:1px solid #1a1e38; background:{bg_cell}; white-space:nowrap;">'
                        f'{arrow} {fmt_fn(fv)}</td>')
            except:
                return f'<td style="padding:{padding}; text-align:right; color:#3a3d5c; border-bottom:1px solid #1a1e38;">—</td>'

        def plain_cell(val, fmt_fn, padding="10px 14px"):
            try:
                return (f'<td style="padding:{padding}; text-align:right; color:#c8cce8; font-weight:500;'
                        f' border-bottom:1px solid #1a1e38; white-space:nowrap;">{fmt_fn(float(val))}</td>')
            except:
                return f'<td style="padding:{padding}; text-align:right; color:#3a3d5c; border-bottom:1px solid #1a1e38;">—</td>'

        # ── Build a lookup: ticker → true weighted avg buy from ORIGINAL df ──
        # Must use df (pre-FIFO) not section_df (post-FIFO) so that selling
        # shares never distorts the displayed weighted average buy price.
        # Example: INTERARCH 4 lots — if the first 2 lots are fully sold and
        # dropped from section_df by compute_net_holdings, the avg would only
        # reflect the 2 remaining lots' prices (wrong).
        # Using original df ensures ALL buy lots are weighted correctly.
        _orig_wtd_avg = {}
        try:
            for _t, _g in df.groupby("Ticker"):
                _total_orig = _g["Shares"].abs().sum()
                if _total_orig > 0:
                    _orig_wtd_avg[_t] = round(
                        (_g["Shares"].abs() * _g["Buy_Price"]).sum() / _total_orig, 4
                    )
        except Exception:
            pass

        # ── GROUP section_df by Ticker for weighted-average display ──────
        # Build a list of (grouped_row_dict, [lot_rows]) per unique ticker
        ticker_groups = []
        for ticker, grp in section_df.groupby("Ticker", sort=False):
            total_qty    = grp["Shares"].sum()
            _abs_qty     = abs(total_qty)
            # Use the original pre-sell weighted avg buy (includes fully-sold lots)
            # Fallback to current grp calc if ticker not found (e.g. short positions)
            if ticker in _orig_wtd_avg:
                wtd_avg_buy = _orig_wtd_avg[ticker]
            else:
                total_cost  = (grp["Shares"].abs() * grp["Buy_Price"]).sum()
                wtd_avg_buy = total_cost / _abs_qty if _abs_qty > 0 else 0.0

            # Aggregated P&L values (sum across lots)
            total_value   = grp["Value"].sum()       if "Value"         in grp.columns else 0.0
            _fallback_cost = (grp["Shares"].abs() * grp["Buy_Price"]).sum()
            total_cb      = grp["Cost_Basis"].sum()  if "Cost_Basis"    in grp.columns else _fallback_cost
            total_unreal  = grp["Unrealized_PnL"].sum() if "Unrealized_PnL" in grp.columns else 0.0
            pnl_pct_g     = (total_unreal / total_cb * 100) if total_cb != 0 else 0.0
            total_daily   = grp["Daily_PnL"].sum()   if "Daily_PnL"    in grp.columns else 0.0
            daily_pct_g   = grp["Daily_PnL_%"].mean() if "Daily_PnL_%"  in grp.columns else 0.0
            cmp_g         = grp["Current_Price"].iloc[0] if "Current_Price" in grp.columns else 0.0
            sold_qty_g    = grp["Sold_Qty"].sum()    if "Sold_Qty"     in grp.columns else 0.0
            atype_g       = str(grp["Asset_Type"].iloc[0]) if "Asset_Type" in grp.columns else "Stock"

            # Individual lots for the expandable section
            lots = []
            for _, lot_row in grp.iterrows():
                lots.append({
                    "buy_date":  str(lot_row.get("Buy_Date", "—")) if "Buy_Date" in grp.columns else "—",
                    "qty":       float(lot_row.get("Shares", 0)),
                    "buy_price": float(lot_row.get("Buy_Price", 0)),
                    "cost":      float(lot_row.get("Cost_Basis", 0)) if "Cost_Basis" in grp.columns else float(lot_row.get("Shares",0)) * float(lot_row.get("Buy_Price",0)),
                })

            # Detect if this group is an open short position
            _grp_is_short = False
            if "Is_Short" in grp.columns:
                _grp_is_short = bool(grp["Is_Short"].astype(str).str.lower().isin(["true","1","yes"]).any())
            elif total_qty < 0:
                _grp_is_short = True

            ticker_groups.append({
                "ticker":     ticker,
                "atype":      atype_g,
                "is_short":   _grp_is_short,
                "total_qty":  abs(total_qty),
                "wtd_avg":    wtd_avg_buy,
                "cmp":        cmp_g,
                "cost_basis": total_cb,
                "value":      total_value,
                "unreal":     total_unreal,
                "pnl_pct":    pnl_pct_g,
                "daily":      total_daily,
                "daily_pct":  daily_pct_g,
                "sold_qty":   sold_qty_g,
                "lots":       lots,
                "multi_lot":  len(grp) > 1,
            })

        # ── DESKTOP: build rows_html ────────────────────────────────────
        rows_html = ""
        for i, g in enumerate(ticker_groups):
            bg        = "#111428" if i % 2 == 0 else "#0d1020"
            hover_bg  = "#1a1e38"
            ticker    = g["ticker"]
            atype     = g["atype"]

            if atype == "Stock":
                badge_col, badge_bg = "#4f7ef8", "rgba(79,126,248,0.15)"
                badge_lbl = "STK"
            elif atype == "ETF":
                badge_col, badge_bg = "#a78bfa", "rgba(167,139,250,0.15)"
                badge_lbl = "ETF"
            elif atype == "F&O":
                badge_col, badge_bg = "#f5c842", "rgba(245,200,66,0.15)"
                badge_lbl = "F&O"
            else:
                badge_col, badge_bg = "#7a7fa8", "rgba(122,127,168,0.15)"
                badge_lbl = atype[:3]

            dot_col = "#22d67b" if g["pnl_pct"] > 0 else ("#f85454" if g["pnl_pct"] < 0 else "#7a7fa8")

            # XIRR cell
            xi = xirr_map.get(ticker)
            if xi is not None:
                try:
                    xi_f = float(xi)
                    xi_col = "#22d67b" if xi_f >= 0 else "#f85454"
                    xi_bg  = "rgba(34,214,123,0.08)" if xi_f >= 0 else "rgba(248,84,84,0.08)"
                    xi_arrow = "▲" if xi_f > 0 else ("▼" if xi_f < 0 else "")
                    xirr_cell = (f'<td style="padding:10px 14px; text-align:right; color:{xi_col}; font-weight:800;'
                                 f' border-bottom:1px solid #1a1e38; background:{xi_bg}; white-space:nowrap;'
                                 f' font-size:12px;">{xi_arrow} {xi_f:+.2f}%</td>')
                except:
                    xirr_cell = '<td style="padding:10px 14px; text-align:right; color:#3a3d5c; border-bottom:1px solid #1a1e38;">—</td>'
            else:
                xirr_cell = '<td style="padding:10px 14px; text-align:right; color:#3a3d5c; border-bottom:1px solid #1a1e38; font-size:11px;">N/A</td>'

            unreal_td   = coloured_cell(g["unreal"],    lambda v: f"₹{v:,.2f}")
            pnl_pct_td  = coloured_cell(g["pnl_pct"],   lambda v: f"{v:+.2f}%")
            daily_td    = coloured_cell(g["daily"],      lambda v: f"₹{v:,.2f}")
            daily_pct_td= coloured_cell(g["daily_pct"], lambda v: f"{v:+.2f}%")

            extra_cells = ""
            if show_sold:
                booked = booked_lookup.get(ticker, 0)
                extra_cells  = plain_cell(g["sold_qty"], lambda v: f"{v:,.2f}")
                extra_cells += coloured_cell(booked, lambda v: f"₹{v:,.2f}")

            # Multi-lot badge next to ticker
            multi_badge = ""
            if g["multi_lot"]:
                multi_badge = (f'<span style="background:rgba(245,200,66,0.15); color:#f5c842; font-size:9px;'
                               f' font-weight:800; padding:1px 6px; border-radius:3px; margin-left:5px;'
                               f' letter-spacing:0.4px;">{len(g["lots"])} lots</span>')
            # Short sell badge
            short_badge = ""
            if g.get("is_short"):
                short_badge = ('<span style="background:rgba(248,84,84,0.15);color:#f85454;font-size:9px;'
                               'font-weight:800;padding:1px 7px;border-radius:3px;margin-left:5px;'
                               'border:1px solid rgba(248,84,84,0.3);letter-spacing:0.5px;">SHORT SELL</span>')

            rows_html += f"""
            <tr style="background:{bg}; transition: background 0.15s;" onmouseover="this.style.background='{hover_bg}'" onmouseout="this.style.background='{bg}'">
              <td style="padding:10px 16px; border-bottom:1px solid #1a1e38; white-space:nowrap;">
                <div style="display:flex; align-items:center; gap:8px;">
                  <span style="width:4px; height:32px; border-radius:2px; background:{dot_col}; display:inline-block; flex-shrink:0;"></span>
                  <div>
                    <div style="color:#f0f2ff; font-weight:800; font-size:13px; letter-spacing:0.3px;">{ticker}{multi_badge}{short_badge}</div>
                    <span style="background:{badge_bg}; color:{badge_col}; font-size:9px; font-weight:800;
                                 padding:1px 6px; border-radius:3px; letter-spacing:0.5px;">{badge_lbl}</span>
                  </div>
                </div>
              </td>
              {plain_cell(g["total_qty"],  lambda v: f"{v:,.2f}")}
              {plain_cell(g["wtd_avg"],    lambda v: f"₹{v:,.2f}")}
              {plain_cell(g["cmp"],        lambda v: f"₹{v:,.2f}")}
              {plain_cell(g["cost_basis"], lambda v: f"₹{v:,.2f}")}
              {plain_cell(g["value"],      lambda v: f"₹{v:,.2f}")}
              {unreal_td}
              {pnl_pct_td}
              {daily_td}
              {daily_pct_td}
              {xirr_cell}
              {extra_cells}
            </tr>"""

            # ── Expandable lots sub-row (only for multi-lot tickers) ─────
            if g["multi_lot"]:
                lot_rows_html = ""
                for lot in g["lots"]:
                    lot_pnl = lot["cost"] - lot["cost"]  # placeholder; cost only shown
                    lot_rows_html += f"""
                    <tr style="background:#0b0e20;">
                      <td style="padding:6px 14px 6px 40px; color:#a0a4c8; font-size:11px; border-bottom:1px solid #13162a; white-space:nowrap;">
                        📅 {lot["buy_date"]}
                      </td>
                      <td style="padding:6px 14px; text-align:right; color:#c8cce8; font-size:11px; border-bottom:1px solid #13162a;">{lot["qty"]:,.2f}</td>
                      <td style="padding:6px 14px; text-align:right; color:#f5c842; font-size:11px; font-weight:700; border-bottom:1px solid #13162a;">₹{lot["buy_price"]:,.2f}</td>
                      <td style="padding:6px 14px; text-align:right; color:#a0a4c8; font-size:11px; border-bottom:1px solid #13162a;">—</td>
                      <td style="padding:6px 14px; text-align:right; color:#c8cce8; font-size:11px; border-bottom:1px solid #13162a;">₹{lot["cost"]:,.2f}</td>
                      <td colspan="6" style="border-bottom:1px solid #13162a;"></td>
                    </tr>"""

                # Unique id for toggle
                row_uid = f"lots_{section_key}_{ticker.replace('.','_')}"
                rows_html += f"""
            <tr id="{row_uid}" style="display:none;">
              <td colspan="12" style="padding:0; background:#090b18; border-bottom:2px solid #252849;">
                <table style="width:100%; border-collapse:collapse; font-size:11px;">
                  <thead>
                    <tr style="background:#0f1225;">
                      <th style="padding:6px 14px 6px 40px; text-align:left; color:#5a5f88; font-size:9px; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">Buy Date</th>
                      <th style="padding:6px 14px; text-align:right; color:#5a5f88; font-size:9px; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">Qty</th>
                      <th style="padding:6px 14px; text-align:right; color:#5a5f88; font-size:9px; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">Buy Price ₹</th>
                      <th style="padding:6px 14px; text-align:right; color:#5a5f88; font-size:9px; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">CMP ₹</th>
                      <th style="padding:6px 14px; text-align:right; color:#5a5f88; font-size:9px; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">Lot Cost</th>
                      <th colspan="7"></th>
                    </tr>
                  </thead>
                  <tbody>{lot_rows_html}</tbody>
                </table>
              </td>
            </tr>
            <script>
            (function(){{
                var row = document.getElementById("{row_uid}");
                if (!row) return;
                var trigger = row.previousElementSibling;
                if (!trigger) return;
                trigger.style.cursor = "pointer";
                var expanded = false;
                trigger.addEventListener("click", function(){{
                    expanded = !expanded;
                    row.style.display = expanded ? "table-row" : "none";
                }});
            }})();
            </script>"""

        # ── MOBILE: build mobile_cards_html ────────────────────────────
        mobile_cards_html = ""
        for i, g in enumerate(ticker_groups):
            ticker    = g["ticker"]
            atype_c   = g["atype"]
            if atype_c == "Stock":
                badge_col_c, badge_bg_c = "#4f7ef8", "rgba(79,126,248,0.18)"
                badge_lbl_c = "STK"
            elif atype_c == "ETF":
                badge_col_c, badge_bg_c = "#a78bfa", "rgba(167,139,250,0.18)"
                badge_lbl_c = "ETF"
            elif atype_c == "F&O":
                badge_col_c, badge_bg_c = "#f5c842", "rgba(245,200,66,0.18)"
                badge_lbl_c = "F&O"
            else:
                badge_col_c, badge_bg_c = "#7a7fa8", "rgba(122,127,168,0.18)"
                badge_lbl_c = atype_c[:3]

            unreal_v   = g["unreal"]
            pnl_pct_cv = g["pnl_pct"]
            daily_v    = g["daily"]
            daily_pct_v= g["daily_pct"]
            cmp_v      = g["cmp"]
            avg_v      = g["wtd_avg"]
            qty_v      = g["total_qty"]
            invested_v = g["cost_basis"]

            pnl_col_c  = "#22d67b" if unreal_v  >= 0 else "#f85454"
            dpnl_col_c = "#22d67b" if daily_v   >= 0 else "#f85454"
            dot_col_c  = "#22d67b" if pnl_pct_cv > 0 else ("#f85454" if pnl_pct_cv < 0 else "#7a7fa8")
            pnl_arrow  = "▲" if unreal_v  > 0 else ("▼" if unreal_v  < 0 else "")
            day_arrow  = "▲" if daily_v   > 0 else ("▼" if daily_v   < 0 else "")

            card_bg = "linear-gradient(135deg,#111428 0%,#0d1020 100%)" if i % 2 == 0 else "linear-gradient(135deg,#0f1128 0%,#0c0e1e 100%)"

            lots_badge = f'<span style="background:rgba(245,200,66,0.15);color:#f5c842;font-size:9px;font-weight:800;padding:1px 6px;border-radius:3px;margin-left:4px;">{len(g["lots"])} lots</span>' if g["multi_lot"] else ""

            # Lots detail rows for mobile expand
            lots_detail_html = ""
            if g["multi_lot"]:
                card_uid = f"mcard_{section_key}_{ticker.replace('.','_')}"
                for lot in g["lots"]:
                    lots_detail_html += f"""
<div style="display:flex; justify-content:space-between; align-items:center;
            padding:7px 12px; border-bottom:1px solid #1a1e38; background:#090b18;">
  <div>
    <div style="font-size:10px; color:#a0a4c8; font-weight:700;">📅 {lot["buy_date"]}</div>
    <div style="font-size:10px; color:#5a5f88; margin-top:2px;">Lot Cost: ₹{lot["cost"]:,.2f}</div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:11px; font-weight:700; color:#f5c842;">₹{lot["buy_price"]:,.2f}</div>
    <div style="font-size:10px; color:#7a7fa8;">{lot["qty"]:,.2f} qty</div>
  </div>
</div>"""
                lots_section = f"""
  <div id="{card_uid}_lots" style="display:none; border-top:1px solid #252849; margin-top:8px; border-radius:0 0 10px 10px; overflow:hidden;">
    <div style="padding:6px 12px; background:#0f1225; font-size:9px; color:#5a5f88; font-weight:700; text-transform:uppercase; letter-spacing:0.8px;">
      All Purchase Lots
    </div>
    {lots_detail_html}
  </div>
  <div onclick="(function(){{ var el=document.getElementById('{card_uid}_lots'); var btn=document.getElementById('{card_uid}_btn'); if(el.style.display==='none'){{el.style.display='block';btn.textContent='▲ Hide lots';}}else{{el.style.display='none';btn.textContent='▼ {len(g["lots"])} lots · tap to expand';}} }})();"
       id="{card_uid}_btn"
       style="margin-top:8px; padding:6px 10px; background:rgba(245,200,66,0.08); border:1px solid rgba(245,200,66,0.25);
              border-radius:6px; color:#f5c842; font-size:10px; font-weight:700; cursor:pointer; text-align:center;">
    ▼ {len(g["lots"])} lots · tap to expand
  </div>"""
            else:
                lots_section = ""

            mobile_cards_html += f"""
<div style="background:{card_bg}; border:1px solid #1e2240; border-left:3px solid {dot_col_c};
            border-radius:12px; padding:14px 14px 10px 14px; margin-bottom:10px;">
  <!-- Row 1: Ticker + badge | Unrealized P&L -->
  <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:10px;">
    <div style="display:flex; align-items:center; gap:8px;">
      <div>
        <div style="color:#f0f2ff; font-weight:800; font-size:14px; letter-spacing:0.4px; line-height:1.2;">
          {ticker}{lots_badge}
        </div>
        <span style="background:{badge_bg_c}; color:{badge_col_c}; font-size:9px; font-weight:800;
                     padding:2px 7px; border-radius:4px; letter-spacing:0.5px; margin-top:3px; display:inline-block;">
          {badge_lbl_c}
        </span>
      </div>
    </div>
    <div style="text-align:right;">
      <div style="color:{pnl_col_c}; font-weight:800; font-size:14px;">{pnl_arrow} ₹{abs(unreal_v):,.2f}</div>
      <div style="color:{pnl_col_c}; font-size:11px; font-weight:700; opacity:0.85;">({pnl_pct_cv:+.2f}%)</div>
    </div>
  </div>
  <!-- Row 2: Key stats grid -->
  <div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; margin-bottom:8px;">
    <div style="background:#0b0d1e; border-radius:8px; padding:8px 10px;">
      <div style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700; margin-bottom:3px;">Invested</div>
      <div style="font-size:12px; font-weight:700; color:#c8cce8;">₹{invested_v:,.2f}</div>
    </div>
    <div style="background:#0b0d1e; border-radius:8px; padding:8px 10px;">
      <div style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700; margin-bottom:3px;">Wtd. Avg Buy</div>
      <div style="font-size:12px; font-weight:700; color:#f5c842;">₹{avg_v:,.2f}</div>
    </div>
    <div style="background:#0b0d1e; border-radius:8px; padding:8px 10px;">
      <div style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700; margin-bottom:3px;">Total Qty</div>
      <div style="font-size:12px; font-weight:700; color:#c8cce8;">{qty_v:,.2f}</div>
    </div>
  </div>
  <!-- Row 3: CMP + Today P&L -->
  <div style="display:flex; justify-content:space-between; align-items:center;
              border-top:1px solid #1a1e38; padding-top:8px; margin-top:2px;">
    <div>
      <span style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">LTP &nbsp;</span>
      <span style="font-size:13px; font-weight:800; color:#f0f2ff;">₹{cmp_v:,.2f}</span>
    </div>
    <div style="text-align:right;">
      <span style="font-size:9px; color:#5a5f88; text-transform:uppercase; letter-spacing:0.8px; font-weight:700;">Today &nbsp;</span>
      <span style="font-size:12px; font-weight:800; color:{dpnl_col_c};">{day_arrow} ₹{abs(daily_v):,.2f}
        <span style="font-size:10px;">({daily_pct_v:+.2f}%)</span>
      </span>
    </div>
  </div>
  {lots_section}
</div>"""

        table_html = f"""
<style>
/* ── Desktop: show table, hide cards ── */
.port-desktop-table {{ display: block; }}
.port-mobile-cards  {{ display: none;  }}

/* ── Mobile (≤ 640px): show cards, hide table ── */
@media (max-width: 640px) {{
    .port-desktop-table {{ display: none !important; }}
    .port-mobile-cards  {{ display: block !important; }}
}}
</style>

<!-- DESKTOP TABLE -->
<div class="port-desktop-table">
  <div style="overflow-x:auto; border-radius:14px; border:1px solid #2a2e52;
              margin-top:12px; background:#0a0c1e;
              box-shadow: 0 4px 32px rgba(79,126,248,0.08), 0 1px 3px rgba(0,0,0,0.4);">
    <table style="width:100%; border-collapse:collapse; font-size:12.5px;
                  font-family:'Inter',-apple-system,sans-serif; min-width:900px;">
      <thead>{header}</thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>

<!-- MOBILE CARDS -->
<div class="port-mobile-cards" style="margin-top:10px;">
  {mobile_cards_html}
</div>"""

        st.markdown(table_html, unsafe_allow_html=True)

    # =====================================================
    # SECTION TABS — fully dynamic, only show tabs with actual holdings
    # =====================================================

    # ── Config: all possible tab types in display order ──────────────────
    _TAB_CONFIG = [
        # (asset_type_group,        tab_label_fn,                        emoji, label,           section_key)
        ("Stock",         lambda n: f"🏦 Stocks ({n})",                  "🏦", "Stock",          "stocks"),
        ("ETF",           lambda n: f"📊 ETFs ({n})",                    "📊", "ETF",            "etfs"),
        ("Liquid ETF",    lambda n: f"💧 Liquid ETF ({n})",              "💧", "Liquid ETF",     "letf"),
        ("Commodity ETF", lambda n: f"🪙 Commodity ETF ({n})",           "🪙", "Commodity ETF",  "cetf"),
        ("International ETF", lambda n: f"🌏 Intl ETF ({n})",           "🌏", "Intl ETF",       "ietf"),
        ("F&O",           lambda n: f"⚡ F&O ({n})",                     "⚡", "F&O",            "fno"),
        ("Index F&O",     lambda n: f"📈 Index F&O ({n})",               "📈", "Index F&O",      "ifno"),
        ("Currency F&O",  lambda n: f"💱 Currency F&O ({n})",            "💱", "Currency F&O",   "cfno"),
        ("REIT/InvIT",    lambda n: f"🏢 REIT/InvIT ({n})",              "🏢", "REIT/InvIT",     "reit"),
        ("SGB",           lambda n: f"🥇 SGB ({n})",                     "🥇", "SGB",            "sgb"),
        ("Bond/NCD",      lambda n: f"📄 Bond/NCD ({n})",                "📄", "Bond/NCD",       "bond"),
        ("Mutual Fund",   lambda n: f"🌱 Mutual Fund ({n})",             "🌱", "Mutual Fund",    "mf"),
        ("Preference Share", lambda n: f"🔵 Pref Share ({n})",           "🔵", "Pref Share",     "pref"),
        ("Unlisted Share",   lambda n: f"🔒 Unlisted ({n})",               "🔒", "Unlisted Share", "unlisted"),
    ]

    # ── Build list of tabs that actually have holdings ────────────────────
    _active_tabs = []
    for (_atype, _label_fn, _emoji, _label, _key) in _TAB_CONFIG:
        # Use full calc for Unlisted, calc_listed for all others
        _calc_src = unlisted_df if _atype == "Unlisted Share" else calc_listed
        _subset = _calc_src[_calc_src["Asset_Type"] == _atype].copy() if _atype != "Unlisted Share" else unlisted_df.copy()
        if not _subset.empty:
            _active_tabs.append({
                "atype":    _atype,
                "label":    _label_fn(len(_subset)),
                "emoji":    _emoji,
                "name":     _label,
                "key":      _key,
                "df":       _subset,
                "invested": _subset["Cost_Basis"].sum(),
            })

    if not _active_tabs:
        st.info("No holdings found.")
    else:
        _tab_labels   = [t["label"] for t in _active_tabs]
        _tab_objects  = st.tabs(_tab_labels)

        def _render_mini_bar(tab_info):
            """Render the invested/value/P&L context strip at the top of each tab."""
            tdf      = tab_info["df"]
            name     = tab_info["name"]
            emoji    = tab_info["emoji"]
            invested = tab_info["invested"]
            val      = tdf["Value"].sum()          if not tdf.empty else 0.0
            pnl      = tdf["Unrealized_PnL"].sum() if not tdf.empty else 0.0
            dpnl     = tdf["Daily_PnL"].sum()      if not tdf.empty else 0.0
            _is_unlisted = tab_info.get("atype") == "Unlisted Share"
            pnl_col  = "#22d67b" if pnl  >= 0 else "#f85454"
            dpnl_col = "#22d67b" if dpnl >= 0 else "#f85454"
            pnl_bg   = "#0f1a14" if pnl  >= 0 else "#1a0f0f"
            dpnl_bg  = "#0f1a14" if dpnl >= 0 else "#1a0f0f"
            # ── 6th tile: Booked Profit — active + fully-sold tickers ────
            _tab_atype       = tab_info.get("atype", "")
            _active_tickers  = set(tdf["Ticker"].tolist()) if not tdf.empty else set()
            _orig_tab_mask   = df["Asset_Type"] == _tab_atype if "Asset_Type" in df.columns else pd.Series([True]*len(df))
            _all_tab_tickers = set(df.loc[_orig_tab_mask, "Ticker"].tolist()) | _active_tickers
            _tab_booked      = sum(v for k, v in (booked_pnl_map or {}).items() if k in _all_tab_tickers)
            bpnl_col = "#22d67b" if _tab_booked >= 0 else "#f85454"
            bpnl_bg  = "#0f1a14" if _tab_booked >= 0 else "#1a0f0f"
            st.markdown(f"""
    <div class="mini-bar-wrap">
      <div style="background:#131628;border:1px solid #252849;border-radius:10px;padding:10px 18px;flex:1;min-width:110px;">
        <div style="font-size:9px;color:#6b7299;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">{emoji} {name} Invested</div>
        <div style="font-size:15px;font-weight:800;color:#f0f2ff;">{fmt_inr(invested)}</div>
      </div>
      <div style="background:#131628;border:1px solid #4f7ef844;border-radius:10px;padding:10px 18px;flex:1;min-width:110px;">
        <div style="font-size:9px;color:#6b7299;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">💰 Total Invested</div>
        <div style="font-size:15px;font-weight:800;color:#4f7ef8;">{fmt_inr(total_invested)}</div>
      </div>
      <div style="background:#131628;border:1px solid #252849;border-radius:10px;padding:10px 18px;flex:1;min-width:110px;">
        <div style="font-size:9px;color:#6b7299;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">📦 {name} Value</div>
        <div style="font-size:15px;font-weight:800;color:#4f7ef8;">{fmt_inr(val)}</div>
      </div>
      {"" if _is_unlisted else f'''
      <div style="background:{pnl_bg};border:1px solid {pnl_col}44;border-radius:10px;padding:10px 18px;flex:1;min-width:110px;">
        <div style="font-size:9px;color:#6b7299;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">📈 {name} P&L</div>
        <div style="font-size:15px;font-weight:800;color:{pnl_col};">{fmt_inr(pnl)}</div>
      </div>
      <div style="background:{dpnl_bg};border:1px solid {dpnl_col}44;border-radius:10px;padding:10px 18px;flex:1;min-width:110px;">
        <div style="font-size:9px;color:#6b7299;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">📅 Today's P&L</div>
        <div style="font-size:15px;font-weight:800;color:{dpnl_col};">{fmt_inr(dpnl)}</div>
      </div>'''}
      {"" if not _is_unlisted else '''
      <div style="background:#1a1020;border:1px solid #a78bfa44;border-radius:10px;padding:10px 18px;flex:2;min-width:180px;">
        <div style="font-size:9px;color:#a78bfa;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">🔒 P&L Excluded from Total</div>
        <div style="font-size:11px;font-weight:600;color:#8878aa;">Unlisted / Pre-IPO holdings are tracked separately and do not affect portfolio P&L totals.</div>
      </div>'''}
      {"" if _is_unlisted else f'''
      <div style="background:{bpnl_bg};border:1px solid {bpnl_col}44;border-radius:10px;padding:10px 18px;flex:1;min-width:110px;">
        <div style="font-size:9px;color:#6b7299;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px;">✅ Booked Profit</div>
        <div style="font-size:15px;font-weight:800;color:{bpnl_col};">{fmt_inr(_tab_booked)}</div>
      </div>'''}
    </div>
    """, unsafe_allow_html=True)

        for _tab_obj, _tab_info in zip(_tab_objects, _active_tabs):
            with _tab_obj:
                _render_mini_bar(_tab_info)

                # ── Tab heading ──────────────────────────────────────────
                _hint = "Click Qty or Avg Buy ₹ to edit inline · saves automatically"
                if _tab_info["atype"] in ("F&O", "Index F&O", "Currency F&O"):
                    _hint = "Enter tickers as Angel One NFO symbols"
                st.markdown(f"""
    <div style="margin-bottom:8px;">
      <span style="font-size:15px;font-weight:700;color:#f0f2ff;">{_tab_info['emoji']} {_tab_info['name']} Holdings</span>
      <span style="font-size:11px;color:#7a7fa8;margin-left:8px;">{_hint}</span>
    </div>
    """, unsafe_allow_html=True)

                # ── F&O debug panel ──────────────────────────────────────
                if _tab_info["atype"] in ("F&O", "Index F&O", "Currency F&O"):
                    st.info(
                        "💡 **For correct F&O prices**, enter tickers exactly as Angel One NFO symbols: "
                        "`ABB26MAY25FUT`, `DRREDDY26MAY25FUT`, `NIFTY26MAY2524000CE`"
                    )
                    with st.expander("🔍 F&O Token Resolution Debug", expanded=False):
                        debug_df = debug_fno_resolution(df, nfo_map)
                        st.dataframe(debug_df, use_container_width=True, hide_index=True)
                        st.caption("If 'Resolved NFO Symbol' shows ❌ NOT FOUND, rename the ticker to the exact Angel One NFO symbol.")
                        st.markdown("**🔎 Search NFO Symbols:**")
                        search_q = st.text_input("Type a symbol name (e.g. ABB, DRREDDY)", key=f"nfo_search_{_tab_info['key']}")
                        if search_q:
                            q = search_q.upper().strip()
                            hits = [(k, v["token"]) for k, v in nfo_map.items() if q in k][:30]
                            if hits:
                                st.dataframe(pd.DataFrame(hits, columns=["NFO Symbol", "Token"]), use_container_width=True, hide_index=True)
                            else:
                                st.warning("No matches found.")

                if _tab_info["atype"] == "Unlisted Share":
                    # ── Unlisted Share special notice ────────────────────
                    st.markdown("""
<div style="background:#1a1020;border:1px solid #a78bfa44;border-radius:10px;
            padding:10px 16px;margin-bottom:12px;display:flex;align-items:center;gap:10px;">
  <span style="font-size:18px;">🔒</span>
  <div>
    <div style="color:#a78bfa;font-size:12px;font-weight:700;text-transform:uppercase;
                letter-spacing:0.8px;margin-bottom:2px;">Unlisted / Pre-IPO Shares</div>
    <div style="color:#8878aa;font-size:11px;line-height:1.5;">
      These holdings are <b style="color:#f0e0ff;">excluded from Total P&amp;L and Daily P&amp;L</b> 
      shown at the top — buy price may be zero or unknown, so P&amp;L is not calculated.<br>
      CMP is fetched automatically from <b>unlistedzone.com</b> (or Angel One if the stock is listed).
      If auto-fetch fails, use the <b>Manual CMP Override</b> in the sidebar — your Buy Price is never changed.
    </div>
  </div>
</div>
""", unsafe_allow_html=True)
                render_section_table(_tab_info["df"], show_sold=True, section_key=_tab_info["key"])

    # =====================================================
    # SELL TRADE HISTORY — Asset-classified with diversification
    # =====================================================

    # ── Section heading ────────────────────────────────────────────
    st.markdown("""
<div style="margin:20px 0 8px 0;">
  <span style="font-size:15px;font-weight:700;color:#f0f2ff;">💰 Sell Trade History</span>
  <span style="font-size:11px;color:#7a7fa8;margin-left:8px;">Realized gains &amp; losses · Asset-classified</span>
</div>
""", unsafe_allow_html=True)

    # ── Dev only: Edit Sell Trades expander ───────────────────────
    if _is_dev and not trades_df.empty:
        st.markdown('<div class="collapsible-section">', unsafe_allow_html=True)
        with st.expander("✏️  Edit Sell Trades — Qty & Sell Price", expanded=False):
            st.caption("Click any **Sell Qty** or **Sell Price ₹** cell to edit · Booked P&L recalculates automatically · Changes save on edit")
            _th_edit = trades_df.copy()
            _th_edit["Booked_PnL"]        = pd.to_numeric(_th_edit["Booked_PnL"],        errors="coerce")
            _th_edit["Sell_Price"]        = pd.to_numeric(_th_edit["Sell_Price"],         errors="coerce")
            _th_edit["Buy_Price_At_Sell"] = pd.to_numeric(_th_edit["Buy_Price_At_Sell"],  errors="coerce")
            _th_edit["Sell_Qty"]          = pd.to_numeric(_th_edit["Sell_Qty"],           errors="coerce")
            _sell_col_cfg = {
                "Ticker":           st.column_config.TextColumn("Ticker",             disabled=True),
                "Asset_Type":       st.column_config.TextColumn("Asset Type",         disabled=True),
                "Sell_Qty":         st.column_config.NumberColumn("Sell Qty",         min_value=0.0001, step=1.0,  format="%.2f"),
                "Sell_Price":       st.column_config.NumberColumn("Sell Price ₹",     min_value=0.0001, step=0.01, format="%.2f"),
                "Sell_Date":        st.column_config.TextColumn("Sell Date",          disabled=True),
                "Buy_Price_At_Sell":st.column_config.NumberColumn("Wtd. Avg Price ₹", disabled=True, format="%.2f"),
                "Booked_PnL":       st.column_config.NumberColumn("Booked P&L ₹",    disabled=True, format="%.2f"),
            }
            _edit_cols = [c for c in ["Ticker","Asset_Type","Sell_Qty","Sell_Price","Sell_Date","Buy_Price_At_Sell","Booked_PnL"] if c in _th_edit.columns]
            _sell_edited = st.data_editor(
                _th_edit[_edit_cols].copy(),
                column_config=_sell_col_cfg,
                use_container_width=True,
                hide_index=True,
                key="sell_trade_editor",
                num_rows="fixed",
            )
            _sell_changed = []
            for _si, (_sorig_idx, _sorig_row) in enumerate(trades_df.iterrows()):
                _new_qty   = float(_sell_edited.iloc[_si]["Sell_Qty"])
                _new_price = float(_sell_edited.iloc[_si]["Sell_Price"])
                _old_qty   = float(_sorig_row["Sell_Qty"])
                _old_price = float(_sorig_row["Sell_Price"])
                if abs(_new_qty - _old_qty) > 0.0001 or abs(_new_price - _old_price) > 0.0001:
                    _sell_changed.append((_sorig_idx, _new_qty, _new_price))
            if _sell_changed:
                for _si_idx, _nq, _np in _sell_changed:
                    _avg_buy = float(trades_df.at[_si_idx, "Buy_Price_At_Sell"])
                    _new_booked = round((_np - _avg_buy) * _nq, 2)
                    trades_df.at[_si_idx, "Sell_Qty"]   = round(_nq, 4)
                    trades_df.at[_si_idx, "Sell_Price"]  = round(_np, 4)
                    trades_df.at[_si_idx, "Booked_PnL"]  = _new_booked
                trades_df.to_csv(TRADES_FILE, index=False)
                _local_sync_to_gh(TRADES_FILE, TRADES_FILE, 'Auto-save: trades')
                st.success(f"✅ {len(_sell_changed)} sell trade(s) updated. Refreshing...")
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Sell Trade History table + diversification ─────────────────
    if trades_df.empty:
        st.info("No sell trades recorded yet.")
    else:
        _th = trades_df.copy()
        _th["Booked_PnL"]        = pd.to_numeric(_th["Booked_PnL"],        errors="coerce").fillna(0)
        _th["Sell_Price"]        = pd.to_numeric(_th["Sell_Price"],         errors="coerce").fillna(0)
        _th["Buy_Price_At_Sell"] = pd.to_numeric(_th["Buy_Price_At_Sell"],  errors="coerce").fillna(0)
        _th["Sell_Qty"]          = pd.to_numeric(_th["Sell_Qty"],           errors="coerce").fillna(0)

        # ── Ensure Asset_Type is populated ────────────────────────
        if "Asset_Type" not in _th.columns:
            _th["Asset_Type"] = ""
        _th["Asset_Type"] = _th.apply(
            lambda r: (str(r["Asset_Type"]).strip()
                       if pd.notna(r.get("Asset_Type")) and str(r.get("Asset_Type","")).strip() not in ("","nan","None")
                       else (_pf_asset_map.get(str(r["Ticker"]).strip()) or _classify_ticker_to_asset_type(str(r["Ticker"])))),
            axis=1
        )

        # ── Detect short positions: Buy_Price_At_Sell == 0 ─────────
        # Short = sold without a prior buy (e.g. options written, short F&O)
        # For these: Live P&L = (Sell_Price - CMP) × Qty  (profit when CMP falls)
        def _is_short(row):
            return float(row.get("Buy_Price_At_Sell", 0)) == 0

        # ── Fetch live CMP for all trade tickers ──────────────────
        # Re-use the already-loaded equity_map / nfo_map from the price section above.
        # Run concurrently so multiple short positions don't stack latency.
        _trade_tickers_unique = _th["Ticker"].unique().tolist()
        _trade_cmp_map = {}

        def _fetch_trade_cmp(tkr):
            _atype = _th[_th["Ticker"] == tkr]["Asset_Type"].iloc[0] if not _th[_th["Ticker"] == tkr].empty else "Stock"
            # Check portfolio prices dict first (already fetched)
            if tkr in prices and prices[tkr]:
                return tkr, prices[tkr]
            # Otherwise fetch fresh
            _p = get_angel_one_price(tkr, equity_map, nfo_map, _atype, bse_map=bse_map)
            return tkr, _p

        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as _tex:
            for _tk, _tp in _tex.map(_fetch_trade_cmp, _trade_tickers_unique):
                _trade_cmp_map[_tk] = _tp

        # ── Compute live P&L for each row ─────────────────────────
        def _live_pnl(row):
            """
            Short position  (Buy_Price_At_Sell == 0):
                Live P&L = (Sell_Price − CMP) × Qty   → profit when CMP < Sell_Price
            Long close      (Buy_Price_At_Sell > 0):
                Booked P&L is final — CMP shown informatively only.
            """
            sell_px  = float(row.get("Sell_Price", 0))
            qty      = float(row.get("Sell_Qty", 0))
            buy_px   = float(row.get("Buy_Price_At_Sell", 0))
            cmp      = _trade_cmp_map.get(str(row.get("Ticker","")))
            if _is_short(row) and cmp and cmp > 0:
                return round((sell_px - cmp) * qty, 2)
            return float(row.get("Booked_PnL", 0))  # for closed longs, booked is final

        _th["CMP"]      = _th["Ticker"].map(_trade_cmp_map)
        _th["Live_PnL"] = _th.apply(_live_pnl, axis=1)

        # ── Asset-type badge colours ───────────────────────────────
        _ATYPE_COLORS = {
            "Stock":         ("#4f7ef8", "rgba(79,126,248,0.15)"),
            "F&O":           ("#f5c842", "rgba(245,200,66,0.15)"),
            "ETF":           ("#22d67b", "rgba(34,214,123,0.15)"),
            "Liquid ETF":    ("#22d67b", "rgba(34,214,123,0.12)"),
            "Commodity ETF": ("#fb923c", "rgba(251,146,60,0.15)"),
            "Mutual Fund":   ("#a78bfa", "rgba(167,139,250,0.15)"),
            "Bond/NCD":      ("#38bdf8", "rgba(56,189,248,0.15)"),
            "SGB":           ("#f5c842", "rgba(245,200,66,0.12)"),
            "REIT/InvIT":    ("#34d399", "rgba(52,211,153,0.15)"),
            "Commodity":     ("#fb923c", "rgba(251,146,60,0.15)"),
            "Crypto":        ("#f472b6", "rgba(244,114,182,0.15)"),
            "Unlisted Share":("#a78bfa", "rgba(167,139,250,0.18)"),
            "Preference Share":("#60a5fa","rgba(96,165,250,0.15)"),
        }
        def _atype_badge_html(atype):
            col, bg = _ATYPE_COLORS.get(atype, ("#7a7fa8", "rgba(122,127,168,0.15)"))
            emoji = asset_type_emoji(atype)
            return (f'<span style="background:{bg};color:{col};font-size:9px;font-weight:800;'
                    f'padding:2px 7px;border-radius:4px;letter-spacing:0.5px;'
                    f'border:1px solid {col}33;">{emoji} {atype.upper()}</span>')

        # ── Has any short position? → show live P&L column ────────
        _has_shorts = any(_is_short(r) for _, r in _th.iterrows())
        _total_live_pnl = float(_th["Live_PnL"].sum())

        # ── Diversification summary cards ─────────────────────────
        _ac_groups = _th.groupby("Asset_Type").agg(
            Trades=("Ticker", "count"),
            Total_PnL=("Live_PnL", "sum"),    # use live P&L for shorts
            Sell_Value=("Sell_Price", lambda x: (x * _th.loc[x.index, "Sell_Qty"]).sum()),
        ).reset_index().sort_values("Total_PnL", ascending=False)

        _total_val_all = float((_th["Sell_Price"] * _th["Sell_Qty"]).sum())

        _div_cards_html = '<div style="display:flex;flex-wrap:wrap;gap:10px;margin-bottom:18px;">'
        for _, _acr in _ac_groups.iterrows():
            _ac_name = str(_acr["Asset_Type"])
            _ac_pnl  = float(_acr["Total_PnL"])
            _ac_val  = float(_acr["Sell_Value"])
            _ac_pct  = (_ac_val / _total_val_all * 100) if _total_val_all > 0 else 0
            _ac_col, _ac_bg = _ATYPE_COLORS.get(_ac_name, ("#7a7fa8", "rgba(122,127,168,0.10)"))
            _pnl_col = "#22d67b" if _ac_pnl >= 0 else "#f85454"
            _arrow = "▲" if _ac_pnl > 0 else ("▼" if _ac_pnl < 0 else "")
            _div_cards_html += f"""
<div style="background:#131628;border:1px solid {_ac_col}44;border-left:4px solid {_ac_col};
            border-radius:12px;padding:12px 16px;min-width:150px;flex:1 1 140px;">
  <div style="font-size:10px;color:{_ac_col};font-weight:800;letter-spacing:0.7px;
              text-transform:uppercase;margin-bottom:6px;">{asset_type_emoji(_ac_name)} {_ac_name}</div>
  <div style="font-size:18px;font-weight:900;color:{_pnl_col};line-height:1.1;">
    {_arrow} ₹{abs(_ac_pnl):,.0f}
  </div>
  <div style="font-size:10px;color:#7a7fa8;margin-top:4px;">
    {int(_acr['Trades'])} trade{'s' if _acr['Trades']>1 else ''} &nbsp;·&nbsp;
    {_ac_pct:.1f}% of volume
  </div>
</div>"""
        _div_cards_html += "</div>"
        st.markdown(_div_cards_html, unsafe_allow_html=True)

        # ── Styled HTML sell table ─────────────────────────────────
        def _sv_plain(val, fmt_fn, padding="10px 14px"):
            try:
                return (f'<td style="padding:{padding};text-align:right;color:#c8cce8;font-weight:500;'
                        f'border-bottom:1px solid #1a1e38;white-space:nowrap;">{fmt_fn(float(val))}</td>')
            except:
                return f'<td style="padding:{padding};text-align:right;color:#3a3d5c;border-bottom:1px solid #1a1e38;">—</td>'

        def _sv_pnl(val, is_live=False, padding="10px 14px"):
            try:
                fv  = float(val)
                col = "#22d67b" if fv >= 0 else "#f85454"
                bg  = "rgba(34,214,123,0.06)" if fv >= 0 else "rgba(248,84,84,0.06)"
                arr = "▲" if fv > 0 else ("▼" if fv < 0 else "")
                live_dot = ('<span style="display:inline-block;width:6px;height:6px;border-radius:50%;'
                            f'background:{col};margin-left:4px;vertical-align:middle;'
                            'animation:blink 1.2s ease-in-out infinite;"></span>'
                            if is_live else "")
                return (f'<td style="padding:{padding};text-align:right;color:{col};font-weight:700;'
                        f'border-bottom:1px solid #1a1e38;background:{bg};white-space:nowrap;">'
                        f'{arr} ₹{abs(fv):,.2f}{live_dot}</td>')
            except:
                return f'<td style="padding:{padding};text-align:right;color:#3a3d5c;border-bottom:1px solid #1a1e38;">—</td>'

        _sell_rows_html = ""
        for _i, (_, _row) in enumerate(_th.iterrows()):
            _rbg        = "#111428" if _i % 2 == 0 else "#0d1020"
            _hover_bg   = "#1a1e38"
            _atype      = str(_row.get("Asset_Type", "Stock"))
            _badge_html = _atype_badge_html(_atype)
            _short      = _is_short(_row)
            _cmp_val    = _trade_cmp_map.get(str(_row.get("Ticker","")))
            _live_pnl_v = float(_row.get("Live_PnL", _row.get("Booked_PnL", 0)))
            _bar_col    = "#22d67b" if _live_pnl_v >= 0 else "#f85454"

            # ── Avg Buy / Short label cell ─────────────────────────
            if _short:
                _avg_buy_td = ('<td style="padding:10px 14px;text-align:right;border-bottom:1px solid #1a1e38;">'
                               '<span style="background:rgba(245,84,84,0.12);color:#f85454;font-size:9px;'
                               'font-weight:800;padding:2px 8px;border-radius:4px;letter-spacing:0.5px;'
                               'border:1px solid rgba(245,84,84,0.3);">SHORT SELL</span></td>')
            else:
                _avg_buy_val = float(_row.get("Buy_Price_At_Sell", 0))
                _avg_buy_td  = _sv_plain(_avg_buy_val, lambda v: f"₹{v:,.2f}")

            # ── CMP cell ───────────────────────────────────────────
            if _cmp_val and _cmp_val > 0:
                _cmp_change_col = "#f85454" if (_short and _cmp_val > float(_row.get("Sell_Price",0))) else (
                                  "#22d67b" if (_short and _cmp_val < float(_row.get("Sell_Price",0))) else "#c8cce8")
                _cmp_td = (f'<td style="padding:10px 14px;text-align:right;color:{_cmp_change_col};font-weight:600;'
                           f'border-bottom:1px solid #1a1e38;white-space:nowrap;">'
                           f'₹{_cmp_val:,.2f}'
                           + (' <span style="display:inline-block;width:5px;height:5px;border-radius:50%;'
                              f'background:{_cmp_change_col};vertical-align:middle;'
                              'animation:blink 1.2s ease-in-out infinite;"></span>' if _short else "")
                           + '</td>')
            else:
                _cmp_td = '<td style="padding:10px 14px;text-align:right;color:#3a3d5c;border-bottom:1px solid #1a1e38;">—</td>'

            _sell_rows_html += f"""
            <tr style="background:{_rbg};transition:background 0.15s;"
                onmouseover="this.style.background='{_hover_bg}'"
                onmouseout="this.style.background='{_rbg}'">
              <td style="padding:10px 16px;border-bottom:1px solid #1a1e38;white-space:nowrap;">
                <div style="display:flex;align-items:center;gap:8px;">
                  <span style="width:4px;height:32px;border-radius:2px;background:{_bar_col};
                               display:inline-block;flex-shrink:0;"></span>
                  <div>
                    <div style="color:#f0f2ff;font-weight:800;font-size:13px;
                                letter-spacing:0.3px;">{str(_row.get("Ticker",""))}</div>
                    {_badge_html}
                  </div>
                </div>
              </td>
              {_sv_plain(_row.get("Sell_Qty"),   lambda v: f"{v:,.2f}")}
              {_sv_plain(_row.get("Sell_Price"), lambda v: f"₹{v:,.2f}")}
              <td style="padding:10px 14px;text-align:right;color:#8888aa;font-size:11px;
                         border-bottom:1px solid #1a1e38;white-space:nowrap;">{str(_row.get("Sell_Date",""))}</td>
              {_avg_buy_td}
              {_cmp_td}
              {_sv_pnl(_live_pnl_v, is_live=_short)}
            </tr>"""

        _pnl_footer_col = "#22d67b" if _total_live_pnl >= 0 else "#f85454"
        _pnl_footer_bg  = "rgba(34,214,123,0.06)" if _total_live_pnl >= 0 else "rgba(248,84,84,0.06)"
        _sell_table_html = f"""
<style>
@keyframes blink {{
  0%, 100% {{ opacity: 1; }}
  50%       {{ opacity: 0.2; }}
}}
</style>
<div class="port-desktop-table">
  <div style="overflow-x:auto;border-radius:14px;border:1px solid #2a2e52;
              margin-top:4px;background:#0a0c1e;
              box-shadow:0 4px 32px rgba(79,126,248,0.08),0 1px 3px rgba(0,0,0,0.4);">
    <table style="width:100%;border-collapse:collapse;font-size:12.5px;
                  font-family:'Inter',-apple-system,sans-serif;min-width:820px;">
      <thead>
        <tr style="background:linear-gradient(90deg,#181b2e 0%,#1a1d33 100%);
                   color:#6b7299;font-size:10px;text-transform:uppercase;
                   letter-spacing:0.9px;font-weight:700;">
          <th style="padding:12px 16px;text-align:left;border-bottom:2px solid #2a2e52;white-space:nowrap;">Holding · Asset Class</th>
          <th style="padding:12px 14px;text-align:right;border-bottom:2px solid #2a2e52;white-space:nowrap;">Sell Qty</th>
          <th style="padding:12px 14px;text-align:right;border-bottom:2px solid #2a2e52;white-space:nowrap;">Sell Price ₹</th>
          <th style="padding:12px 14px;text-align:right;border-bottom:2px solid #2a2e52;white-space:nowrap;">Sell Date</th>
          <th style="padding:12px 14px;text-align:right;border-bottom:2px solid #2a2e52;white-space:nowrap;">Avg Buy / Type</th>
          <th style="padding:12px 14px;text-align:right;border-bottom:2px solid #2a2e52;white-space:nowrap;">
            CMP ₹{'&nbsp;<span style="color:#22d67b;font-size:8px;">● LIVE</span>' if _has_shorts else ""}
          </th>
          <th style="padding:12px 14px;text-align:right;border-bottom:2px solid #2a2e52;white-space:nowrap;">
            {'Live P&amp;L' if _has_shorts else 'Booked P&amp;L'}
          </th>
        </tr>
      </thead>
      <tbody>{_sell_rows_html}</tbody>
      <tfoot>
        <tr style="background:#181b2e;">
          <td colspan="6" style="padding:10px 16px;color:#7a7fa8;font-size:11px;
              font-weight:700;border-top:2px solid #2a2e52;text-transform:uppercase;
              letter-spacing:0.7px;">{'Live' if _has_shorts else 'Total Realized'} P&amp;L</td>
          <td style="padding:10px 14px;text-align:right;font-weight:900;font-size:14px;
              color:{_pnl_footer_col};border-top:2px solid #2a2e52;
              background:{_pnl_footer_bg};white-space:nowrap;">
            {'▲' if _total_live_pnl > 0 else ('▼' if _total_live_pnl < 0 else '')} ₹{abs(_total_live_pnl):,.2f}
          </td>
        </tr>
      </tfoot>
    </table>
  </div>
</div>"""
        st.markdown(_sell_table_html, unsafe_allow_html=True)

    # =====================================================
    # P&L CHART — STOCKS + F&O ONLY (NO ETF)
    # =====================================================

    st.markdown('<div class="portfolio-mobile-expander">', unsafe_allow_html=True)
    with st.expander("📊 P&L Overview — Stocks & F&O · Unrealized gains & allocation", expanded=True):
        chart_df = calc_listed[calc_listed["Asset_Type"].isin(["Stock", "F&O"])].copy()
        chart_df = chart_df.dropna(subset=["Unrealized_PnL"])

        if not chart_df.empty:
            chart_df = chart_df.sort_values("Unrealized_PnL", ascending=False)
            DONUT_COLORS = ["#4f7ef8","#22d67b","#f5c842","#a78bfa","#fb923c",
                            "#38bdf8","#f472b6","#34d399","#facc15","#818cf8"]
            total_val  = chart_df["Value"].sum()
            total_pnl  = chart_df["Unrealized_PnL"].sum()
            cost_base  = total_val - total_pnl
            total_pct  = (total_pnl / cost_base * 100) if cost_base != 0 else 0
            pnl_col    = "#22d67b" if total_pnl >= 0 else "#f85454"
            pnl_arrow  = "▲" if total_pnl > 0 else "▼"
            max_abs    = chart_df["Unrealized_PnL"].abs().max() or 1

            if _IS_MOBILE:
                # ── MOBILE: pure HTML — no Plotly ──────────────────────────
                holding_rows = ""
                alloc_rows   = ""
                for i, (_, r) in enumerate(chart_df.iterrows()):
                    pv   = float(r.get("Unrealized_PnL", 0))
                    pp   = float(r.get("PnL_%", 0))
                    vv   = float(r.get("Value", 0))
                    av   = float(r.get("Buy_Price", 0))
                    cv   = float(r.get("Current_Price", 0))
                    qv   = float(r.get("Shares", 0))
                    inv  = float(r.get("Cost_Basis", 0))
                    alc  = (vv / total_val * 100) if total_val else 0
                    bw   = abs(pv) / max_abs * 100
                    clr  = "#22d67b" if pv >= 0 else "#f85454"
                    dim  = "rgba(34,214,123,0.12)" if pv >= 0 else "rgba(248,84,84,0.10)"
                    arr  = "▲" if pv > 0 else ("▼" if pv < 0 else "—")
                    dot  = DONUT_COLORS[i % len(DONUT_COLORS)]
                    tkr  = r.get("Ticker", "")
                    atp  = str(r.get("Asset_Type", ""))
                    bdg  = {"Stock":"STK","F&O":"F&O","ETF":"ETF"}.get(atp, atp[:3])

                    holding_rows += f"""
<div style="background:#0e1022;border:1px solid #1e2240;border-radius:12px;
            padding:14px 16px 10px;margin-bottom:10px;position:relative;overflow:hidden;">
  <div style="position:absolute;inset:0;width:{bw:.1f}%;background:{dim};z-index:0;border-radius:12px;"></div>
  <div style="position:relative;z-index:1;">
    <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:10px;">
      <div style="display:flex;align-items:center;gap:10px;">
        <div style="width:9px;height:9px;border-radius:50%;background:{dot};flex-shrink:0;
                    box-shadow:0 0 6px {dot}88;"></div>
        <div>
          <div style="font-size:15px;font-weight:800;color:#f0f2ff;letter-spacing:0.3px;">{tkr}</div>
          <span style="font-size:9px;font-weight:700;color:#5a5f88;text-transform:uppercase;
                       letter-spacing:0.6px;background:#1a1e38;padding:2px 6px;border-radius:4px;">
            {bdg} &nbsp;·&nbsp; {alc:.1f}%
          </span>
        </div>
      </div>
      <div style="text-align:right;">
        <div style="font-size:16px;font-weight:800;color:{clr};line-height:1.1;">
          {arr} ₹{abs(pv):,.0f}
        </div>
        <div style="font-size:12px;font-weight:700;color:{clr};opacity:0.8;margin-top:2px;">
          {pp:+.2f}%
        </div>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:8px;">
      <div style="background:#0a0b18;border-radius:7px;padding:7px 9px;">
        <div style="font-size:8px;color:#454870;text-transform:uppercase;letter-spacing:0.8px;
                    font-weight:700;margin-bottom:2px;">Invested</div>
        <div style="font-size:11px;font-weight:700;color:#b8bcd8;">₹{inv:,.0f}</div>
      </div>
      <div style="background:#0a0b18;border-radius:7px;padding:7px 9px;">
        <div style="font-size:8px;color:#454870;text-transform:uppercase;letter-spacing:0.8px;
                    font-weight:700;margin-bottom:2px;">Avg</div>
        <div style="font-size:11px;font-weight:700;color:#b8bcd8;">₹{av:,.1f}</div>
      </div>
      <div style="background:#0a0b18;border-radius:7px;padding:7px 9px;">
        <div style="font-size:8px;color:#454870;text-transform:uppercase;letter-spacing:0.8px;
                    font-weight:700;margin-bottom:2px;">LTP</div>
        <div style="font-size:11px;font-weight:700;color:#f0f2ff;">₹{cv:,.1f}</div>
      </div>
    </div>
    <div style="height:3px;background:#1a1e38;border-radius:3px;overflow:hidden;">
      <div style="height:100%;width:{bw:.1f}%;background:{clr};border-radius:3px;"></div>
    </div>
  </div>
</div>"""

                    alloc_rows += f"""
  <div style="display:flex;align-items:center;gap:10px;padding:7px 0;
              border-bottom:1px solid #151830;">
    <div style="width:8px;height:8px;border-radius:50%;background:{dot};flex-shrink:0;"></div>
    <span style="font-size:12px;font-weight:700;color:#e0e2f0;">{tkr}</span>
    <span style="font-size:11px;color:#5a5f88;margin-left:auto;">₹{vv/1e5:.2f}L</span>
    <span style="font-size:11px;font-weight:700;color:#b8bcd8;min-width:40px;text-align:right;">
      {alc:.1f}%
    </span>
  </div>"""

                st.markdown(f"""
<div style="background:linear-gradient(135deg,#0f1124 0%,#0b0d1c 100%);
            border:1px solid #252849;border-radius:16px;padding:18px 18px 14px;margin-bottom:14px;">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <div>
      <div style="font-size:10px;color:#454870;text-transform:uppercase;letter-spacing:1.2px;
                  font-weight:700;margin-bottom:5px;">Total Unrealized P&L</div>
      <div style="font-size:26px;font-weight:800;color:{pnl_col};line-height:1.1;">
        {pnl_arrow} ₹{abs(total_pnl):,.0f}
      </div>
      <div style="font-size:12px;font-weight:600;color:{pnl_col};opacity:0.75;margin-top:3px;">
        {total_pct:+.2f}% overall return
      </div>
    </div>
    <div style="text-align:right;">
      <div style="font-size:10px;color:#454870;text-transform:uppercase;letter-spacing:1.2px;
                  font-weight:700;margin-bottom:5px;">Portfolio</div>
      <div style="font-size:22px;font-weight:800;color:#f0f2ff;">₹{total_val/1e5:.2f}L</div>
      <div style="font-size:11px;color:#5a5f88;margin-top:3px;">{len(chart_df)} holdings</div>
    </div>
  </div>
</div>
<div style="font-size:9px;color:#454870;text-transform:uppercase;letter-spacing:1.2px;
            font-weight:700;margin-bottom:10px;padding-left:2px;">▸ P&L per Holding</div>
{holding_rows}
<div style="background:#0e1022;border:1px solid #1e2240;border-radius:12px;
            padding:14px 16px;margin-top:6px;">
  <div style="font-size:9px;color:#454870;text-transform:uppercase;letter-spacing:1.2px;
              font-weight:700;margin-bottom:8px;">▸ Value Allocation</div>
  {alloc_rows}
</div>
""", unsafe_allow_html=True)

            else:
                # ── DESKTOP: Plotly charts ─────────────────────────────────
                pos = chart_df[chart_df["Unrealized_PnL"] >= 0]
                neg = chart_df[chart_df["Unrealized_PnL"] < 0]
                fig_bar = go.Figure()
                for sub_df, clr_main, clr_line in [
                    (pos, "#22d67b", "#00e676"),
                    (neg, "#f85454", "#ff1744"),
                ]:
                    if sub_df.empty:
                        continue
                    fig_bar.add_trace(go.Bar(
                        x=sub_df["Ticker"], y=sub_df["Unrealized_PnL"],
                        marker=dict(color=clr_main, opacity=0.92,
                                    line=dict(color=clr_line, width=1.5)),
                        text=sub_df["PnL_%"].apply(lambda x: f"{x:+.1f}%"),
                        textfont=dict(size=11, color="#ffffff", family="Inter"),
                        textposition="outside",
                        hovertext=sub_df.apply(
                            lambda r: f"<b>{r['Ticker']}</b><br>P&L: ₹{r['Unrealized_PnL']:,.2f}<br>Return: {r['PnL_%']:+.2f}%",
                            axis=1),
                        hoverinfo="text", showlegend=False,
                    ))
                fig_bar.add_hline(y=0, line=dict(color="#454870", width=1.5, dash="dot"))
                fig_bar.update_layout(
                    title=dict(text="<b>Unrealized P&L</b> <span style='font-size:12px;color:#7a7fa8'>per Holding</span>",
                               font=dict(size=15, color="#f0f2ff", family="Inter"), x=0.0, xanchor="left"),
                    xaxis=dict(title="", tickangle=-30,
                               tickfont=dict(size=11, color="#b8bcd8", family="Inter"),
                               showgrid=False, zeroline=False, linecolor="#252849"),
                    yaxis=dict(title="<b>P&L (₹)</b>", title_font=dict(size=11, color="#7a7fa8"),
                               tickfont=dict(size=10, color="#7a7fa8", family="Inter"),
                               gridcolor="rgba(37,40,73,0.7)", zerolinecolor="#454870",
                               zerolinewidth=1.5, tickprefix="₹", separatethousands=True),
                    plot_bgcolor="#0d0f1e", paper_bgcolor="#131628",
                    font=dict(color="#b8bcd8", family="Inter"),
                    bargap=0.38, showlegend=False, height=390,
                    margin=dict(t=55, b=15, l=15, r=15),
                )
                _dc1, _dc2 = st.columns([3, 2])
                with _dc1:
                    st.plotly_chart(fig_bar, use_container_width=True)
                with _dc2:
                    fig_pie = go.Figure(go.Pie(
                        labels=chart_df["Ticker"], values=chart_df["Value"],
                        hole=0.60,
                        marker=dict(colors=DONUT_COLORS[:len(chart_df)],
                                    line=dict(color="#0d0f1e", width=2.5)),
                        textposition="inside", textinfo="percent",
                        textfont=dict(size=11, color="#ffffff", family="Inter"),
                        hovertemplate="<b>%{label}</b><br>Value: ₹%{value:,.0f}<br>Share: %{percent}<extra></extra>",
                        pull=[0.03 if i == 0 else 0 for i in range(len(chart_df))],
                    ))
                    fig_pie.add_annotation(
                        text=f"<b>₹{total_val/1e5:.1f}L</b><br><span style='font-size:10px;color:#7a7fa8'>Portfolio</span>",
                        x=0.5, y=0.5, showarrow=False,
                        font=dict(size=13, color="#f0f2ff", family="Inter"), align="center",
                    )
                    fig_pie.update_layout(
                        title=dict(text="<b>Value Allocation</b>",
                                   font=dict(size=15, color="#f0f2ff", family="Inter"), x=0.0, xanchor="left"),
                        plot_bgcolor="#0d0f1e", paper_bgcolor="#131628",
                        font=dict(color="#b8bcd8", family="Inter"),
                        legend=dict(font=dict(size=11, color="#b8bcd8", family="Inter"),
                                    orientation="v", bgcolor="rgba(0,0,0,0)", borderwidth=0),
                        height=390, margin=dict(t=55, b=15, l=10, r=10),
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)

        else:
            st.info("No Stock or F&O data available for chart.")
    st.markdown('</div>', unsafe_allow_html=True)

# =====================================================
# ANALYTICS TAB SCREENS — nav-gated
# =====================================================

# TAB 1 — FULL ALLOCATION (redesigned)
if _nav_tab == "Allocation":
    pie_data = calc.dropna(subset=["Value"])

    if not pie_data.empty:
        # ── Derived values ──────────────────────────────────────────────
        _ALLOC_COLORS = [
            "#4f7ef8","#22d67b","#f5c842","#a78bfa","#fb923c",
            "#38bdf8","#f472b6","#34d399","#facc15","#818cf8",
            "#e879f9","#2dd4bf","#fbbf24","#60a5fa","#f87171",
        ]
        alloc_tbl = pie_data[["Ticker","Asset_Type","Value","Cost_Basis","Unrealized_PnL","PnL_%"]].copy()
        alloc_tbl["Allocation %"] = (alloc_tbl["Value"] / alloc_tbl["Value"].sum() * 100).round(2)
        alloc_tbl = alloc_tbl.sort_values("Value", ascending=False).reset_index(drop=True)

        total_val  = alloc_tbl["Value"].sum()
        total_cost = alloc_tbl["Cost_Basis"].sum()
        total_pnl  = alloc_tbl["Unrealized_PnL"].sum()
        total_ret  = (total_pnl / total_cost * 100) if total_cost else 0
        n_pos      = (alloc_tbl["Unrealized_PnL"] >= 0).sum()
        n_neg      = (alloc_tbl["Unrealized_PnL"] < 0).sum()
        top_hold   = alloc_tbl.iloc[0]["Ticker"] if len(alloc_tbl) > 0 else "—"
        top_alloc  = alloc_tbl.iloc[0]["Allocation %"] if len(alloc_tbl) > 0 else 0
        pnl_col    = "#22d67b" if total_pnl >= 0 else "#f85454"
        pnl_arrow  = "▲" if total_pnl >= 0 else "▼"

        # ── Section header ──────────────────────────────────────────────
        st.markdown("""
<style>
/* ── Allocation Tab Redesign ── */
.alloc-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 20px;
}
.alloc-header-title {
    font-size: 18px;
    font-weight: 800;
    color: #f0f2ff;
    letter-spacing: -0.3px;
}
.alloc-header-pill {
    font-size: 10px;
    font-weight: 700;
    color: #4f7ef8;
    background: rgba(79,126,248,0.12);
    border: 1px solid rgba(79,126,248,0.25);
    border-radius: 20px;
    padding: 3px 10px;
    text-transform: uppercase;
    letter-spacing: 0.8px;
}
/* ── Summary KPI strip ── */
.alloc-kpi-strip {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
    margin-bottom: 20px;
}
.alloc-kpi {
    background: #131628;
    border: 1px solid #252849;
    border-radius: 14px;
    padding: 14px 16px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s;
}
.alloc-kpi::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    border-radius: 14px 14px 0 0;
}
.alloc-kpi.blue::before  { background: linear-gradient(90deg,#4f7ef8,#818cf8); }
.alloc-kpi.green::before { background: linear-gradient(90deg,#22d67b,#34d399); }
.alloc-kpi.red::before   { background: linear-gradient(90deg,#f85454,#fb923c); }
.alloc-kpi.gold::before  { background: linear-gradient(90deg,#f5c842,#fb923c); }
.alloc-kpi-label {
    font-size: 9px;
    font-weight: 700;
    color: #454870;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-bottom: 6px;
}
.alloc-kpi-value {
    font-size: 17px;
    font-weight: 800;
    color: #f0f2ff;
    line-height: 1.1;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.alloc-kpi-sub {
    font-size: 10px;
    color: #5a5f88;
    margin-top: 3px;
    font-weight: 500;
}
/* ── Holding rows ── */
.alloc-row {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 13px 16px;
    background: #0e1022;
    border: 1px solid #1a1e38;
    border-radius: 12px;
    margin-bottom: 8px;
    transition: border-color 0.15s, background 0.15s;
    position: relative;
    overflow: hidden;
}
.alloc-row:hover { border-color: #2e3355; background: #111326; }
.alloc-dot {
    width: 10px; height: 10px;
    border-radius: 50%;
    flex-shrink: 0;
}
.alloc-ticker {
    font-size: 14px;
    font-weight: 800;
    color: #f0f2ff;
    min-width: 70px;
}
.alloc-badge {
    font-size: 9px;
    font-weight: 700;
    color: #5a5f88;
    background: #151830;
    border-radius: 5px;
    padding: 2px 7px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    flex-shrink: 0;
}
.alloc-bar-wrap {
    flex: 1;
    height: 5px;
    background: #1a1e38;
    border-radius: 5px;
    overflow: hidden;
    min-width: 40px;
}
.alloc-bar-fill {
    height: 100%;
    border-radius: 5px;
}
.alloc-value {
    font-size: 12px;
    font-weight: 700;
    color: #b8bcd8;
    min-width: 76px;
    text-align: right;
    flex-shrink: 0;
}
.alloc-pct-box {
    font-size: 11px;
    font-weight: 800;
    min-width: 46px;
    text-align: right;
    flex-shrink: 0;
}
.alloc-pnl-box {
    font-size: 11px;
    font-weight: 700;
    min-width: 76px;
    text-align: right;
    flex-shrink: 0;
}
/* ── Legend row labels on desktop ── */
.alloc-col-labels {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 0 16px 6px;
    font-size: 9px;
    font-weight: 700;
    color: #3a3d5c;
    text-transform: uppercase;
    letter-spacing: 0.9px;
}
.alloc-col-labels .lbl-ticker { min-width: 70px; }
.alloc-col-labels .lbl-badge  { min-width: 30px; }
.alloc-col-labels .lbl-bar    { flex: 1; }
.alloc-col-labels .lbl-value  { min-width: 76px; text-align: right; }
.alloc-col-labels .lbl-alloc  { min-width: 46px; text-align: right; }
.alloc-col-labels .lbl-pnl    { min-width: 76px; text-align: right; }

/* ── Mobile overrides ── */
@media (max-width: 600px) {
    .alloc-kpi-strip { grid-template-columns: 1fr 1fr; gap: 8px; margin-bottom: 14px; }
    .alloc-kpi { padding: 12px 13px; border-radius: 12px; }
    .alloc-kpi-value { font-size: 14px; }
    .alloc-row { padding: 10px 12px; gap: 8px; }
    .alloc-ticker { font-size: 13px; min-width: 55px; }
    .alloc-badge { display: none; }
    .alloc-pnl-box { display: none; }
    .alloc-col-labels .lbl-badge { display: none; }
    .alloc-col-labels .lbl-pnl   { display: none; }
    .alloc-value { min-width: 60px; font-size: 11px; }
    .alloc-pct-box { min-width: 38px; font-size: 11px; }
    .alloc-header-title { font-size: 16px; }
}
@media (max-width: 400px) {
    .alloc-kpi-value { font-size: 13px; }
    .alloc-kpi-label { font-size: 8px; }
    .alloc-value { min-width: 52px; }
}
</style>

<div class="alloc-header">
  <span class="alloc-header-title">📊 Portfolio Allocation</span>
  <span class="alloc-header-pill">Live Holdings</span>
</div>
""", unsafe_allow_html=True)

        # ── KPI strip ───────────────────────────────────────────────────
        _kpi_pnl_sign = "+" if total_pnl >= 0 else ""
        st.markdown(f"""
<div class="alloc-kpi-strip">
  <div class="alloc-kpi blue">
    <div class="alloc-kpi-label">Portfolio Value</div>
    <div class="alloc-kpi-value">₹{total_val/1e5:.2f}L</div>
    <div class="alloc-kpi-sub">{len(alloc_tbl)} holdings</div>
  </div>
  <div class="alloc-kpi {'green' if total_pnl >= 0 else 'red'}">
    <div class="alloc-kpi-label">Unrealized P&amp;L</div>
    <div class="alloc-kpi-value" style="color:{pnl_col};">
      {pnl_arrow} ₹{abs(total_pnl)/1e5:.2f}L
    </div>
    <div class="alloc-kpi-sub" style="color:{pnl_col};opacity:0.7;">{_kpi_pnl_sign}{total_ret:.2f}% overall</div>
  </div>
  <div class="alloc-kpi green">
    <div class="alloc-kpi-label">Winners / Losers</div>
    <div class="alloc-kpi-value">
      <span style="color:#22d67b;">{n_pos}↑</span>
      <span style="color:#3a3d5c;font-size:13px;font-weight:500;"> / </span>
      <span style="color:#f85454;">{n_neg}↓</span>
    </div>
    <div class="alloc-kpi-sub">out of {len(alloc_tbl)} stocks</div>
  </div>
  <div class="alloc-kpi gold">
    <div class="alloc-kpi-label">Top Holding</div>
    <div class="alloc-kpi-value">{top_hold}</div>
    <div class="alloc-kpi-sub">{top_alloc:.1f}% of portfolio</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Donut chart (full width on mobile, 40% on desktop) + rows ──
        if _IS_MOBILE:
            # Mobile: donut on top, rows below
            fig_alloc_m = go.Figure(go.Pie(
                labels=alloc_tbl["Ticker"], values=alloc_tbl["Value"],
                hole=0.58,
                marker=dict(
                    colors=_ALLOC_COLORS[:len(alloc_tbl)],
                    line=dict(color="#0b0d17", width=2.5)
                ),
                textposition="inside", textinfo="percent",
                textfont=dict(size=10, color="#ffffff", family="Inter"),
                hovertemplate="<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>",
                sort=False,
            ))
            fig_alloc_m.add_annotation(
                text=f"<b>₹{total_val/1e5:.1f}L</b><br><span style='font-size:9px;color:#7a7fa8'>Total</span>",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=12, color="#f0f2ff", family="Inter"), align="center",
            )
            fig_alloc_m.update_layout(
                plot_bgcolor="#0b0d17", paper_bgcolor="#0b0d17",
                font=dict(color="#b8bcd8", family="Inter"),
                legend=dict(
                    orientation="h", font=dict(size=10, color="#b8bcd8"),
                    y=-0.08, x=0.5, xanchor="center",
                    bgcolor="rgba(0,0,0,0)", borderwidth=0,
                ),
                margin=dict(t=10, b=10, l=10, r=10), height=260,
            )
            st.plotly_chart(fig_alloc_m, use_container_width=True)

            # Mobile holding rows
            _mob_rows = ""
            for i, row in alloc_tbl.iterrows():
                _clr  = _ALLOC_COLORS[i % len(_ALLOC_COLORS)]
                _pv   = float(row["Unrealized_PnL"])
                _pp   = float(row["PnL_%"])
                _vv   = float(row["Value"])
                _alc  = float(row["Allocation %"])
                _pclr = "#22d67b" if _pv >= 0 else "#f85454"
                _arr  = "▲" if _pv > 0 else "▼"
                _bar  = _alc / (alloc_tbl["Allocation %"].max() or 1) * 100
                _atp  = str(row["Asset_Type"])
                _bdg  = {"Stock":"STK","F&O":"F&O","ETF":"ETF"}.get(_atp, _atp[:3])
                _mob_rows += f"""
<div class="alloc-row">
  <div class="alloc-dot" style="background:{_clr};box-shadow:0 0 6px {_clr}66;"></div>
  <span class="alloc-ticker">{row['Ticker']}</span>
  <div class="alloc-bar-wrap">
    <div class="alloc-bar-fill" style="width:{_bar:.1f}%;background:{_clr};"></div>
  </div>
  <span class="alloc-value">₹{_vv/1e5:.2f}L</span>
  <span class="alloc-pct-box" style="color:{_clr};">{_alc:.1f}%</span>
  <span class="alloc-pnl-box" style="color:{_pclr};">{_arr} {_pp:+.1f}%</span>
</div>"""
            st.markdown(f"""
<div style="margin-top:4px;">
  <div class="alloc-col-labels">
    <span class="lbl-ticker">Holding</span>
    <span class="lbl-bar" style="flex:1;padding-left:4px;">Weight</span>
    <span class="lbl-value">Value</span>
    <span class="lbl-alloc">Alloc</span>
    <span class="lbl-pnl">P&amp;L</span>
  </div>
  {_mob_rows}
</div>""", unsafe_allow_html=True)

        else:
            # Desktop: donut left + rows right in columns
            _dc1, _dc2 = st.columns([4, 5])

            with _dc1:
                fig_alloc_d = go.Figure(go.Pie(
                    labels=alloc_tbl["Ticker"], values=alloc_tbl["Value"],
                    hole=0.60,
                    marker=dict(
                        colors=_ALLOC_COLORS[:len(alloc_tbl)],
                        line=dict(color="#0b0d17", width=2.5)
                    ),
                    textposition="inside", textinfo="percent",
                    textfont=dict(size=11, color="#ffffff", family="Inter"),
                    hovertemplate="<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>",
                    pull=[0.04 if i == 0 else 0 for i in range(len(alloc_tbl))],
                    sort=False,
                ))
                fig_alloc_d.add_annotation(
                    text=f"<b>₹{total_val/1e5:.1f}L</b><br><span style='font-size:10px;color:#7a7fa8'>Portfolio</span>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=14, color="#f0f2ff", family="Inter"), align="center",
                )
                fig_alloc_d.update_layout(
                    plot_bgcolor="#131628", paper_bgcolor="#131628",
                    font=dict(color="#b8bcd8", family="Inter"),
                    legend=dict(
                        orientation="h", font=dict(size=11, color="#b8bcd8"),
                        bgcolor="rgba(0,0,0,0)", borderwidth=0,
                        x=0.5, xanchor="center", y=-0.06,
                    ),
                    margin=dict(t=10, b=20, l=10, r=10), height=380,
                )
                st.plotly_chart(fig_alloc_d, use_container_width=True)

            with _dc2:
                _desk_rows = ""
                for i, row in alloc_tbl.iterrows():
                    _clr  = _ALLOC_COLORS[i % len(_ALLOC_COLORS)]
                    _pv   = float(row["Unrealized_PnL"])
                    _pp   = float(row["PnL_%"])
                    _vv   = float(row["Value"])
                    _alc  = float(row["Allocation %"])
                    _pclr = "#22d67b" if _pv >= 0 else "#f85454"
                    _arr  = "▲" if _pv > 0 else "▼"
                    _bar  = _alc / (alloc_tbl["Allocation %"].max() or 1) * 100
                    _atp  = str(row["Asset_Type"])
                    _bdg  = {"Stock":"STK","F&O":"F&O","ETF":"ETF"}.get(_atp, _atp[:3])
                    _desk_rows += f"""
<div class="alloc-row">
  <div class="alloc-dot" style="background:{_clr};box-shadow:0 0 6px {_clr}66;"></div>
  <span class="alloc-ticker">{row['Ticker']}</span>
  <span class="alloc-badge">{_bdg}</span>
  <div class="alloc-bar-wrap">
    <div class="alloc-bar-fill" style="width:{_bar:.1f}%;background:{_clr};"></div>
  </div>
  <span class="alloc-value">₹{_vv/1e5:.2f}L</span>
  <span class="alloc-pct-box" style="color:{_clr};">{_alc:.2f}%</span>
  <span class="alloc-pnl-box" style="color:{_pclr};">{_arr} {abs(_pv)/1e3:.1f}k ({_pp:+.1f}%)</span>
</div>"""

                st.markdown(f"""
<div style="padding-top:4px;">
  <div class="alloc-col-labels">
    <span class="lbl-ticker">Holding</span>
    <span class="lbl-badge"></span>
    <span class="lbl-bar" style="padding-left:4px;">Weight</span>
    <span class="lbl-value">Mkt Value</span>
    <span class="lbl-alloc">Alloc %</span>
    <span class="lbl-pnl">Unrealized P&amp;L</span>
  </div>
  {_desk_rows}
</div>""", unsafe_allow_html=True)

    else:
        st.info("No holdings data available for allocation view.")

# =====================================================
# TAB 2 — GAINERS / LOSERS
# =====================================================

if _nav_tab == "Gainers / Losers":

    # ── CSS for the entire Gainers/Losers tab ───────────────────────
    st.markdown("""
<style>
/* ══════════════════════════════════════════════════════
   GAINERS / LOSERS TAB — Full Redesign
   ══════════════════════════════════════════════════════ */

/* ── Section header ── */
.gl-section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 24px 0 14px;
}
.gl-section-title {
    font-size: 16px;
    font-weight: 800;
    color: #f0f2ff;
    letter-spacing: -0.2px;
}
.gl-section-pill {
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 20px;
}
.gl-section-pill.green { color:#22d67b; background:rgba(34,214,123,0.10); border:1px solid rgba(34,214,123,0.22); }
.gl-section-pill.red   { color:#f85454; background:rgba(248,84,84,0.10);  border:1px solid rgba(248,84,84,0.22);  }
.gl-section-pill.blue  { color:#4f7ef8; background:rgba(79,126,248,0.10); border:1px solid rgba(79,126,248,0.22); }
.gl-section-pill.gold  { color:#f5c842; background:rgba(245,200,66,0.10); border:1px solid rgba(245,200,66,0.22); }

/* ── Mover card (individual stock row) ── */
.gl-mover-card {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 13px 16px;
    border-radius: 12px;
    border: 1px solid #1a1e38;
    background: #0d0f1e;
    margin-bottom: 7px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.15s, transform 0.12s;
    cursor: default;
}
.gl-mover-card:hover { border-color: #2e3355; transform: translateX(2px); }

/* Ambient glow strip on left edge */
.gl-mover-card::before {
    content: '';
    position: absolute;
    left: 0; top: 0; bottom: 0;
    width: 3px;
    border-radius: 12px 0 0 12px;
}
.gl-mover-card.gain::before { background: linear-gradient(180deg, #22d67b, #34d39988); }
.gl-mover-card.loss::before { background: linear-gradient(180deg, #f85454, #fb923c88); }

/* Rank badge */
.gl-rank {
    font-size: 11px;
    font-weight: 800;
    color: #3a3d5c;
    min-width: 22px;
    text-align: center;
    flex-shrink: 0;
}
/* Ticker + type */
.gl-ticker-wrap { flex: 1; min-width: 0; }
.gl-ticker {
    font-size: 14px;
    font-weight: 800;
    color: #f0f2ff;
    white-space: nowrap;
}
.gl-type-badge {
    display: inline-block;
    font-size: 8px;
    font-weight: 700;
    letter-spacing: 0.6px;
    text-transform: uppercase;
    padding: 2px 6px;
    border-radius: 4px;
    background: #151830;
    color: #5a5f88;
    margin-top: 2px;
}
/* Heat bar */
.gl-heat-wrap {
    width: 70px;
    flex-shrink: 0;
}
.gl-heat-track {
    height: 4px;
    background: #1a1e38;
    border-radius: 4px;
    overflow: hidden;
    margin-bottom: 3px;
}
.gl-heat-fill {
    height: 100%;
    border-radius: 4px;
}
/* Values */
.gl-pct {
    font-size: 15px;
    font-weight: 800;
    min-width: 62px;
    text-align: right;
    flex-shrink: 0;
    line-height: 1.1;
}
.gl-inr {
    font-size: 11px;
    font-weight: 600;
    color: #5a5f88;
    min-width: 68px;
    text-align: right;
    flex-shrink: 0;
}

/* ── Split columns wrapper for gainers + losers ── */
.gl-split {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 14px;
    margin-bottom: 4px;
}
.gl-panel {
    background: #0e1022;
    border: 1px solid #1a1e38;
    border-radius: 16px;
    padding: 16px 14px;
}
.gl-panel-header {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 12px;
    padding-bottom: 10px;
    border-bottom: 1px solid #151830;
}
.gl-panel-title {
    font-size: 12px;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: 0.8px;
}
.gl-panel-title.green { color: #22d67b; }
.gl-panel-title.red   { color: #f85454; }
.gl-panel-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
}
.gl-panel-count {
    margin-left: auto;
    font-size: 10px;
    color: #3a3d5c;
    font-weight: 600;
}

/* ── Star performer cards (best/worst) ── */
.gl-star-grid {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr 1fr;
    gap: 12px;
    margin-bottom: 6px;
}
.gl-star-card {
    background: #0e1022;
    border: 1px solid #1a1e38;
    border-radius: 16px;
    padding: 20px 16px;
    text-align: center;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s, transform 0.15s;
}
.gl-star-card:hover { transform: translateY(-2px); }
.gl-star-card::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 60px;
    border-radius: 0 0 16px 16px;
    pointer-events: none;
}
.gl-star-card.best-overall  { border-top: 2px solid #22d67b; }
.gl-star-card.worst-overall { border-top: 2px solid #f85454; }
.gl-star-card.best-daily    { border-top: 2px solid #4f7ef8; }
.gl-star-card.worst-daily   { border-top: 2px solid #f5c842; }
.gl-star-card.best-overall::after  { background: radial-gradient(ellipse at bottom,rgba(34,214,123,0.06),transparent 70%); }
.gl-star-card.worst-overall::after { background: radial-gradient(ellipse at bottom,rgba(248,84,84,0.06),transparent 70%); }
.gl-star-card.best-daily::after    { background: radial-gradient(ellipse at bottom,rgba(79,126,248,0.06),transparent 70%); }
.gl-star-card.worst-daily::after   { background: radial-gradient(ellipse at bottom,rgba(245,200,66,0.06),transparent 70%); }
.gl-star-label {
    font-size: 9px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #454870;
    margin-bottom: 8px;
}
.gl-star-ticker {
    font-size: 22px;
    font-weight: 900;
    color: #f0f2ff;
    letter-spacing: -0.5px;
    margin-bottom: 6px;
}
.gl-star-pct {
    font-size: 20px;
    font-weight: 800;
    line-height: 1;
    margin-bottom: 6px;
}
.gl-star-sub {
    font-size: 10px;
    color: #454870;
    font-weight: 500;
}

/* ── Sector rows ── */
.gl-sector-row {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 11px 14px;
    background: #0d0f1e;
    border: 1px solid #1a1e38;
    border-radius: 10px;
    margin-bottom: 6px;
    transition: border-color 0.15s;
}
.gl-sector-row:hover { border-color: #2e3355; }
.gl-sector-dot { width:8px;height:8px;border-radius:50%;flex-shrink:0; }
.gl-sector-name { font-size:13px;font-weight:700;color:#e0e2f0;flex:1;min-width:0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis; }
.gl-sector-count { font-size:10px;color:#454870;font-weight:600;flex-shrink:0; }
.gl-sector-bar-wrap { width:80px;flex-shrink:0; }
.gl-sector-bar-track { height:4px;background:#1a1e38;border-radius:4px;overflow:hidden; }
.gl-sector-bar-fill  { height:100%;border-radius:4px; }
.gl-sector-alloc { font-size:12px;font-weight:800;color:#b8bcd8;min-width:40px;text-align:right;flex-shrink:0; }
.gl-sector-pnl   { font-size:11px;font-weight:700;min-width:72px;text-align:right;flex-shrink:0; }

/* ── Divider ── */
.gl-divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, #252849 20%, #252849 80%, transparent);
    margin: 24px 0;
}

/* ══ MOBILE ══ */
@media (max-width: 600px) {
    .gl-chart-wrap { display: none !important; }
    .gl-split { grid-template-columns: 1fr; gap: 10px; }
    .gl-star-grid { grid-template-columns: 1fr 1fr; gap: 8px; }
    .gl-star-card { padding: 15px 12px; border-radius: 14px; }
    .gl-star-ticker { font-size: 18px; }
    .gl-star-pct    { font-size: 16px; }
    .gl-mover-card  { padding: 10px 12px; gap: 9px; }
    .gl-ticker      { font-size: 13px; }
    .gl-pct         { font-size: 13px; min-width: 52px; }
    .gl-inr         { display: none; }
    .gl-heat-wrap   { display: none; }
    .gl-sector-bar-wrap { display: none; }
    .gl-sector-name { font-size: 12px; }
    .gl-section-title { font-size: 14px; }
    .gl-panel { padding: 12px 10px; }
}
@media (max-width: 400px) {
    .gl-star-grid { grid-template-columns: 1fr 1fr; }
    .gl-star-ticker { font-size: 15px; }
    .gl-star-pct    { font-size: 14px; }
}
</style>
""", unsafe_allow_html=True)

    # ── Sector map (NSE sector classification) ──────────────────────
    SECTOR_MAP = {
        # ── Pharma ──────────────────────────────────────────────────────
        "DRREDDY": "Pharma", "SUNPHARMA": "Pharma", "CIPLA": "Pharma",
        "DIVISLAB": "Pharma", "AUROPHARMA": "Pharma", "LUPIN": "Pharma",
        "BIOCON": "Pharma", "ALKEM": "Pharma", "TORNTPHARM": "Pharma",
        "LINCOLNPHARMA": "Pharma", "LINCOLN": "Pharma",        # Lincoln Pharma
        "OCCLLTD": "Pharma", "OCCLLTDNS": "Pharma",           # OCCL Ltd (ophthalmic/pharma)
        # ── IT ──────────────────────────────────────────────────────────
        "TCS": "IT", "INFY": "IT", "WIPRO": "IT", "HCLTECH": "IT",
        "TECHM": "IT", "LTIM": "IT", "MPHASIS": "IT", "PERSISTENT": "IT",
        "COFORGE": "IT", "OFSS": "IT",
        # ── Banking & Finance ────────────────────────────────────────────
        "HDFCBANK": "Banking", "ICICIBANK": "Banking", "SBIN": "Banking",
        "KOTAKBANK": "Banking", "AXISBANK": "Banking", "INDUSINDBK": "Banking",
        "BANDHANBNK": "Banking", "FEDERALBNK": "Banking", "IDFCFIRSTB": "Banking",
        "BAJFINANCE": "NBFC", "BAJAJFINSV": "NBFC", "CHOLAFIN": "NBFC",
        "MUTHOOTFIN": "NBFC", "MANAPPURAM": "NBFC",
        "GSFC": "Fertilizers & Chemicals",                     # Gujarat State Fertilizers & Chemicals
        # ── Insurance ───────────────────────────────────────────────────
        "HDFCLIFE": "Insurance", "SBILIFE": "Insurance", "ICICIPRULI": "Insurance",
        "LICI": "Insurance", "MAXFINSERV": "Insurance", "NIACL": "Insurance",
        "STARHEALTH": "Insurance",
        # ── Auto ────────────────────────────────────────────────────────
        "MARUTI": "Auto", "TATAMOTORS": "Auto", "M&M": "Auto",
        "BAJAJ-AUTO": "Auto", "HEROMOTOCO": "Auto", "EICHERMOT": "Auto",
        "ASHOKLEY": "Auto", "TVSMOTOR": "Auto",
        # ── Energy & Oil ────────────────────────────────────────────────
        "RELIANCE": "Energy", "ONGC": "Energy", "BPCL": "Energy",
        "IOC": "Energy", "HINDPETRO": "Energy", "GAIL": "Energy",
        "POWERGRID": "Energy", "NTPC": "Energy", "TATAPOWER": "Energy",
        # ── Metals & Mining ─────────────────────────────────────────────
        "TATASTEEL": "Metals", "JSWSTEEL": "Metals", "HINDALCO": "Metals",
        "VEDL": "Metals", "SAIL": "Metals", "COALINDIA": "Metals",
        "NMDC": "Metals", "NATIONALUM": "Metals",
        "EXCELINDUS": "Metals",                                # Excel Industries (specialty chemicals/metal salts)
        "EXCELINDUSNS": "Metals",
        # ── FMCG ────────────────────────────────────────────────────────
        "HINDUNILVR": "FMCG", "ITC": "FMCG", "NESTLEIND": "FMCG",
        "DABUR": "FMCG", "MARICO": "FMCG", "GODREJCP": "FMCG",
        "BRITANNIA": "FMCG", "TATACONSUM": "FMCG",
        # ── Capital Goods / Industrials ──────────────────────────────────
        "ABB": "Capital Goods", "SIEMENS": "Capital Goods", "BHEL": "Capital Goods",
        "HAVELLS": "Capital Goods", "CUMMINSIND": "Capital Goods",
        "THERMAX": "Capital Goods", "INTERARCH": "Capital Goods",
        "VOLTAMP": "Capital Goods", "GVKPIL": "Infra",
        "HEXT": "Capital Goods",                               # Hext Technologies (industrial/engineering)
        "HEXTNS": "Capital Goods",
        # ── Cement ──────────────────────────────────────────────────────
        "ULTRACEMCO": "Cement", "GRASIM": "Cement", "SHREECEM": "Cement",
        "AMBUJACEM": "Cement", "ACC": "Cement",
        # ── Telecom ─────────────────────────────────────────────────────
        "BHARTIARTL": "Telecom", "IDEA": "Telecom",
        # ── Consumer / Retail ────────────────────────────────────────────
        "TITAN": "Consumer", "TRENT": "Consumer", "DMART": "Consumer",
        "NYKAA": "Consumer", "ZOMATO": "Consumer",
        "KALAMANDIR": "Consumer",                              # Kalamandir (fashion retail)
        "KALAMANDIR1": "Consumer", "KALAMANDIRNS": "Consumer",
        "KAMATHOTEL": "Hotels & Hospitality",                  # Kamat Hotels
        "KAMATHOTELNS": "Hotels & Hospitality",
        # ── Realty ──────────────────────────────────────────────────────
        "DLF": "Realty", "GODREJPROP": "Realty", "PRESTIGE": "Realty",
        # ── Chemicals / Fertilizers ──────────────────────────────────────
        "PIDILITIND": "Chemicals", "ATUL": "Chemicals", "DEEPAKNITR": "Chemicals",
        # ── Diversified / Conglomerates ──────────────────────────────────
        "JAYBARMARU": "Diversified",                           # Jay Bharat Maruti (auto ancillary/diversified)
        "JAYBARMARUNSNS": "Diversified",
        # ── F&O Index ───────────────────────────────────────────────────
        "NIFTY": "Index F&O", "BANKNIFTY": "Index F&O", "FINNIFTY": "Index F&O",
    }

    def get_sector(ticker):
        t = ticker.upper().replace(".NS","").replace(".BO","")
        for suf in ["26JUN25FUT","26MAY25FUT","26MAY26FUT","FUT","CE","PE"]:
            t = t.replace(suf, "")
        t = t.strip()

        # ── 0. ETF short-circuit ─────────────────────────────────────
        # NSE classifies ALL ETFs as "Financial Services" — intercept here
        if classify_asset(ticker) == "ETF":
            if any(k in t for k in ["GOLD","SILVER","BSLGOLDETF","AXISGOLD",
                                     "HDFCGOLD","ICICIGOLD","NIPGOLD","SBIGOLD"]):
                return "ETF – Gold/Silver"
            if any(k in t for k in ["LIQUID","OVERNIGHT","GSEC","GILT","DEBT","BOND","TBILL"]):
                return "ETF – Liquid/Debt"
            if "BANK" in t:   return "ETF – Banking"
            if "IT" in t:     return "ETF – IT"
            if "PSU" in t:    return "ETF – PSU"
            if "MIDCAP" in t: return "ETF – Midcap"
            if "INFRA" in t:  return "ETF – Infra"
            if "PHARMA" in t: return "ETF – Pharma"
            return "ETF – Index"

        # ── 1. Static map (instant, no network) ─────────────────────
        if t in SECTOR_MAP:
            return SECTOR_MAP[t]
        for key, sec in SECTOR_MAP.items():
            if t.startswith(key) or key.startswith(t):
                return sec

        # ── 2. Live NSE lookup (cached 24 h per symbol) ─────────────
        live_sec = _fetch_sector_nse(t)
        if live_sec:
            return live_sec

        return "Other"

    # ── Prepare working dataframe ────────────────────────────────────
    gl = calc.copy()
    gl["Sector"] = gl["Ticker"].apply(get_sector)
    gl["Daily_Change_%"] = pd.to_numeric(gl["Daily_PnL_%"], errors="coerce")
    gl["Daily_PnL_₹"]   = pd.to_numeric(gl["Daily_PnL"],   errors="coerce")
    gl["Overall_%"]     = pd.to_numeric(gl["PnL_%"],        errors="coerce")
    gl["Overall_PnL_₹"] = pd.to_numeric(gl["Unrealized_PnL"], errors="coerce")

    daily_valid   = gl.dropna(subset=["Daily_Change_%"])
    overall_valid = gl.dropna(subset=["Overall_%"])

    # ── helper: render mover card rows ──────────────────────────────
    def _mover_rows(df_rows, pct_col, inr_col, max_abs):
        html = ""
        for rank, (_, r) in enumerate(df_rows.iterrows(), 1):
            pct  = float(r[pct_col])
            inr  = float(r[inr_col]) if pd.notna(r.get(inr_col)) else 0
            clr  = "#22d67b" if pct >= 0 else "#f85454"
            cls  = "gain" if pct >= 0 else "loss"
            arr  = "▲" if pct > 0 else "▼"
            bar  = abs(pct) / (max_abs or 1) * 100
            atp  = str(r.get("Asset_Type",""))
            bdg  = {"Stock":"STK","F&O":"F&O","ETF":"ETF"}.get(atp, atp[:3])
            html += f"""
<div class="gl-mover-card {cls}">
  <span class="gl-rank">#{rank}</span>
  <div class="gl-ticker-wrap">
    <div class="gl-ticker">{r['Ticker']}</div>
    <span class="gl-type-badge">{bdg}</span>
  </div>
  <div class="gl-heat-wrap">
    <div class="gl-heat-track">
      <div class="gl-heat-fill" style="width:{bar:.1f}%;background:{clr};"></div>
    </div>
  </div>
  <span class="gl-pct" style="color:{clr};">{arr} {abs(pct):.2f}%</span>
  <span class="gl-inr" style="color:{clr};">₹{abs(inr):,.0f}</span>
</div>"""
        return html

    # ══ SECTION 1 — Today's Movers ══════════════════════════════════
    st.markdown("""
<div class="gl-section-header">
  <span class="gl-section-title">📅 Today's Movers</span>
  <span class="gl-section-pill blue">Live Session</span>
</div>""", unsafe_allow_html=True)

    if daily_valid.empty:
        st.info("Daily price data unavailable — market may be closed or prices not yet fetched.")
    else:
        dg = daily_valid[daily_valid["Daily_Change_%"] >= 0].nlargest(5, "Daily_Change_%")
        dl = daily_valid[daily_valid["Daily_Change_%"] < 0].nsmallest(5, "Daily_Change_%")
        d_max = daily_valid["Daily_Change_%"].abs().max() or 1

        g_rows = _mover_rows(dg, "Daily_Change_%", "Daily_PnL_₹", d_max) if not dg.empty else \
            '<div style="color:#454870;font-size:12px;padding:10px 0;">No gainers today</div>'
        l_rows = _mover_rows(dl, "Daily_Change_%", "Daily_PnL_₹", d_max) if not dl.empty else \
            '<div style="color:#454870;font-size:12px;padding:10px 0;">No losers today</div>'

        st.markdown(f"""
<div class="gl-split">
  <div class="gl-panel">
    <div class="gl-panel-header">
      <div class="gl-panel-dot" style="background:#22d67b;box-shadow:0 0 6px #22d67b88;"></div>
      <span class="gl-panel-title green">Gainers</span>
      <span class="gl-panel-count">{len(dg)} stocks</span>
    </div>
    {g_rows}
  </div>
  <div class="gl-panel">
    <div class="gl-panel-header">
      <div class="gl-panel-dot" style="background:#f85454;box-shadow:0 0 6px #f8545488;"></div>
      <span class="gl-panel-title red">Losers</span>
      <span class="gl-panel-count">{len(dl)} stocks</span>
    </div>
    {l_rows}
  </div>
</div>""", unsafe_allow_html=True)

        # Daily waterfall bar chart
        daily_chart = daily_valid.sort_values("Daily_Change_%")
        fig_daily = go.Figure(go.Bar(
            x=daily_chart["Ticker"],
            y=daily_chart["Daily_Change_%"],
            marker=dict(
                color=daily_chart["Daily_Change_%"].apply(lambda x: "#22d67b" if x >= 0 else "#f85454"),
                opacity=0.9,
                line=dict(
                    color=daily_chart["Daily_Change_%"].apply(lambda x: "#00e676" if x >= 0 else "#ff1744"),
                    width=1.2
                )
            ),
            text=daily_chart["Daily_Change_%"].apply(lambda x: f"{x:+.2f}%"),
            textfont=dict(size=10, color="#ffffff", family="Inter"),
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>Daily: %{y:+.2f}%<extra></extra>"
        ))
        fig_daily.add_hline(y=0, line=dict(color="#454870", width=1.5, dash="dot"))
        fig_daily.update_layout(
            title=dict(
                text="<b>Today's % Change</b> <span style='font-size:11px;color:#7a7fa8'>— All Holdings</span>",
                font=dict(size=14, color="#f0f2ff", family="Inter"), x=0, xanchor="left"
            ),
            xaxis=dict(title="", tickangle=-30, tickfont=dict(size=11, color="#b8bcd8", family="Inter"),
                       showgrid=False, zeroline=False, linecolor="#252849"),
            yaxis=dict(title="<b>Change %</b>", title_font=dict(size=11, color="#7a7fa8"),
                       tickfont=dict(size=10, color="#7a7fa8", family="Inter"),
                       gridcolor="rgba(37,40,73,0.6)", zerolinecolor="#454870", zerolinewidth=1.5,
                       ticksuffix="%"),
            plot_bgcolor="#0d0f1e", paper_bgcolor="#0e1022",
            font=dict(color="#b8bcd8", family="Inter"),
            bargap=0.35, showlegend=False, height=340,
            margin=dict(t=50, b=15, l=15, r=15),
        )
        if not _IS_MOBILE:
            st.plotly_chart(fig_daily, use_container_width=True)

    st.markdown('<div class="gl-divider"></div>', unsafe_allow_html=True)

    # ══ SECTION 2 — Overall Performance ═════════════════════════════
    st.markdown("""
<div class="gl-section-header">
  <span class="gl-section-title">📊 Overall Performance</span>
  <span class="gl-section-pill gold">Since Buy</span>
</div>""", unsafe_allow_html=True)

    if not overall_valid.empty:
        og = overall_valid[overall_valid["Overall_%"] >= 0].nlargest(5, "Overall_%")
        ol = overall_valid[overall_valid["Overall_%"] < 0].nsmallest(5, "Overall_%")
        o_max = overall_valid["Overall_%"].abs().max() or 1

        og_rows = _mover_rows(og, "Overall_%", "Overall_PnL_₹", o_max) if not og.empty else \
            '<div style="color:#454870;font-size:12px;padding:10px 0;">No overall gainers</div>'
        ol_rows = _mover_rows(ol, "Overall_%", "Overall_PnL_₹", o_max) if not ol.empty else \
            '<div style="color:#454870;font-size:12px;padding:10px 0;">No overall losers</div>'

        st.markdown(f"""
<div class="gl-split">
  <div class="gl-panel">
    <div class="gl-panel-header">
      <div class="gl-panel-dot" style="background:#22d67b;box-shadow:0 0 6px #22d67b88;"></div>
      <span class="gl-panel-title green">Overall Gainers</span>
      <span class="gl-panel-count">top {len(og)}</span>
    </div>
    {og_rows}
  </div>
  <div class="gl-panel">
    <div class="gl-panel-header">
      <div class="gl-panel-dot" style="background:#f85454;box-shadow:0 0 6px #f8545488;"></div>
      <span class="gl-panel-title red">Overall Losers</span>
      <span class="gl-panel-count">top {len(ol)}</span>
    </div>
    {ol_rows}
  </div>
</div>""", unsafe_allow_html=True)
    else:
        st.info("No overall P&L data available.")

    st.markdown('<div class="gl-divider"></div>', unsafe_allow_html=True)

    # ══ SECTION 3 — Star Performer Cards ════════════════════════════
    st.markdown("""
<div class="gl-section-header">
  <span class="gl-section-title">🎖️ Star Performers</span>
  <span class="gl-section-pill blue">Highlights</span>
</div>""", unsafe_allow_html=True)

    if not overall_valid.empty:
        best  = overall_valid.loc[overall_valid["Overall_%"].idxmax()]
        worst = overall_valid.loc[overall_valid["Overall_%"].idxmin()]
        bc = "#22d67b" if best["Overall_%"]  >= 0 else "#f85454"
        wc = "#22d67b" if worst["Overall_%"] >= 0 else "#f85454"

        bd_html = wd_html = ""
        if not daily_valid.empty:
            best_d  = daily_valid.loc[daily_valid["Daily_Change_%"].idxmax()]
            worst_d = daily_valid.loc[daily_valid["Daily_Change_%"].idxmin()]
            bdc = "#22d67b" if best_d["Daily_Change_%"]  >= 0 else "#f85454"
            wdc = "#22d67b" if worst_d["Daily_Change_%"] >= 0 else "#f85454"
            bd_html = f"""
  <div class="gl-star-card best-daily">
    <div class="gl-star-label">📅 Best Today</div>
    <div class="gl-star-ticker">{best_d['Ticker']}</div>
    <div class="gl-star-pct" style="color:{bdc};">{best_d['Daily_Change_%']:+.2f}%</div>
    <div class="gl-star-sub">{safe_inr(best_d['Daily_PnL_₹'])} daily</div>
  </div>"""
            wd_html = f"""
  <div class="gl-star-card worst-daily">
    <div class="gl-star-label">📅 Worst Today</div>
    <div class="gl-star-ticker">{worst_d['Ticker']}</div>
    <div class="gl-star-pct" style="color:{wdc};">{worst_d['Daily_Change_%']:+.2f}%</div>
    <div class="gl-star-sub">{safe_inr(worst_d['Daily_PnL_₹'])} daily</div>
  </div>"""

        st.markdown(f"""
<div class="gl-star-grid">
  <div class="gl-star-card best-overall">
    <div class="gl-star-label">🏆 Best Overall</div>
    <div class="gl-star-ticker">{best['Ticker']}</div>
    <div class="gl-star-pct" style="color:{bc};">{best['Overall_%']:+.2f}%</div>
    <div class="gl-star-sub">{safe_inr(best['Overall_PnL_₹'])} unrealized</div>
  </div>
  <div class="gl-star-card worst-overall">
    <div class="gl-star-label">📉 Worst Overall</div>
    <div class="gl-star-ticker">{worst['Ticker']}</div>
    <div class="gl-star-pct" style="color:{wc};">{worst['Overall_%']:+.2f}%</div>
    <div class="gl-star-sub">{safe_inr(worst['Overall_PnL_₹'])} unrealized</div>
  </div>
  {bd_html}
  {wd_html}
</div>""", unsafe_allow_html=True)

    st.markdown('<div class="gl-divider"></div>', unsafe_allow_html=True)

    # ══ SECTION 4 — Sector Diversification ══════════════════════════
    st.markdown("""
<div class="gl-section-header">
  <span class="gl-section-title">🗂️ Sector Diversification</span>
  <span class="gl-section-pill blue">Breakdown</span>
</div>""", unsafe_allow_html=True)

    sec_data = gl.dropna(subset=["Value"]).copy()
    sec_data["Value"] = pd.to_numeric(sec_data["Value"], errors="coerce")
    sec_group = sec_data.groupby("Sector").agg(
        Value=("Value","sum"),
        PnL=("Overall_PnL_₹","sum"),
        Count=("Ticker","count")
    ).reset_index()
    sec_group["Allocation_%"] = (sec_group["Value"] / sec_group["Value"].sum() * 100).round(2)
    sec_group = sec_group.sort_values("Value", ascending=False).reset_index(drop=True)

    _SEC_COLORS = ["#4f7ef8","#22d67b","#f5c842","#a78bfa","#fb923c",
                   "#38bdf8","#f472b6","#34d399","#facc15","#818cf8",
                   "#e879f9","#2dd4bf","#fbbf24","#60a5fa","#f87171"]

    # Donut + bar chart side by side (hidden on mobile via CSS)
    # Build both figures first
    fig_sec = go.Figure(go.Pie(
        labels=sec_group["Sector"], values=sec_group["Value"],
        hole=0.58,
        marker=dict(colors=_SEC_COLORS[:len(sec_group)],
                    line=dict(color="#0b0d17", width=2.5)),
        textposition="inside", textinfo="percent",
        textfont=dict(size=10, color="#ffffff", family="Inter"),
        hovertemplate="<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>",
        sort=False,
    ))
    fig_sec.add_annotation(
        text=f"<b>{len(sec_group)}</b><br><span style='font-size:9px;color:#7a7fa8'>Sectors</span>",
        x=0.5, y=0.5, showarrow=False,
        font=dict(size=13, color="#f0f2ff", family="Inter"), align="center",
    )
    fig_sec.update_layout(
        plot_bgcolor="#0e1022", paper_bgcolor="#0e1022",
        font=dict(color="#b8bcd8", family="Inter"),
        legend=dict(orientation="h", font=dict(size=10, color="#b8bcd8"),
                    y=-0.06, x=0.5, xanchor="center",
                    bgcolor="rgba(0,0,0,0)", borderwidth=0),
        margin=dict(t=10, b=20, l=10, r=10),
        height=260 if _IS_MOBILE else 340,
    )

    fig_sec_bar = go.Figure(go.Bar(
        x=sec_group["Sector"],
        y=sec_group["PnL"],
        marker=dict(
            color=sec_group["PnL"].apply(lambda x: "#22d67b" if x >= 0 else "#f85454"),
            opacity=0.88,
            line=dict(
                color=sec_group["PnL"].apply(lambda x: "#00e676" if x >= 0 else "#ff1744"),
                width=1.2
            )
        ),
        text=sec_group["PnL"].apply(lambda x: f"₹{x/1e3:+.1f}k"),
        textfont=dict(size=10, color="#ffffff", family="Inter"),
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>P&L: ₹%{y:,.0f}<extra></extra>"
    ))
    fig_sec_bar.add_hline(y=0, line=dict(color="#454870", width=1.5, dash="dot"))
    fig_sec_bar.update_layout(
        title=dict(
            text="<b>P&L by Sector</b>",
            font=dict(size=14, color="#f0f2ff", family="Inter"), x=0, xanchor="left"
        ),
        xaxis=dict(title="", tickangle=-30, tickfont=dict(size=10, color="#b8bcd8", family="Inter"),
                   showgrid=False, zeroline=False, linecolor="#252849"),
        yaxis=dict(title_font=dict(size=11, color="#7a7fa8"),
                   tickfont=dict(size=10, color="#7a7fa8", family="Inter"),
                   gridcolor="rgba(37,40,73,0.6)", zerolinecolor="#454870", zerolinewidth=1.5,
                   tickprefix="₹", separatethousands=True),
        plot_bgcolor="#0d0f1e", paper_bgcolor="#0e1022",
        font=dict(color="#b8bcd8", family="Inter"),
        bargap=0.35, showlegend=False, height=340,
        margin=dict(t=45, b=15, l=15, r=15),
    )

    # Render: mobile → donut only; desktop → donut + P&L bar side by side
    if _IS_MOBILE:
        st.plotly_chart(fig_sec, use_container_width=True)
    else:
        sc1, sc2 = st.columns([4, 5])
        with sc1:
            st.plotly_chart(fig_sec, use_container_width=True)
        with sc2:
            st.plotly_chart(fig_sec_bar, use_container_width=True)

    # Sector list rows (replacing the dataframe)
    sec_max_alloc = sec_group["Allocation_%"].max() or 1
    sec_pnl_abs   = sec_group["PnL"].abs().max() or 1
    sec_rows_html = ""
    for i, sr in sec_group.iterrows():
        s_clr  = _SEC_COLORS[i % len(_SEC_COLORS)]
        s_pclr = "#22d67b" if sr["PnL"] >= 0 else "#f85454"
        s_arr  = "▲" if sr["PnL"] > 0 else "▼"
        s_bar  = sr["Allocation_%"] / sec_max_alloc * 100
        sec_rows_html += f"""
<div class="gl-sector-row">
  <div class="gl-sector-dot" style="background:{s_clr};box-shadow:0 0 5px {s_clr}66;"></div>
  <span class="gl-sector-name">{sr['Sector']}</span>
  <span class="gl-sector-count">{int(sr['Count'])} stock{'s' if sr['Count']>1 else ''}</span>
  <div class="gl-sector-bar-wrap">
    <div class="gl-sector-bar-track">
      <div class="gl-sector-bar-fill" style="width:{s_bar:.1f}%;background:{s_clr};"></div>
    </div>
  </div>
  <span class="gl-sector-alloc">{sr['Allocation_%']:.1f}%</span>
  <span class="gl-sector-pnl" style="color:{s_pclr};">{s_arr} ₹{abs(sr['PnL'])/1e3:.1f}k</span>
</div>"""

    st.markdown(sec_rows_html, unsafe_allow_html=True)

# =====================================================
# TAB 3 — MARKET
# =====================================================

if _nav_tab == "Market":

    # ═══════════════════════════════════════════════════════
    # Shared data-fetch functions (same for both layouts)
    # ═══════════════════════════════════════════════════════
    # ── Helper: fetch NSE market data ────────────────────────────────
    @st.cache_data(ttl=300, show_spinner=False)
    def fetch_nse_gainers_losers():
        """Fetch top gainers & losers from NSE equity market in parallel."""
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://www.nseindia.com/",
        }

        def _fetch_one(endpoint):
            try:
                s = requests.Session()
                s.get("https://www.nseindia.com", headers=headers, timeout=8)
                r = s.get(endpoint, headers=headers, timeout=12)
                r.raise_for_status()
                data = r.json()
                rows = (data.get("NIFTY", {}).get("data", []) or
                        data.get("allSec", {}).get("data", []) or
                        (data if isinstance(data, list) else []))[:10]
                return [
                    {
                        "Symbol":   item.get("symbol", ""),
                        "LTP (₹)":  item.get("ltp", item.get("lastPrice", 0)),
                        "Change %": item.get("pChange", item.get("perChange", 0)),
                        "Change ₹": item.get("change", item.get("netPrice", 0)),
                        "Volume":   item.get("totalTradedVolume", item.get("tradedQuantity", 0)),
                    }
                    for item in rows
                ]
            except Exception:
                return []

        gainers_url = "https://www.nseindia.com/api/live-analysis-variations?index=gainers"
        losers_url  = "https://www.nseindia.com/api/live-analysis-variations?index=loosers"

        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
            fut_g = ex.submit(_fetch_one, gainers_url)
            fut_l = ex.submit(_fetch_one, losers_url)
            gainers = fut_g.result()
            losers  = fut_l.result()

        return gainers, losers

    @st.cache_data(ttl=300, show_spinner=False)
    def fetch_nse_indices():
        """Fetch NIFTY 50, BANKNIFTY, SENSEX indices from NSE."""
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://www.nseindia.com/",
            "Accept": "application/json",
        }
        indices = {}
        try:
            s = requests.Session()
            s.get("https://www.nseindia.com", headers=headers, timeout=10)
            r = s.get(
                "https://www.nseindia.com/api/allIndices",
                headers=headers, timeout=15
            )
            r.raise_for_status()
            for idx in r.json().get("data", []):
                name = idx.get("indexSymbol", "")
                if name in ("NIFTY 50", "NIFTY BANK", "NIFTY MIDCAP 100",
                            "NIFTY IT", "INDIA VIX"):
                    indices[name] = {
                        "last":   idx.get("last", 0),
                        "change": idx.get("change", 0),
                        "pChange": idx.get("percentChange", 0),
                    }
        except Exception:
            pass
        # Fallback: Angel One for NIFTY & BANKNIFTY
        if smartApi and "NIFTY 50" not in indices:
            try:
                nifty_info = resolve_token("Nifty 50", equity_map, nfo_map) or \
                             {"exch_seg": "NSE", "symbol": "Nifty 50", "token": "99926000"}
                nd = smartApi.ltpData(nifty_info["exch_seg"], nifty_info["symbol"], nifty_info["token"])
                if nd and nd.get("status"):
                    indices["NIFTY 50"] = {"last": nd["data"]["ltp"], "change": 0, "pChange": 0}
            except Exception:
                pass
            try:
                bn_info = resolve_token("Nifty Bank", equity_map, nfo_map) or \
                          {"exch_seg": "NSE", "symbol": "Nifty Bank", "token": "99926009"}
                bd = smartApi.ltpData(bn_info["exch_seg"], bn_info["symbol"], bn_info["token"])
                if bd and bd.get("status"):
                    indices["NIFTY BANK"] = {"last": bd["data"]["ltp"], "change": 0, "pChange": 0}
            except Exception:
                pass
        return indices


    # ════════════════════════════════════════════════════════════════
    # MOBILE WATCHLIST UI  — only shown on phones (≤ 640 px)
    # ════════════════════════════════════════════════════════════════
    st.markdown('<div class="mkt-mobile">', unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════
    # MOBILE MARKET PAGE  — clean, no ticker tape, no search bar
    # ════════════════════════════════════════════════════════════════

    # ── Mobile Market CSS ─────────────────────────────────────────
    st.markdown("""
<style>
/* ── Show/hide: mkt-mobile hidden on desktop, shown on mobile ── */
.mkt-mobile { display: none; }
@media (max-width: 640px) {
    .mkt-mobile { display: block !important; }
}

/* ── Market page title ── */
.mkt-title {
    font-size:20px;font-weight:900;color:#f0f2ff;
    padding:14px 14px 6px;letter-spacing:-0.3px;
    display:flex;align-items:center;gap:8px;
}
/* ── Status bar ── */
.mkt-status-bar {
    display:flex;align-items:center;gap:8px;
    padding:8px 14px;background:#060810;
    border-top:1px solid #131628;border-bottom:1px solid #131628;
    font-size:11px;color:#454870;font-weight:600;
}
/* ── Section header ── */
.mkt-sec-hdr {
    font-size:10px;font-weight:900;color:#454870;
    text-transform:uppercase;letter-spacing:1.4px;
    padding:14px 14px 6px;
}
/* ── Index card ── */
.idx-grid {
    display:grid;grid-template-columns:1fr 1fr;
    gap:8px;padding:6px 10px 10px;
}
.idx-card-m {
    background:#131628;border:1px solid #1e2240;
    border-radius:12px;padding:13px 12px;
    position:relative;overflow:hidden;
}
.idx-card-m::before {
    content:'';position:absolute;top:0;left:0;right:0;height:2px;
}
.idx-card-m.up::before   { background:linear-gradient(90deg,#22d67b,#00ff87); }
.idx-card-m.down::before { background:linear-gradient(90deg,#f85454,#ff2d55); }
.idx-card-m.flat::before { background:#252849; }
.idx-card-name  { font-size:9px;font-weight:800;color:#7a7fa8;text-transform:uppercase;
                  letter-spacing:.8px;margin-bottom:6px; }
.idx-card-price { font-size:17px;font-weight:900;color:#f0f2ff;line-height:1; }
.idx-card-chg   { font-size:11px;font-weight:700;margin-top:5px;display:flex;gap:5px;align-items:center; }
.idx-card-abs   { font-size:10px;color:#454870;margin-top:2px; }
/* ── Mover section ── */
.mkt-divider {
    height:6px;background:#060810;
    border-top:1px solid #131628;border-bottom:1px solid #131628;
    margin:6px 0;
}
.mv-row-m {
    display:flex;align-items:center;gap:0;
    padding:10px 14px;border-bottom:1px solid #0f1120;background:#0b0d17;
}
.mv-rank-m  { font-size:11px;color:#2a2e52;font-weight:700;min-width:22px; }
.mv-sym-m   { font-size:13px;font-weight:900;color:#f0f2ff;flex:1; }
.mv-ltp-m   { font-size:12px;color:#7a7fa8;font-weight:600;min-width:62px;text-align:right; }
.mv-pct-m   { font-size:13px;font-weight:900;min-width:60px;text-align:right; }
/* ── Refresh button wrapper ── */
.mkt-refresh-wrap {
    padding:14px;
}
/* ── VIX special card ── */
.vix-card {
    margin:0 10px 10px;background:#111428;border:1px solid #252849;
    border-radius:12px;padding:12px 14px;
    display:flex;justify-content:space-between;align-items:center;
}
.vix-label  { font-size:10px;font-weight:800;color:#7a7fa8;text-transform:uppercase;letter-spacing:1px; }
.vix-val    { font-size:22px;font-weight:900;color:#f5c842;line-height:1.1; }
.vix-sub    { font-size:10px;color:#454870;margin-top:3px; }
.vix-right  { text-align:right; }
</style>
""", unsafe_allow_html=True)

    # ── Fetch data ────────────────────────────────────────────────
    try:
        _m_indices = fetch_nse_indices()
    except Exception:
        _m_indices = {}
    try:
        _m_gainers, _m_losers = fetch_nse_gainers_losers()
    except Exception:
        _m_gainers, _m_losers = [], []

    # ── Market status bar ─────────────────────────────────────────
    _mkt_open_m   = _is_market_open_now()
    _mkt_stat_lbl = "● MARKET OPEN" if _mkt_open_m else "● MARKET CLOSED"
    _mkt_stat_col = "#22d67b"       if _mkt_open_m else "#f85454"
    _now_ist_str  = datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime("%I:%M %p IST")
    st.markdown(f"""
<div class="mkt-status-bar">
  <span style="width:7px;height:7px;border-radius:50%;background:{_mkt_stat_col};
               display:inline-block;flex-shrink:0;box-shadow:0 0 6px {_mkt_stat_col};"></span>
  <span style="color:{_mkt_stat_col};font-weight:900;">{_mkt_stat_lbl}</span>
  <span style="color:#2a2e52;">|</span>
  <span>NSE &nbsp;·&nbsp; {_now_ist_str}</span>
</div>
""", unsafe_allow_html=True)

    # ── Page title ────────────────────────────────────────────────
    st.markdown('''<div class="mkt-title">📊 Market</div>''', unsafe_allow_html=True)

    # ── Index grid cards ─────────────────────────────────────────
    st.markdown('<div class="mkt-sec-hdr">Indices</div>', unsafe_allow_html=True)

    _IDX_CONF = [
        ("NIFTY 50",       "Nifty 50"),
        ("NIFTY BANK",     "Bank Nifty"),
        ("NIFTY MIDCAP 100","MidCap 100"),
        ("NIFTY IT",       "Nifty IT"),
    ]
    _idx_cards_html = '<div class="idx-grid">'
    for _ikey, _ilabel in _IDX_CONF:
        if _ikey in _m_indices:
            _id   = _m_indices[_ikey]
            _ip   = float(_id.get("last",    0) or 0)
            _ic   = float(_id.get("change",  0) or 0)
            _ipc  = float(_id.get("pChange", 0) or 0)
            _cls  = "up" if _ipc >= 0 else "down"
            _clr  = "#22d67b" if _ipc >= 0 else "#f85454"
            _arr  = "▲" if _ipc >= 0 else "▼"
            _idx_cards_html += f"""
<div class="idx-card-m {_cls}">
  <div class="idx-card-name">{_ilabel}</div>
  <div class="idx-card-price">{_ip:,.2f}</div>
  <div class="idx-card-chg" style="color:{_clr};">
    <span>{_arr}</span><span>{abs(_ipc):.2f}%</span>
  </div>
  <div class="idx-card-abs">({_ic:+.2f})</div>
</div>"""
        else:
            _idx_cards_html += f'''
<div class="idx-card-m flat">
  <div class="idx-card-name">{_ilabel}</div>
  <div class="idx-card-price" style="color:#454870;">—</div>
</div>'''
    _idx_cards_html += '</div>'
    st.markdown(_idx_cards_html, unsafe_allow_html=True)

    # ── VIX card ─────────────────────────────────────────────────
    if "INDIA VIX" in _m_indices:
        _vd   = _m_indices["INDIA VIX"]
        _vval = float(_vd.get("last", 0) or 0)
        _vpc  = float(_vd.get("pChange", 0) or 0)
        _vcol = "#f85454" if _vval > 20 else ("#f5c842" if _vval > 14 else "#22d67b")
        _varr = "▲" if _vpc >= 0 else "▼"
        _vsen = "High Volatility" if _vval > 20 else ("Moderate" if _vval > 14 else "Low Volatility")
        st.markdown(f"""
<div class="vix-card">
  <div>
    <div class="vix-label">India VIX</div>
    <div style="font-size:10px;color:#454870;margin-top:3px;">Fear &amp; Volatility Index</div>
  </div>
  <div class="vix-right">
    <div class="vix-val" style="color:{_vcol};">{_vval:.2f}</div>
    <div style="font-size:10px;font-weight:700;color:{_vcol};">{_varr}{abs(_vpc):.2f}% · {_vsen}</div>
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Top Gainers ───────────────────────────────────────────────
    st.markdown('<div class="mkt-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="mkt-sec-hdr">🟢 Top Gainers</div>', unsafe_allow_html=True)

    if _m_gainers:
        _max_g = max(abs(float(r.get("Change %", 0) or 0)) for r in _m_gainers) or 1
        _gainers_html = ""
        for _gi, _gr in enumerate(_m_gainers[:8]):
            _gs   = _gr.get("Symbol", "")
            _gltp = float(_gr.get("LTP (₹)", 0) or 0)
            _gpc  = float(_gr.get("Change %", 0) or 0)
            _gbw  = abs(_gpc) / _max_g * 80
            _gainers_html += f"""
<div class="mv-row-m">
  <div class="mv-rank-m">{_gi+1}</div>
  <div style="flex:1;">
    <div class="mv-sym-m">{_gs}</div>
    <div style="height:3px;background:#1a1e38;border-radius:3px;overflow:hidden;margin-top:5px;width:80%;">
      <div style="height:100%;width:{_gbw:.0f}%;background:#22d67b;border-radius:3px;"></div>
    </div>
  </div>
  <div class="mv-ltp-m">₹{_gltp:,.2f}</div>
  <div class="mv-pct-m" style="color:#22d67b;">▲{abs(_gpc):.2f}%</div>
</div>"""
        st.markdown(_gainers_html, unsafe_allow_html=True)
    else:
        st.markdown('<div style="padding:16px 14px;color:#454870;font-size:12px;">Data unavailable</div>', unsafe_allow_html=True)

    # ── Top Losers ────────────────────────────────────────────────
    st.markdown('<div class="mkt-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="mkt-sec-hdr">🔴 Top Losers</div>', unsafe_allow_html=True)

    if _m_losers:
        _max_l = max(abs(float(r.get("Change %", 0) or 0)) for r in _m_losers) or 1
        _losers_html = ""
        for _li, _lr in enumerate(_m_losers[:8]):
            _ls   = _lr.get("Symbol", "")
            _lltp = float(_lr.get("LTP (₹)", 0) or 0)
            _lpc  = float(_lr.get("Change %", 0) or 0)
            _lbw  = abs(_lpc) / _max_l * 80
            _losers_html += f"""
<div class="mv-row-m">
  <div class="mv-rank-m">{_li+1}</div>
  <div style="flex:1;">
    <div class="mv-sym-m">{_ls}</div>
    <div style="height:3px;background:#1a1e38;border-radius:3px;overflow:hidden;margin-top:5px;width:80%;">
      <div style="height:100%;width:{_lbw:.0f}%;background:#f85454;border-radius:3px;"></div>
    </div>
  </div>
  <div class="mv-ltp-m">₹{_lltp:,.2f}</div>
  <div class="mv-pct-m" style="color:#f85454;">▼{abs(_lpc):.2f}%</div>
</div>"""
        st.markdown(_losers_html, unsafe_allow_html=True)
    else:
        st.markdown('<div style="padding:16px 14px;color:#454870;font-size:12px;">Data unavailable</div>', unsafe_allow_html=True)

    # ── Refresh button ────────────────────────────────────────────
    st.markdown('<div class="mkt-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="mkt-refresh-wrap">', unsafe_allow_html=True)
    if st.button("🔄 Refresh", key="refresh_market_mobile", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    # ── End mobile wrapper ────────────────────────────────────────────
    st.markdown('</div>', unsafe_allow_html=True)

# =====================================================
# NEWS HELPERS
# =====================================================

RSS_FEEDS = {
    "Moneycontrol Results":      "https://www.moneycontrol.com/rss/results.xml",
    "Moneycontrol Corp Actions": "https://www.moneycontrol.com/rss/corporateactions.xml",
    "ET Earnings":               "https://economictimes.indiatimes.com/markets/earnings/rssfeeds/2139228.cms",
    "ET Markets Stocks":         "https://economictimes.indiatimes.com/markets/stocks/rssfeeds/2146842.cms",
    "BS Companies":              "https://www.business-standard.com/rss/companies-101.rss",
    "Mint Companies":            "https://www.livemint.com/rss/companies",
    "CNBC-TV18 Markets":         "https://www.cnbctv18.com/commonfeeds/v1/cne/rss/market.xml",
    "NDTV Profit Markets":       "https://www.ndtvprofit.com/rss/markets",
    "NSE Announcements":         "https://nsearchives.nseindia.com/content/RSS/Online_announcements.xml",
    "BSE Notices":               "https://www.bseindia.com/data/xml/notices.xml",
}

HIGH_PRIORITY = [
    "dividend", "interim dividend", "final dividend", "special dividend",
    "results", "earnings", "bonus", "bonus share", "stock split", "share split",
    "buyback", "buy back", "merger", "acquisition", "takeover", "demerger",
    "board meeting", "record date", "ex-date", "ex-dividend",
    "sebi order", "sebi notice", "sebi penalty",
    "order win", "guidance", "bulk deal", "block deal",
    "stake sale", "rights issue", "fund raise", "qip", "preferential allotment",
    "upgrade", "downgrade", "capex", "expansion", "rating change",
    "net profit", "net loss", "revenue growth", "quarterly result",
    "esop", "promoter stake", "pledge", "insider trading",
]

# Headlines matching these patterns are roundup/listicle articles that mention
# many companies — they should never be tagged to a specific stock.
_ROUNDUP_BLOCK_PATTERNS = [
    r"full list", r"here('s| is| are) the", r"these companies", r"among companies",
    r"list of companies", r"list of stocks", r"companies to (declare|post|report|announce)",
    r"stocks to watch", r"top (gainers|losers|stocks|picks)",
    r"q\d results.*:.*among", r"among.*to (declare|post|report)",
    r"others to (post|report|declare)", r"and (many )?others",
    r"see (details|full list)", r"next week in stock market",
    r"this week.*dividend", r"upcoming.*ex.?date.*these",
]
_ROUNDUP_RE = re.compile("|".join(_ROUNDUP_BLOCK_PATTERNS), re.IGNORECASE)

def _is_roundup_article(title: str, summary: str = "") -> bool:
    """Return True if the article is a market roundup listing multiple companies."""
    text = (title + " " + summary).lower()
    return bool(_ROUNDUP_RE.search(text))

def make_hash(text):
    return hashlib.md5(re.sub(r'\s+', ' ', text.lower().strip()).encode()).hexdigest()

def is_recent_news(pub_text):
    try:
        if not pub_text:
            return False
        if str(pub_text).isdigit():
            published_time = datetime.fromtimestamp(int(pub_text), tz=timezone.utc)
        else:
            try:
                published_time = parsedate_to_datetime(pub_text)
            except Exception:
                published_time = pd.to_datetime(pub_text, utc=True).to_pydatetime()
            if published_time.tzinfo is None:
                published_time = published_time.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - published_time) <= timedelta(days=7)
    except:
        return True

# ── Company name aliases for news matching ──────────────────────────────────
_NEWS_TICKER_ALIASES = {
    "KALAMANDIR":    ("Sai Silks Kalamandir",        ["sai silks kalamandir", "sai silks"]),
    "GAYAPROJ":      ("Gayatri Projects",             ["gayatri projects"]),
    "JAYBARMARU":    ("Jay Bharat Maruti",            ["jay bharat maruti"]),
    "KAMATHOTEL":    ("Kamat Hotels India",           ["kamat hotels", "kamat hotel"]),
    "OCCLLTD":       ("Oriental Carbon Chemicals",    ["oriental carbon", "oriental carbon & chemicals", "oriental carbon and chemicals", "occl"]),
    "HEXT":          ("Hexaware Technologies",        ["hexaware technologies", "hexaware tech", "hexaware limited", "hexaware"]),
    "LINCOLNPHARMA": ("Lincoln Pharmaceuticals",      ["lincoln pharma", "lincoln pharmaceuticals", "lincoln pharma ltd"]),
    "LINCOLN":       ("Lincoln Pharmaceuticals",      ["lincoln pharma", "lincoln pharmaceuticals", "lincoln pharma ltd"]),
    "GVPIL":         ("GE Power India",               ["ge power india", "alstom india"]),
    "INTERARCH":     ("Interarch Building Solutions", ["interarch building solutions", "interarch building products", "interarch"]),
    "VOLTAMP":       ("Voltamp Transformers",         ["voltamp transformers", "voltamp"]),
    "ENRIN":         ("Siemens Energy India",         ["siemens energy india"]),
    "RECLTD":        ("REC Limited",                  ["rec ltd", "rec limited", "rural electrification corporation"]),
    "MAWANASUG":     ("Mawana Sugars",                ["mawana sugar", "mawana sugars"]),
    "INDUSINDBK":    ("IndusInd Bank",                ["indusind bank", "indusind"]),
    "EXCELINDUS":    ("Excel Industries",             ["excel industries"]),
    "CHENNPETRO":    ("Chennai Petroleum",            ["chennai petroleum", "cpcl", "chennai petro corporation"]),
    "BCG":           ("Brightcom Group",              ["brightcom group", "brightcom"]),
    "CEINSYS":       ("Ceinsys Tech",                 ["ceinsys tech", "ceinsys technologies", "ceinsys"]),
}

# Per-ticker hard disambiguation: (required_any, blocked_any)
_NEWS_DISAMBIGUATION = {
    "LINCOLN": (
        ["lincoln pharma", "lincoln pharmaceuticals", "lincolnpharma", "lincoln pharma ltd"],
        ["lincoln national", "lincoln educational", "lincoln financial", "lincoln electric",
         "lincoln motors", "lincoln center", "lincoln memorial", "lincoln university",
         "lincoln park", "lincoln international", "lincoln lab", "nasdaq:linc",
         "nyse:lnc", "lnc)", "linc)", "lincoln nebraska"],
    ),
    "BCG": (
        ["brightcom group", "brightcom technologies"],
        ["boston consulting", "bcg matrix", "bcg report", "boston group"],
    ),
}

_NEWS_GOOGLE_OVERRIDES = {
    "LINCOLNPHARMA": "Lincoln Pharmaceuticals India NSE",
    "LINCOLN":       "Lincoln Pharmaceuticals India NSE",
    "GAYAPROJ":      "Gayatri Projects India NSE",
    "HEXT":          "Hexaware Technologies India NSE",
    "OCCLLTD":       "Oriental Carbon Chemicals India NSE",
    "BCG":           "Brightcom Group India NSE",
    "ITC":           "ITC Limited India NSE results",
}

def _get_news_company_name(ticker: str) -> str:
    if ticker in _NEWS_GOOGLE_OVERRIDES:
        return _NEWS_GOOGLE_OVERRIDES[ticker]
    if ticker in _NEWS_TICKER_ALIASES:
        return _NEWS_TICKER_ALIASES[ticker][0]
    return ticker

def _get_news_aliases(ticker: str) -> list:
    aliases = [ticker.lower()]
    if ticker in _NEWS_TICKER_ALIASES:
        _, extra = _NEWS_TICKER_ALIASES[ticker]
        aliases += extra
    return aliases

def _passes_news_disambiguation(text: str, ticker: str) -> bool:
    tl = text.lower()
    for fragment, (required, blocked) in _NEWS_DISAMBIGUATION.items():
        if fragment in ticker.upper():
            if any(b in tl for b in blocked):
                return False
            if required and not any(r in tl for r in required):
                return False
    return True

def matches_portfolio(text, ticker_names):
    """Return True if any portfolio stock appears in the news text."""
    text_lower = text.lower()
    for name in ticker_names:
        if not name or len(name) < 2:
            continue
        for alias in _get_news_aliases(name):
            if len(alias) <= 4:
                if re.search(r'(?<![a-z0-9])' + re.escape(alias) + r'(?![a-z0-9])', text_lower):
                    if _passes_news_disambiguation(text, name):
                        return True
            else:
                if alias in text_lower:
                    if _passes_news_disambiguation(text, name):
                        return True
    return False

def _tag_stock(text: str, clean_names: list) -> str:
    text_lower = text.lower()
    for name in clean_names:
        for alias in _get_news_aliases(name):
            if len(alias) <= 4:
                if re.search(r'(?<![a-z0-9])' + re.escape(alias) + r'(?![a-z0-9])', text_lower):
                    if _passes_news_disambiguation(text, name):
                        return name
            else:
                if alias in text_lower:
                    if _passes_news_disambiguation(text, name):
                        return name
    return ""

def get_clean_ticker_names(tickers):
    return list({t.replace(".NS","").replace(".BO","").upper() for t in tickers})

def fetch_rss(feed_name, url):
    articles = []
    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        feed = feedparser.parse(resp.content)
        for entry in feed.entries[:50]:
            articles.append({
                "title":     entry.get("title","").strip(),
                "summary":   entry.get("summary",""),
                "published": entry.get("published",""),
                "link":      entry.get("link",""),
                "source":    feed_name,
            })
    except:
        pass
    return articles

def fetch_google_news_query(query, force_ticker=None):
    articles = []
    try:
        url = f"https://news.google.com/rss/search?q={requests.utils.quote(query)}&hl=en-IN&gl=IN&ceid=IN:en"
        feed = feedparser.parse(requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"}).content)
        # Build aliases for the force_ticker so we can validate title relevance
        _ft_aliases = _get_news_aliases(force_ticker) if force_ticker else []
        for entry in feed.entries[:20]:
            title = entry.get("title","").strip()
            title = re.sub(r'\s+-\s+[\w\s\.]+$', '', title).strip()
            summary = entry.get("summary","")
            if not title:
                continue
            # Drop roundup/listicle articles — they mention dozens of companies
            if _is_roundup_article(title, summary):
                continue
            # When force_ticker is set, require at least one alias in the TITLE
            # (not just anywhere in the article) to prevent tagging unrelated news
            if force_ticker and _ft_aliases:
                title_lower = title.lower()
                if not any(alias in title_lower for alias in _ft_aliases):
                    continue
            art = {
                "title":     title,
                "summary":   summary,
                "published": entry.get("published",""),
                "link":      entry.get("link",""),
                "source":    "Google News",
            }
            if force_ticker:
                art["_force_ticker"] = force_ticker
            articles.append(art)
    except:
        pass
    return articles

def fetch_bse_corporate_news(clean_names: list) -> list:
    articles = []
    try:
        url = ("https://api.bseindia.com/BseIndiaAPI/api/AnnSubCategoryGetData/w"
               "?strCat=-1&strPrevDate=&strScrip=&strSearch=P"
               "&strToDate=&strType=C&subcategory=-1")
        headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.bseindia.com/"}
        resp = requests.get(url, headers=headers, timeout=12)
        data = resp.json()
        for ann in (data.get("Table") or [])[:300]:
            title  = ann.get("HEADLINE","").strip()
            scrip  = ann.get("SLONGNAME","")
            date_s = ann.get("News_submission_dt","")
            attach = ann.get("ATTACHMENTNAME","")
            link   = (f"https://www.bseindia.com/xml-data/corpfiling/AttachLive/{attach}"
                      if attach else "https://www.bseindia.com/corporates/ann.html")
            if not title:
                continue
            # Drop roundup articles
            if _is_roundup_article(title, scrip):
                continue
            combined = (title + " " + scrip).lower()
            matched = _tag_stock(combined, clean_names)
            if not matched:
                continue
            articles.append({
                "title":         f"{scrip}: {title}" if scrip else title,
                "summary":       scrip,
                "published":     date_s,
                "link":          link,
                "source":        "BSE Corporate",
                "_force_ticker": matched,
            })
    except:
        pass
    return articles

def fetch_yfinance_portfolio_news(portfolio_tickers: list) -> list:
    if not _YF_AVAILABLE:
        return []
    articles = []
    for ticker_raw in portfolio_tickers:
        clean = ticker_raw.replace(".NS","").replace(".BO","").upper()
        yf_base = ANGEL_ONE_ALIASES.get(clean, clean)
        suffix = ".BO" if ticker_raw.upper().endswith(".BO") else ".NS"
        yf_sym = yf_base + suffix
        try:
            ticker_obj = yf.Ticker(yf_sym)
            news_items = ticker_obj.news or []
            for item in news_items[:10]:
                content = item.get("content") or {}
                title   = (content.get("title") or item.get("title","")).strip()
                title   = re.sub(r'\s+-\s+[\w\s\.]+$', '', title).strip()
                url_s   = ((content.get("canonicalUrl") or {}).get("url") or item.get("link",""))
                pub     = (content.get("pubDate") or item.get("providerPublishTime") or item.get("published",""))
                if not title or not _passes_news_disambiguation(title, clean):
                    continue
                articles.append({
                    "title":         title,
                    "summary":       "",
                    "published":     pub,
                    "link":          url_s,
                    "source":        "Yahoo Finance",
                    "_force_ticker": clean,
                })
        except:
            pass
    return articles

def fetch_all_news(portfolio_tickers):
    clean_names = get_clean_ticker_names(portfolio_tickers)
    all_news = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=25) as executor:
        futures = []
        for feed_name, url in RSS_FEEDS.items():
            futures.append(executor.submit(fetch_rss, feed_name, url))
        for name in clean_names:
            sname = _get_news_company_name(name)
            futures.append(executor.submit(fetch_google_news_query, sname, name))
            futures.append(executor.submit(fetch_google_news_query, f"{sname} NSE results", name))
            futures.append(executor.submit(fetch_google_news_query, f"{sname} dividend bonus", name))
            futures.append(executor.submit(fetch_google_news_query, f"{sname} quarterly profit India", name))
            futures.append(executor.submit(fetch_google_news_query, f"{sname} order win acquisition merger", name))
        futures.append(executor.submit(fetch_bse_corporate_news, clean_names))
        futures.append(executor.submit(fetch_yfinance_portfolio_news, portfolio_tickers))
        for future in concurrent.futures.as_completed(futures):
            try:
                r = future.result()
                if isinstance(r, list):
                    all_news.extend(r)
            except:
                pass
    return all_news

def get_filtered_news(portfolio_tickers):
    clean_names = get_clean_ticker_names(portfolio_tickers)
    raw_news = fetch_all_news(portfolio_tickers)
    filtered = []
    seen = set()
    for item in raw_news:
        title   = item.get("title","")
        summary = item.get("summary","")
        text    = title + " " + summary
        if not title or title in ("[Removed]","None",""):
            continue
        if not is_recent_news(item.get("published","")):
            continue
        # Drop roundup/listicle articles — these mention many companies and
        # get incorrectly tagged to whichever stock triggered the query
        if _is_roundup_article(title, summary):
            continue
        h = make_hash(title)
        if h in seen:
            continue
        if "_force_ticker" in item:
            stock_tag = item["_force_ticker"]
            # Even for force_ticker, require alias appears in the TITLE
            # to prevent roundup articles being tagged to a specific stock
            _aliases = _get_news_aliases(stock_tag)
            title_lower = title.lower()
            if not any(alias in title_lower for alias in _aliases):
                continue
            if not _passes_news_disambiguation(text, stock_tag):
                continue
        else:
            if not matches_portfolio(text, clean_names):
                continue
            stock_tag = _tag_stock(text, clean_names)
        seen.add(h)
        priority = "🟡 NORMAL"
        for word in HIGH_PRIORITY:
            if word in text.lower():
                priority = "🔴 HIGH"
                break
        item["priority"]  = priority
        item["stock_tag"] = stock_tag
        pub_text = item.get("published","")
        pub_dt = None
        try:
            if pub_text:
                if str(pub_text).isdigit():
                    pub_dt = datetime.fromtimestamp(int(pub_text), tz=timezone.utc)
                else:
                    try:
                        pub_dt = parsedate_to_datetime(str(pub_text))
                    except Exception:
                        pub_dt = pd.to_datetime(pub_text, utc=True).to_pydatetime()
                    if pub_dt.tzinfo is None:
                        pub_dt = pub_dt.replace(tzinfo=timezone.utc)
        except Exception:
            pub_dt = None
        item["published_dt"] = pub_dt
        filtered.append(item)
    def sort_key(x):
        dt = x.get("published_dt")
        return dt if dt is not None else datetime.min.replace(tzinfo=timezone.utc)
    return sorted(filtered, key=sort_key, reverse=True)



def get_results_calendar(portfolio_tickers):
    """
    Fetch upcoming corporate actions from NSE event-calendar + corporate actions API.
    Shows all corporate action types with Ex-Date, Record Date, and Details columns.
    Deduplicates strictly: one row per (nse_symbol + exact_date + canonical_event).
    """
    rows = []
    seen = set()
    clean_names = get_clean_ticker_names(portfolio_tickers)

    # Priority tiers for different event types
    HIGH_PRIORITY_EVENTS = {
        "Financial Results", "Dividend", "Interim Dividend", "Final Dividend",
        "Special Dividend", "Bonus Issue", "Stock Split", "Buyback", "Rights Issue",
        "Merger / Demerger", "Acquisition", "Delisting",
    }

    # Events that have an Ex-Date (shareholder eligibility date)
    EX_DATE_EVENTS = {
        "Dividend", "Interim Dividend", "Final Dividend", "Special Dividend",
        "Bonus Issue", "Stock Split", "Rights Issue",
    }

    def symbol_matches(nse_sym, clean_list):
        nse_norm = re.sub(r"[^A-Z0-9]", "", nse_sym.strip().upper())
        for c in clean_list:
            c_norm = re.sub(r"[^A-Z0-9]", "", c.strip().upper())
            if nse_norm == c_norm:
                return c
            if len(c_norm) >= 3 and len(nse_norm) >= 3:
                if nse_norm.startswith(c_norm) or c_norm.startswith(nse_norm):
                    return c
        return None

    def canonical_event(ev):
        """Map raw NSE event description → clean corporate action label."""
        el = str(ev).strip().lower()

        # ── Results ───────────────────────────────────────────────
        if any(x in el for x in ("financial result", "quarterly result",
                                  "annual result", "half yearly result",
                                  "unaudited result", "audited result")):
            return "Financial Results"

        # ── Dividends ─────────────────────────────────────────────
        if "interim dividend" in el:
            return "Interim Dividend"
        if "final dividend" in el:
            return "Final Dividend"
        if "special dividend" in el:
            return "Special Dividend"
        if "dividend" in el:
            return "Dividend"

        # ── Bonus / Split ─────────────────────────────────────────
        if "bonus" in el:
            return "Bonus Issue"
        if "stock split" in el or "sub-division" in el or "sub division" in el:
            return "Stock Split"
        if "face value" in el and "split" in el:
            return "Stock Split"

        # ── Capital actions ───────────────────────────────────────
        if "rights issue" in el or "rights entitlement" in el:
            return "Rights Issue"
        if "buyback" in el or "buy back" in el or "buy-back" in el:
            return "Buyback"

        # ── Corporate restructuring ───────────────────────────────
        if "merger" in el or "amalgamation" in el or "demerger" in el:
            return "Merger / Demerger"
        if "acquisition" in el or "takeover" in el:
            return "Acquisition"
        if "scheme of arrangement" in el:
            return "Scheme of Arrangement"
        if "delisting" in el:
            return "Delisting"

        # ── General meetings ──────────────────────────────────────
        if "agm" in el or "annual general meeting" in el:
            return "AGM"
        if "egm" in el or "extra ordinary general" in el:
            return "EGM"

        # ── Warrants / ESOP ───────────────────────────────────────
        if "warrant" in el:
            return "Warrants"
        if "esop" in el or "employee stock" in el:
            return "ESOP"

        # ── Preferential allotment ────────────────────────────────
        if "preferential" in el or "preferential allotment" in el:
            return "Preferential Allotment"

        # ── Fundraising ───────────────────────────────────────────
        if "ncd" in el or "non-convertible debenture" in el:
            return "NCD Issue"
        if "qip" in el or "qualified institutional" in el:
            return "QIP"
        if "ipo" in el or "fpo" in el:
            return "IPO / FPO"

        # ── Record / Ex-date notices ──────────────────────────────
        if "record date" in el:
            return "Record Date"
        if "ex-date" in el or "ex date" in el:
            return "Ex-Date Notice"

        # ── Drop bare board meeting intimations (no action info) ──
        if el.strip() in ("board meeting", "board meeting intimation"):
            return None

        # ── Everything else — keep as-is but Title-case it ────────
        cleaned = str(ev).strip()
        if len(cleaned) > 60:
            cleaned = cleaned[:57] + "…"
        return cleaned if cleaned else None

    def parse_date_safe(d):
        if not d or str(d).strip() in ("", "-", "nan", "None"):
            return None
        try:
            return pd.to_datetime(str(d).strip(), dayfirst=True)
        except Exception:
            return None

    def fmt_date(ts):
        if ts is None:
            return "—"
        try:
            return ts.strftime("%d-%b-%Y")
        except Exception:
            return "—"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.nseindia.com/",
    }

    today = pd.Timestamp.now().normalize()
    window_end = today + pd.Timedelta(days=90)  # look 90 days ahead for ex-dates

    # ── Build a lookup: symbol → list of corporate action details from NSE CA API ──
    # This gives us ex-date, record date, and details like ratio / amount
    ca_lookup = {}   # nse_sym_upper → list of {ex_date, record_date, details, event_type}

    # Also track which clean_names are BSE-only (have .BO in original tickers)
    bse_only_names = set()
    for orig_t in portfolio_tickers:
        t_up = str(orig_t).upper()
        if t_up.endswith(".BO"):
            bse_only_names.add(t_up.replace(".BO", ""))

    bse_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.bseindia.com/",
        "Origin": "https://www.bseindia.com",
    }

    def fetch_ca_for_symbol(sym):
        """Fetch NSE corporate actions for a single symbol."""
        result = []
        try:
            s = requests.Session()
            s.get("https://www.nseindia.com", headers=headers, timeout=10)
            url = f"https://www.nseindia.com/api/corporates-corporateActions?index=equities&symbol={sym}&subject=all"
            r = s.get(url, headers=headers, timeout=15)
            if r.status_code == 200:
                data = r.json()
                items = data if isinstance(data, list) else data.get("data", [])
                for it in items:
                    ex_dt  = parse_date_safe(it.get("exDate") or it.get("ex_date") or it.get("exdate"))
                    rec_dt = parse_date_safe(it.get("recordDate") or it.get("record_date") or it.get("recDate"))
                    purpose_raw = str(it.get("subject") or it.get("purpose") or "").strip()
                    details = str(it.get("remarks") or it.get("details") or it.get("value") or "").strip()
                    if not details:
                        details = purpose_raw
                    result.append({
                        "ex_date":     ex_dt,
                        "record_date": rec_dt,
                        "purpose_raw": purpose_raw,
                        "details":     details,
                        "event":       canonical_event(purpose_raw),
                        "source":      "NSE",
                    })
        except Exception:
            pass
        return sym, result

    def fetch_bse_ca_for_symbol(sym_clean):
        """Fetch BSE corporate actions for a BSE-listed stock.
        BSE CA API uses scrip codes; we try by symbol name via the BSE search."""
        result = []
        try:
            sb = requests.Session()
            sb.get("https://www.bseindia.com", headers=bse_headers, timeout=10)
            # BSE corporate actions endpoint (works with scrip code or symbol)
            # First try to resolve scrip code via BSE search
            search_url = f"https://api.bseindia.com/BseIndiaAPI/api/ListofScripData/w?Group=&Scripcode=&industry=&segment=Equity&status=Active"
            # Directly use BSE corporate actions by symbol name
            ca_url = (
                f"https://api.bseindia.com/BseIndiaAPI/api/CorporateAction/w?"
                f"scripcode=&qtyfrom=&qtyto=&exdfrom=&exdto=&segment=Equity&type=all&pageno=1"
            )
            # Use BSE's announcements API filtered by company name
            ann_url = (
                f"https://api.bseindia.com/BseIndiaAPI/api/AnnSubCategoryGetData/w?"
                f"strCat=-1&strPrevDate=&strScrip=&strSearch=P&strToDate=&strType=C"
                f"&subcategory=-1"
            )
            # Most reliable: BSE corporate actions search by company code
            # First get the BSE scrip code from the company name
            find_url = f"https://api.bseindia.com/BseIndiaAPI/api/ddlbywatchlist/w?categoryid=0&industry=0&mktcapid=0&statusid=A"
            # Simpler: directly query BSE corporate actions with symbol
            direct_url = (
                f"https://www.bseindia.com/corporates/corp_act.html"
            )
            # Most practical BSE CA endpoint:
            bse_ca_url = (
                f"https://api.bseindia.com/BseIndiaAPI/api/CorporateAction/w?"
                f"scripcode=&qtyfrom=&qtyto=&exdfrom=&exdto=&segment=Equity"
                f"&type=all&pageno=1&pagesize=50&companyname={sym_clean}"
            )
            r = sb.get(bse_ca_url, headers=bse_headers, timeout=15)
            if r.status_code == 200:
                data = r.json()
                items = (
                    data.get("Table", [])
                    or data.get("data", [])
                    or (data if isinstance(data, list) else [])
                )
                for it in items:
                    # BSE CA fields
                    ex_dt  = parse_date_safe(
                        it.get("Ex_Date") or it.get("ExDate") or it.get("ex_date")
                        or it.get("exDate") or it.get("EX_DATE")
                    )
                    rec_dt = parse_date_safe(
                        it.get("RD_Date") or it.get("RecordDate") or it.get("record_date")
                        or it.get("recordDate")
                    )
                    purpose_raw = str(
                        it.get("Purpose") or it.get("purpose")
                        or it.get("ActionType") or it.get("action_type") or ""
                    ).strip()
                    details = str(
                        it.get("Remarks") or it.get("remarks")
                        or it.get("Details") or it.get("details") or ""
                    ).strip()
                    if not details:
                        details = purpose_raw
                    result.append({
                        "ex_date":     ex_dt,
                        "record_date": rec_dt,
                        "purpose_raw": purpose_raw,
                        "details":     details,
                        "event":       canonical_event(purpose_raw),
                        "source":      "BSE",
                    })
        except Exception:
            pass
        return sym_clean, result

    # Fetch CA details for all portfolio symbols concurrently
    nse_symbols_for_ca = [
        re.sub(r"[^A-Z0-9]", "", c.strip().upper()) for c in clean_names
        if c.strip().upper() not in bse_only_names   # skip BSE-only for NSE fetch
    ]
    bse_symbols_for_ca = list(bse_only_names)

    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as ex_pool:
        nse_futures = {ex_pool.submit(fetch_ca_for_symbol, sym): sym for sym in nse_symbols_for_ca}
        bse_futures = {ex_pool.submit(fetch_bse_ca_for_symbol, sym): sym for sym in bse_symbols_for_ca}
        for fut in concurrent.futures.as_completed({**nse_futures, **bse_futures}):
            try:
                sym, ca_list = fut.result()
                ca_lookup[sym] = ca_lookup.get(sym, []) + ca_list
            except Exception:
                pass

    # ── Source 1: NSE event-calendar (board meeting dates / announcement dates) ──
    try:
        session = requests.Session()
        session.get("https://www.nseindia.com", headers=headers, timeout=10)
        resp = session.get(
            "https://www.nseindia.com/api/event-calendar",
            headers=headers, timeout=15
        )
        resp.raise_for_status()
        data = resp.json()
        items = data if isinstance(data, list) else data.get("data", [])

        for item in items:
            nse_sym = str(item.get("symbol", "")).upper().strip()
            matched = symbol_matches(nse_sym, clean_names)
            if not matched:
                continue

            date_str  = str(item.get("date", "")).strip()
            event_raw = str(item.get("purpose", "")).strip()

            ev_date = parse_date_safe(date_str)
            if ev_date is None:
                continue

            # Only today → +30 days for board meeting / announcement date
            if ev_date < today or ev_date > today + pd.Timedelta(days=30):
                continue

            event = canonical_event(event_raw)
            if event is None:
                continue

            # Try to find matching ex-date from CA lookup
            nse_norm = re.sub(r"[^A-Z0-9]", "", nse_sym)
            ca_records = ca_lookup.get(nse_norm, [])
            ex_date_str    = "—"
            record_date_str = "—"
            details_str    = ""

            for ca in ca_records:
                if ca.get("event") == event:
                    ex_dt  = ca.get("ex_date")
                    rec_dt = ca.get("record_date")
                    # Only use future / very recent ex-dates
                    if ex_dt is not None and ex_dt >= today - pd.Timedelta(days=5):
                        ex_date_str    = fmt_date(ex_dt)
                        record_date_str = fmt_date(rec_dt)
                        details_str    = ca.get("details", "")
                        break

            # ── Strict dedup ──
            key = f"{nse_sym}|{ev_date.date()}|{event}"
            if key in seen:
                continue
            seen.add(key)

            days_away  = (ev_date - today).days
            days_label = "Today" if days_away == 0 else f"In {days_away}d"
            priority   = "🔴 HIGH" if event in HIGH_PRIORITY_EVENTS else "🟡 MEDIUM"

            rows.append({
                "Ticker":       matched,
                "Event":        event,
                "Board/Ann. Date": ev_date.strftime("%d-%b-%Y"),
                "Ex-Date":      ex_date_str,
                "Record Date":  record_date_str,
                "Details":      details_str[:60] if details_str else "—",
                "Days Away":    days_label,
                "Priority":     priority,
                "Source":       "NSE",
                "_sort_ts":     ev_date,
            })

    except Exception:
        pass

    # ── Source 2: Corporate actions API directly (catches ex-dates in next 90 days) ──
    # This covers cases where ex-date is upcoming but board meeting was in the past.
    # Works for both NSE and BSE stocks since ca_lookup contains both.
    for sym_clean in clean_names:
        nse_norm = re.sub(r"[^A-Z0-9]", "", sym_clean.strip().upper())
        ca_records = ca_lookup.get(nse_norm, [])
        is_bse = sym_clean.upper() in bse_only_names

        for ca in ca_records:
            event = ca.get("event")
            if event is None:
                continue

            ex_dt  = ca.get("ex_date")
            rec_dt = ca.get("record_date")

            # Use ex-date as the primary date for this entry
            primary_dt = ex_dt or rec_dt
            if primary_dt is None:
                continue
            if primary_dt < today or primary_dt > window_end:
                continue

            src_label = ca.get("source", "BSE CA" if is_bse else "NSE CA")
            key = f"{nse_norm}|{primary_dt.date()}|{event}|exdate"
            if key in seen:
                continue
            seen.add(key)

            days_away  = (primary_dt - today).days
            days_label = "Today" if days_away == 0 else f"In {days_away}d"
            priority   = "🔴 HIGH" if event in HIGH_PRIORITY_EVENTS else "🟡 MEDIUM"

            # Restore .BO suffix in Ticker column for BSE stocks
            ticker_display = (sym_clean + ".BO") if is_bse else sym_clean

            rows.append({
                "Ticker":          ticker_display,
                "Event":           event,
                "Board/Ann. Date": "—",
                "Ex-Date":         fmt_date(ex_dt),
                "Record Date":     fmt_date(rec_dt),
                "Details":         ca.get("details", "—")[:60] if ca.get("details") else "—",
                "Days Away":       days_label,
                "Priority":        priority,
                "Source":          src_label,
                "_sort_ts":        primary_dt,
            })

    df_out = pd.DataFrame(rows)
    if not df_out.empty:
        try:
            df_out = df_out.sort_values(["_sort_ts", "Ticker"]).drop(columns=["_sort_ts"])
        except Exception:
            df_out = df_out.drop(columns=["_sort_ts"], errors="ignore")
        df_out = df_out.reset_index(drop=True)
    else:
        df_out = pd.DataFrame(columns=[
            "Ticker", "Event", "Board/Ann. Date", "Ex-Date",
            "Record Date", "Details", "Days Away", "Priority", "Source"
        ])
    return df_out


# =====================================================
# TAB 4 — PORTFOLIO NEWS
# =====================================================

if _nav_tab == "Portfolio News":
    st.subheader("📰 Portfolio News")

    col_refresh, col_time, col_stock = st.columns([1, 2, 2])

    with col_refresh:
        if st.button("🔄 Refresh News", key="refresh_news"):
            st.cache_data.clear()
            st.session_state.pop("_ticker_news", None)
            st.session_state.pop("_ticker_cal_df", None)
            st.rerun()

    with col_time:
        time_filter = st.selectbox(
            "🕐 Time Range",
            ["Today", "Last 3 Days", "Last 7 Days"],
            key="news_time_filter"
        )

    with col_stock:
        # Stock filter populated after fetching
        pass

    with st.spinner("📡 Fetching latest news for your portfolio..."):
        news_items = get_filtered_news(tickers)
        # Store in session_state so the ticker banner can read it without re-fetching
        st.session_state["_ticker_news"] = news_items

    if not news_items:
        st.info("No recent news found for your portfolio stocks. Try refreshing or check back later.")
    else:
        # ── Time filter ───────────────────────────────────────────────
        now_utc = datetime.now(timezone.utc)
        if time_filter == "Today":
            cutoff = now_utc - timedelta(hours=24)
        elif time_filter == "Last 3 Days":
            cutoff = now_utc - timedelta(days=3)
        else:
            cutoff = now_utc - timedelta(days=7)

        def passes_time(item):
            dt = item.get("published_dt")
            if dt is None:
                return True   # no date info → keep it
            return dt >= cutoff

        time_filtered = [n for n in news_items if passes_time(n)]

        # ── Stock filter ──────────────────────────────────────────────
        stock_tags = sorted(set(n.get("stock_tag", "") for n in time_filtered if n.get("stock_tag")))
        with col_stock:
            stock_filter = st.selectbox(
                "📌 Filter by Stock",
                ["All"] + stock_tags,
                key="news_stock_filter"
            )

        if stock_filter != "All":
            time_filtered = [n for n in time_filtered if n.get("stock_tag") == stock_filter]

        high_news   = [n for n in time_filtered if n.get("priority") == "🔴 HIGH"]
        normal_news = [n for n in time_filtered if n.get("priority") != "🔴 HIGH"]

        # ── Summary bar ───────────────────────────────────────────────
        st.markdown(f"""
<div style="display:flex;gap:16px;padding:10px 16px;background:#0f0f23;
            border-radius:8px;border:1px solid #2a2a5a;margin-bottom:14px;">
  <span style="color:#ff5252;font-weight:700;font-size:14px;">🔴 HIGH &nbsp;{len(high_news)}</span>
  <span style="color:#444466;">|</span>
  <span style="color:#ffaa00;font-weight:600;font-size:14px;">🟡 NORMAL &nbsp;{len(normal_news)}</span>
  <span style="color:#444466;">|</span>
  <span style="color:#8888aa;font-size:13px;">⏱ {time_filter} · {len(time_filtered)} total</span>
</div>
""", unsafe_allow_html=True)

        # ── Helper: render one news card ──────────────────────────────
        IST = timezone(timedelta(hours=5, minutes=30))

        def format_exact_time(dt):
            """Return exact date & time in IST, e.g. '11 May 2026, 03:45 PM'"""
            if dt is None:
                return "Time unknown"
            dt_ist = dt.astimezone(IST)
            return dt_ist.strftime("%d %b %Y, %I:%M %p")

        def news_card(item, border_color, pri_label, pri_color):
            stock_tag = item.get("stock_tag", "")
            title     = item.get("title", "No Title")
            summary   = item.get("summary", "")
            source    = item.get("source", "")
            link      = item.get("link", "#")
            pub_dt    = item.get("published_dt")
            pub_str   = format_exact_time(pub_dt)

            tag_html = (
                f'<span style="background:#1e1e3e;color:#8888aa;padding:2px 7px;'
                f'border-radius:4px;font-size:11px;margin-right:5px;">📌 {stock_tag}</span>'
            ) if stock_tag else ""

            return f"""
<div style="border-left:3px solid {border_color};padding:10px 13px;margin-bottom:10px;
            background:#0f0f23;border-radius:6px;">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">
    <div>{tag_html}<span style="color:{pri_color};font-size:11px;font-weight:600;">{pri_label}</span></div>
    <span style="color:#8888aa;font-size:11px;">🕐 {pub_str}</span>
  </div>
  <a href="{link}" target="_blank"
     style="color:#90caf9;font-size:14px;font-weight:600;text-decoration:none;line-height:1.4;">{title}</a>
  <div style="color:#6666aa;font-size:11px;margin-top:3px;">{source}</div>
  <div style="color:#8888aa;font-size:12px;margin-top:5px;">{summary[:180] + '…' if len(summary) > 180 else summary}</div>
</div>"""

        # ── Sort news newest-first ─────────────────────────────────────
        def sort_key(n):
            dt = n.get("published_dt")
            return dt if dt is not None else datetime.min.replace(tzinfo=timezone.utc)

        high_news   = sorted(high_news,   key=sort_key, reverse=True)
        normal_news = sorted(normal_news, key=sort_key, reverse=True)

        # ── Two-column layout ─────────────────────────────────────────
        col_high, col_normal = st.columns(2, gap="medium")

        with col_high:
            st.markdown(
                f'<div style="color:#ff5252;font-weight:700;font-size:15px;'
                f'border-bottom:2px solid #ff5252;padding-bottom:6px;margin-bottom:12px;">'
                f'🔴 HIGH PRIORITY &nbsp;<span style="font-size:12px;color:#cc4444;">({len(high_news)})</span></div>',
                unsafe_allow_html=True
            )
            if not high_news:
                st.markdown(
                    '<div style="color:#454870;font-size:13px;padding:12px;">No high-priority news in this period.</div>',
                    unsafe_allow_html=True
                )
            else:
                for item in high_news:
                    st.markdown(news_card(item, "#f85454", "🔴 HIGH", "#f85454"), unsafe_allow_html=True)

        with col_normal:
            st.markdown(
                f'<div style="color:#ffaa00;font-weight:700;font-size:15px;'
                f'border-bottom:2px solid #ffaa00;padding-bottom:6px;margin-bottom:12px;">'
                f'🟡 NORMAL &nbsp;<span style="font-size:12px;color:#cc8800;">({len(normal_news)})</span></div>',
                unsafe_allow_html=True
            )
            if not normal_news:
                st.markdown(
                    '<div style="color:#454870;font-size:13px;padding:12px;">No normal news in this period.</div>',
                    unsafe_allow_html=True
                )
            else:
                for item in normal_news:
                    st.markdown(news_card(item, "#444466", "🟡 NORMAL", "#ffaa00"), unsafe_allow_html=True)


# =====================================================
# TAB 5 — RESULTS CALENDAR
# =====================================================

if _nav_tab == "Corporate Actions":
    st.subheader("📅 Corporate Actions Calendar")

    st.caption(
        "Shows upcoming corporate actions for your portfolio — "
        "**Ex-Date** is the key date: you must hold shares *before* this date "
        "to be eligible for dividend / bonus / split benefits."
    )

    if st.button("🔄 Refresh Calendar", key="refresh_calendar"):
        st.cache_data.clear()
        st.session_state.pop("_ticker_news", None)
        st.session_state.pop("_ticker_cal_df", None)

    with st.spinner("📡 Fetching corporate actions & ex-dates from NSE..."):
        cal_df = get_results_calendar(tickers)
        st.session_state["_ticker_cal_df"] = cal_df

    if cal_df.empty:
        st.info("No upcoming corporate actions found for your portfolio in the next 90 days.")
    else:
        today_ts = pd.Timestamp.now().normalize()

        # ── Shared stats (used by both desktop and mobile) ────────────
        ex_date_soon = 0
        urgent_rows  = []
        if "Ex-Date" in cal_df.columns:
            for _, row in cal_df.iterrows():
                try:
                    v = row.get("Ex-Date", "—")
                    if v and v != "—":
                        ex_ts = pd.to_datetime(v, dayfirst=True)
                        if today_ts <= ex_ts <= today_ts + pd.Timedelta(days=7):
                            ex_date_soon += 1
                            urgent_rows.append(row)
                except Exception:
                    pass

        total       = len(cal_df)
        high_count  = len(cal_df[cal_df["Priority"] == "🔴 HIGH"])
        med_count   = len(cal_df[cal_df["Priority"] == "🟡 MEDIUM"])
        today_count = len(cal_df[cal_df["Days Away"] == "Today"])

        # ── Filters (shared — Streamlit widgets must render once) ──────
        col_ef, col_pf, col_sf = st.columns(3)
        with col_ef:
            event_types  = ["All"] + sorted(cal_df["Event"].unique().tolist())
            event_filter = st.selectbox("Filter by Event Type", event_types, key="cal_event_filter")
        with col_pf:
            priority_filter = st.selectbox("Filter by Priority", ["All", "🔴 HIGH", "🟡 MEDIUM"], key="cal_priority_filter")
        with col_sf:
            ticker_list   = ["All"] + sorted(cal_df["Ticker"].unique().tolist())
            ticker_filter = st.selectbox("Filter by Stock", ticker_list, key="cal_ticker_filter")

        display_df = cal_df.copy()
        if event_filter != "All":
            display_df = display_df[display_df["Event"] == event_filter]
        if priority_filter != "All":
            display_df = display_df[display_df["Priority"] == priority_filter]
        if ticker_filter != "All":
            display_df = display_df[display_df["Ticker"] == ticker_filter]

        # ══════════════════════════════════════════════════════════════
        # DESKTOP — unchanged original layout
        # ══════════════════════════════════════════════════════════════

        st.markdown(f"""
<div class="hdr-desktop" style="gap:16px;flex-wrap:wrap;padding:10px 16px;background:#0f0f23;
            border-radius:8px;border:1px solid #2a2a5a;margin-bottom:14px;">
  <span style="color:#ff5252;font-weight:700;font-size:14px;">🔴 HIGH &nbsp;{high_count}</span>
  <span style="color:#444466;">|</span>
  <span style="color:#ffaa00;font-weight:600;font-size:14px;">🟡 MEDIUM &nbsp;{med_count}</span>
  <span style="color:#444466;">|</span>
  <span style="color:#00e676;font-weight:600;font-size:14px;">📌 Today &nbsp;{today_count}</span>
  <span style="color:#444466;">|</span>
  <span style="color:#90caf9;font-weight:600;font-size:14px;">⚡ Ex-Date ≤7d &nbsp;{ex_date_soon}</span>
  <span style="color:#444466;">|</span>
  <span style="color:#8888aa;font-size:13px;">📋 {total} total events · next 90 days</span>
</div>
""", unsafe_allow_html=True)

        # Desktop alert banner
        if urgent_rows:
            alert_html = '<div class="hdr-desktop" style="background:#1a0a0a;border:1px solid #ff5252;border-radius:8px;padding:10px 14px;margin-bottom:14px;flex-direction:column;">'
            alert_html += '<div style="color:#ff5252;font-weight:700;font-size:13px;margin-bottom:8px;">⚡ Ex-Dates in next 7 days — act before these dates to be eligible!</div>'
            for r in urgent_rows:
                alert_html += (
                    f'<div style="color:#ffcccc;font-size:12px;margin-bottom:3px;">'
                    f'📌 <b>{r["Ticker"]}</b> — {r["Event"]} &nbsp;|&nbsp; '
                    f'Ex-Date: <span style="color:#ff9999;font-weight:700;">{r["Ex-Date"]}</span>'
                    f'{(" &nbsp;|&nbsp; " + r["Details"]) if r.get("Details") and r["Details"] != "—" else ""}'
                    f'</div>'
                )
            alert_html += '</div>'
            st.markdown(alert_html, unsafe_allow_html=True)

        st.caption(f"Showing {len(display_df)} of {total} corporate actions")

        def style_row(row):
            styles = [""] * len(row)
            days_val     = row.get("Days Away", "")
            priority_val = row.get("Priority", "")
            ex_date_val  = row.get("Ex-Date", "—")
            col_idx      = list(row.index)
            ex_urgent = False
            ex_today  = False
            try:
                if ex_date_val and ex_date_val != "—":
                    ex_ts2 = pd.to_datetime(ex_date_val, dayfirst=True)
                    days_to_ex = (ex_ts2 - today_ts).days
                    if days_to_ex == 0:   ex_today  = True
                    elif days_to_ex <= 3: ex_urgent = True
            except Exception:
                pass
            for i, col in enumerate(col_idx):
                if col == "Ex-Date":
                    if ex_today:    styles[i] = "color: #ff5252; font-weight: 800; background-color: #2a0a0a"
                    elif ex_urgent: styles[i] = "color: #ffaa00; font-weight: 700"
                    elif ex_date_val and ex_date_val != "—": styles[i] = "color: #90caf9; font-weight: 600"
                    else:           styles[i] = "color: #555577"
                elif col == "Days Away":
                    if days_val == "Today": styles[i] = "color: #ff5252; font-weight: 700"
                    else:
                        try:
                            d = int(str(days_val).replace("In ", "").replace("d", ""))
                            if d <= 3:   styles[i] = "color: #ffaa00; font-weight: 600"
                            elif d <= 7: styles[i] = "color: #f0f2ff; font-weight: 500"
                            else:        styles[i] = "color: #aaaacc"
                        except Exception:
                            styles[i] = "color: #aaaacc"
                elif col == "Priority":
                    styles[i] = "color: #ff5252; font-weight: 700" if priority_val == "🔴 HIGH" else "color: #ffaa00; font-weight: 600"
                elif col == "Event":
                    styles[i] = "font-weight: 600"
                elif col == "Record Date":
                    styles[i] = "color: #aaffcc" if row.get("Record Date", "—") != "—" else "color: #555577"
            return styles

        styled_cal = display_df.style.apply(style_row, axis=1)
        st.dataframe(styled_cal, use_container_width=True, hide_index=True)

        st.markdown("""
<div class="hdr-desktop" style="font-size:11px;color:#666688;margin-top:6px;padding:6px 10px;
            background:#0b0d17;border-radius:6px;border:1px solid #252849;flex-direction:column;">
  <b style="color:#8888aa;">Legend:</b> &nbsp;
  <span style="color:#ff5252;">■</span> Ex-Date = Today &nbsp;|&nbsp;
  <span style="color:#ffaa00;">■</span> Ex-Date ≤ 3 days &nbsp;|&nbsp;
  <span style="color:#90caf9;">■</span> Ex-Date upcoming &nbsp;|&nbsp;
  <span style="color:#aaffcc;">■</span> Record Date available &nbsp;&nbsp;
  <b style="color:#8888aa;">Note:</b> Buy shares <i>before</i> Ex-Date to receive dividend/bonus/split benefits.
</div>
""", unsafe_allow_html=True)

        st.markdown("---")
        col_summary, col_exdates = st.columns(2)

        with col_summary:
            st.markdown("**📊 Corporate Action Summary**")
            event_counts = cal_df["Event"].value_counts().reset_index()
            event_counts.columns = ["Event Type", "Count"]
            st.dataframe(event_counts, use_container_width=True, hide_index=True)

        with col_exdates:
            st.markdown("**📌 Upcoming Ex-Dates (all events)**")
            if "Ex-Date" in cal_df.columns:
                ex_rows = cal_df[cal_df["Ex-Date"] != "—"].copy()
                if not ex_rows.empty:
                    try:
                        ex_rows["_ex_sort"] = pd.to_datetime(ex_rows["Ex-Date"], dayfirst=True, errors="coerce")
                        ex_rows = ex_rows.sort_values("_ex_sort").drop(columns=["_ex_sort"])
                    except Exception:
                        pass
                    st.dataframe(
                        ex_rows[["Ticker", "Event", "Ex-Date", "Record Date", "Details"]],
                        use_container_width=True, hide_index=True
                    )
                else:
                    st.info("No ex-dates available from NSE at this time.")

        # ══════════════════════════════════════════════════════════════
        # MOBILE — completely new card-based layout
        # ══════════════════════════════════════════════════════════════

        # ── 1. Stat pills grid ─────────────────────────────────────────
        st.markdown(f"""
<div class="hdr-mobile">
  <div style="font-size:10px;color:#6b7099;font-weight:700;text-transform:uppercase;
              letter-spacing:1px;margin-bottom:12px;display:flex;align-items:center;gap:8px;">
    <span style="display:inline-block;width:20px;height:2px;background:linear-gradient(90deg,#4f7ef8,transparent);border-radius:2px;"></span>
    Portfolio CA Feed &nbsp;·&nbsp; {total} Events · Next 90 Days
    <span style="display:inline-block;width:20px;height:2px;background:linear-gradient(90deg,transparent,#4f7ef8);border-radius:2px;"></span>
  </div>
  <div class="ca-stat-grid">
    <div class="ca-stat-pill" style="border-top:3px solid #f85454;">
      <div style="font-size:18px;margin-bottom:2px;position:relative;z-index:1;">🔴</div>
      <div class="ca-stat-count" style="color:#f85454;">{high_count}</div>
      <div class="ca-stat-label">High Priority</div>
    </div>
    <div class="ca-stat-pill" style="border-top:3px solid #ffaa00;">
      <div style="font-size:18px;margin-bottom:2px;position:relative;z-index:1;">🟡</div>
      <div class="ca-stat-count" style="color:#ffaa00;">{med_count}</div>
      <div class="ca-stat-label">Medium</div>
    </div>
    <div class="ca-stat-pill" style="border-top:3px solid #22d67b;">
      <div style="font-size:18px;margin-bottom:2px;position:relative;z-index:1;">📌</div>
      <div class="ca-stat-count" style="color:#22d67b;">{today_count}</div>
      <div class="ca-stat-label">Today</div>
    </div>
    <div class="ca-stat-pill" style="border-top:3px solid #4f7ef8;">
      <div style="font-size:18px;margin-bottom:2px;position:relative;z-index:1;">⚡</div>
      <div class="ca-stat-count" style="color:#4f7ef8;">{ex_date_soon}</div>
      <div class="ca-stat-label">Ex-Date ≤7d</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── 2. Urgent alert banner ─────────────────────────────────────
        if urgent_rows:
            mob_alert = '<div class="hdr-mobile ca-alert-mob">'
            mob_alert += f'<div class="alert-title"><span style="font-size:15px;">🚨</span> Act Before Ex-Date &nbsp;·&nbsp; {len(urgent_rows)} Stocks in 7 Days</div>'
            for r in urgent_rows:
                days_str = str(r.get("Days Away", ""))
                mob_alert += f"""
<div class="alert-row">
  <div>
    <div class="alert-ticker">{r["Ticker"]}</div>
    <div class="alert-event">{r["Event"]}</div>
  </div>
  <div style="text-align:right;">
    <div class="alert-date">{r["Ex-Date"]}</div>
    <div class="alert-daysaway">{days_str}</div>
  </div>
</div>"""
            mob_alert += '</div>'
            st.markdown(mob_alert, unsafe_allow_html=True)

        # ── 3. Cards list ──────────────────────────────────────────────
        # Event-type accent colours
        _CA_COLORS = {
            "Dividend":              "#90caf9",
            "Bonus":                 "#00e676",
            "Stock Split":           "#69f0ae",
            "Rights Issue":          "#ffcc02",
            "Buyback":               "#ff9800",
            "Merger / Demerger":     "#ce93d8",
            "Acquisition":           "#f48fb1",
            "Delisting":             "#f85454",
            "AGM":                   "#8888aa",
            "EGM":                   "#bbbbdd",
            "Financial Results":     "#ffd54f",
            "Preferential Allotment":"#ffab40",
            "NCD Issue":             "#80cbc4",
            "QIP":                   "#a5d6a7",
            "ESOP":                  "#ce93d8",
        }

        def _ex_date_color(ex_val):
            try:
                if ex_val and ex_val != "—":
                    ex_ts3 = pd.to_datetime(ex_val, dayfirst=True)
                    d = (ex_ts3 - today_ts).days
                    if d == 0:   return "#f85454", "⚠️ TODAY"
                    elif d <= 3: return "#ffaa00", f"In {d}d"
                    elif d <= 7: return "#f0f2ff", f"In {d}d"
                    elif d > 0:  return "#90caf9", f"In {d}d"
                    else:        return "#555577", "Passed"
            except Exception:
                pass
            return "#555577", "—"

        if display_df.empty:
            st.markdown(
                '<div class="hdr-mobile" style="text-align:center;padding:32px 16px;color:#454870;font-size:13px;">'
                '📭 No corporate actions match your filters.</div>',
                unsafe_allow_html=True
            )
        else:
            # Group by priority for visual separation
            high_df   = display_df[display_df["Priority"] == "🔴 HIGH"]
            other_df  = display_df[display_df["Priority"] != "🔴 HIGH"]

            def _render_ca_cards(df_section, section_label, section_color):
                if df_section.empty:
                    return
                st.markdown(
                    f'<div class="hdr-mobile ca-section-hdr" style="color:{section_color};">'
                    f'{section_label} <span style="color:#333660;font-weight:500;font-size:9px;">({len(df_section)})</span></div>',
                    unsafe_allow_html=True
                )
                for _, row in df_section.iterrows():
                    ticker      = row.get("Ticker", "—")
                    event       = row.get("Event", "—")
                    ex_date     = row.get("Ex-Date", "—")
                    rec_date    = row.get("Record Date", "—")
                    details     = row.get("Details", "—")
                    days_away   = str(row.get("Days Away", "—"))
                    priority    = row.get("Priority", "")
                    ann_date    = row.get("Announcement Date", "—") if "Announcement Date" in row.index else "—"

                    ev_color    = _CA_COLORS.get(event, "#7a7fa8")
                    ex_color, ex_badge = _ex_date_color(ex_date)

                    # Priority badge styles
                    if priority == "🔴 HIGH":
                        pri_bg    = "rgba(248,84,84,0.18)"
                        pri_color = "#f87171"
                        pri_icon  = "🔴"
                        pri_text  = "HIGH"
                        card_glow = f"0 0 0 1px rgba(248,84,84,0.15), 0 6px 28px rgba(0,0,0,0.45)"
                    else:
                        pri_bg    = "rgba(255,170,0,0.15)"
                        pri_color = "#ffb733"
                        pri_icon  = "🟡"
                        pri_text  = "MED"
                        card_glow = f"0 6px 28px rgba(0,0,0,0.4)"

                    # Ex-date chip colour + bg
                    if ex_badge == "⚠️ TODAY":
                        chip_bg  = "rgba(248,84,84,0.2)"
                        chip_col = "#f85454"
                    elif "3" in days_away or "2" in days_away or "1" in days_away:
                        chip_bg  = "rgba(255,170,0,0.18)"
                        chip_col = "#ffaa00"
                    elif ex_color == "#90caf9":
                        chip_bg  = "rgba(144,202,249,0.12)"
                        chip_col = "#90caf9"
                    else:
                        chip_bg  = "rgba(85,85,119,0.15)"
                        chip_col = "#555577"

                    # Subtle event-coloured left border glow
                    ev_rgba = ev_color.replace("#", "")
                    ev_r = int(ev_rgba[0:2], 16) if len(ev_rgba) == 6 else 100
                    ev_g = int(ev_rgba[2:4], 16) if len(ev_rgba) == 6 else 100
                    ev_b = int(ev_rgba[4:6], 16) if len(ev_rgba) == 6 else 100

                    details_html = ""
                    if details and details != "—":
                        details_html = f'<div class="ca-card-footer"><span style="opacity:0.5;">📋</span>&nbsp;{details}</div>'

                    rec_col = "#6ee7a0" if rec_date != "—" else "#333660"

                    st.markdown(f"""
<div class="hdr-mobile ca-card-mob"
     style="border-top:3px solid {ev_color};
            box-shadow:{card_glow}, 0 0 0 1px rgba({ev_r},{ev_g},{ev_b},0.08);">

  <!-- ── Header: Ticker + Event type + Priority badge ── -->
  <div class="ca-card-header">
    <div class="ca-ticker-block">
      <span class="ca-ticker">{ticker}</span>
      <span class="ca-company" style="color:{ev_color};">
        <span style="display:inline-block;width:6px;height:6px;border-radius:50%;
                     background:{ev_color};margin-right:5px;vertical-align:middle;
                     box-shadow:0 0 6px {ev_color}55;"></span>{event}
      </span>
    </div>
    <span class="ca-priority-badge"
          style="background:{pri_bg};color:{pri_color};
                 border:1px solid {pri_color}30;
                 box-shadow:0 2px 8px {pri_color}20;">
      {pri_icon}&nbsp;{pri_text}
    </span>
  </div>

  <!-- ── Body: data fields ── -->
  <div class="ca-card-body">
    <div class="ca-field">
      <span class="ca-field-label">Ex-Date</span>
      <span class="ca-field-value" style="color:{ex_color};">{ex_date}</span>
      <span class="ca-ex-chip" style="background:{chip_bg};color:{chip_col};
                                      border:1px solid {chip_col}30;">
        ⏱ {ex_badge}
      </span>
    </div>
    <div class="ca-field">
      <span class="ca-field-label">Days Away</span>
      <span class="ca-field-value" style="color:{ex_color};font-size:18px;
             letter-spacing:-0.5px;">{days_away}</span>
    </div>
    <div class="ca-field">
      <span class="ca-field-label">Record Date</span>
      <span class="ca-field-value" style="color:{rec_col};">{rec_date}</span>
    </div>
    <div class="ca-field">
      <span class="ca-field-label">Announced</span>
      <span class="ca-field-value" style="color:#555577;">{ann_date}</span>
    </div>
  </div>

  {details_html}
</div>""", unsafe_allow_html=True)

            _render_ca_cards(high_df,  "🔴 HIGH PRIORITY", "#f85454")
            _render_ca_cards(other_df, "🟡 MEDIUM / OTHER", "#ffaa00")

        # ── 4. Mobile legend ───────────────────────────────────────────
        st.markdown("""
<div class="hdr-mobile ca-legend-mob">
  <div style="font-size:10px;font-weight:800;color:#7a7fa8;text-transform:uppercase;
              letter-spacing:0.8px;margin-bottom:8px;">📖 Guide</div>
  <div style="display:flex;flex-wrap:wrap;gap:6px 14px;">
    <span><span style="color:#f85454;">●</span> Ex-Date Today</span>
    <span><span style="color:#ffaa00;">●</span> ≤3 days</span>
    <span><span style="color:#90caf9;">●</span> Upcoming</span>
    <span><span style="color:#6ee7a0;">●</span> Record Date set</span>
    <span><span style="color:#00e676;">●</span> Bonus</span>
    <span><span style="color:#90caf9;">●</span> Dividend</span>
    <span><span style="color:#ffcc02;">●</span> Rights</span>
    <span><span style="color:#ff9800;">●</span> Buyback</span>
  </div>
  <div style="margin-top:10px;padding-top:10px;border-top:1px solid #1a1e38;
              font-size:10px;color:#454870;line-height:1.7;">
    ⚠️ <b style="color:#7a7fa8;">Important:</b> You must hold shares
    <b style="color:#c0c4e8;">before</b> Ex-Date to receive dividend, bonus or split benefits.
  </div>
</div>
""", unsafe_allow_html=True)
# =====================================================
# TAB 6 — XIRR & RETURNS
# =====================================================

if _nav_tab == "XIRR & Returns":
    # ── Beautiful header ──────────────────────────────────────────────
    st.markdown("""
<div style="margin-bottom:20px;">
  <span style="font-size:16px;font-weight:800;color:#f0f2ff;">📈 XIRR &amp; Annualised Returns</span>
  <span style="font-size:11px;color:#7a7fa8;margin-left:10px;">
    XIRR = time-weighted annualised return accounting for exact buy dates &amp; cashflows
  </span>
</div>
""", unsafe_allow_html=True)

    # ── Portfolio-level XIRR — big hero card ──────────────────────────
    try:
        port_cf = []
        for _, row in calc.iterrows():
            bd = pd.to_datetime(row.get("Buy_Date",""), errors="coerce")
            if not pd.isna(bd):
                bd = bd.date()
                port_cf.append((bd, -float(row["Cost_Basis"])))
        if not trades_df.empty:
            for _, tr in trades_df.iterrows():
                sd = pd.to_datetime(tr["Sell_Date"], errors="coerce")
                if not pd.isna(sd):
                    port_cf.append((sd.date(), float(tr["Sell_Qty"]) * float(tr["Sell_Price"])))
        port_cf.append((datetime.today().date(), float(total_value)))
        port_xirr = xirr(port_cf)
        if port_xirr is not None:
            px_pct = port_xirr * 100
            # Compute earliest holding date for short-hold warning
            port_dates = [cf[0] for cf in port_cf if cf[1] < 0]
            min_port_days = (datetime.today().date() - min(port_dates)).days if port_dates else 365
            px_color = "#22d67b" if px_pct >= 0 else "#f85454"
            px_glow  = "rgba(34,214,123,0.18)" if px_pct >= 0 else "rgba(248,84,84,0.18)"
            px_border= "#22d67b55" if px_pct >= 0 else "#f8545455"
            px_grad  = "linear-gradient(135deg,#0f1a14 0%,#0c1a22 100%)" if px_pct >= 0 else "linear-gradient(135deg,#1a0f0f 0%,#1a0c14 100%)"
            px_arrow = "▲" if px_pct >= 0 else "▼"
            px_label = ("Exceptional returns 🔥" if px_pct > 50 else
                        "Outperforming market 🎯" if px_pct > 18 else
                        "Beating fixed income 📊" if px_pct > 7 else
                        "Underperforming 📉" if px_pct < 0 else "Moderate returns ⚖️")
            # Simple absolute return for comparison
            abs_ret = (total_pnl / total_invested * 100) if total_invested else 0
            abs_color = "#22d67b" if abs_ret >= 0 else "#f85454"
            short_note_html = (
                '<div style="font-size:10px;color:#f5c842;margin-top:8px;padding:7px 12px;'
                'background:rgba(245,200,66,0.06);border-radius:8px;border-left:3px solid #f5c842;">'
                '⚠️ Avg. holding &lt; 30 days — annualised XIRR is indicative, not conclusive.'
                '</div>'
            ) if min_port_days < 30 else ""
            st.markdown(f"""
<div style="background:{px_grad}; border:1px solid {px_border};
            border-radius:20px; padding:24px 28px; margin-bottom:24px;
            box-shadow: 0 0 40px {px_glow}, 0 2px 8px rgba(0,0,0,0.3);">
  <div style="display:flex; align-items:flex-start; justify-content:space-between; flex-wrap:wrap; gap:16px;">
    <div style="flex:1; min-width:160px;">
      <div style="font-size:11px; color:#8888aa; text-transform:uppercase; letter-spacing:1.5px;
                  font-weight:700; margin-bottom:6px;">🏆 Portfolio XIRR — Annualised Return</div>
      <div style="font-size:52px; font-weight:900; color:{px_color}; line-height:1;
                  letter-spacing:-2px;">{px_arrow} {px_pct:+.2f}%</div>
      <div style="font-size:12px; color:{px_color}; margin-top:8px; font-weight:600; opacity:0.8;">
        {px_label}
      </div>
      <div style="font-size:11px; color:#555577; margin-top:4px;">
        Weighted across all buy dates · Includes sell cashflows
      </div>
      {short_note_html}
    </div>
    <div style="display:flex; flex-direction:column; gap:12px; min-width:160px;">
      <div style="background:rgba(255,255,255,0.04); border:1px solid #2a2e52;
                  border-radius:12px; padding:14px 16px; text-align:center;">
        <div style="font-size:10px; color:#7a7fa8; text-transform:uppercase; letter-spacing:0.8px; margin-bottom:4px;">Absolute Return</div>
        <div style="font-size:20px; font-weight:800; color:{abs_color};">{'▲' if abs_ret >= 0 else '▼'} {abs_ret:+.2f}%</div>
      </div>
      <div style="background:rgba(255,255,255,0.04); border:1px solid #2a2e52;
                  border-radius:12px; padding:14px 16px; text-align:center;">
        <div style="font-size:10px; color:#7a7fa8; text-transform:uppercase; letter-spacing:0.8px; margin-bottom:4px;">Total Unrealized P&amp;L</div>
        <div style="font-size:18px; font-weight:800; color:{abs_color};">₹{total_pnl:,.0f}</div>
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)
    except Exception:
        pass

    # ── Per-holding XIRR cards ────────────────────────────────────────
    st.markdown("""
<div style="font-size:13px; font-weight:700; color:#8888aa; text-transform:uppercase;
            letter-spacing:1px; margin:4px 0 14px 0;">Per-Holding XIRR Breakdown</div>
""", unsafe_allow_html=True)

    xirr_rows = []
    for _, row in calc.iterrows():
        t = row["Ticker"]
        xi = xirr_map.get(t)
        xirr_rows.append({
            "Ticker":       t,
            "Asset Type":   row.get("Asset_Type",""),
            "Cost Basis ₹": round(float(row.get("Cost_Basis",0)),2),
            "Value ₹":      round(float(row.get("Value",0)),2),
            "Unreal P&L ₹": round(float(row.get("Unrealized_PnL",0)),2),
            "P&L %":        round(float(row.get("PnL_%",0)),2),
            "XIRR %":       xi if xi is not None else None,
            "_buy_date":    row.get("Buy_Date",""),
        })
    xirr_table = pd.DataFrame(xirr_rows)

    # Render beautiful cards in a grid
    cols_per_row = 3
    card_rows = [xirr_rows[i:i+cols_per_row] for i in range(0, len(xirr_rows), cols_per_row)]
    for cgroup in card_rows:
        cols = st.columns(len(cgroup))
        for col, item in zip(cols, cgroup):
            xi = item["XIRR %"]
            pnl_p = item["P&L %"]
            unreal = item["Unreal P&L ₹"]
            atype  = item["Asset Type"]

            # Colors
            if xi is not None:
                xi_f = float(xi)
                xi_col = "#22d67b" if xi_f >= 0 else "#f85454"
                xi_bg  = "rgba(34,214,123,0.07)" if xi_f >= 0 else "rgba(248,84,84,0.07)"
                xi_border = "#22d67b44" if xi_f >= 0 else "#f8545444"
                xi_str = f"{'▲' if xi_f > 0 else '▼'} {xi_f:+.2f}%"
                xi_label = ("Exceptional 🔥" if xi_f > 50 else
                            "Excellent ✨" if xi_f > 25 else
                            "Strong 📈" if xi_f > 15 else
                            "Good" if xi_f > 10 else
                            "Moderate" if xi_f > 0 else "Loss 📉")
                xi_bar_w = min(100, abs(xi_f) / 2)   # 0–200% maps to 0–100% bar width
                xi_bar_col = xi_col
            else:
                xi_f = None
                xi_col = "#7a7fa8"
                xi_bg  = "rgba(122,127,168,0.05)"
                xi_border = "#2a2e52"
                xi_str = "N/A"
                xi_label = "No data"
                xi_bar_w = 0
                xi_bar_col = "#3a3d5c"

            pnl_col   = "#22d67b" if float(pnl_p) >= 0 else "#f85454"
            unr_col   = "#22d67b" if float(unreal) >= 0 else "#f85454"

            # Badge
            if atype == "Stock": badge, badge_c = "STOCK", "#4f7ef8"
            elif atype == "ETF": badge, badge_c = "ETF",   "#a78bfa"
            elif atype == "F&O": badge, badge_c = "F&O",   "#f5c842"
            else: badge, badge_c = atype[:4], "#7a7fa8"

            # Buy date display
            bd_str = "—"
            short_hold_note = ""
            try:
                bd = pd.to_datetime(item.get("_buy_date",""), errors="coerce")
                if not pd.isna(bd):
                    bd_str = bd.strftime("%d %b %Y")
                    days_held = (datetime.today() - bd).days
                    bd_str += f" · {days_held}d held"
                    if days_held < 30 and xi_f is not None:
                        short_hold_note = '<div style="font-size:9px;color:#f5c842;margin-top:4px;">⚠️ Holding < 30 days · annualised XIRR is indicative</div>'
            except: pass

            col.markdown(f"""
<div style="background:{xi_bg}; border:1px solid {xi_border}; border-radius:14px;
            padding:16px; margin-bottom:10px; position:relative; overflow:hidden;">
  <div style="position:absolute; top:0; left:0; right:0; height:3px;
              background:{xi_col}; border-radius:14px 14px 0 0;"></div>
  <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:10px;">
    <div>
      <div style="font-size:14px; font-weight:800; color:#f0f2ff; letter-spacing:0.3px;">{item["Ticker"]}</div>
      <span style="background:{badge_c}22; color:{badge_c}; font-size:9px; font-weight:800;
                   padding:2px 7px; border-radius:4px;">{badge}</span>
    </div>
    <div style="text-align:right;">
      <div style="font-size:22px; font-weight:900; color:{xi_col}; line-height:1.1;">{xi_str}</div>
      <div style="font-size:9px; color:{xi_col}; font-weight:700; opacity:0.7;">{xi_label}</div>
    </div>
  </div>
  <div style="background:rgba(0,0,0,0.2); border-radius:4px; height:4px; margin-bottom:12px; overflow:hidden;">
    <div style="width:{xi_bar_w}%; background:{xi_bar_col}; height:100%; border-radius:4px;
                transition:width 0.5s;"></div>
  </div>
  <div style="display:grid; grid-template-columns:1fr 1fr; gap:6px; font-size:11px;">
    <div style="color:#6b7299;">Cost Basis</div>
    <div style="color:#c8cce8; text-align:right; font-weight:600;">₹{item["Cost Basis ₹"]:,.0f}</div>
    <div style="color:#6b7299;">Curr. Value</div>
    <div style="color:#c8cce8; text-align:right; font-weight:600;">₹{item["Value ₹"]:,.0f}</div>
    <div style="color:#6b7299;">Unreal P&amp;L</div>
    <div style="color:{unr_col}; text-align:right; font-weight:700;">₹{unreal:+,.0f}</div>
    <div style="color:#6b7299;">Abs. Return</div>
    <div style="color:{pnl_col}; text-align:right; font-weight:700;">{pnl_p:+.2f}%</div>
  </div>
  <div style="margin-top:8px; font-size:10px; color:#454870; border-top:1px solid #1a1e38; padding-top:8px;">
    📅 {bd_str}
  </div>
  {short_hold_note}
</div>
""", unsafe_allow_html=True)

    # XIRR bar chart
    st.markdown("---")
    xi_chart_data = [(r["Ticker"], r["XIRR %"]) for r in xirr_rows if r["XIRR %"] is not None]
    if xi_chart_data:
        xi_df = pd.DataFrame(xi_chart_data, columns=["Ticker", "XIRR %"])
        xi_df = xi_df.sort_values("XIRR %", ascending=False)
        xi_colors = ["#22d67b" if v >= 0 else "#f85454" for v in xi_df["XIRR %"]]
        fig_xi = go.Figure(go.Bar(
            x=xi_df["Ticker"],
            y=xi_df["XIRR %"],
            marker_color=xi_colors,
            marker_line_width=0,
            text=xi_df["XIRR %"].apply(lambda v: f"{v:+.2f}%"),
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>XIRR: %{y:+.2f}%<extra></extra>"
        ))
        fig_xi.update_layout(
            title=dict(text="XIRR % per Holding (Annualised Return)", font=dict(size=14, color="#f0f2ff")),
            xaxis_title="", yaxis_title="XIRR %",
            plot_bgcolor="#0a0c1e", paper_bgcolor="#111428",
            font=dict(color="#b8bcd8", family="Inter"),
            yaxis=dict(gridcolor="#1a1e38", zerolinecolor="#4f7ef8", zerolinewidth=2,
                       tickformat="+.1f", ticksuffix="%"),
            xaxis=dict(gridcolor="#1a1e38", tickangle=-30),
            showlegend=False, height=400,
            margin=dict(t=50, b=10, l=10, r=10)
        )
        st.plotly_chart(fig_xi, use_container_width=True)

    # Beta breakdown
    st.markdown("---")
    st.markdown("""
<div style="font-size:13px; font-weight:700; color:#8888aa; text-transform:uppercase;
            letter-spacing:1px; margin-bottom:14px;">📐 Portfolio Beta Breakdown</div>
""", unsafe_allow_html=True)

    beta_color = "#f85454" if portfolio_beta > 1.2 else "#ffaa00" if portfolio_beta > 1.0 else "#22d67b"
    beta_label = ('⚠️ High risk — portfolio moves more than market' if portfolio_beta > 1.2
                  else '🟡 Slightly aggressive' if portfolio_beta > 1.0
                  else '✅ Defensive / stable')
    st.markdown(f"""
<div style="background:rgba(245,200,66,0.06); border:1px solid {beta_color}44;
            border-radius:14px; padding:18px 24px; margin-bottom:16px;
            display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:12px;">
  <div>
    <div style="font-size:10px; color:#8888aa; text-transform:uppercase; letter-spacing:1px; margin-bottom:4px;">Portfolio Beta</div>
    <div style="font-size:42px; font-weight:900; color:{beta_color}; line-height:1;">{portfolio_beta:.3f}</div>
  </div>
  <div style="font-size:13px; color:{beta_color}; font-weight:600; max-width:280px;">{beta_label}</div>
  <div style="font-size:11px; color:#454870;">Beta > 1 → amplifies market moves<br>Beta &lt; 1 → defensive, dampens volatility</div>
</div>
""", unsafe_allow_html=True)
    st.dataframe(beta_df[["Ticker","Beta","Weight","Weighted_Beta"]].style.format({
        "Beta": "{:.3f}", "Weight": "{:.2%}", "Weighted_Beta": "{:.4f}"
    }), use_container_width=True, hide_index=True)


# =====================================================
# TAB 7 — CAPITAL GAINS
# =====================================================

if _nav_tab == "Capital Gains":
    st.subheader("⚖️ Capital Gains Report — STCG & LTCG")
    st.caption("STCG = holding ≤ 365 days | LTCG = holding > 365 days")

    if cap_gains_df.empty:
        st.info("No sell trades recorded. Record sells from the sidebar to see capital gains.")
    else:
        stcg_df = cap_gains_df[cap_gains_df["Gain Type"] == "STCG"]
        ltcg_df = cap_gains_df[cap_gains_df["Gain Type"] == "LTCG"]

        stcg_total = stcg_df["Booked P&L ₹"].sum() if not stcg_df.empty else 0
        ltcg_total = ltcg_df["Booked P&L ₹"].sum() if not ltcg_df.empty else 0

        cg1, cg2 = st.columns(2)
        def _cg_card(col, label, amount, count, rate_note):
            color = "#22d67b" if amount >= 0 else "#f85454"
            col.markdown(f"""
<div style="background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:16px;text-align:center;">
  <div style="font-size:12px;color:#8888aa;">{label}</div>
  <div style="font-size:24px;font-weight:700;color:{color};margin-top:4px;">₹{amount:,.2f}</div>
  <div style="font-size:11px;color:#8888aa;margin-top:2px;">{count} trades &nbsp;|&nbsp; {rate_note}</div>
</div>
""", unsafe_allow_html=True)

        _cg_card(cg1, "📅 Short-Term Capital Gains (STCG)", stcg_total, len(stcg_df), "Taxed at 20%")
        _cg_card(cg2, "📆 Long-Term Capital Gains (LTCG)", ltcg_total, len(ltcg_df), "₹1.25L exempt, 12.5% above")

        st.markdown("---")
        # Filter
        filter_type = st.radio("Show:", ["All", "STCG Only", "LTCG Only"], horizontal=True, key="cg_filter")
        disp_cg = cap_gains_df.copy()
        if filter_type == "STCG Only":
            disp_cg = stcg_df
        elif filter_type == "LTCG Only":
            disp_cg = ltcg_df

        def _cg_style(val):
            try:
                return "color: #00e676; font-weight:600" if float(val) >= 0 else "color: #ff5252; font-weight:600"
            except:
                return ""

        st.dataframe(
            disp_cg.style.map(_cg_style, subset=["Booked P&L ₹"]),
            use_container_width=True, hide_index=True
        )

        st.markdown("---")
        # Tax estimate
        st.subheader("🧾 Estimated Tax Liability")
        ltcg_taxable = max(0, ltcg_total - 125000)
        stcg_tax  = max(0, stcg_total) * 0.20
        ltcg_tax  = max(0, ltcg_taxable) * 0.125
        total_tax = stcg_tax + ltcg_tax

        tc1, tc2, tc3 = st.columns(3)
        tc1.metric("STCG Tax (20%)",   f"₹{stcg_tax:,.2f}")
        tc2.metric("LTCG Tax (12.5%)", f"₹{ltcg_tax:,.2f}", delta=f"After ₹1.25L exemption")
        tc3.metric("Total Est. Tax",   f"₹{total_tax:,.2f}")
        st.caption("⚠️ This is an estimate only. Consult a CA for accurate tax computation.")


# =====================================================
# TAB 8 — REPORTS & EXPORT
# =====================================================

if _nav_tab == "Reports & Export":

    # ══════════════════════════════════════════════════════════════════
    # MOBILE VERSION — Stunning card-based UI
    # ══════════════════════════════════════════════════════════════════
    if _IS_MOBILE:
        # ── Mobile CSS ──────────────────────────────────────────────
        st.markdown("""
<style>
/* ─── Mobile Reports — CSS ─────────────────────────── */
@keyframes shimmer {
  0%   { background-position: -400px 0; }
  100% { background-position:  400px 0; }
}
@keyframes pulse-glow {
  0%, 100% { box-shadow: 0 0 12px 2px rgba(79,126,248,0.25); }
  50%       { box-shadow: 0 0 22px 6px rgba(79,126,248,0.55); }
}
@keyframes float-up {
  0%   { opacity:0; transform:translateY(18px); }
  100% { opacity:1; transform:translateY(0); }
}
@keyframes spin-slow {
  0%   { transform: rotate(0deg); }
  100% { transform: rotate(360deg); }
}

/* Hero banner */
.rpt-hero {
    background: linear-gradient(135deg, #0f1535 0%, #1a1040 50%, #0f2035 100%);
    border: 1px solid #2e3355;
    border-radius: 20px;
    padding: 28px 20px 22px;
    margin-bottom: 20px;
    text-align: center;
    position: relative;
    overflow: hidden;
    animation: float-up 0.5s ease both;
}
.rpt-hero::before {
    content: '';
    position: absolute;
    inset: 0;
    background: linear-gradient(90deg,
        transparent 0%, rgba(79,126,248,0.06) 50%, transparent 100%);
    background-size: 400px 100%;
    animation: shimmer 3s infinite linear;
}
.rpt-hero-icon {
    font-size: 44px;
    display: block;
    margin-bottom: 8px;
    filter: drop-shadow(0 0 12px rgba(79,126,248,0.6));
}
.rpt-hero-title {
    font-size: 22px;
    font-weight: 800;
    background: linear-gradient(90deg, #7eb8ff, #a78bfa, #67e8f9);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    letter-spacing: -0.3px;
    margin-bottom: 4px;
}
.rpt-hero-sub {
    font-size: 12px;
    color: #7a7fa8;
    letter-spacing: 0.4px;
}
.rpt-hero-client {
    display: inline-block;
    background: rgba(79,126,248,0.15);
    border: 1px solid rgba(79,126,248,0.35);
    border-radius: 20px;
    padding: 4px 14px;
    font-size: 12px;
    font-weight: 600;
    color: #90b8ff;
    margin-top: 10px;
}

/* Section header */
.rpt-section-hdr {
    display: flex;
    align-items: center;
    gap: 9px;
    margin: 22px 0 12px;
    animation: float-up 0.4s ease both;
}
.rpt-section-hdr-icon {
    width: 34px; height: 34px;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 16px;
    flex-shrink: 0;
}
.rpt-section-hdr-text { font-size: 14px; font-weight: 700; color: #e0e4ff; }
.rpt-section-hdr-sub  { font-size: 11px; color: #7a7fa8; margin-top: 1px; }

/* Export cards */
.rpt-export-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
    margin-bottom: 6px;
}
.rpt-export-card {
    background: #131628;
    border: 1px solid #252849;
    border-radius: 16px;
    padding: 18px 14px 16px;
    text-align: center;
    position: relative;
    overflow: hidden;
    animation: float-up 0.45s ease both;
    transition: border-color 0.2s, transform 0.15s;
}
.rpt-export-card:active { transform: scale(0.97); }
.rpt-export-card::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 3px;
    border-radius: 0 0 16px 16px;
}
.rpt-card-pdf::after  { background: linear-gradient(90deg,#f85454,#ff9f7a); }
.rpt-card-xl::after   { background: linear-gradient(90deg,#22d67b,#00e5c0); }
.rpt-card-wa::after   { background: linear-gradient(90deg,#25d366,#128c7e); }
.rpt-card-email::after{ background: linear-gradient(90deg,#4f7ef8,#a78bfa); }

.rpt-card-icon {
    font-size: 30px;
    display: block;
    margin-bottom: 8px;
    filter: drop-shadow(0 2px 8px rgba(0,0,0,0.4));
}
.rpt-card-label {
    font-size: 12px;
    font-weight: 700;
    color: #d0d4f0;
    margin-bottom: 3px;
    line-height: 1.3;
}
.rpt-card-desc {
    font-size: 10px;
    color: #5a5e80;
    line-height: 1.4;
}

/* Action pill buttons */
.rpt-action-strip {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    margin-bottom: 14px;
}
.rpt-pill {
    flex: 1;
    min-width: 130px;
    background: linear-gradient(135deg, #1a1d35, #141628);
    border: 1px solid #2e3355;
    border-radius: 40px;
    padding: 11px 14px;
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 12px;
    font-weight: 600;
    color: #c0c4e0;
    cursor: pointer;
    animation: float-up 0.5s ease both;
}
.rpt-pill-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
}

/* Date-range section */
.rpt-daterange-card {
    background: linear-gradient(135deg, #131628, #0f1535);
    border: 1px solid #252849;
    border-left: 3px solid #7c4dff;
    border-radius: 14px;
    padding: 16px 16px 12px;
    margin-bottom: 16px;
    animation: float-up 0.5s ease both;
}
.rpt-daterange-label {
    font-size: 11px;
    font-weight: 600;
    color: #7a7fa8;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    margin-bottom: 10px;
}

/* Glowing metric pill */
.rpt-metric-pill {
    background: linear-gradient(135deg,#0f2035,#1a0f35);
    border: 1px solid #2e3355;
    border-radius: 14px;
    padding: 14px 16px;
    margin: 10px 0;
    animation: pulse-glow 3s ease infinite;
}
.rpt-metric-label { font-size: 10px; color:#7a7fa8; text-transform:uppercase; letter-spacing:0.5px; }
.rpt-metric-value { font-size: 22px; font-weight:800; color:#22d67b; margin: 2px 0; }
.rpt-metric-sub   { font-size: 10px; color:#5a5e80; }

/* Disclaimer card */
.rpt-disclaimer {
    background: #0d0f1e;
    border: 1px solid #1e2238;
    border-left: 3px solid #f5c842;
    border-radius: 12px;
    padding: 12px 14px;
    margin-top: 16px;
    font-size: 10px;
    color: #5a5e80;
    line-height: 1.6;
    animation: float-up 0.6s ease both;
}
.rpt-disclaimer b { color: #a0a440; }
</style>
""", unsafe_allow_html=True)

        # ── Hero Banner ──────────────────────────────────────────────
        st.markdown(f"""
<div class="rpt-hero">
  <span class="rpt-hero-icon">📊</span>
  <div class="rpt-hero-title">Reports & Export</div>
  <div class="rpt-hero-sub">Generate · Download · Share</div>
  <div class="rpt-hero-client">📁 {selected_client}</div>
</div>
""", unsafe_allow_html=True)

        # ── Export Cards ─────────────────────────────────────────────
        st.markdown("""
<div class="rpt-section-hdr">
  <div class="rpt-section-hdr-icon" style="background:rgba(248,84,84,0.12);">⬇️</div>
  <div>
    <div class="rpt-section-hdr-text">Download Reports</div>
    <div class="rpt-section-hdr-sub">PDF &amp; Excel formats</div>
  </div>
</div>
  <div class="rpt-export-grid">
  <div class="rpt-export-card rpt-card-pdf">
    <span class="rpt-card-icon">🖥️</span>
    <div class="rpt-card-label">HTML Dashboard</div>
    <div class="rpt-card-desc">Interactive · AI Analysis · Riskometer · Charts</div>
  </div>
  <div class="rpt-export-card rpt-card-xl">
    <span class="rpt-card-icon">📊</span>
    <div class="rpt-card-label">Excel Export</div>
    <div class="rpt-card-desc">Multi-sheet: Holdings, Trades, P&amp;L</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Generate Both (single AI+LTP run) ──────────────────────────────
        if st.button("⚡  Generate Both Dashboards (PC + Mobile)", key="gen_both", use_container_width=True):
            with st.spinner("Building dashboards — AI analysis running once for both…"):
                try:
                    html_bytes = generate_html_report(
                        selected_client, calc, summary_dict,
                        trades_df, booked_pnl_map
                    )
                    if html_bytes:
                        st.session_state["_html_report_bytes"]  = html_bytes
                        st.session_state["_html_report_client"] = selected_client
                        mob_bytes = generate_html_report_mobile(
                            selected_client, calc, summary_dict,
                            trades_df, booked_pnl_map,
                            _prebuilt_desktop_bytes=html_bytes
                        )
                        if mob_bytes:
                            st.session_state["_html_mob_bytes"]  = mob_bytes
                            st.session_state["_html_mob_client"] = selected_client
                        st.success("✅ Both dashboards ready — save below.")
                    else:
                        st.warning("Dashboard generation failed.")
                except Exception as e:
                    st.error(f"Dashboard error: {e}")

        st.markdown("---")

        # HTML Dashboard button
        if st.button("📥  Generate PC HTML Dashboard", key="gen_html", use_container_width=True):
            with st.spinner("Building institutional HTML dashboard…"):
                try:
                    html_bytes = generate_html_report(
                        selected_client, calc, summary_dict,
                        trades_df, booked_pnl_map
                    )
                    if html_bytes:
                        st.session_state["_html_report_bytes"]  = html_bytes
                        st.session_state["_html_report_client"] = selected_client
                        st.success("✅ Dashboard ready — tap Save below. Open .html in any browser.")
                    else:
                        st.warning("Dashboard generation failed.")
                except Exception as e:
                    st.error(f"Dashboard error: {e}")

        if st.session_state.get("_html_report_bytes") and st.session_state.get("_html_report_client") == selected_client:
            st.download_button(
                "💾  Tap to Save HTML Dashboard",
                data=st.session_state["_html_report_bytes"],
                file_name=f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.html",
                mime="text/html",
                key="dl_html",
                use_container_width=True,
            )

        st.markdown("---")

        # Mobile HTML Dashboard button — reuses desktop bytes if available
        if st.button("📱  Generate Mobile HTML Dashboard", key="gen_html_mob", use_container_width=True):
            with st.spinner("Building mobile-optimised card dashboard…"):
                try:
                    _cached = st.session_state.get("_html_report_bytes") \
                              if st.session_state.get("_html_report_client") == selected_client else None
                    mob_bytes = generate_html_report_mobile(
                        selected_client, calc, summary_dict,
                        trades_df, booked_pnl_map,
                        _prebuilt_desktop_bytes=_cached
                    )
                    if mob_bytes:
                        st.session_state["_html_mob_bytes"]  = mob_bytes
                        st.session_state["_html_mob_client"] = selected_client
                        if _cached:
                            st.success("✅ Mobile dashboard ready (reused PC build — instant).")
                        else:
                            st.success("✅ Mobile dashboard ready — tap Save below. Open .html on your phone.")
                    else:
                        st.warning("Mobile dashboard generation failed.")
                except Exception as e:
                    st.error(f"Mobile dashboard error: {e}")

        if st.session_state.get("_html_mob_bytes") and st.session_state.get("_html_mob_client") == selected_client:
            st.download_button(
                "💾  Tap to Save Mobile Dashboard",
                data=st.session_state["_html_mob_bytes"],
                file_name=f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}_mobile.html",
                mime="text/html",
                key="dl_html_mob",
                use_container_width=True,
            )

        # Excel button
        if st.button("📊  Generate Excel Dashboard", key="gen_xlsx", use_container_width=True):
            with st.spinner("Building 5-sheet Excel dashboard…"):
                try:
                    xl_bytes = generate_excel_report(
                        selected_client, calc, trades_df, summary_dict
                    )
                    if xl_bytes:
                        st.download_button(
                            "⬇️  Tap to Save Excel Dashboard",
                            data=xl_bytes,
                            file_name=f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_xlsx",
                            use_container_width=True,
                        )
                        st.success("✅ 5-sheet dashboard ready — Cover, Holdings, AI Analysis, Capital Gains, Trade History")
                    else:
                        st.warning("Excel generation failed. Install openpyxl: `pip install openpyxl`")
                except Exception as e:
                    st.error(f"Excel error: {e}")
                    st.info("Run: `pip install openpyxl` and restart.")
        # ── WhatsApp ────────────────────────────────────────────────
        st.markdown("""
<div class="rpt-section-hdr" style="margin-top:24px;">
  <div class="rpt-section-hdr-icon" style="background:rgba(37,211,102,0.12);">💬</div>
  <div>
    <div class="rpt-section-hdr-text">WhatsApp Update</div>
    <div class="rpt-section-hdr-sub">One-tap message to client</div>
  </div>
</div>
""", unsafe_allow_html=True)

        if st.button("📲  Generate WhatsApp Message", key="gen_wa", use_container_width=True):
            wa_msg = generate_whatsapp_message(selected_client, summary_dict, calc)
            st.text_area("Copy & send on WhatsApp:", value=wa_msg, height=300, key="wa_output")
            st.markdown("""
<div style="background:rgba(37,211,102,0.08);border:1px solid rgba(37,211,102,0.2);
            border-radius:10px;padding:10px 12px;margin-top:6px;
            font-size:11px;color:#25d366;text-align:center;">
  📋 Long-press the text above → Select All → Copy → Open WhatsApp → Paste
</div>
""", unsafe_allow_html=True)

        # ── Email ────────────────────────────────────────────────────
        st.markdown("""
<div class="rpt-section-hdr" style="margin-top:24px;">
  <div class="rpt-section-hdr-icon" style="background:rgba(79,126,248,0.12);">📧</div>
  <div>
    <div class="rpt-section-hdr-text">Email to Client</div>
    <div class="rpt-section-hdr-sub">Send PDF directly via Gmail</div>
  </div>
</div>
""", unsafe_allow_html=True)

        with st.expander("⚙️  Configure & Send Email"):
            email_to   = st.text_input("Client Email", placeholder="client@example.com", key="email_to")
            email_from = st.text_input("Your Gmail", placeholder="you@gmail.com", key="email_from")
            email_pass = st.text_input("App Password", type="password", key="email_pass",
                                       help="Gmail App Password — not your main password.")
            if st.button("📤  Send Report", key="send_email", use_container_width=True):
                if not email_to or not email_from or not email_pass:
                    st.error("Fill in all fields.")
                else:
                    with st.spinner("Generating HTML dashboard and sending…"):
                        try:
                            html_bytes = generate_html_report(
                                selected_client, calc, summary_dict,
                                trades_df, booked_pnl_map
                            )
                            if not html_bytes:
                                st.error("Dashboard generation failed.")
                            else:
                                msg = MIMEMultipart()
                                msg["From"]    = email_from
                                msg["To"]      = email_to
                                msg["Subject"] = f"Portfolio Dashboard — {selected_client} — {datetime.now().strftime('%d %b %Y')}"
                                body = (f"Dear {selected_client},\n\n"
                                        f"Please find attached your interactive portfolio dashboard as of "
                                        f"{datetime.now().strftime('%d %b %Y')}.\n\n"
                                        f"Open the .html file in any browser for full charts and AI analysis.\n\n"
                                        f"Regards,\nNortheast Broking Services Limited")
                                msg.attach(MIMEText(body, "plain"))
                                part = MIMEBase("text", "html")
                                part.set_payload(html_bytes)
                                encoders.encode_base64(part)
                                fname = f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.html"
                                part.add_header("Content-Disposition", f"attachment; filename={fname}")
                                msg.attach(part)
                                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                                    server.login(email_from, email_pass)
                                    server.sendmail(email_from, email_to, msg.as_string())
                                st.success(f"✅ Dashboard sent to {email_to}!")
                        except Exception as e:
                            st.error(f"Email failed: {e}")
                            st.info("Use Gmail App Password. Enable 2FA → App Passwords in Google account.")

        # ── Date-Range P&L ───────────────────────────────────────────
        st.markdown("""
<div class="rpt-section-hdr" style="margin-top:24px;">
  <div class="rpt-section-hdr-icon" style="background:rgba(124,77,255,0.12);">📅</div>
  <div>
    <div class="rpt-section-hdr-text">P&amp;L Statement</div>
    <div class="rpt-section-hdr-sub">Filter by date range</div>
  </div>
</div>
""", unsafe_allow_html=True)

        if not trades_df.empty:
            st.markdown('<div class="rpt-daterange-card"><div class="rpt-daterange-label">Select Period</div>', unsafe_allow_html=True)
            start_d = st.date_input("From", value=date(datetime.today().year, 4, 1), key="dr_start")
            end_d   = st.date_input("To",   value=datetime.today().date(),           key="dr_end")
            st.markdown('</div>', unsafe_allow_html=True)

            if st.button("📊  Generate P&L Statement", key="dr_gen", use_container_width=True):
                t_filtered = trades_df.copy()
                t_filtered["Sell_Date_dt"] = pd.to_datetime(t_filtered["Sell_Date"], errors="coerce")
                t_filtered = t_filtered[
                    (t_filtered["Sell_Date_dt"].dt.date >= start_d) &
                    (t_filtered["Sell_Date_dt"].dt.date <= end_d)
                ]
                if t_filtered.empty:
                    st.info("No trades in selected date range.")
                else:
                    total_booked = t_filtered["Booked_PnL"].sum()
                    pnl_color = "#22d67b" if total_booked >= 0 else "#f85454"
                    pnl_sign  = "+" if total_booked >= 0 else ""
                    st.markdown(f"""
<div class="rpt-metric-pill">
  <div class="rpt-metric-label">Total Booked P&amp;L · {len(t_filtered)} trades</div>
  <div class="rpt-metric-value" style="color:{pnl_color};">{pnl_sign}₹{total_booked:,.2f}</div>
  <div class="rpt-metric-sub">{start_d.strftime('%d %b %Y')} → {end_d.strftime('%d %b %Y')}</div>
</div>
""", unsafe_allow_html=True)
                    t_filtered["Booked_PnL"] = pd.to_numeric(t_filtered["Booked_PnL"], errors="coerce")
                    st.dataframe(
                        t_filtered.drop(columns=["Sell_Date_dt"]).style
                            .map(style_pnl, subset=["Booked_PnL"]),
                        use_container_width=True, hide_index=True
                    )
        else:
            st.info("No sell trades recorded yet.")

        # ── Disclaimer ───────────────────────────────────────────────
        st.markdown("""
<div class="rpt-disclaimer">
  ⚠️ <b>Disclaimer:</b> Reports are generated for informational purposes only.
  Verify all figures with your broker before making any investment decisions.
  Northeast Broking Services Limited is not liable for errors in exported data.
</div>
""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    # DESKTOP VERSION — Professional two-column layout
    # ══════════════════════════════════════════════════════════════════
    else:
        st.subheader("📄 Reports & Export")

        # ── Desktop CSS ──────────────────────────────────────────────
        st.markdown("""
<style>
/* ─── Desktop Reports ─── */
.rpt-desk-card {
    background: #131628;
    border: 1px solid #252849;
    border-radius: 14px;
    padding: 22px 24px 20px;
    height: 100%;
    transition: border-color 0.2s, box-shadow 0.2s;
}
.rpt-desk-card:hover {
    border-color: #4f7ef8;
    box-shadow: 0 4px 24px rgba(79,126,248,0.12);
}
.rpt-desk-card-title {
    font-size: 15px;
    font-weight: 700;
    color: #e0e4ff;
    margin-bottom: 5px;
    display: flex;
    align-items: center;
    gap: 8px;
}
.rpt-desk-card-desc {
    font-size: 12px;
    color: #7a7fa8;
    line-height: 1.55;
    margin-bottom: 16px;
}
.rpt-desk-divider {
    height: 1px;
    background: linear-gradient(90deg, #252849 0%, transparent 100%);
    margin: 24px 0;
}
.rpt-desk-section-title {
    font-size: 14px;
    font-weight: 700;
    color: #c8ccec;
    margin-bottom: 14px;
    padding-bottom: 8px;
    border-bottom: 1px solid #252849;
    display: flex;
    align-items: center;
    gap: 8px;
}
.rpt-desk-disclaimer {
    background: #0d0f1e;
    border: 1px solid #1e2238;
    border-left: 3px solid #f5c842;
    border-radius: 10px;
    padding: 12px 16px;
    margin-top: 20px;
    font-size: 11px;
    color: #666688;
    line-height: 1.6;
}
</style>
""", unsafe_allow_html=True)

        # ── Row 1: Excel + PDF ───────────────────────────────────────
        st.markdown("""
<style>
.rpt-card-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    margin-bottom: 6px;
}
.rpt-dl-card {
    background: #0d1b2a;
    border: 1px solid #1565C0;
    border-radius: 14px;
    padding: 20px 22px 16px;
}
.rpt-dl-card-title {
    font-size: 15px;
    font-weight: 700;
    color: #e3f2fd;
    margin-bottom: 6px;
}
.rpt-dl-badge {
    display: inline-block;
    background: #1565C0;
    color: #fff;
    font-size: 10px;
    font-weight: 600;
    padding: 2px 8px;
    border-radius: 20px;
    margin-bottom: 10px;
    letter-spacing: 0.5px;
}
.rpt-dl-feat {
    font-size: 11px;
    color: #78909c;
    line-height: 1.7;
}
.rpt-dl-feat span {
    color: #90caf9;
    font-weight: 600;
}
</style>
<div class="rpt-card-grid">
  <div class="rpt-dl-card" style="border-color:#1565C0;">
    <div class="rpt-dl-card-title">📊 Professional Excel Dashboard</div>
    <div class="rpt-dl-badge">5 SHEETS · COLOUR-CODED · AI ANALYSIS</div>
    <div class="rpt-dl-feat">
      <span>Sheet 1</span> Cover Page — KPI cards + Asset Allocation<br>
      <span>Sheet 2</span> Holdings — Sorted by value, XIRR, P&amp;L colouring<br>
      <span>Sheet 3</span> AI Analysis — Riskometer, Mistakes, Recommendations<br>
      <span>Sheet 4</span> Capital Gains — STCG / LTCG breakdown<br>
      <span>Sheet 5</span> Trade History — Complete sell log
    </div>
  </div>
  <div class="rpt-dl-card" style="border-color:#4f7ef8;">
    <div class="rpt-dl-card-title">🖥️ Interactive HTML Dashboard</div>
    <div class="rpt-dl-badge">INSTITUTIONAL · AI SUMMARY · RISKOMETER · CHARTS</div>
    <div class="rpt-dl-feat">
      <span>Tab 1</span> Portfolio Overview — KPIs, allocation donut, P&amp;L bar<br>
      <span>Tab 2</span> AI Analysis — Narrative summary, scorecard, health bars<br>
      <span>Tab 3</span> Riskometer — Gauge, risk factor table, concentration view<br>
      <span>Tab 4</span> Holdings — Full sortable table with XIRR<br>
      <span>Tab 5</span> Trade History &amp; Capital Gains — STCG/LTCG breakdown<br>
      <span>Bonus</span> Print-to-PDF directly from browser
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        r1, r2 = st.columns(2)

        with r1:
            if st.button("📥 Generate Excel Dashboard", key="gen_xlsx", use_container_width=True,
                         help="Generates a 5-sheet professional Excel workbook with AI analysis"):
                with st.spinner("Building professional Excel dashboard..."):
                    try:
                        xl_bytes = generate_excel_report(
                            selected_client, calc, trades_df, summary_dict
                        )
                        if xl_bytes:
                            fname = f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.xlsx"
                            st.download_button(
                                "⬇️ Download Excel Dashboard",
                                data=xl_bytes,
                                file_name=fname,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_xlsx",
                                use_container_width=True,
                            )
                            st.success("✅ 5-sheet Excel dashboard ready! Includes AI analysis, riskometer & recommendations.")
                        else:
                            st.warning("Excel generation failed — check server logs.")
                    except Exception as e:
                        st.error(f"Excel error: {e}")
                        st.info("Ensure openpyxl is installed: `pip install openpyxl`")

        with r2:
            # ── Generate Both (single AI+LTP run) ─────────────────────────
            if st.button("⚡ Generate Both Dashboards (PC + Mobile)", key="gen_both", use_container_width=True):
                with st.spinner("Building dashboards — AI analysis runs once for both…"):
                    try:
                        html_bytes = generate_html_report(
                            selected_client, calc, summary_dict,
                            trades_df, booked_pnl_map
                        )
                        if html_bytes:
                            st.session_state["_html_report_bytes"]  = html_bytes
                            st.session_state["_html_report_client"] = selected_client
                            mob_bytes = generate_html_report_mobile(
                                selected_client, calc, summary_dict,
                                trades_df, booked_pnl_map,
                                _prebuilt_desktop_bytes=html_bytes
                            )
                            if mob_bytes:
                                st.session_state["_html_mob_bytes"]  = mob_bytes
                                st.session_state["_html_mob_client"] = selected_client
                            st.success("✅ Both dashboards ready — save below.")
                        else:
                            st.warning("Dashboard generation failed — check logs.")
                    except Exception as e:
                        st.error(f"Dashboard error: {e}")

            st.markdown("---")

            if st.button("📥 Generate PC HTML Dashboard", key="gen_html", use_container_width=True):
                with st.spinner("Building institutional HTML dashboard..."):
                    try:
                        html_bytes = generate_html_report(
                            selected_client, calc, summary_dict,
                            trades_df, booked_pnl_map
                        )
                        if html_bytes:
                            st.session_state["_html_report_bytes"]  = html_bytes
                            st.session_state["_html_report_client"] = selected_client
                            st.success("✅ Dashboard ready! Click Save below. Open .html in any browser.")
                        else:
                            st.warning("Dashboard generation failed — check logs.")
                    except Exception as e:
                        st.error(f"Dashboard error: {e}")

            if st.session_state.get("_html_report_bytes") and st.session_state.get("_html_report_client") == selected_client:
                st.download_button(
                    "💾 Save HTML Dashboard",
                    data=st.session_state["_html_report_bytes"],
                    file_name=f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.html",
                    mime="text/html",
                    key="dl_html",
                    use_container_width=True,
                )

            st.markdown("---")

            # Mobile — reuses desktop bytes from session state if available
            if st.button("📱 Generate Mobile HTML Dashboard", key="gen_html_mob", use_container_width=True):
                with st.spinner("Building mobile card dashboard..."):
                    try:
                        _cached = st.session_state.get("_html_report_bytes") \
                                  if st.session_state.get("_html_report_client") == selected_client else None
                        mob_bytes = generate_html_report_mobile(
                            selected_client, calc, summary_dict,
                            trades_df, booked_pnl_map,
                            _prebuilt_desktop_bytes=_cached
                        )
                        if mob_bytes:
                            st.session_state["_html_mob_bytes"]  = mob_bytes
                            st.session_state["_html_mob_client"] = selected_client
                            if _cached:
                                st.success("✅ Mobile dashboard ready (reused PC build — instant)!")
                            else:
                                st.success("✅ Mobile dashboard ready! Send to client's phone.")
                        else:
                            st.warning("Mobile dashboard generation failed.")
                    except Exception as e:
                        st.error(f"Mobile dashboard error: {e}")

            if st.session_state.get("_html_mob_bytes") and st.session_state.get("_html_mob_client") == selected_client:
                st.download_button(
                    "💾 Save Mobile Dashboard",
                    data=st.session_state["_html_mob_bytes"],
                    file_name=f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}_mobile.html",
                    mime="text/html",
                    key="dl_html_mob",
                    use_container_width=True,
                )

        st.markdown('<div class="rpt-desk-divider"></div>', unsafe_allow_html=True)

        # ── Row 2: WhatsApp + Email ───────────────────────────────────
        wa_col, em_col = st.columns(2)

        with wa_col:
            st.markdown('<div class="rpt-desk-section-title">💬 WhatsApp Portfolio Update</div>', unsafe_allow_html=True)
            st.caption("One-click message ready to copy and send to client on WhatsApp.")
            if st.button("📲 Generate WhatsApp Message", key="gen_wa", use_container_width=True):
                wa_msg = generate_whatsapp_message(selected_client, summary_dict, calc)
                st.text_area("Copy & send on WhatsApp:", value=wa_msg, height=300, key="wa_output")
                st.caption("📋 Select all text above → Copy → Paste in WhatsApp")

        with em_col:
            st.markdown('<div class="rpt-desk-section-title">📧 Email HTML Dashboard to Client</div>', unsafe_allow_html=True)
            with st.expander("Configure Email & Send"):
                email_to   = st.text_input("Client Email", placeholder="client@example.com", key="email_to")
                email_from = st.text_input("Your Gmail", placeholder="you@gmail.com", key="email_from")
                email_pass = st.text_input("App Password (Gmail)", type="password", key="email_pass",
                                           help="Use Gmail App Password — not your main password.")
                if st.button("📤 Send Report by Email", key="send_email", use_container_width=True):
                    if not email_to or not email_from or not email_pass:
                        st.error("Fill in all fields.")
                    else:
                        with st.spinner("Generating HTML dashboard and sending email..."):
                            try:
                                html_bytes = generate_html_report(
                                    selected_client, calc, summary_dict,
                                    trades_df, booked_pnl_map
                                )
                                if not html_bytes:
                                    st.error("Dashboard generation failed.")
                                else:
                                    msg = MIMEMultipart()
                                    msg["From"]    = email_from
                                    msg["To"]      = email_to
                                    msg["Subject"] = f"Portfolio Dashboard — {selected_client} — {datetime.now().strftime('%d %b %Y')}"
                                    body = (f"Dear {selected_client},\n\n"
                                            f"Please find attached your interactive portfolio dashboard as of "
                                            f"{datetime.now().strftime('%d %b %Y')}.\n\n"
                                            f"Open the .html file in any browser to view charts, AI analysis and riskometer.\n\n"
                                            f"Regards,\nNortheast Broking Services Limited")
                                    msg.attach(MIMEText(body, "plain"))
                                    part = MIMEBase("text", "html")
                                    part.set_payload(html_bytes)
                                    encoders.encode_base64(part)
                                    fname = f"Portfolio_{selected_client.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.html"
                                    part.add_header("Content-Disposition", f"attachment; filename={fname}")
                                    msg.attach(part)
                                    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                                        server.login(email_from, email_pass)
                                        server.sendmail(email_from, email_to, msg.as_string())
                                    st.success(f"✅ Dashboard sent to {email_to}!")
                            except Exception as e:
                                st.error(f"Email failed: {e}")
                                st.info("Make sure you're using a Gmail App Password. Enable 2FA → App Passwords in Google account.")

        st.markdown('<div class="rpt-desk-divider"></div>', unsafe_allow_html=True)

        # ── Date Range P&L Statement ──────────────────────────────────
        st.markdown('<div class="rpt-desk-section-title">📅 Date-Range P&L Statement</div>', unsafe_allow_html=True)
        st.caption("Filter sell trades by date range to generate monthly / quarterly statements.")

        if not trades_df.empty:
            dr1, dr2 = st.columns(2)
            start_d = dr1.date_input("From Date", value=date(datetime.today().year, 4, 1), key="dr_start")
            end_d   = dr2.date_input("To Date",   value=datetime.today().date(),           key="dr_end")
            if st.button("📊 Generate Statement", key="dr_gen"):
                t_filtered = trades_df.copy()
                t_filtered["Sell_Date_dt"] = pd.to_datetime(t_filtered["Sell_Date"], errors="coerce")
                t_filtered = t_filtered[
                    (t_filtered["Sell_Date_dt"].dt.date >= start_d) &
                    (t_filtered["Sell_Date_dt"].dt.date <= end_d)
                ]
                if t_filtered.empty:
                    st.info("No trades in selected date range.")
                else:
                    total_booked = t_filtered["Booked_PnL"].sum()
                    wins = (pd.to_numeric(t_filtered["Booked_PnL"], errors="coerce") > 0).sum()
                    losses = len(t_filtered) - wins
                    win_pct = (wins / len(t_filtered) * 100) if len(t_filtered) > 0 else 0
                    total_col = "#22d67b" if total_booked >= 0 else "#f85454"

                    # Summary KPI cards
                    st.markdown(f"""
<style>
.dr-kpi-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px;}}
.dr-kpi{{background:#0f1228;border:1px solid #1e2440;border-radius:12px;padding:18px 16px;text-align:center;}}
.dr-kpi-val{{font-size:20px;font-weight:900;line-height:1.1;margin-bottom:4px;}}
.dr-kpi-lbl{{font-size:10px;color:#5a5e80;text-transform:uppercase;letter-spacing:0.7px;font-weight:600;}}
.dr-trade-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:12px;margin-top:4px;}}
.dr-trade-card{{background:#0b0e1e;border:1px solid #1a1d36;border-radius:10px;padding:14px;}}
.dr-trade-ticker{{font-size:14px;font-weight:800;color:#e0e4ff;margin-bottom:6px;}}
.dr-trade-row{{display:flex;justify-content:space-between;margin-bottom:3px;font-size:11.5px;}}
.dr-trade-lbl{{color:#5a5e80;}}
.dr-trade-val{{color:#9098bc;font-weight:600;}}
@media(max-width:700px){{.dr-kpi-row{{grid-template-columns:repeat(2,1fr);}}}}
</style>
<div class="dr-kpi-row">
  <div class="dr-kpi"><div class="dr-kpi-val" style="color:{total_col};">₹{total_booked:,.0f}</div><div class="dr-kpi-lbl">Total Booked P&L</div></div>
  <div class="dr-kpi"><div class="dr-kpi-val" style="color:#e0e4ff;">{len(t_filtered)}</div><div class="dr-kpi-lbl">Total Trades</div></div>
  <div class="dr-kpi"><div class="dr-kpi-val" style="color:#22d67b;">{wins}</div><div class="dr-kpi-lbl">Winners</div></div>
  <div class="dr-kpi"><div class="dr-kpi-val" style="color:#f85454;">{losses}</div><div class="dr-kpi-lbl">Losers · Win Rate {win_pct:.0f}%</div></div>
</div>
""", unsafe_allow_html=True)

                    # Trade cards
                    cards_html = '<div class="dr-trade-grid">'
                    for _, row in t_filtered.iterrows():
                        bv = float(pd.to_numeric(row.get("Booked_PnL", 0), errors="coerce") or 0)
                        bc = "#22d67b" if bv >= 0 else "#f85454"
                        sp = f'₹{float(row.get("Sell_Price",0) or 0):,.2f}'
                        ab = f'₹{float(row.get("Buy_Price_At_Sell",0) or 0):,.2f}'
                        qty = f'{float(row.get("Sell_Qty",0) or 0):,.2f}'
                        cards_html += f"""
<div class="dr-trade-card" style="border-left:3px solid {bc};">
  <div class="dr-trade-ticker">{row.get('Ticker','—')} <span style="font-size:10px;font-weight:500;color:#5a5e80;margin-left:6px;">{row.get('Sell_Date','—')}</span></div>
  <div class="dr-trade-row"><span class="dr-trade-lbl">Qty</span><span class="dr-trade-val">{qty}</span></div>
  <div class="dr-trade-row"><span class="dr-trade-lbl">Sell Price</span><span class="dr-trade-val">{sp}</span></div>
  <div class="dr-trade-row"><span class="dr-trade-lbl">Avg Buy</span><span class="dr-trade-val">{ab}</span></div>
  <div class="dr-trade-row"><span class="dr-trade-lbl">Booked P&L</span><span style="font-weight:800;color:{bc};">₹{bv:,.0f}</span></div>
</div>"""
                    cards_html += '</div>'
                    st.markdown(cards_html, unsafe_allow_html=True)

                    # Download CSV still available
                    csv_dl = t_filtered.drop(columns=["Sell_Date_dt"], errors="ignore").to_csv(index=False).encode("utf-8")
                    st.download_button("⬇️ Download Statement CSV", data=csv_dl,
                                       file_name=f"PnL_{start_d}_{end_d}.csv", mime="text/csv", key="dl_dr_csv")
        else:
            st.info("No sell trades recorded yet.")

        st.markdown("""
<div class="rpt-desk-disclaimer">
  ⚠️ <b>Disclaimer:</b> Reports are generated for informational purposes only.
  Always verify figures with your broker before making any investment decisions.
  Northeast Broking Services Limited is not responsible for errors in exported data.
</div>
""", unsafe_allow_html=True)


# =====================================================
# TAB 9 — CRR RECONCILIATION
# =====================================================

if _nav_tab == "CRR Reconciliation":
    st.subheader("📁 CRR File Upload & Reconciliation")
    st.caption("Upload your Angel One CRR (Client Registration Report) or holdings export. "
               "The app will compare it against your portfolio and flag mismatches.")

    crr_file = st.file_uploader(
        "Upload CRR / Holdings file (.xlsx or .csv)",
        type=["xlsx","xls","csv"],
        key="crr_upload"
    )

    if crr_file is not None:
        try:
            if crr_file.name.endswith(".csv"):
                crr_raw = pd.read_csv(crr_file)
            else:
                # Try to detect header row (first row with 'symbol' or 'scrip' etc.)
                crr_raw_all = pd.read_excel(crr_file, header=None)
                header_idx = 0
                for i, row in crr_raw_all.iterrows():
                    row_str = " ".join(str(v).lower() for v in row.values)
                    if any(k in row_str for k in ["symbol","scrip","ticker","isin","qty","quantity"]):
                        header_idx = i
                        break
                crr_raw = pd.read_excel(crr_file, header=header_idx)

            st.success(f"✅ File loaded: {crr_file.name} — {len(crr_raw)} rows")
            st.markdown("**Preview (first 5 rows):**")
            st.dataframe(crr_raw.head(), use_container_width=True, hide_index=True)

            recon_df, err = reconcile_crr(crr_raw, df)
            if err:
                st.error(f"Reconciliation error: {err}")
            else:
                st.markdown("---")
                st.subheader("🔍 Reconciliation Result")

                match   = recon_df[recon_df["Status"] == "✅ Match"]
                mismatch= recon_df[recon_df["Status"].str.startswith("🔴")]
                crr_only= recon_df[recon_df["Status"].str.contains("CRR Only")]
                port_only=recon_df[recon_df["Status"].str.contains("Portfolio Only")]

                rc1,rc2,rc3,rc4 = st.columns(4)
                rc1.metric("✅ Match",             len(match))
                rc2.metric("🔴 Qty Mismatch",      len(mismatch))
                rc3.metric("❌ In CRR Only",        len(crr_only))
                rc4.metric("⚠️ In Portfolio Only", len(port_only))

                def _recon_style(val):
                    if "Match" in str(val):      return "color:#00e676;font-weight:600"
                    if "Mismatch" in str(val):   return "color:#ff5252;font-weight:600"
                    if "CRR Only" in str(val):   return "color:#ff9800;font-weight:600"
                    if "Portfolio" in str(val):  return "color:#ffaa00;font-weight:600"
                    return ""

                st.dataframe(
                    recon_df.style.map(_recon_style, subset=["Status"]),
                    use_container_width=True, hide_index=True
                )

                if not mismatch.empty:
                    st.warning("🔴 Qty mismatches found — review and correct your portfolio quantities.")
                if not crr_only.empty:
                    st.warning("❌ Some holdings in CRR are not in your portfolio — add them from the sidebar.")
                if not port_only.empty:
                    st.info("⚠️ Some holdings in your portfolio are not in the CRR — verify with broker.")

        except Exception as e:
            st.error(f"Failed to read file: {e}")
            st.info("Make sure the file is a valid Excel or CSV. Close the file in Excel before uploading.")
    else:
        st.info("Upload a CRR or holdings export file from Angel One to begin reconciliation.")
        st.markdown("""
**How to get CRR from Angel One:**
1. Login to Angel One → Reports → Client Registration Report
2. Download as Excel / CSV
3. Upload here

**Columns the app looks for:**
- Symbol / Scrip / Ticker / ISIN
- Qty / Quantity / Net Qty / Holdings
- Avg Cost / Avg Buy Price (optional)
""")

# =====================================================
# MODULE-LEVEL CACHED FETCH FUNCTIONS — ALL CORPORATE ACTIONS
# These MUST be at module level so @st.cache_data persists across reruns.
# =====================================================

@st.cache_data(ttl=1800, show_spinner=False)
def _mktca_fetch_nse(from_date_str: str, to_date_str: str) -> list:
    """
    Fetch NSE corporate actions with multiple fallback strategies.
    NSE blocks plain requests; we try several approaches + fallback to
    Tickertape / Screener public data.
    """
    _UA_LIST = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    ]
    results = []

    # ── Strategy 1: NSE API with full browser headers + cookie warmup ──
    for ua in _UA_LIST:
        if results:
            break
        try:
            _HDR = {
                "User-Agent": ua,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-IN,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                "X-Requested-With": "XMLHttpRequest",
                "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124"',
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": '"Windows"',
                "sec-fetch-dest": "empty",
                "sec-fetch-mode": "cors",
                "sec-fetch-site": "same-origin",
            }
            s = requests.Session()
            # Warm up cookies — visit home and the corporate actions page
            s.get("https://www.nseindia.com", headers=_HDR, timeout=10)
            s.get(
                "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                headers=_HDR, timeout=10
            )
            _to_ts    = pd.to_datetime(to_date_str, dayfirst=True)
            _from_ts  = _to_ts - pd.Timedelta(days=180)
            wide_from = _from_ts.strftime("%d-%m-%Y")
            url = (
                f"https://www.nseindia.com/api/corporates-corporateActions"
                f"?index=equities&from_date={wide_from}&to_date={to_date_str}&subject=all"
            )
            r = s.get(url, headers=_HDR, timeout=20)
            if r.status_code == 200:
                data = r.json()
                raw = data if isinstance(data, list) else data.get("data", [])
                if raw:
                    results = raw
        except Exception:
            pass

    # ── Strategy 2: Per-event-type fetch (NSE sometimes rejects subject=all
    #    but accepts individual subjects — fetch each and merge) ──
    if not results:
        _SUBJECTS = ["dividend", "bonus", "splits", "rights", "buyback",
                     "agm", "egm", "merger", "amalgamation", "delisting",
                     "scheme of arrangement", "demerger"]
        try:
            _HDR2 = {
                "User-Agent": _UA_LIST[0],
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-IN,en;q=0.9",
                "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                "X-Requested-With": "XMLHttpRequest",
            }
            s2 = requests.Session()
            s2.get("https://www.nseindia.com", headers=_HDR2, timeout=10)
            s2.get(
                "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                headers=_HDR2, timeout=10
            )
            _to_ts   = pd.to_datetime(to_date_str, dayfirst=True)
            _from_ts = _to_ts - pd.Timedelta(days=180)
            wide_from2 = _from_ts.strftime("%d-%m-%Y")
            seen2 = set()
            for subj in _SUBJECTS:
                try:
                    r2 = s2.get(
                        f"https://www.nseindia.com/api/corporates-corporateActions"
                        f"?index=equities&from_date={wide_from2}&to_date={to_date_str}&subject={subj}",
                        headers=_HDR2, timeout=15
                    )
                    if r2.status_code == 200:
                        data2 = r2.json()
                        raw2 = data2 if isinstance(data2, list) else data2.get("data", [])
                        for item in raw2:
                            k = (
                                str(item.get("symbol","")).upper(),
                                str(item.get("exDate","") or item.get("ex_date","")),
                                str(item.get("subject",""))[:20],
                            )
                            if k not in seen2:
                                seen2.add(k)
                                results.append(item)
                except Exception:
                    pass
        except Exception:
            pass

    # ── Strategy 3: NSE Board-Meeting / Corporate-Filings endpoint ─────────
    # This endpoint reliably returns bm_dt (board meeting / announcement date)
    # and nd_startDt / nd_endDt (record date window).  We use it to ENRICH
    # existing results (fill in missing ann/record dates) AND as a standalone
    # fallback if nothing else worked.
    try:
        _HDR3 = {
            "User-Agent": _UA_LIST[0],
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-IN,en;q=0.9",
            "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-actions",
            "X-Requested-With": "XMLHttpRequest",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
        }
        s3 = requests.Session()
        s3.get("https://www.nseindia.com", headers=_HDR3, timeout=10)
        s3.get(
            "https://www.nseindia.com/companies-listing/corporate-filings-actions",
            headers=_HDR3, timeout=10,
        )
        _to_ts3  = pd.to_datetime(to_date_str, dayfirst=True)
        _from_ts3 = _to_ts3 - pd.Timedelta(days=180)
        wide_from3 = _from_ts3.strftime("%d-%m-%Y")

        # NSE corporate-filings-actions endpoint (returns bm_dt, nd_startDt, exDate)
        r3 = s3.get(
            f"https://www.nseindia.com/api/corporate-filings-actions"
            f"?index=equities&from_date={wide_from3}&to_date={to_date_str}",
            headers=_HDR3, timeout=20,
        )
        bm_items = []
        if r3.status_code == 200:
            d3 = r3.json()
            bm_items = d3 if isinstance(d3, list) else d3.get("data", [])

        if bm_items:
            # Build a lookup: symbol+exDate → {bm_dt, nd_startDt}
            bm_lookup = {}
            for bm in bm_items:
                _sym = str(bm.get("symbol","") or bm.get("Symbol","")).strip().upper()
                _ex  = str(bm.get("exDate","") or bm.get("ex_date","")).strip()
                if _sym:
                    bm_lookup[(_sym, _ex)] = bm

            # Enrich existing results that have missing bm_dt or recordDate
            for item in results:
                _sym = str(item.get("symbol","") or item.get("Symbol","")).strip().upper()
                _ex  = str(item.get("exDate","") or item.get("ex_date","")).strip()
                _match = bm_lookup.get((_sym, _ex)) or bm_lookup.get((_sym, ""))
                if _match:
                    # Fill announcement date (bm_dt) if missing
                    if not item.get("bm_dt") and not item.get("annDate"):
                        item["bm_dt"] = (_match.get("bm_dt") or
                                         _match.get("announceDate") or
                                         _match.get("anDt") or "")
                    # Fill record date (nd_startDt) if missing
                    if not item.get("recordDate") and not item.get("record_date"):
                        item["recordDate"] = (_match.get("nd_startDt") or
                                              _match.get("ndStartDt") or
                                              _match.get("nd_endDt") or
                                              _match.get("recDate") or "")

            # Also add bm_items as standalone rows if results is still empty
            if not results:
                seen3 = set()
                for bm in bm_items:
                    k3 = (
                        str(bm.get("symbol","")).upper(),
                        str(bm.get("exDate","") or bm.get("ex_date","")),
                        str(bm.get("subject",""))[:20],
                    )
                    if k3 not in seen3:
                        seen3.add(k3)
                        results.append(bm)
    except Exception:
        pass

    # ── Strategy 4: NSE corporate-filings-actions endpoint with all event types ──
    # More reliable for Bonus, Split, Merger, Demerger; always includes bm_dt
    if not results:
        try:
            _HDR4 = {
                "User-Agent": _UA_LIST[0],
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-IN,en;q=0.9",
                "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                "X-Requested-With": "XMLHttpRequest",
                "sec-fetch-dest": "empty",
                "sec-fetch-mode": "cors",
                "sec-fetch-site": "same-origin",
            }
            s4 = requests.Session()
            s4.get("https://www.nseindia.com", headers=_HDR4, timeout=10)
            s4.get("https://www.nseindia.com/companies-listing/corporate-filings-actions",
                   headers=_HDR4, timeout=10)
            _to_ts4   = pd.to_datetime(to_date_str, dayfirst=True)
            _from_ts4 = _to_ts4 - pd.Timedelta(days=180)
            wide_from4 = _from_ts4.strftime("%d-%m-%Y")
            r4 = s4.get(
                f"https://www.nseindia.com/api/corporate-filings-actions"
                f"?index=equities&from_date={wide_from4}&to_date={to_date_str}",
                headers=_HDR4, timeout=20,
            )
            if r4.status_code == 200:
                d4 = r4.json()
                items4 = d4 if isinstance(d4, list) else d4.get("data", [])
                seen4 = set()
                for it4 in items4:
                    k4 = (
                        str(it4.get("symbol","")).upper(),
                        str(it4.get("exDate","") or it4.get("ex_date","")),
                        str(it4.get("subject",""))[:20],
                    )
                    if k4 not in seen4:
                        seen4.add(k4)
                        results.append(it4)
        except Exception:
            pass

    return results


@st.cache_data(ttl=1800, show_spinner=False)
def _mktca_fetch_bse(from_date_str: str, to_date_str: str) -> list:
    """
    Fetch BSE corporate actions with multiple endpoint fallbacks.
    Tries main BSE API, alternate BSE endpoint, and BSE XML feed.
    """
    _HDR = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-IN,en-US;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://www.bseindia.com/corporates/corporate_act.aspx",
        "Origin": "https://www.bseindia.com",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
    }
    results = []

    # ── Strategy 1: Main BSE API endpoint (paginated) ──
    try:
        sb = requests.Session()
        sb.get("https://www.bseindia.com", headers=_HDR, timeout=10)
        sb.get(
            "https://www.bseindia.com/corporates/corporate_act.aspx",
            headers=_HDR, timeout=10
        )
        for page in range(1, 8):
            url = (
                f"https://api.bseindia.com/BseIndiaAPI/api/CorporateAction/w?"
                f"scripcode=&qtyfrom=&qtyto="
                f"&exdfrom={from_date_str}&exdto={to_date_str}"
                f"&segment=Equity&type=all&pageno={page}&pagesize=500"
            )
            r = sb.get(url, headers=_HDR, timeout=20)
            if r.status_code == 200:
                data = r.json()
                items = (
                    data.get("Table", [])
                    or data.get("data", [])
                    or (data if isinstance(data, list) else [])
                )
                results.extend(items)
                if len(items) < 500:
                    break
            else:
                break
    except Exception:
        pass

    # ── Strategy 2: BSE alternate API (different base path) ──
    if not results:
        try:
            _HDR2 = dict(_HDR)
            _HDR2["Referer"] = "https://www.bseindia.com/"
            sb2 = requests.Session()
            sb2.get("https://www.bseindia.com", headers=_HDR2, timeout=8)
            r2 = sb2.get(
                f"https://api.bseindia.com/BseIndiaAPI/api/CorporateActioin/w?"
                f"scripcode=&from={from_date_str}&to={to_date_str}"
                f"&segment=Equity&type=all&pageno=1&pagesize=500",
                headers=_HDR2, timeout=15
            )
            if r2.status_code == 200:
                data2 = r2.json()
                items2 = (
                    data2.get("Table", [])
                    or data2.get("data", [])
                    or (data2 if isinstance(data2, list) else [])
                )
                results.extend(items2)
        except Exception:
            pass

    # ── Strategy 3: BSE CorpActionDetail API — carries ANNOUNCEMENT_DATE and ND_StartDate ──
    # Enrich existing results with announcement date and record date fields that
    # the main CorporateAction endpoint sometimes omits.
    try:
        _HDR3 = dict(_HDR)
        _HDR3["Referer"] = "https://www.bseindia.com/corporates/corporate_act.aspx"
        sb3 = requests.Session()
        sb3.get("https://www.bseindia.com", headers=_HDR3, timeout=8)
        sb3.get("https://www.bseindia.com/corporates/corporate_act.aspx", headers=_HDR3, timeout=8)

        # BSE detailed corporate-actions endpoint (includes ANNOUNCEMENT_DATE, ND_StartDate)
        r3 = sb3.get(
            f"https://api.bseindia.com/BseIndiaAPI/api/CorporateAction/w?"
            f"scripcode=&qtyfrom=&qtyto="
            f"&exdfrom={from_date_str}&exdto={to_date_str}"
            f"&segment=Equity&type=all&pageno=1&pagesize=1000",
            headers=_HDR3, timeout=20,
        )
        if r3.status_code == 200:
            d3 = r3.json()
            detail_items = (
                d3.get("Table", []) or d3.get("data", [])
                or (d3 if isinstance(d3, list) else [])
            )
            # Build lookup by SCRIP_CD + Ex_Date for enrichment
            detail_lookup = {}
            for di in detail_items:
                _scrip = str(di.get("SCRIP_CD") or di.get("scripcode") or
                             di.get("short_name") or "").strip().upper()
                _ex    = str(di.get("Ex_Date") or di.get("ExDate") or "").strip()
                if _scrip:
                    detail_lookup[(_scrip, _ex)] = di
                    # Also index by short_name in case symbol differs
                    _sn = str(di.get("short_name") or "").strip().upper()
                    if _sn and _sn != _scrip:
                        detail_lookup[(_sn, _ex)] = di

            for item in results:
                _scrip = str(
                    item.get("SCRIP_CD") or item.get("scripcode") or
                    item.get("short_name") or item.get("ShortName") or ""
                ).strip().upper()
                _ex = str(item.get("Ex_Date") or item.get("ExDate") or "").strip()
                _match = detail_lookup.get((_scrip, _ex)) or detail_lookup.get((_scrip, ""))
                if _match:
                    # Backfill record date if missing
                    if not item.get("RD_Date") and not item.get("RecordDate"):
                        item["RD_Date"] = (
                            _match.get("ND_StartDate") or _match.get("nd_startDate") or
                            _match.get("RD_Date") or _match.get("RecordDate") or
                            _match.get("BC_StartDate") or ""
                        )
                    # Backfill announcement date if missing
                    if not item.get("AnnDate") and not item.get("ANNOUNCEMENT_DATE"):
                        item["AnnDate"] = (
                            _match.get("ANNOUNCEMENT_DATE") or _match.get("AnnDate") or
                            _match.get("Announcement_Date") or _match.get("ann_date") or
                            _match.get("DT_TM") or _match.get("FilingDate") or ""
                        )
    except Exception:
        pass

    return results

# =====================================================
# MODULE-LEVEL — EX-DATE RADAR FETCH (BSE reliable ex-dates)
# =====================================================

@st.cache_data(ttl=900, show_spinner=False)
def _exradar_fetch_bse(from_date_str: str, to_date_str: str) -> list:
    """BSE ex-date API — filters strictly by ex-date range, with multi-UA fallback."""
    _UA_LIST = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    ]
    results = []
    for ua in _UA_LIST:
        if results:
            break
        try:
            _HDR = {
                "User-Agent": ua,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-IN,en-US;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Referer": "https://www.bseindia.com/corporates/corporate_act.aspx",
                "Origin": "https://www.bseindia.com",
                "sec-fetch-dest": "empty",
                "sec-fetch-mode": "cors",
                "sec-fetch-site": "same-site",
            }
            sb = requests.Session()
            sb.get("https://www.bseindia.com", headers=_HDR, timeout=10)
            sb.get(
                "https://www.bseindia.com/corporates/corporate_act.aspx",
                headers=_HDR, timeout=10
            )
            for page in range(1, 11):
                url = (
                    f"https://api.bseindia.com/BseIndiaAPI/api/CorporateAction/w?"
                    f"scripcode=&qtyfrom=&qtyto="
                    f"&exdfrom={from_date_str}&exdto={to_date_str}"
                    f"&segment=Equity&type=all&pageno={page}&pagesize=500"
                )
                r = sb.get(url, headers=_HDR, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    items = (
                        data.get("Table", [])
                        or data.get("data", [])
                        or (data if isinstance(data, list) else [])
                    )
                    results.extend(items)
                    if len(items) < 500:
                        break
                else:
                    break
        except Exception:
            pass
    return results


@st.cache_data(ttl=900, show_spinner=False)
def _exradar_fetch_nse(from_date_str: str, to_date_str: str) -> list:
    """NSE corporate actions with multi-UA fallback, cookie warmup, and per-subject fallback."""
    _UA_LIST = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    ]
    results = []

    # ── Strategy 1: subject=all with proper warmup ──
    for ua in _UA_LIST:
        if results:
            break
        try:
            _HDR = {
                "User-Agent": ua,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-IN,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                "X-Requested-With": "XMLHttpRequest",
                "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124"',
                "sec-ch-ua-mobile": "?0",
                "sec-fetch-dest": "empty",
                "sec-fetch-mode": "cors",
                "sec-fetch-site": "same-origin",
            }
            s = requests.Session()
            s.get("https://www.nseindia.com", headers=_HDR, timeout=10)
            s.get(
                "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                headers=_HDR, timeout=10
            )
            url = (
                f"https://www.nseindia.com/api/corporates-corporateActions"
                f"?index=equities&from_date={from_date_str}&to_date={to_date_str}&subject=all"
            )
            r = s.get(url, headers=_HDR, timeout=20)
            if r.status_code == 200:
                data = r.json()
                raw = data if isinstance(data, list) else data.get("data", [])
                if raw:
                    results = raw
        except Exception:
            pass

    # ── Strategy 2: per-event-type fetch (NSE sometimes rejects subject=all) ──
    if not results:
        _SUBJECTS = ["dividend", "bonus", "splits", "rights", "buyback",
                     "agm", "egm", "merger", "amalgamation", "delisting",
                     "scheme of arrangement", "demerger"]
        try:
            _HDR2 = {
                "User-Agent": _UA_LIST[0],
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-IN,en;q=0.9",
                "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                "X-Requested-With": "XMLHttpRequest",
            }
            s2 = requests.Session()
            s2.get("https://www.nseindia.com", headers=_HDR2, timeout=10)
            s2.get(
                "https://www.nseindia.com/companies-listing/corporate-filings-actions",
                headers=_HDR2, timeout=10
            )
            seen2 = set()
            for subj in _SUBJECTS:
                try:
                    r2 = s2.get(
                        f"https://www.nseindia.com/api/corporates-corporateActions"
                        f"?index=equities&from_date={from_date_str}&to_date={to_date_str}&subject={subj}",
                        headers=_HDR2, timeout=15
                    )
                    if r2.status_code == 200:
                        data2 = r2.json()
                        raw2 = data2 if isinstance(data2, list) else data2.get("data", [])
                        for item in raw2:
                            k = (
                                str(item.get("symbol","")).upper(),
                                str(item.get("exDate","") or item.get("ex_date","")),
                                str(item.get("subject",""))[:20],
                            )
                            if k not in seen2:
                                seen2.add(k)
                                results.append(item)
                except Exception:
                    pass
        except Exception:
            pass

    return results


# =====================================================
# TAB 10 — ALL CORPORATE ACTIONS (MARKET-WIDE NSE + BSE)
# =====================================================

if _nav_tab == "All Corporate Actions":

    st.markdown("""
<div style="background:linear-gradient(135deg,#0f0f23,#1a1a3e);border:1px solid #2a2a5a;
            border-radius:12px;padding:16px 24px;margin-bottom:18px;">
  <div style="font-size:20px;font-weight:800;color:#f0f2ff;">🏦 All Corporate Actions — Market-Wide (NSE &amp; BSE)</div>
  <div style="font-size:12px;color:#8888aa;margin-top:4px;">
    Dividend · Bonus · Split · Buyback · Rights Issue · AGM/EGM · Merger · Delisting
    — <b style="color:#00e676;">All listed stocks</b>, not just your portfolio · Your holdings highlighted in green
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Portfolio tickers for highlighting ───────────────────────────
    portfolio_syms = set(
        t.upper().replace(".NS","").replace(".BO","")
        for t in (list(df["Ticker"].unique()) if not df.empty else [])
        if not any(x in t.upper() for x in ["FUT","CE","PE"])
    )

    # ── Controls row 1 ───────────────────────────────────────────────
    ctrl1, ctrl2, ctrl3, ctrl4, ctrl5 = st.columns([1.1, 1.3, 1.3, 1.3, 1.5])
    with ctrl1:
        if st.button("🔄 Refresh", key="refresh_mkt_ca"):
            st.cache_data.clear()
            st.rerun()
    with ctrl2:
        ca_exchange = st.selectbox(
            "🏦 Exchange",
            ["Both NSE & BSE", "NSE Only", "BSE Only"],
            key="ca_mkt_exchange"
        )
    with ctrl3:
        ca_window = st.selectbox(
            "📅 Date Window",
            ["Today", "This Week", "Next 7 Days", "Next 30 Days",
             "Next 90 Days", "Last 30 Days", "Last 90 Days", "Last 1 Year"],
            index=2,
            key="ca_mkt_window"
        )
    with ctrl4:
        ca_mkt_event = st.selectbox(
            "🏷 Event Type",
            ["All", "Dividend", "Interim Dividend", "Final Dividend", "Special Dividend",
             "Bonus Issue", "Stock Split", "Rights Issue", "Buyback",
             "Merger / Demerger", "AGM", "EGM", "Financial Results",
             "Delisting", "Preferential Allotment", "NCD Issue", "QIP", "ESOP"],
            key="ca_mkt_event"
        )
    with ctrl5:
        portfolio_only = st.toggle(
            "⭐ My Portfolio Only",
            value=False,
            key="ca_portfolio_only_toggle",
            help="When ON: shows only your portfolio stocks. When OFF: shows ALL listed stocks."
        )

    # ── Date helpers (fast, no I/O) ─────────────────────────────────
    def _parse_dt_mkt(v):
        if not v or str(v).strip() in ("", "-", "nan", "None", "null"):
            return None
        try:
            return pd.to_datetime(str(v).strip(), dayfirst=True)
        except Exception:
            return None

    def _fmt_dt_mkt(ts):
        if ts is None:
            return "—"
        try:
            return ts.strftime("%d-%b-%Y")
        except Exception:
            return "—"

    def _canonical_mkt(ev):
        el = str(ev).strip().lower()
        if any(x in el for x in ("financial result", "quarterly result",
                                  "annual result", "unaudited result", "audited result")):
            return "Financial Results"
        if "interim dividend" in el:    return "Interim Dividend"
        if "final dividend" in el:      return "Final Dividend"
        if "special dividend" in el:    return "Special Dividend"
        if "dividend" in el:            return "Dividend"
        if "bonus" in el:               return "Bonus Issue"
        if "stock split" in el or "sub-division" in el or "sub division" in el:
            return "Stock Split"
        if "face value" in el and "split" in el: return "Stock Split"
        if "rights issue" in el or "rights entitlement" in el: return "Rights Issue"
        if "buyback" in el or "buy back" in el or "buy-back" in el: return "Buyback"
        if "merger" in el or "amalgamation" in el or "demerger" in el: return "Merger / Demerger"
        if "acquisition" in el or "takeover" in el: return "Acquisition"
        if "scheme of arrangement" in el: return "Scheme of Arrangement"
        if "delisting" in el:           return "Delisting"
        if "agm" in el or "annual general meeting" in el: return "AGM"
        if "egm" in el or "extra ordinary general" in el: return "EGM"
        if "warrant" in el:             return "Warrants"
        if "esop" in el or "employee stock" in el: return "ESOP"
        if "preferential" in el:        return "Preferential Allotment"
        if "ncd" in el or "non-convertible debenture" in el: return "NCD Issue"
        if "qip" in el or "qualified institutional" in el: return "QIP"
        if "ipo" in el or "fpo" in el:  return "IPO / FPO"
        raw = str(ev).strip()
        return (raw[:57] + "…" if len(raw) > 60 else raw) or "Other"

    def _status_mkt(primary_dt, today_ts):
        if primary_dt is None:
            return "—"
        if primary_dt.date() == today_ts.date():
            return "📌 Today"
        if primary_dt > today_ts:
            d = (primary_dt - today_ts).days
            if d <= 7:   return f"⚡ In {d}d"
            if d <= 30:  return f"🔜 In {d}d"
            return f"🗓 In {d}d"
        else:
            d = (today_ts - primary_dt).days
            return f"✅ {d}d ago"

    # ── Compute date range from window selection ──────────────────────
    today_ts3 = pd.Timestamp.now().normalize()

    window_ranges = {
        "Today":         (today_ts3, today_ts3),
        "This Week":     (today_ts3, today_ts3 + pd.Timedelta(days=6 - today_ts3.weekday())),
        "Next 7 Days":   (today_ts3, today_ts3 + pd.Timedelta(days=7)),
        "Next 30 Days":  (today_ts3, today_ts3 + pd.Timedelta(days=30)),
        "Next 90 Days":  (today_ts3, today_ts3 + pd.Timedelta(days=90)),
        "Last 30 Days":  (today_ts3 - pd.Timedelta(days=30), today_ts3),
        "Last 90 Days":  (today_ts3 - pd.Timedelta(days=90), today_ts3),
        "Last 1 Year":   (today_ts3 - pd.Timedelta(days=365), today_ts3),
    }
    w_from, w_to = window_ranges.get(ca_window, (today_ts3, today_ts3 + pd.Timedelta(days=30)))

    # NSE uses DD-MM-YYYY, BSE uses DD/MM/YYYY
    nse_from_str = w_from.strftime("%d-%m-%Y")
    nse_to_str   = w_to.strftime("%d-%m-%Y")
    bse_from_str = w_from.strftime("%d/%m/%Y")
    bse_to_str   = w_to.strftime("%d/%m/%Y")

    # ── Fetch NSE + BSE in parallel using module-level cached functions ──
    fetch_nse = ca_exchange in ("Both NSE & BSE", "NSE Only")
    fetch_bse = ca_exchange in ("Both NSE & BSE", "BSE Only")

    all_market_rows = []
    seen_mkt = set()

    with st.spinner("📡 Fetching corporate actions from NSE & BSE..."):

        # Fire both requests concurrently
        nse_raw, bse_raw = [], []
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as _tp:
            fut_nse = _tp.submit(_mktca_fetch_nse, nse_from_str, nse_to_str) if fetch_nse else None
            fut_bse = _tp.submit(_mktca_fetch_bse, bse_from_str, bse_to_str) if fetch_bse else None
            if fut_nse:
                try:
                    nse_raw = fut_nse.result()
                except Exception:
                    nse_raw = []
            if fut_bse:
                try:
                    bse_raw = fut_bse.result()
                except Exception:
                    bse_raw = []

        # ── Process NSE rows ─────────────────────────────────────────
        for it in nse_raw:
            sym = str(it.get("symbol","") or it.get("Symbol","")).strip().upper()
            if not sym:
                continue
            ex_dt  = _parse_dt_mkt(it.get("exDate") or it.get("ex_date") or it.get("exdate")
                                   or it.get("ExDate") or it.get("EX_DATE"))
            rec_dt = _parse_dt_mkt(it.get("recordDate") or it.get("record_date")
                                   or it.get("RecordDate") or it.get("recDate")
                                   or it.get("rec_date") or it.get("nd_startDt")
                                   or it.get("bcStartDt") or it.get("bc_startDt"))
            ann_dt = _parse_dt_mkt(it.get("bm_dt") or it.get("anDt")
                                   or it.get("announceDate") or it.get("annDate")
                                   or it.get("AnnDate") or it.get("announcement_date")
                                   or it.get("boardMeetingDate") or it.get("boardDate")
                                   or it.get("filingDate") or it.get("date"))
            purpose_raw = str(it.get("subject") or it.get("purpose") or "").strip()
            comp    = str(it.get("comp") or it.get("companyName") or it.get("company") or sym).strip()
            details = str(it.get("remarks") or it.get("details") or it.get("value") or "").strip()
            if not details:
                details = purpose_raw
            ev = _canonical_mkt(purpose_raw)
            # Use ex-date for window filtering if available; else record/announcement date
            filter_dt = ex_dt or rec_dt or ann_dt
            if filter_dt is None:
                continue
            if not (w_from <= filter_dt <= w_to):
                continue
            primary_dt = ex_dt or rec_dt or ann_dt
            key = f"NSE|{sym}|{primary_dt.date()}|{ev}"
            if key in seen_mkt:
                continue
            seen_mkt.add(key)
            all_market_rows.append({
                "Symbol":             sym,
                "Company":            comp[:40] if comp else sym,
                "Exchange":           "NSE",
                "Event Type":         ev,
                "Ex-Date":            _fmt_dt_mkt(ex_dt),
                "Record Date":        _fmt_dt_mkt(rec_dt),
                "Announcement Date":  _fmt_dt_mkt(ann_dt),
                "Details":            details[:80] if details else "—",
                "Status":             _status_mkt(primary_dt, today_ts3),
                "In Portfolio":       "⭐ Yes" if sym in portfolio_syms else "",
                "_primary_dt":        primary_dt,
            })

        # ── Process BSE rows ─────────────────────────────────────────
        for it in bse_raw:
            sym = str(
                it.get("short_name") or it.get("ShortName") or
                it.get("SCRIP_CD") or it.get("scripcode") or
                it.get("Symbol") or ""
            ).strip().upper()
            comp = str(
                it.get("Scrip_Name") or it.get("CompanyName") or
                it.get("company_name") or sym
            ).strip()
            if not sym:
                continue
            ex_dt  = _parse_dt_mkt(it.get("Ex_Date") or it.get("ExDate") or
                                    it.get("ex_date") or it.get("exDate") or
                                    it.get("EX_DATE") or it.get("EXDATE"))
            rec_dt = _parse_dt_mkt(it.get("RD_Date") or it.get("RecordDate") or
                                    it.get("record_date") or it.get("RECORD_DATE") or
                                    it.get("ND_StartDate") or it.get("nd_startDate") or
                                    it.get("BC_StartDate") or it.get("bc_startdate") or
                                    it.get("recDate") or it.get("rec_date"))
            ann_dt = _parse_dt_mkt(it.get("AnnDate") or it.get("ann_date") or
                                    it.get("AnnouncementDate") or it.get("ANNOUNCEMENT_DATE") or
                                    it.get("Announcement_Date") or it.get("ann_dt") or
                                    it.get("BoardMeetingDate") or it.get("board_meeting_date") or
                                    it.get("FilingDate") or it.get("filing_date") or
                                    it.get("DT_TM") or it.get("DTTM"))
            purpose_raw = str(it.get("Purpose") or it.get("purpose") or
                               it.get("ActionType") or "").strip()
            details = str(it.get("Remarks") or it.get("remarks") or
                          it.get("Details") or "").strip()
            if not details:
                details = purpose_raw
            ev = _canonical_mkt(purpose_raw)
            filter_dt = ex_dt or rec_dt or ann_dt
            if filter_dt is None:
                continue
            if not (w_from <= filter_dt <= w_to):
                continue
            primary_dt = ex_dt or rec_dt or ann_dt
            key = f"BSE|{sym}|{primary_dt.date()}|{ev}"
            if key in seen_mkt:
                continue
            seen_mkt.add(key)
            in_portfolio = (
                sym in portfolio_syms or
                any(sym.startswith(p) or p.startswith(sym[:5])
                    for p in portfolio_syms if len(p) >= 5)
            )
            all_market_rows.append({
                "Symbol":             sym,
                "Company":            comp[:40] if comp else sym,
                "Exchange":           "BSE",
                "Event Type":         ev,
                "Ex-Date":            _fmt_dt_mkt(ex_dt),
                "Record Date":        _fmt_dt_mkt(rec_dt),
                "Announcement Date":  _fmt_dt_mkt(ann_dt),
                "Details":            details[:80] if details else "—",
                "Status":             _status_mkt(primary_dt, today_ts3),
                "In Portfolio":       "⭐ Yes" if in_portfolio else "",
                "_primary_dt":        primary_dt,
            })

        # ── Build dataframe ───────────────────────────────────────────────
    if not all_market_rows:
        st.markdown("""
<div style="background:#0a1a2a;border:1px solid #1a4a7a;border-radius:10px;padding:16px 20px;margin:8px 0;">
  <div style="font-size:15px;font-weight:700;color:#90caf9;">📡 NSE/BSE returned no data</div>
  <div style="font-size:12px;color:#8888aa;margin-top:8px;line-height:1.7;">
    BSE and NSE use browser fingerprinting to block automated requests. This is a known limitation
    when running from a server or phone.<br><br>
    <b style="color:#fff;">Quick fixes:</b><br>
    <b style="color:#ffaa00;">•</b> Click <b style="color:#fff;">🔄 Refresh</b> — clears cache and retries with different headers<br>
    <b style="color:#ffaa00;">•</b> Change the <b style="color:#fff;">Date Window</b> (try "Last 90 Days" or "Last 1 Year")<br>
    <b style="color:#ffaa00;">•</b> Try <b style="color:#fff;">BSE Only</b> (BSE is more lenient than NSE)<br>
    <b style="color:#ffaa00;">•</b> Wait 2–5 minutes then refresh (IP rate limiting by exchange)<br><br>
    <span style="color:#666688;font-size:11px;">
      ⚠️ NSE blocks non-browser traffic from cloud/VPS IPs. If running on Streamlit Cloud or a VPS,
      this tab may be unreliable. Running locally on your PC gives best results.
    </span>
  </div>
</div>
""", unsafe_allow_html=True)
    else:
        mkt_ca_df = pd.DataFrame(all_market_rows)
        mkt_ca_df = (
            mkt_ca_df
            .sort_values("_primary_dt")
            .drop(columns=["_primary_dt"])
            .reset_index(drop=True)
        )

        # ── Apply portfolio-only filter ───────────────────────────────
        if portfolio_only:
            mkt_ca_df = mkt_ca_df[mkt_ca_df["In Portfolio"] == "⭐ Yes"]

        # ── Apply event type filter ───────────────────────────────────
        if ca_mkt_event != "All":
            # "Dividend" filter should include Interim/Final/Special dividends too
            if ca_mkt_event == "Dividend":
                mkt_ca_df = mkt_ca_df[mkt_ca_df["Event Type"].str.contains("Dividend", case=False, na=False)]
            else:
                mkt_ca_df = mkt_ca_df[mkt_ca_df["Event Type"] == ca_mkt_event]

        # ── Summary metrics ───────────────────────────────────────────
        total_mkt     = len(mkt_ca_df)
        port_count    = len(mkt_ca_df[mkt_ca_df["In Portfolio"] == "⭐ Yes"])
        upcoming_mkt  = len(mkt_ca_df[mkt_ca_df["Status"].str.contains("In |Today", na=False)])
        div_count     = len(mkt_ca_df[mkt_ca_df["Event Type"].str.contains("Dividend", na=False)])
        bonus_count   = len(mkt_ca_df[mkt_ca_df["Event Type"] == "Bonus Issue"])
        split_count   = len(mkt_ca_df[mkt_ca_df["Event Type"] == "Stock Split"])
        buyback_count = len(mkt_ca_df[mkt_ca_df["Event Type"] == "Buyback"])

        st.markdown(f"""
<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px;">
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">Total</div>
    <div style="font-size:22px;font-weight:800;color:#fff;">{total_mkt}</div>
  </div>
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">My Stocks</div>
    <div style="font-size:22px;font-weight:800;color:#ffd700;">{port_count}</div>
  </div>
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">Upcoming</div>
    <div style="font-size:22px;font-weight:800;color:#90caf9;">{upcoming_mkt}</div>
  </div>
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">Dividends</div>
    <div style="font-size:22px;font-weight:800;color:#00e676;">{div_count}</div>
  </div>
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">Bonus</div>
    <div style="font-size:22px;font-weight:800;color:#69f0ae;">{bonus_count}</div>
  </div>
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">Splits</div>
    <div style="font-size:22px;font-weight:800;color:#ffcc02;">{split_count}</div>
  </div>
  <div style="flex:1 1 90px;min-width:80px;background:#0f0f23;border:1px solid #2a2a5a;border-radius:10px;padding:10px 8px;text-align:center;">
    <div style="font-size:10px;color:#8888aa;text-transform:uppercase;letter-spacing:1px;">Buybacks</div>
    <div style="font-size:22px;font-weight:800;color:#ff9800;">{buyback_count}</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Urgent ex-date alerts (portfolio stocks only) ─────────────
        urgent_port = mkt_ca_df[
            (mkt_ca_df["In Portfolio"] == "⭐ Yes") &
            (mkt_ca_df["Ex-Date"] != "—")
        ].copy()
        urgent_list = []
        for _, row in urgent_port.iterrows():
            try:
                ex_ts = pd.to_datetime(row["Ex-Date"], dayfirst=True)
                diff  = (ex_ts - today_ts3).days
                if 0 <= diff <= 7:
                    urgent_list.append((row, diff))
            except Exception:
                pass

        if urgent_list:
            ah = (
                '<div style="background:#1a0808;border:1px solid #ff5252;border-radius:8px;'
                'padding:12px 16px;margin-bottom:14px;">'
                f'<div style="color:#ff5252;font-weight:700;font-size:13px;margin-bottom:8px;">'
                f'⚡ {len(urgent_list)} Ex-Date(s) in next 7 days for YOUR portfolio stocks!'
                f'</div>'
            )
            for row, diff in sorted(urgent_list, key=lambda x: x[1]):
                label = "TODAY" if diff == 0 else f"In {diff}d"
                det = str(row.get("Details",""))
                ah += (
                    f'<div style="color:#ffcccc;font-size:12px;padding:3px 0;border-bottom:1px solid #2a1010;">'
                    f'⭐ <b>{row["Symbol"]}</b> — {row["Company"][:30]} &nbsp;·&nbsp; {row["Event Type"]}'
                    f' &nbsp;·&nbsp; Ex-Date: <span style="color:#ff9999;font-weight:800;">'
                    f'{row["Ex-Date"]} ({label})</span>'
                    f'{(" &nbsp;·&nbsp; " + det[:50]) if det and det != "—" else ""}'
                    f'</div>'
                )
            ah += '</div>'
            st.markdown(ah, unsafe_allow_html=True)

        # ── Second filter row: search box ─────────────────────────────
        search_sym = st.text_input(
            "🔍 Search by Symbol or Company Name",
            placeholder="e.g. RELIANCE, HDFC, ITC...",
            key="ca_mkt_search"
        )
        if search_sym.strip():
            q = search_sym.strip().upper()
            mkt_ca_df = mkt_ca_df[
                mkt_ca_df["Symbol"].str.upper().str.contains(q, na=False) |
                mkt_ca_df["Company"].str.upper().str.contains(q, na=False)
            ]

        st.caption(
            f"Showing **{len(mkt_ca_df)}** corporate actions · "
            f"Window: **{ca_window}** · "
            f"⭐ = your portfolio stock"
        )

        # ── Color-coding ──────────────────────────────────────────────
        EMOJI_COLORS_MKT = {
            "Dividend":              "#90caf9",
            "Interim Dividend":      "#90caf9",
            "Final Dividend":        "#64b5f6",
            "Special Dividend":      "#42a5f5",
            "Bonus Issue":           "#22d67b",
            "Stock Split":           "#69f0ae",
            "Rights Issue":          "#ffcc02",
            "Buyback":               "#ff9800",
            "Merger / Demerger":     "#ce93d8",
            "Acquisition":           "#f48fb1",
            "Delisting":             "#f85454",
            "AGM":                   "#8888aa",
            "EGM":                   "#bbbbdd",
            "Financial Results":     "#ffd54f",
            "Preferential Allotment":"#ffab40",
            "NCD Issue":             "#80cbc4",
            "QIP":                   "#a5d6a7",
            "ESOP":                  "#ce93d8",
        }

        def _style_mkt_row(row):
            styles = [""] * len(row)
            idx = list(row.index)
            ev     = str(row.get("Event Type",""))
            stat   = str(row.get("Status",""))
            ex_v   = row.get("Ex-Date","—")
            in_pf  = str(row.get("In Portfolio",""))
            ev_col = EMOJI_COLORS_MKT.get(ev, "#e0e0e0")
            is_portfolio = in_pf == "⭐ Yes"

            for i, col in enumerate(idx):
                base = "background:#0d1a12;" if is_portfolio else ""
                if col == "Symbol":
                    styles[i] = (base + "color:#00e676;font-weight:800") if is_portfolio else "color:#b8bcd8;font-weight:600"
                elif col == "Company":
                    styles[i] = (base + "color:#aaffaa") if is_portfolio else "color:#b8bcd8"
                elif col == "In Portfolio":
                    styles[i] = "color:#00e676;font-weight:800;font-size:14px"
                elif col == "Event Type":
                    styles[i] = base + f"color:{ev_col};font-weight:700"
                elif col == "Ex-Date":
                    try:
                        if ex_v and ex_v != "—":
                            ex_ts = pd.to_datetime(ex_v, dayfirst=True)
                            d = (ex_ts - today_ts3).days
                            if d == 0:
                                styles[i] = "color:#ff5252;font-weight:800;background:#2a0a0a"
                            elif d <= 3:
                                styles[i] = "color:#ff7043;font-weight:700"
                            elif d <= 7:
                                styles[i] = "color:#ffaa00;font-weight:700"
                            elif d > 0:
                                styles[i] = "color:#90caf9;font-weight:600"
                            else:
                                styles[i] = "color:#454870"
                        else:
                            styles[i] = "color:#3a3d5c"
                    except Exception:
                        styles[i] = "color:#3a3d5c"
                elif col == "Status":
                    if "Today" in stat:   styles[i] = "color:#ff5252;font-weight:800"
                    elif "⚡" in stat:    styles[i] = "color:#ffaa00;font-weight:700"
                    elif "🔜" in stat:    styles[i] = "color:#f0f2ff;font-weight:600"
                    elif "🗓" in stat:    styles[i] = "color:#8888aa"
                    elif "ago" in stat:   styles[i] = "color:#454870"
                elif col == "Exchange":
                    styles[i] = "color:#8888aa;font-size:11px"
                elif col in ("Record Date", "Announcement Date"):
                    styles[i] = "color:#8888aa"
                elif col == "Details":
                    styles[i] = "color:#b8bcd8;font-size:11px"
            return styles

        # ── Mobile card layout (phone only) ──────────────────────────
        st.markdown("""
<style>
/* Hide mobile cards on desktop, show table */
.allca-mobile-cards { display: none; }
.allca-desktop-table { display: block; }
@media (max-width: 640px) {
    .allca-mobile-cards  { display: block !important; }
    .allca-desktop-table { display: none !important; }
}

/* ── Card wrapper ── */
.allca-card {
    background: linear-gradient(145deg, #13162a 0%, #0d1020 100%);
    border: 1px solid #1c2040;
    border-radius: 18px;
    margin-bottom: 10px;
    overflow: hidden;
    box-shadow: 0 4px 18px rgba(0,0,0,0.4), 0 1px 0 rgba(255,255,255,0.03) inset;
    position: relative;
}
/* Coloured top accent bar */
.allca-card::before {
    content: "";
    display: block;
    height: 2px;
    position: absolute;
    top: 0; left: 0; right: 0;
    opacity: 0.85;
}
/* Header row: ticker + event badge */
.allca-card-hdr {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 12px 14px 10px 14px;
    border-bottom: 1px solid rgba(255,255,255,0.05);
    background: rgba(255,255,255,0.02);
}
.allca-ticker {
    font-size: 16px;
    font-weight: 900;
    color: #eef0ff;
    letter-spacing: 0.4px;
    line-height: 1;
}
.allca-company {
    font-size: 10.5px;
    color: #7a7fa8;
    margin-top: 3px;
    font-weight: 500;
    max-width: 180px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.allca-badge {
    font-size: 9.5px;
    font-weight: 800;
    padding: 4px 10px;
    border-radius: 50px;
    text-transform: uppercase;
    letter-spacing: 0.4px;
    white-space: nowrap;
}
/* Body: 2-col data grid */
.allca-card-body {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px 8px;
    padding: 12px 14px 12px 14px;
}
.allca-field {
    display: flex;
    flex-direction: column;
    gap: 2px;
}
.allca-flabel {
    font-size: 8px;
    text-transform: uppercase;
    letter-spacing: 0.9px;
    color: #3a3d5c;
    font-weight: 700;
}
.allca-fvalue {
    font-size: 12px;
    font-weight: 700;
    color: #c0c4e8;
    line-height: 1.3;
}
/* Footer: details */
.allca-card-footer {
    padding: 8px 14px;
    background: rgba(0,0,0,0.25);
    font-size: 10px;
    color: #555577;
    border-top: 1px solid rgba(255,255,255,0.04);
    line-height: 1.5;
}
/* Portfolio highlight */
.allca-card.portfolio-stock {
    border-color: rgba(0,230,118,0.25);
    background: linear-gradient(145deg, #0d1a12 0%, #0a1410 100%);
}
.allca-card.portfolio-stock .allca-ticker { color: #00e676; }
/* Status chips */
.allca-status-today { color: #ff5252; font-weight: 800; }
.allca-status-soon  { color: #ffaa00; font-weight: 700; }
.allca-status-upcoming { color: #90caf9; font-weight: 600; }
.allca-status-past  { color: #454870; }
</style>
""", unsafe_allow_html=True)

        # ── Build mobile cards ────────────────────────────────────────
        _mobile_cards_html = '<div class="allca-mobile-cards">'
        for _, _row in mkt_ca_df.iterrows():
            _sym      = str(_row.get("Symbol",   "—"))
            _company  = str(_row.get("Company",  "—"))
            _ev       = str(_row.get("Event Type","—"))
            _exdate   = str(_row.get("Ex-Date",  "—"))
            _recdate  = str(_row.get("Record Date", "—"))
            _anndate  = str(_row.get("Announcement Date", "—"))
            _exch     = str(_row.get("Exchange", "—"))
            _status   = str(_row.get("Status",   "—"))
            _details  = str(_row.get("Details",  ""))
            _in_pf    = str(_row.get("In Portfolio", ""))
            _is_pf    = (_in_pf == "⭐ Yes")

            # Event type colour (reuse same palette)
            _ev_col   = EMOJI_COLORS_MKT.get(_ev, "#e0e0e0")
            _pf_cls   = "portfolio-stock" if _is_pf else ""
            _ticker_prefix = "⭐ " if _is_pf else ""

            # Status colour class
            if "Today" in _status:
                _stat_cls = "allca-status-today"
            elif "⚡" in _status:
                _stat_cls = "allca-status-soon"
            elif "🔜" in _status or "In " in _status:
                _stat_cls = "allca-status-upcoming"
            elif "ago" in _status:
                _stat_cls = "allca-status-past"
            else:
                _stat_cls = "allca-status-upcoming"

            _details_footer = ""
            if _details and _details not in ("—", "", "None"):
                _details_footer = f'<div class="allca-card-footer">📋 {_details[:80]}{"…" if len(_details)>80 else ""}</div>'

            _mobile_cards_html += f"""
<div class="allca-card {_pf_cls}" style="--accent:{_ev_col};">
  <style>
    .allca-card[style*="--accent:{_ev_col}"]::before {{ background:{_ev_col}; }}
  </style>
  <div class="allca-card-hdr">
    <div>
      <div class="allca-ticker">{_ticker_prefix}{_sym}</div>
      <div class="allca-company">{_company}</div>
    </div>
    <div class="allca-badge" style="background:{_ev_col}22;color:{_ev_col};border:1px solid {_ev_col}44;">{_ev}</div>
  </div>
  <div class="allca-card-body">
    <div class="allca-field">
      <div class="allca-flabel">Ex-Date</div>
      <div class="allca-fvalue" style="color:{('#ff5252' if _exdate != '—' and 'Today' in _status else ('#ffaa00' if '⚡' in _status else '#90caf9')) if _exdate != '—' else '#3a3d5c'};">{_exdate}</div>
    </div>
    <div class="allca-field">
      <div class="allca-flabel">Status</div>
      <div class="allca-fvalue {_stat_cls}">{_status}</div>
    </div>
    <div class="allca-field">
      <div class="allca-flabel">Record Date</div>
      <div class="allca-fvalue">{_recdate}</div>
    </div>
    <div class="allca-field">
      <div class="allca-flabel">Exchange</div>
      <div class="allca-fvalue" style="color:#8888aa;font-size:11px;">{_exch}</div>
    </div>
    <div class="allca-field">
      <div class="allca-flabel">Announced</div>
      <div class="allca-fvalue" style="font-size:11px;color:#7a7fa8;">{_anndate}</div>
    </div>
    <div class="allca-field">
      <div class="allca-flabel">Portfolio</div>
      <div class="allca-fvalue" style="color:{'#00e676' if _is_pf else '#3a3d5c'};">{'⭐ Yes' if _is_pf else 'No'}</div>
    </div>
  </div>
  {_details_footer}
</div>"""
        _mobile_cards_html += '</div>'
        st.markdown(_mobile_cards_html, unsafe_allow_html=True)

        # ── Desktop table (unchanged) ─────────────────────────────────
        st.markdown('<div class="allca-desktop-table">', unsafe_allow_html=True)
        styled_mkt = mkt_ca_df.style.apply(_style_mkt_row, axis=1)
        st.dataframe(styled_mkt, use_container_width=True, hide_index=True, height=520)
        st.markdown('</div>', unsafe_allow_html=True)

        # ── Legend ────────────────────────────────────────────────────
        st.markdown("""
<div style="font-size:11px;color:#666688;margin-top:6px;padding:8px 12px;
            background:#0b0d17;border-radius:6px;border:1px solid #252849;line-height:2;">
  <b style="color:#8888aa;">Color Guide:</b> &nbsp;
  <span style="color:#00e676;font-weight:700;">■ Green row = your portfolio stock</span> &nbsp;|&nbsp;
  <span style="color:#90caf9;">■ Dividend</span> &nbsp;
  <span style="color:#00e676;">■ Bonus</span> &nbsp;
  <span style="color:#69f0ae;">■ Split</span> &nbsp;
  <span style="color:#ffcc02;">■ Rights</span> &nbsp;
  <span style="color:#ff9800;">■ Buyback</span> &nbsp;
  <span style="color:#ce93d8;">■ Merger</span> &nbsp;
  <span style="color:#ff5252;">■ Delisting / Ex-Date Today</span> &nbsp;
  <span style="color:#ffaa00;">■ Ex-Date ≤7d</span>
  <br>
  <b style="color:#8888aa;">⭐ Portfolio highlight:</b> Rows with green background are stocks you hold.
  Must <b>hold shares before Ex-Date</b> to receive dividend/bonus/split benefits.
</div>
""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Charts ────────────────────────────────────────────────────
        if len(mkt_ca_df) > 0:
            ch1, ch2 = st.columns(2)

            with ch1:
                st.markdown("**📊 Event Type Breakdown**")
                ev_counts = mkt_ca_df["Event Type"].value_counts().reset_index()
                ev_counts.columns = ["Event Type", "Count"]
                fig_ev = go.Figure(go.Bar(
                    x=ev_counts["Event Type"],
                    y=ev_counts["Count"],
                    marker_color=[EMOJI_COLORS_MKT.get(e, "#7a7fa8") for e in ev_counts["Event Type"]],
                    text=ev_counts["Count"],
                    textposition="outside",
                    hovertemplate="<b>%{x}</b><br>Count: %{y}<extra></extra>"
                ))
                fig_ev.update_layout(
                    xaxis_title="", yaxis_title="Count",
                    plot_bgcolor="#0b0d17", paper_bgcolor="#0b0d17",
                    font=dict(color="#b8bcd8", family="Inter"),
                    xaxis=dict(tickangle=-35, gridcolor="#1e2238"),
                    yaxis=dict(gridcolor="#1e2238"),
                    height=320, showlegend=False, margin=dict(t=10, b=10)
                )
                st.plotly_chart(fig_ev, use_container_width=True)

            with ch2:
                st.markdown("**📅 Timeline — Actions by Date**")
                try:
                    timeline_df = mkt_ca_df.copy()
                    timeline_df["_dt_parsed"] = pd.to_datetime(
                        timeline_df["Ex-Date"].replace("—", None), dayfirst=True, errors="coerce"
                    )
                    timeline_df = timeline_df.dropna(subset=["_dt_parsed"])
                    if not timeline_df.empty:
                        timeline_df["Date"] = timeline_df["_dt_parsed"].dt.strftime("%d-%b")
                        date_counts = timeline_df.groupby("Date").size().reset_index(name="Count")
                        # Sort by actual date
                        date_counts["_sort"] = pd.to_datetime(
                            timeline_df.groupby("Date")["_dt_parsed"].first().values
                        )
                        date_counts = date_counts.sort_values("_sort")
                        fig_tl = go.Figure(go.Bar(
                            x=date_counts["Date"],
                            y=date_counts["Count"],
                            marker_color="#7c4dff",
                            text=date_counts["Count"],
                            textposition="outside",
                            hovertemplate="<b>%{x}</b><br>Actions: %{y}<extra></extra>"
                        ))
                        fig_tl.update_layout(
                            xaxis_title="Ex-Date", yaxis_title="# Actions",
                            plot_bgcolor="#0b0d17", paper_bgcolor="#0b0d17",
                            font=dict(color="#b8bcd8", family="Inter"),
                            xaxis=dict(tickangle=-35, gridcolor="#1e2238"),
                            yaxis=dict(gridcolor="#1e2238"),
                            height=320, showlegend=False, margin=dict(t=10, b=10)
                        )
                        st.plotly_chart(fig_tl, use_container_width=True)
                    else:
                        st.info("No ex-dates available for timeline.")
                except Exception:
                    st.info("Timeline chart unavailable.")

        # ── Download ──────────────────────────────────────────────────
        st.markdown("---")
        dl1, dl2 = st.columns(2)
        with dl1:
            csv_mkt = mkt_ca_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download Filtered Data (CSV)",
                data=csv_mkt,
                file_name=f"MarketCA_{ca_window.replace(' ','_')}_{datetime.now().strftime('%d%b%Y')}.csv",
                mime="text/csv",
                key="dl_mkt_ca_csv"
            )
        with dl2:
            port_only_df = mkt_ca_df[mkt_ca_df["In Portfolio"] == "⭐ Yes"]
            if not port_only_df.empty:
                csv_port = port_only_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "⭐ Download My Portfolio Actions Only (CSV)",
                    data=csv_port,
                    file_name=f"PortfolioCA_{datetime.now().strftime('%d%b%Y')}.csv",
                    mime="text/csv",
                    key="dl_port_ca_csv"
                )

        st.markdown("""
<div style="background:#0b0d17;border:1px solid #252849;border-radius:8px;
            padding:10px 16px;margin-top:8px;">
  <div style="font-size:11px;color:#666688;">
    ⚠️ <b>Disclaimer:</b> Data fetched live from NSE and BSE public APIs.
    Always verify ex-dates and record dates with your broker before making investment decisions.
    Northeast Broking Services Limited is not responsible for any errors in exchange-provided data.
  </div>
</div>
""", unsafe_allow_html=True)

# =========================================================
# MASTER IMPORT TAB
# =========================================================
if _nav_tab == "Master Import" and _is_dev:
    show_master_import_tab(
        clients_dict_ref=clients_dict,
        save_clients_fn=save_clients,
        hash_pw_fn=hash_password,
        client_portfolio_file_fn=client_portfolio_file,
        client_trades_file_fn=client_trades_file,
        dev_code_ref=DEV_CODE,
    )